from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path
import tempfile
from types import SimpleNamespace
from unittest.mock import patch

from django.contrib.auth.models import Group
from django.contrib.auth import get_user_model
from django.contrib.messages import get_messages
from django.core import mail
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
import zipfile

from .models import (
    AditivoContratoTerceiro,
    CentroDeTrabalho,
    Cliente,
    BM,
    Contrato,
    ContratoTerceiros,
    DocumentoBM,
    DocumentoContratoTerceiro,
    EmpresaTerceira,
    Evento,
    OS,
    PropostaFornecedor,
    RegistroAuditoria,
    SolicitacaoContrato,
    SolicitacaoOrdemServico,
    SolicitacaoProspeccao,
)
from .forms import (
    BMForm,
    ClienteForm,
    ContratoForm,
    EventoEntregaForm,
    FornecedorForm,
    NFClienteForm,
    NFForm,
    OrdemServicoForm,
    RegistroEntregaOSForm,
)
from .views import (
    build_weekly_supply_report,
    can_user_request_contract_addendum,
    can_user_manage_event_delivery,
    can_user_manage_os_delivery,
    can_user_manage_supplier_choice,
    criar_contrato_se_aprovado,
    criar_contrato_se_aprovado_minuta,
    format_request_line,
    get_week_ranges,
    is_request_concluded,
    send_request_notification_to_management,
    user_shares_center_with_coordinator,
)


User = get_user_model()


class BaseUserTestCase(TestCase):
    def create_user(self, username, grupo, password="senha123", email=None):
        return User.objects.create_user(
            username=username,
            password=password,
            grupo=grupo,
            email=f"{username}@example.com" if email is None else email,
            first_name=username.capitalize(),
        )

    def create_center(self, codigo="CT1", nome="Centro 1"):
        return CentroDeTrabalho.objects.create(codigo=codigo, nome=nome)

    def create_client(self, nome="Cliente Teste", cpf_cnpj=None):
        if cpf_cnpj is None:
            cpf_cnpj = f"00.000.000/0001-{Cliente.objects.count() + 1:02d}"
        return Cliente.objects.create(nome=nome, cpf_cnpj=cpf_cnpj)

    def create_contract(
        self,
        codigo="PRJ-001",
        cliente=None,
        coordenador=None,
        lider_contrato=None,
        valor_total=Decimal("1000.00"),
    ):
        return Contrato.objects.create(
            cod_projeto=codigo,
            cliente=cliente or self.create_client(),
            coordenador=coordenador,
            lider_contrato=lider_contrato,
            valor_total=valor_total,
            objeto="Objeto do contrato",
            status="ativo",
        )

    def create_supplier(self, nome="Fornecedor Teste", cpf_cnpj=None):
        if cpf_cnpj is None:
            cpf_cnpj = f"11.111.111/0001-{EmpresaTerceira.objects.count() + 1:02d}"
        email_base = nome.lower().replace(" ", "")
        return EmpresaTerceira.objects.create(
            nome=nome,
            cpf_cnpj=cpf_cnpj,
            informacoes_bancarias="Banco XPTO",
            email=f"{email_base}{EmpresaTerceira.objects.count() + 1}@example.com",
        )

    def create_supplier_contract(
        self,
        cod_projeto=None,
        empresa_terceira=None,
        coordenador=None,
        lider_contrato=None,
        guarda_chuva=False,
        status="ativo",
        num_contrato="CT-001",
    ):
        return ContratoTerceiros.objects.create(
            cod_projeto=cod_projeto or self.create_contract(),
            empresa_terceira=empresa_terceira or self.create_supplier(),
            coordenador=coordenador,
            lider_contrato=lider_contrato,
            guarda_chuva=guarda_chuva,
            num_contrato=num_contrato,
            objeto="Objeto contrato fornecedor",
            status=status,
        )

    def create_os_request(
        self,
        contrato=None,
        solicitante=None,
        lider_contrato=None,
        coordenador=None,
        titulo="OS Solicitada",
        status="pendente_lider",
    ):
        contrato = contrato or self.create_supplier_contract(guarda_chuva=True)
        return SolicitacaoOrdemServico.objects.create(
            contrato=contrato,
            cod_projeto=contrato.cod_projeto,
            solicitante=solicitante or lider_contrato or self.create_user("solicitante_os", "lider_contrato"),
            lider_contrato=lider_contrato,
            coordenador=coordenador or contrato.coordenador,
            titulo=titulo,
            descricao="Descrição da solicitação de OS",
            valor_previsto=Decimal("250.00"),
            prazo_execucao=date(2026, 4, 30),
            status=status,
        )

    def create_os(
        self,
        contrato=None,
        solicitacao=None,
        coordenador=None,
        lider_contrato=None,
        titulo="OS Cadastrada",
        status="em_execucao",
        prazo_execucao=None,
    ):
        contrato = contrato or self.create_supplier_contract(guarda_chuva=True)
        return OS.objects.create(
            contrato=contrato,
            solicitacao=solicitacao,
            cod_projeto=contrato.cod_projeto,
            coordenador=coordenador or contrato.coordenador,
            lider_contrato=lider_contrato or contrato.lider_contrato,
            titulo=titulo,
            descricao="Descrição da OS",
            valor=Decimal("350.00"),
            prazo_execucao=prazo_execucao or date(2026, 4, 20),
            status=status,
        )

    def create_event(
        self,
        contrato=None,
        prospeccao=None,
        empresa_terceira=None,
        data_prevista=None,
        realizado=False,
    ):
        contrato = contrato or self.create_supplier_contract()
        empresa_terceira = empresa_terceira or contrato.empresa_terceira
        return Evento.objects.create(
            contrato_terceiro=contrato,
            prospeccao=prospeccao,
            empresa_terceira=empresa_terceira,
            descricao="Entrega prevista",
            data_prevista=data_prevista or date(2026, 4, 18),
            valor_previsto=Decimal("500.00"),
            realizado=realizado,
        )


class UppercaseNormalizationTests(BaseUserTestCase):
    def test_cliente_e_salvo_em_maiusculas(self):
        cliente = Cliente.objects.create(
            nome="Clínica São José",
            razao_social="Clínica são josé ltda",
            cpf_cnpj="00.000.000/0001-99",
            endereco="Rua das flores, 123",
            ponto_focal="João da Silva",
            observacao="cliente estratégico",
        )

        self.assertEqual(cliente.nome, "CLÍNICA SÃO JOSÉ")
        self.assertEqual(cliente.razao_social, "CLÍNICA SÃO JOSÉ LTDA")
        self.assertEqual(cliente.endereco, "RUA DAS FLORES, 123")
        self.assertEqual(cliente.ponto_focal, "JOÃO DA SILVA")
        self.assertEqual(cliente.observacao, "CLIENTE ESTRATÉGICO")

    def test_fornecedor_e_salvo_em_maiusculas(self):
        fornecedor = EmpresaTerceira.objects.create(
            nome="Ação engenharia",
            setor_de_atuacao="manutenção elétrica",
            cpf_cnpj="11.111.111/0001-99",
            endereco="av. brasil",
            numero="bloco b",
            bairro="centro",
            municipio="são luís",
            estado="ma",
            cep="65000-000",
            informacoes_bancarias="banco do brasil ag 1234",
            ponto_focal="Márcia Souza",
            ponto_focal2="José Neto",
            observacao="fornecedor homologado",
        )

        self.assertEqual(fornecedor.nome, "AÇÃO ENGENHARIA")
        self.assertEqual(fornecedor.setor_de_atuacao, "MANUTENÇÃO ELÉTRICA")
        self.assertEqual(fornecedor.endereco, "AV. BRASIL")
        self.assertEqual(fornecedor.numero, "BLOCO B")
        self.assertEqual(fornecedor.bairro, "CENTRO")
        self.assertEqual(fornecedor.municipio, "SÃO LUÍS")
        self.assertEqual(fornecedor.estado, "MA")
        self.assertEqual(fornecedor.informacoes_bancarias, "BANCO DO BRASIL AG 1234")
        self.assertEqual(fornecedor.ponto_focal, "MÁRCIA SOUZA")
        self.assertEqual(fornecedor.ponto_focal2, "JOSÉ NETO")
        self.assertEqual(fornecedor.observacao, "FORNECEDOR HOMOLOGADO")

    def test_codigo_do_projeto_e_salvo_em_maiusculas(self):
        contrato = Contrato.objects.create(
            cod_projeto="prj-aç-001",
            cliente=self.create_client(cpf_cnpj="00.000.000/0001-88"),
            objeto="Objeto",
            status="ativo",
        )

        self.assertEqual(contrato.cod_projeto, "PRJ-AÇ-001")

    def test_forms_tambem_persistem_dados_em_maiusculas(self):
        cliente_form = ClienteForm(
            data={
                "nome": "Cliente teste",
                "razao_social": "Razão social teste",
                "cpf_cnpj": "00.000.000/0001-77",
                "endereco": "rua alfa",
                "telefone": "",
                "email": "cliente@example.com",
                "ponto_focal": "Ana cláudia",
                "email_focal": "",
                "telefone_focal": "",
                "observacao": "observação geral",
            }
        )
        self.assertTrue(cliente_form.is_valid(), cliente_form.errors)
        cliente = cliente_form.save()

        fornecedor_form = FornecedorForm(
            data={
                "nome": "Fornecedor teste",
                "setor_de_atuacao": "obras civis",
                "cpf_cnpj": "11.111.111/0001-77",
                "endereco": "avenida beta",
                "numero": "sala 2",
                "bairro": "centro",
                "municipio": "belém",
                "estado": "pa",
                "cep": "66000-000",
                "telefone": "",
                "email": "fornecedor@example.com",
                "informacoes_bancarias": "caixa ag 55",
                "guarda_chuva": False,
                "ponto_focal": "Carlos eduardo",
                "email_focal": "",
                "telefone_focal": "",
                "ponto_focal2": "Bia lima",
                "email_focal2": "",
                "telefone_focal2": "",
                "observacao": "fornecedor novo",
            }
        )
        self.assertTrue(fornecedor_form.is_valid(), fornecedor_form.errors)
        fornecedor = fornecedor_form.save()

        contrato_form = ContratoForm(
            data={
                "cod_projeto": "abc-ç-01",
                "cliente": cliente.pk,
                "coordenador": "",
                "data_inicio": "",
                "data_fim": "",
                "valor_total": "",
                "status": "ativo",
                "objeto": "Objeto teste",
                "proposta": "",
                "lider_contrato": "",
                "observacao": "",
            }
        )
        self.assertTrue(contrato_form.is_valid(), contrato_form.errors)
        contrato = contrato_form.save()

        self.assertEqual(cliente.nome, "CLIENTE TESTE")
        self.assertEqual(fornecedor.nome, "FORNECEDOR TESTE")
        self.assertEqual(contrato.cod_projeto, "ABC-Ç-01")


class MoneyFieldNormalizationTests(BaseUserTestCase):
    def setUp(self):
        self.coordenador = self.create_user("coordenador_moeda", "coordenador")
        self.lider = self.create_user("lider_moeda", "lider_contrato")
        self.contrato_cliente = self.create_contract(
            codigo="PRJ-MONEY",
            cliente=self.create_client(cpf_cnpj="00.000.000/0001-65"),
            coordenador=self.coordenador,
            lider_contrato=self.lider,
        )
        self.contrato_fornecedor = self.create_supplier_contract(
            cod_projeto=self.contrato_cliente,
            coordenador=self.coordenador,
            lider_contrato=self.lider,
            guarda_chuva=True,
            num_contrato="CT-MONEY",
        )
        self.solicitacao_os = self.create_os_request(
            contrato=self.contrato_fornecedor,
            solicitante=self.lider,
            lider_contrato=self.lider,
            coordenador=self.coordenador,
            titulo="OS Money",
            status="aprovada",
        )
        self.evento = self.create_event(contrato=self.contrato_fornecedor)
        self.os = self.create_os(
            contrato=self.contrato_fornecedor,
            solicitacao=self.solicitacao_os,
            coordenador=self.coordenador,
            lider_contrato=self.lider,
            titulo="OS Money",
        )
        self.bm = BM.objects.create(
            contrato=self.contrato_fornecedor,
            evento=self.evento,
            numero_bm=1,
            parcela_paga=1,
            valor_pago=Decimal("500.00"),
            data_pagamento=date(2026, 4, 20),
        )

    def test_forms_de_valor_aceitam_formato_brasileiro(self):
        evento_form = EventoEntregaForm(
            instance=self.evento,
            data={
                "observacao": "Entrega validada",
                "caminho_evidencia": "Pasta/Arquivo",
                "justificativa": "",
                "avaliacao": "Aprovado",
                "data_entrega": "2026-04-20",
                "realizado": "on",
                "com_atraso": "",
                "valor_pago": "1.234,56",
                "data_pagamento": "2026-04-22",
            },
        )
        self.assertTrue(evento_form.is_valid(), evento_form.errors)
        self.assertEqual(evento_form.cleaned_data["valor_pago"], Decimal("1234.56"))

        registro_os_form = RegistroEntregaOSForm(
            instance=self.os,
            data={
                "caminho_evidencia": "Evidencia/OS",
                "avaliacao": "Aprovado",
                "data_entrega": "2026-04-20",
                "realizado": "on",
                "com_atraso": "",
                "valor_pago": "2.345,67",
                "data_pagamento": "2026-04-23",
                "observacao": "Entrega OS",
            },
        )
        self.assertTrue(registro_os_form.is_valid(), registro_os_form.errors)
        self.assertEqual(registro_os_form.cleaned_data["valor_pago"], Decimal("2345.67"))

        bm_form = BMForm(
            instance=self.bm,
            data={
                "numero_bm": 1,
                "parcela_paga": 1,
                "valor_pago": "3.456,78",
                "data_pagamento": "2026-04-24",
                "data_inicial_medicao": "",
                "data_final_medicao": "",
                "observacao": "BM atualizado",
            },
        )
        self.assertTrue(bm_form.is_valid(), bm_form.errors)
        self.assertEqual(bm_form.cleaned_data["valor_pago"], Decimal("3456.78"))

        nf_form = NFForm(
            evento=self.evento,
            data={
                "bm": self.bm.id,
                "valor_pago": "4.567,89",
                "parcela_paga": 1,
                "data_pagamento": "2026-04-25",
                "observacao": "NF fornecedor",
                "financeiro_autorizou": "on",
                "nf_dentro_prazo": "on",
            },
        )
        self.assertTrue(nf_form.is_valid(), nf_form.errors)
        self.assertEqual(nf_form.cleaned_data["valor_pago"], Decimal("4567.89"))

        nf_cliente_form = NFClienteForm(
            data={
                "valor_pago": "5.678,90",
                "parcela_paga": 2,
                "data_emissao": "2026-04-24",
                "data_pagamento": "2026-04-26",
                "observacao": "NF cliente",
            },
        )
        self.assertTrue(nf_cliente_form.is_valid(), nf_cliente_form.errors)
        self.assertEqual(nf_cliente_form.cleaned_data["valor_pago"], Decimal("5678.90"))

        ordem_servico_form = OrdemServicoForm(
            data={
                "contrato": self.contrato_fornecedor.id,
                "solicitacao": self.solicitacao_os.id,
                "cod_projeto": self.contrato_cliente.id,
                "coordenador": self.coordenador.id,
                "lider_contrato": self.lider.id,
                "titulo": "OS com mascara",
                "descricao": "Descricao da OS",
                "valor": "6.789,01",
                "prazo_execucao": "2026-05-10",
                "status": "em_execucao",
            },
        )
        self.assertTrue(ordem_servico_form.is_valid(), ordem_servico_form.errors)
        self.assertEqual(ordem_servico_form.cleaned_data["valor"], Decimal("6789.01"))

    def test_evento_entrega_tambem_aceita_formato_antigo_com_ponto_decimal(self):
        form = EventoEntregaForm(
            instance=self.evento,
            data={
                "observacao": "Entrega validada",
                "caminho_evidencia": "Pasta/Arquivo",
                "justificativa": "",
                "avaliacao": "Aprovado",
                "data_entrega": "2026-04-20",
                "realizado": "on",
                "com_atraso": "",
                "valor_pago": "480.00",
                "data_pagamento": "2026-04-22",
            },
        )
        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data["valor_pago"], Decimal("480.00"))


class MultipleCoordenadoresContratoTests(BaseUserTestCase):
    def setUp(self):
        self.coordenador_principal = self.create_user("coordprincipal", "coordenador")
        self.coordenador_extra = self.create_user("coordextra", "coordenador")
        self.gerente_lider = self.create_user("gerentelider", "gerente_lider")
        self.centro_a = self.create_center("CTA", "Centro A")
        self.centro_b = self.create_center("CTB", "Centro B")
        self.coordenador_principal.centros.add(self.centro_a)
        self.coordenador_extra.centros.add(self.centro_b)
        self.gerente_lider.centros.add(self.centro_b)
        self.cliente = self.create_client(cpf_cnpj="00.000.000/0001-66")

    def test_contrato_cliente_sincroniza_coordenador_principal_na_lista_de_coordenadores(self):
        contrato = Contrato.objects.create(
            cod_projeto="PRJ-MULTI-01",
            cliente=self.cliente,
            coordenador=self.coordenador_principal,
            objeto="Contrato com múltiplos coordenadores",
            status="ativo",
        )

        self.assertIn(self.coordenador_principal, contrato.coordenadores.all())

    def test_form_de_contrato_persiste_coordenadores_adicionais(self):
        form = ContratoForm(
            data={
                "cod_projeto": "PRJ-MULTI-02",
                "cliente": self.cliente.pk,
                "coordenador": self.coordenador_principal.pk,
                "coordenadores": [self.coordenador_principal.pk, self.coordenador_extra.pk],
                "data_inicio": "",
                "data_fim": "",
                "valor_total": "",
                "status": "ativo",
                "objeto": "Objeto",
                "proposta": "",
                "lider_contrato": "",
                "observacao": "",
            }
        )

        self.assertTrue(form.is_valid(), form.errors)
        contrato = form.save()

        self.assertCountEqual(
            contrato.coordenadores.values_list("pk", flat=True),
            [self.coordenador_principal.pk, self.coordenador_extra.pk],
        )

    def test_coordenador_adicional_aparece_na_lista_de_contratos_cliente(self):
        contrato = Contrato.objects.create(
            cod_projeto="PRJ-MULTI-03",
            cliente=self.cliente,
            coordenador=self.coordenador_principal,
            objeto="Objeto",
            status="ativo",
        )
        contrato.coordenadores.add(self.coordenador_extra)

        self.client.force_login(self.coordenador_extra)
        response = self.client.get(reverse("lista_contratos"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, contrato.cod_projeto)

    def test_coordenador_adicional_aparece_na_lista_de_contratos_fornecedor(self):
        contrato_base = Contrato.objects.create(
            cod_projeto="PRJ-MULTI-04",
            cliente=self.cliente,
            coordenador=self.coordenador_principal,
            objeto="Objeto",
            status="ativo",
        )
        fornecedor = self.create_supplier(cpf_cnpj="11.111.111/0001-66")
        contrato_fornecedor = ContratoTerceiros.objects.create(
            cod_projeto=contrato_base,
            empresa_terceira=fornecedor,
            coordenador=self.coordenador_principal,
            objeto="Contrato fornecedor",
            status="ativo",
            num_contrato="CF-MULTI-01",
        )
        contrato_fornecedor.coordenadores.add(self.coordenador_extra)

        self.client.force_login(self.coordenador_extra)
        response = self.client.get(reverse("lista_contratos_fornecedores"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, contrato_fornecedor.num_contrato)

    def test_gerente_lider_com_centro_de_coordenador_adicional_ve_contrato_fornecedor(self):
        contrato_base = Contrato.objects.create(
            cod_projeto="PRJ-MULTI-05",
            cliente=self.cliente,
            coordenador=self.coordenador_principal,
            objeto="Objeto",
            status="ativo",
        )
        fornecedor = self.create_supplier(cpf_cnpj="11.111.111/0001-65")
        contrato_fornecedor = ContratoTerceiros.objects.create(
            cod_projeto=contrato_base,
            empresa_terceira=fornecedor,
            coordenador=self.coordenador_principal,
            objeto="Contrato fornecedor",
            status="ativo",
            num_contrato="CF-MULTI-02",
        )
        contrato_fornecedor.coordenadores.add(self.coordenador_extra)

        self.client.force_login(self.gerente_lider)
        response = self.client.get(reverse("lista_contratos_fornecedores"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, contrato_fornecedor.num_contrato)

class GuiaPermissoesTests(BaseUserTestCase):
    def test_guia_permissoes_exige_login(self):
        response = self.client.get(reverse("guia_permissoes"))

        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse("login"), response.url)

    def test_usuario_suprimento_ve_todos_os_grupos(self):
        user = self.create_user("supri", "suprimento")
        self.client.force_login(user)

        response = self.client.get(reverse("guia_permissoes"))

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["can_view_all_groups"])
        self.assertEqual(len(response.context["group_guides"]), len(User.GRUPOS_CHOICES))
        self.assertContains(response, "Coordenador de Contrato")
        self.assertContains(response, "Gerente de Contratos")
        self.assertContains(response, "Fornecedor")

    def test_usuario_nao_suprimento_ve_apenas_o_proprio_grupo(self):
        user = self.create_user("lider", "lider_contrato")
        self.client.force_login(user)

        response = self.client.get(reverse("guia_permissoes"))

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context["can_view_all_groups"])
        self.assertEqual(response.context["current_group_item"]["key"], "lider_contrato")
        self.assertContains(response, "Seu grupo")
        self.assertContains(response, "Líder de Contrato")
        self.assertNotContains(response, "Gerente de Contratos")


class RankingFornecedoresTests(BaseUserTestCase):
    def setUp(self):
        self.suprimento = self.create_user("supranking", "suprimento")
        self.gerente_contrato = self.create_user("gcranking", "gerente_contrato")
        self.outro_gerente_contrato = self.create_user("gcoutranking", "gerente_contrato")
        self.gerente_lider = self.create_user("glranking", "gerente_lider")
        self.coordenador = self.create_user("coordranking", "coordenador")
        self.outro_coordenador = self.create_user("coordoutranking", "coordenador")
        self.centro = self.create_center("CTRANK", "Centro Ranking")
        self.outro_centro = self.create_center("CTOUTRANK", "Centro Outro Ranking")
        self.coordenador.centros.add(self.centro)
        self.outro_coordenador.centros.add(self.outro_centro)
        self.gerente_lider.centros.add(self.centro)

        self.fornecedor_ativo = self.create_supplier("Fornecedor Ativo")
        self.contrato_base_ativo = self.create_contract(
            codigo="PRJ-RANK-ATV",
            coordenador=self.coordenador,
            lider_contrato=self.gerente_contrato,
        )
        self.contrato_ativo = self.create_supplier_contract(
            cod_projeto=self.contrato_base_ativo,
            empresa_terceira=self.fornecedor_ativo,
            coordenador=self.coordenador,
            lider_contrato=self.gerente_contrato,
            status="ativo",
            num_contrato="CTRANK1",
        )
        self.contrato_ativo.valor_total = Decimal("1000.00")
        self.contrato_ativo.save(update_fields=["valor_total"])
        self.evento_ativo = self.create_event(
            contrato=self.contrato_ativo,
            empresa_terceira=self.fornecedor_ativo,
            data_prevista=timezone.localdate(),
        )
        self.evento_ativo.valor_previsto = Decimal("1000.00")
        self.evento_ativo.valor_pago = Decimal("400.00")
        self.evento_ativo.save(update_fields=["valor_previsto", "valor_pago"])

        self.contrato_base_encerrado = self.create_contract(
            codigo="PRJ-RANK-END",
            coordenador=self.coordenador,
            lider_contrato=self.gerente_contrato,
        )
        self.contrato_encerrado = self.create_supplier_contract(
            cod_projeto=self.contrato_base_encerrado,
            empresa_terceira=self.fornecedor_ativo,
            coordenador=self.coordenador,
            lider_contrato=self.gerente_contrato,
            status="encerrado",
            num_contrato="CTRANK2",
        )
        self.contrato_encerrado.valor_total = Decimal("9000.00")
        self.contrato_encerrado.save(update_fields=["valor_total"])
        self.evento_encerrado_mesmo_fornecedor = self.create_event(
            contrato=self.contrato_encerrado,
            empresa_terceira=self.fornecedor_ativo,
            data_prevista=timezone.localdate(),
        )
        self.evento_encerrado_mesmo_fornecedor.valor_previsto = Decimal("9000.00")
        self.evento_encerrado_mesmo_fornecedor.valor_pago = Decimal("8000.00")
        self.evento_encerrado_mesmo_fornecedor.save(update_fields=["valor_previsto", "valor_pago"])

        self.fornecedor_so_encerrado = self.create_supplier("Fornecedor Encerrado")
        self.contrato_base_so_encerrado = self.create_contract(
            codigo="PRJ-RANK-OLD",
            coordenador=self.coordenador,
            lider_contrato=self.gerente_contrato,
        )
        self.contrato_so_encerrado = self.create_supplier_contract(
            cod_projeto=self.contrato_base_so_encerrado,
            empresa_terceira=self.fornecedor_so_encerrado,
            coordenador=self.coordenador,
            lider_contrato=self.gerente_contrato,
            status="encerrado",
            num_contrato="CTRANK3",
        )
        self.contrato_so_encerrado.valor_total = Decimal("5000.00")
        self.contrato_so_encerrado.save(update_fields=["valor_total"])
        self.evento_so_encerrado = self.create_event(
            contrato=self.contrato_so_encerrado,
            empresa_terceira=self.fornecedor_so_encerrado,
            data_prevista=timezone.localdate(),
        )
        self.evento_so_encerrado.valor_previsto = Decimal("5000.00")
        self.evento_so_encerrado.valor_pago = Decimal("3000.00")
        self.evento_so_encerrado.save(update_fields=["valor_previsto", "valor_pago"])

        self.fornecedor_outro_escopo = self.create_supplier("Fornecedor Outro Escopo")
        self.contrato_base_outro_escopo = self.create_contract(
            codigo="PRJ-RANK-OTH",
            coordenador=self.outro_coordenador,
            lider_contrato=self.outro_gerente_contrato,
        )
        self.contrato_outro_escopo = self.create_supplier_contract(
            cod_projeto=self.contrato_base_outro_escopo,
            empresa_terceira=self.fornecedor_outro_escopo,
            coordenador=self.outro_coordenador,
            lider_contrato=self.outro_gerente_contrato,
            status="ativo",
            num_contrato="CTRANK4",
        )
        self.contrato_outro_escopo.valor_total = Decimal("2000.00")
        self.contrato_outro_escopo.save(update_fields=["valor_total"])
        self.evento_outro_escopo = self.create_event(
            contrato=self.contrato_outro_escopo,
            empresa_terceira=self.fornecedor_outro_escopo,
            data_prevista=timezone.localdate(),
        )
        self.evento_outro_escopo.valor_previsto = Decimal("2000.00")
        self.evento_outro_escopo.valor_pago = Decimal("1000.00")
        self.evento_outro_escopo.save(update_fields=["valor_previsto", "valor_pago"])

    def test_ranking_fornecedores_exclui_fornecedor_com_apenas_contrato_encerrado(self):
        self.client.force_login(self.suprimento)

        response = self.client.get(reverse("ranking_fornecedores"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "FORNECEDOR ATIVO")
        self.assertNotContains(response, "FORNECEDOR ENCERRADO")

    def test_ranking_fornecedores_ignora_eventos_de_contratos_encerrados(self):
        self.client.force_login(self.suprimento)

        response = self.client.get(reverse("ranking_fornecedores"))

        self.assertEqual(response.status_code, 200)
        dados_por_fornecedor = {item["fornecedor"]: item for item in response.context["dados"]}
        self.assertIn("FORNECEDOR ATIVO", dados_por_fornecedor)
        self.assertEqual(dados_por_fornecedor["FORNECEDOR ATIVO"]["valor_total_contratos"], Decimal("1000.00"))
        self.assertEqual(dados_por_fornecedor["FORNECEDOR ATIVO"]["valor_previsto"], Decimal("1000.00"))
        self.assertEqual(dados_por_fornecedor["FORNECEDOR ATIVO"]["valor_pago"], Decimal("400.00"))

    def test_ranking_fornecedores_restringe_gerente_contrato_ao_proprio_vinculo(self):
        self.client.force_login(self.gerente_contrato)

        response = self.client.get(reverse("ranking_fornecedores"))

        self.assertEqual(response.status_code, 200)
        dados = response.context["dados"]
        self.assertEqual(len(dados), 1)
        self.assertEqual(dados[0]["fornecedor"], "FORNECEDOR ATIVO")
        self.assertNotContains(response, "FORNECEDOR OUTRO ESCOPO")

    def test_ranking_fornecedores_restringe_gerente_lider_ao_mesmo_centro(self):
        self.client.force_login(self.gerente_lider)

        response = self.client.get(reverse("ranking_fornecedores"))

        self.assertEqual(response.status_code, 200)
        dados = response.context["dados"]
        self.assertEqual(len(dados), 1)
        self.assertEqual(dados[0]["fornecedor"], "FORNECEDOR ATIVO")
        self.assertNotContains(response, "FORNECEDOR OUTRO ESCOPO")

    def test_ranking_fornecedores_informa_projetos_ativos(self):
        self.client.force_login(self.suprimento)

        response = self.client.get(reverse("ranking_fornecedores"))

        self.assertEqual(response.status_code, 200)
        dados_por_fornecedor = {item["fornecedor"]: item for item in response.context["dados"]}
        self.assertEqual(dados_por_fornecedor["FORNECEDOR ATIVO"]["projetos_ativos"], ["PRJ-RANK-ATV"])
        self.assertContains(response, "PRJ-RANK-ATV")

    def test_ranking_fornecedores_nao_duplica_mesmo_contrato_por_multiplos_coordenadores_do_mesmo_centro(self):
        coordenador_extra_mesmo_centro = self.create_user("coordrankextra", "coordenador")
        coordenador_extra_mesmo_centro.centros.add(self.centro)
        self.contrato_ativo.coordenadores.add(self.coordenador, coordenador_extra_mesmo_centro)

        self.client.force_login(self.gerente_lider)

        response = self.client.get(reverse("ranking_fornecedores"))

        self.assertEqual(response.status_code, 200)
        dados_por_fornecedor = {item["fornecedor"]: item for item in response.context["dados"]}
        self.assertEqual(dados_por_fornecedor["FORNECEDOR ATIVO"]["valor_total_contratos"], Decimal("1000.00"))
        self.assertEqual(dados_por_fornecedor["FORNECEDOR ATIVO"]["valor_previsto"], Decimal("1000.00"))
        self.assertEqual(dados_por_fornecedor["FORNECEDOR ATIVO"]["valor_pago"], Decimal("400.00"))


class ClienteAjaxTests(BaseUserTestCase):
    def test_add_cliente_permite_cadastro_para_lider_contrato(self):
        user = self.create_user("lidercliente", "lider_contrato")
        self.client.force_login(user)

        response = self.client.post(
            reverse("add_cliente"),
            {
                "nome": "Cliente Ajax",
                "cpf_cnpj": "12.345.678/0001-99",
                "email": "clienteajax@example.com",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["nome"], "CLIENTE AJAX")
        self.assertTrue(Cliente.objects.filter(nome="CLIENTE AJAX").exists())

    def test_add_cliente_bloqueia_usuario_sem_permissao(self):
        user = self.create_user("coordenadorcliente", "coordenador")
        self.client.force_login(user)

        response = self.client.post(
            reverse("add_cliente"),
            {
                "nome": "Cliente Bloqueado",
                "cpf_cnpj": "98.765.432/0001-11",
            },
        )

        self.assertEqual(response.status_code, 403)
        self.assertFalse(Cliente.objects.filter(nome="Cliente Bloqueado").exists())

    def test_tela_nova_solicitacao_prospeccao_exibe_modal_de_cliente(self):
        user = self.create_user("liderprosp", "lider_contrato")
        self.client.force_login(user)

        response = self.client.get(reverse("nova_solicitacao_prospeccao"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="modalCliente"')
        self.assertContains(response, reverse("add_cliente"))

    def test_tela_nova_solicitacao_contrato_exibe_modal_de_cliente(self):
        user = self.create_user("lidercontr", "lider_contrato")
        self.client.force_login(user)

        response = self.client.get(reverse("nova_solicitacao_contrato"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="modalCliente"')
        self.assertContains(response, reverse("add_cliente"))

    def test_cliente_criado_via_ajax_aparece_em_nova_solicitacao_prospeccao(self):
        user = self.create_user("liderprospajax", "lider_contrato")
        self.client.force_login(user)

        create_response = self.client.post(
            reverse("add_cliente"),
            {
                "nome": "Cliente Prospeccao Ajax",
                "cpf_cnpj": "11.222.333/0001-44",
                "email": "clienteprospajax@example.com",
            },
        )

        self.assertEqual(create_response.status_code, 200)

        response = self.client.get(reverse("nova_solicitacao_prospeccao"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "CLIENTE PROSPECCAO AJAX")
        self.assertContains(response, 'id="clienteSelect"')

    def test_cliente_criado_via_ajax_aparece_em_nova_solicitacao_contrato(self):
        user = self.create_user("lidercontrajax", "lider_contrato")
        self.client.force_login(user)

        create_response = self.client.post(
            reverse("add_cliente"),
            {
                "nome": "Cliente Contratacao Ajax",
                "cpf_cnpj": "55.666.777/0001-88",
                "email": "clientecontrajax@example.com",
            },
        )

        self.assertEqual(create_response.status_code, 200)

        response = self.client.get(reverse("nova_solicitacao_contrato"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "CLIENTE CONTRATACAO AJAX")
        self.assertContains(response, 'id="clienteSelect"')


class ReportSuprimentoViewTests(BaseUserTestCase):
    def test_report_suprimento_bloqueia_usuario_de_outro_grupo(self):
        user = self.create_user("lider", "lider_contrato")
        self.client.force_login(user)

        response = self.client.post(reverse("enviar_report_suprimento"), follow=True)

        self.assertRedirects(response, reverse("home"))
        messages = [message.message for message in get_messages(response.wsgi_request)]
        self.assertIn("Você não tem permissão para isso!", messages)
        self.assertEqual(len(mail.outbox), 0)

    def test_report_suprimento_requer_post(self):
        user = self.create_user("supri", "suprimento")
        self.client.force_login(user)

        response = self.client.get(reverse("enviar_report_suprimento"))

        self.assertRedirects(response, reverse("home"))
        self.assertEqual(len(mail.outbox), 0)

    def test_report_suprimento_exige_email_no_usuario(self):
        user = self.create_user("supri", "suprimento", email="")
        self.client.force_login(user)

        response = self.client.post(reverse("enviar_report_suprimento"), follow=True)

        self.assertRedirects(response, reverse("home"))
        messages = [message.message for message in get_messages(response.wsgi_request)]
        self.assertIn("Seu usuário não possui e-mail cadastrado.", messages)
        self.assertEqual(len(mail.outbox), 0)

    def test_report_suprimento_post_envia_email_para_o_proprio_usuario(self):
        user = self.create_user("supri", "suprimento", email="suprimento@example.com")
        self.client.force_login(user)

        response = self.client.post(reverse("enviar_report_suprimento"), follow=True)

        self.assertRedirects(response, reverse("home"))
        self.assertEqual(len(mail.outbox), 1)
        email = mail.outbox[0]
        self.assertEqual(email.to, ["suprimento@example.com"])
        self.assertEqual(email.subject, "Report Semanal de Suprimentos")

        messages = [message.message for message in get_messages(response.wsgi_request)]
        self.assertIn("Report enviado para suprimento@example.com.", messages)


class PermissionHelperTests(BaseUserTestCase):
    def test_user_shares_center_with_coordinator_returns_true_when_centers_intersect(self):
        centro = self.create_center()
        gerente_lider = self.create_user("gl", "gerente_lider")
        coordenador = self.create_user("coord", "coordenador")
        gerente_lider.centros.add(centro)
        coordenador.centros.add(centro)

        self.assertTrue(user_shares_center_with_coordinator(gerente_lider, coordenador))

    def test_user_shares_center_with_coordinator_returns_false_without_intersection(self):
        centro_a = self.create_center("CTA", "Centro A")
        centro_b = self.create_center("CTB", "Centro B")
        gerente_lider = self.create_user("gl", "gerente_lider")
        coordenador = self.create_user("coord", "coordenador")
        gerente_lider.centros.add(centro_a)
        coordenador.centros.add(centro_b)

        self.assertFalse(user_shares_center_with_coordinator(gerente_lider, coordenador))

    def test_can_user_manage_supplier_choice_respects_group_rules(self):
        centro = self.create_center()
        lider = self.create_user("lider", "lider_contrato")
        outro_lider = self.create_user("outro", "lider_contrato")
        gerente_lider = self.create_user("gl", "gerente_lider")
        gerente_contrato = self.create_user("gc", "gerente_contrato")
        coordenador = self.create_user("coord", "coordenador")
        gerente_lider.centros.add(centro)
        coordenador.centros.add(centro)
        contrato = self.create_contract(codigo="PRJ-SUP", coordenador=coordenador, lider_contrato=lider)
        solicitacao = SolicitacaoProspeccao.objects.create(
            contrato=contrato,
            coordenador=coordenador,
            lider_contrato=lider,
            descricao="Solicitação teste",
            status="Triagem realizada",
        )

        self.assertTrue(can_user_manage_supplier_choice(lider, solicitacao))
        self.assertFalse(can_user_manage_supplier_choice(outro_lider, solicitacao))
        self.assertTrue(can_user_manage_supplier_choice(gerente_contrato, solicitacao))
        self.assertTrue(can_user_manage_supplier_choice(gerente_lider, solicitacao))

    def test_can_user_manage_supplier_choice_blocks_gerente_lider_outside_center(self):
        centro_a = self.create_center("CTA", "Centro A")
        centro_b = self.create_center("CTB", "Centro B")
        lider = self.create_user("lider", "lider_contrato")
        gerente_lider = self.create_user("gl", "gerente_lider")
        coordenador = self.create_user("coord", "coordenador")
        gerente_lider.centros.add(centro_a)
        coordenador.centros.add(centro_b)
        contrato = self.create_contract(codigo="PRJ-NO", coordenador=coordenador, lider_contrato=lider)
        solicitacao = SolicitacaoProspeccao.objects.create(
            contrato=contrato,
            coordenador=coordenador,
            lider_contrato=lider,
            descricao="Solicitação teste",
            status="Triagem realizada",
        )

        self.assertFalse(can_user_manage_supplier_choice(gerente_lider, solicitacao))

    def test_can_user_manage_event_delivery_respects_group_rules(self):
        centro = self.create_center()
        lider = self.create_user("lider", "lider_contrato")
        outro_lider = self.create_user("outro", "lider_contrato")
        gerente_lider = self.create_user("gl", "gerente_lider")
        gerente_contrato = self.create_user("gc", "gerente_contrato")
        coordenador = self.create_user("coord", "coordenador")
        gerente_lider.centros.add(centro)
        coordenador.centros.add(centro)
        contrato = SimpleNamespace(lider_contrato=lider, coordenador=coordenador)

        self.assertTrue(can_user_manage_event_delivery(lider, contrato))
        self.assertFalse(can_user_manage_event_delivery(outro_lider, contrato))
        self.assertTrue(can_user_manage_event_delivery(gerente_contrato, contrato))
        self.assertTrue(can_user_manage_event_delivery(gerente_lider, contrato))

    def test_can_user_manage_os_delivery_respects_group_rules(self):
        centro = self.create_center()
        lider = self.create_user("lider", "lider_contrato")
        outro_lider = self.create_user("outro", "lider_contrato")
        gerente_lider = self.create_user("gl", "gerente_lider")
        gerente_contrato = self.create_user("gc", "gerente_contrato")
        coordenador = self.create_user("coord", "coordenador")
        gerente_lider.centros.add(centro)
        coordenador.centros.add(centro)
        os_obj = SimpleNamespace(lider_contrato=lider, coordenador=coordenador)

        self.assertTrue(can_user_manage_os_delivery(lider, os_obj))
        self.assertFalse(can_user_manage_os_delivery(outro_lider, os_obj))
        self.assertTrue(can_user_manage_os_delivery(gerente_contrato, os_obj))
        self.assertTrue(can_user_manage_os_delivery(gerente_lider, os_obj))
        self.assertFalse(can_user_manage_os_delivery(None, os_obj))


class HelperFunctionTests(BaseUserTestCase):
    def test_get_week_ranges_returns_expected_boundaries(self):
        reference = date(2026, 4, 24)  # sexta-feira

        weeks = get_week_ranges(reference)

        self.assertEqual(weeks["current_week_start"], date(2026, 4, 20))
        self.assertEqual(weeks["current_week_end"], date(2026, 4, 26))
        self.assertEqual(weeks["previous_week_start"], date(2026, 4, 13))
        self.assertEqual(weeks["previous_week_end"], date(2026, 4, 19))
        self.assertEqual(weeks["next_week_start"], date(2026, 4, 27))
        self.assertEqual(weeks["next_week_end"], date(2026, 5, 3))

    def test_is_request_concluded_by_status(self):
        contrato = self.create_contract(codigo="PRJ-HELP")
        guarda_chuva = self.create_supplier_contract(
            cod_projeto=contrato,
            guarda_chuva=True,
            num_contrato="CT-HELP",
        )
        solicitante = self.create_user("solicitante", "lider_contrato")

        prospeccao = SolicitacaoProspeccao(contrato=contrato, status="Onboarding")
        solicitacao_contrato = SolicitacaoContrato(status="Onboarding")
        solicitacao_os = SolicitacaoOrdemServico(
            contrato=guarda_chuva,
            solicitante=solicitante,
            titulo="OS",
            descricao="Descrição",
            status="finalizada",
        )
        solicitacao_os_pendente = SolicitacaoOrdemServico(
            contrato=guarda_chuva,
            solicitante=solicitante,
            titulo="OS pendente",
            descricao="Descrição",
            status="pendente_lider",
        )

        self.assertTrue(is_request_concluded(prospeccao))
        self.assertTrue(is_request_concluded(solicitacao_contrato))
        self.assertTrue(is_request_concluded(solicitacao_os))
        self.assertFalse(is_request_concluded(solicitacao_os_pendente))

    def test_format_request_line_uses_type_id_and_status(self):
        contrato = self.create_contract(codigo="PRJ-LINE")
        solicitacao = SolicitacaoProspeccao.objects.create(
            contrato=contrato,
            descricao="Comprar serviço especializado",
            status="Triagem realizada",
        )

        line = format_request_line("Prospecção", solicitacao)

        self.assertIn("Prospecção", line)
        self.assertIn(f"#{solicitacao.id}", line)
        self.assertIn("Triagem realizada", line)

    def test_send_request_notification_to_management_sends_only_to_valid_recipients(self):
        self.create_user("diretor", "diretoria", email="diretor@example.com")
        self.create_user("gerente_contrato", "gerente_contrato", email="gc@example.com")
        self.create_user("sem_email", "diretoria", email="")
        self.create_user("lider", "lider_contrato", email="lider@example.com")

        send_request_notification_to_management("Nova solicitação", "Mensagem de teste")

        self.assertEqual(len(mail.outbox), 1)
        email = mail.outbox[0]
        self.assertEqual(email.subject, "Nova solicitação")
        self.assertCountEqual(email.to, ["diretor@example.com", "gc@example.com"])


class WeeklyReportBuilderTests(BaseUserTestCase):
    def test_build_weekly_supply_report_includes_expected_sections_and_records(self):
        reference_date = date(2026, 4, 24)
        weeks = get_week_ranges(reference_date)

        suprimento = self.create_user("supri", "suprimento", email="supri@example.com")
        coordenador = self.create_user("coord", "coordenador")
        lider = self.create_user("lider", "lider_contrato")
        cliente = self.create_client("Cliente Report", "22.222.222/0001-22")
        contrato = self.create_contract(
            codigo="PRJ-REPORT",
            cliente=cliente,
            coordenador=coordenador,
            lider_contrato=lider,
        )
        fornecedor = self.create_supplier("Fornecedor Report", "33.333.333/0001-33")
        contrato_terceiro = self.create_supplier_contract(
            cod_projeto=contrato,
            empresa_terceira=fornecedor,
            coordenador=coordenador,
            lider_contrato=lider,
            guarda_chuva=True,
            status="encerrado",
            num_contrato="CTR-REPORT",
        )
        contrato_terceiro.data_fim = weeks["previous_week_end"]
        contrato_terceiro.save(update_fields=["data_fim"])

        prospeccao = SolicitacaoProspeccao.objects.create(
            contrato=contrato,
            coordenador=coordenador,
            lider_contrato=lider,
            descricao="Prospecção semanal",
            status="Em análise",
        )
        SolicitacaoProspeccao.objects.filter(pk=prospeccao.pk).update(
            data_solicitacao=timezone.make_aware(
                datetime.combine(weeks["previous_week_start"], datetime.min.time())
            )
        )
        prospeccao.refresh_from_db()

        solicitacao_pendente = SolicitacaoContrato.objects.create(
            contrato=contrato,
            coordenador=coordenador,
            lider_contrato=lider,
            descricao="Contratação pendente geral",
            status="Em análise",
        )
        SolicitacaoContrato.objects.filter(pk=solicitacao_pendente.pk).update(
            data_solicitacao=timezone.make_aware(
                datetime.combine(weeks["current_week_start"], datetime.min.time())
            )
        )

        solicitacao_os = SolicitacaoOrdemServico.objects.create(
            contrato=contrato_terceiro,
            cod_projeto=contrato,
            solicitante=lider,
            lider_contrato=lider,
            coordenador=coordenador,
            titulo="OS semanal",
            descricao="OS criada na semana anterior",
            status="finalizada",
        )
        SolicitacaoOrdemServico.objects.filter(pk=solicitacao_os.pk).update(
            criado_em=timezone.make_aware(
                datetime.combine(weeks["previous_week_start"], datetime.min.time())
            )
        )

        Evento.objects.create(
            empresa_terceira=fornecedor,
            contrato_terceiro=contrato_terceiro,
            descricao="Entrega semana anterior",
            data_prevista=weeks["previous_week_end"],
            data_entrega=weeks["previous_week_end"],
            realizado=True,
        )
        Evento.objects.create(
            empresa_terceira=fornecedor,
            contrato_terceiro=contrato_terceiro,
            descricao="Entrega semana atual",
            data_prevista=weeks["current_week_start"],
            realizado=False,
        )
        Evento.objects.create(
            empresa_terceira=fornecedor,
            contrato_terceiro=contrato_terceiro,
            descricao="Evento atrasado sem entrega",
            data_prevista=weeks["previous_week_start"],
            realizado=False,
        )
        contrato_proxima_semana = self.create_supplier_contract(
            cod_projeto=contrato,
            empresa_terceira=fornecedor,
            coordenador=coordenador,
            lider_contrato=lider,
            guarda_chuva=False,
            status="ativo",
            num_contrato="CTR-NEXT",
        )
        contrato_proxima_semana.data_fim = weeks["next_week_start"]
        contrato_proxima_semana.save(update_fields=["data_fim"])

        contrato_atrasado = self.create_supplier_contract(
            cod_projeto=contrato,
            empresa_terceira=fornecedor,
            coordenador=coordenador,
            lider_contrato=lider,
            guarda_chuva=False,
            status="ativo",
            num_contrato="CTR-OVERDUE",
        )
        contrato_atrasado.data_fim = weeks["previous_week_start"]
        contrato_atrasado.save(update_fields=["data_fim"])

        with patch("gestao_contratos.views.timezone.localdate", return_value=reference_date):
            html = build_weekly_supply_report(suprimento)

        self.assertIn("Report Semanal de Suprimentos", html)
        self.assertIn("Semana anterior", html)
        self.assertIn("Semana atual", html)
        self.assertIn("Próxima semana", html)
        self.assertIn("Pendências gerais", html)
        self.assertIn("Solicitação para PRJ-REPORT - CLIENTE REPORT", html)
        self.assertIn("OS semanal", html)
        self.assertIn("Entrega semana anterior", html)
        self.assertIn("Entrega semana atual", html)
        self.assertIn("Contratação", html)
        self.assertIn("Em análise", html)
        self.assertIn("Evento atrasado sem entrega", html)
        self.assertIn("FORNECEDOR REPORT", html)
        self.assertIn("Solicitações criadas", html)
        self.assertIn("Todas as solicitações pendentes", html)
        self.assertIn("Eventos previstos que não foram entregues", html)
        self.assertIn("Contratos com data fim vencida e ainda não encerrados", html)
        self.assertIn("Contratos previstos para finalizar na próxima semana", html)
        self.assertIn("Ativo", html)


class ApproveFornecedorGerenteTests(BaseUserTestCase):
    def setUp(self):
        self.centro = self.create_center()
        self.coordenador = self.create_user("coord", "coordenador")
        self.lider = self.create_user("lider", "lider_contrato")
        self.gerente_contrato = self.create_user("gc", "gerente_contrato")
        self.gerente_lider = self.create_user("gl", "gerente_lider")
        self.suprimento = self.create_user("supri", "suprimento")
        self.diretor = self.create_user("diretor", "diretoria")
        self.coordenador.centros.add(self.centro)
        self.gerente_lider.centros.add(self.centro)
        self.contrato = self.create_contract(
            codigo="PRJ-APR",
            coordenador=self.coordenador,
            lider_contrato=self.lider,
        )
        self.fornecedor = self.create_supplier("Fornecedor Aprovação", "44.444.444/0001-44")
        self.solicitacao = SolicitacaoProspeccao.objects.create(
            contrato=self.contrato,
            coordenador=self.coordenador,
            lider_contrato=self.lider,
            descricao="Solicitação para aprovação",
            aprovado=True,
            triagem_realizada=True,
            fornecedor_escolhido=self.fornecedor,
            status="Fornecedor selecionado",
        )

    def test_aprovar_fornecedor_gerente_bloqueia_usuario_sem_permissao(self):
        self.client.force_login(self.suprimento)

        response = self.client.post(
            reverse("aprovar_fornecedor_gerente", args=[self.solicitacao.pk]),
            {"acao": "aprovar"},
            follow=True,
        )

        self.assertRedirects(response, reverse("home"))
        self.solicitacao.refresh_from_db()
        self.assertEqual(self.solicitacao.aprovacao_fornecedor_gerente, "pendente")

    def test_aprovar_fornecedor_gerente_como_gerente_contrato(self):
        self.client.force_login(self.gerente_contrato)

        response = self.client.post(
            reverse("aprovar_fornecedor_gerente", args=[self.solicitacao.pk]),
            {"acao": "aprovar"},
            follow=True,
        )

        self.assertRedirects(response, reverse("lista_solicitacoes"))
        self.solicitacao.refresh_from_db()
        self.assertEqual(self.solicitacao.aprovacao_fornecedor_gerente, "aprovado")
        self.assertIsNotNone(self.solicitacao.aprocacao_fornecedor_gerente_em)

    def test_aprovar_fornecedor_gerente_como_gerente_lider_fora_do_centro(self):
        outro_centro = self.create_center("CT2", "Centro 2")
        gerente_lider_sem_escopo = self.create_user("gl2", "gerente_lider")
        gerente_lider_sem_escopo.centros.add(outro_centro)
        self.client.force_login(gerente_lider_sem_escopo)

        response = self.client.post(
            reverse("aprovar_fornecedor_gerente", args=[self.solicitacao.pk]),
            {"acao": "aprovar"},
            follow=True,
        )

        self.assertRedirects(response, reverse("home"))
        self.solicitacao.refresh_from_db()
        self.assertEqual(self.solicitacao.aprovacao_fornecedor_gerente, "pendente")

    def test_reprovar_fornecedor_gerente_limpa_escolha_e_solicita_nova_triagem(self):
        self.client.force_login(self.gerente_contrato)

        response = self.client.post(
            reverse("aprovar_fornecedor_gerente", args=[self.solicitacao.pk]),
            {"acao": "reprovar", "justificativa": "Fornecedor fora do escopo"},
            follow=True,
        )

        self.assertRedirects(response, reverse("lista_solicitacoes"))
        self.solicitacao.refresh_from_db()
        self.assertEqual(self.solicitacao.aprovacao_fornecedor_gerente, "reprovado")
        self.assertEqual(self.solicitacao.status, "Fornecedor reprovado pela gerência")
        self.assertFalse(self.solicitacao.triagem_realizada)
        self.assertIsNone(self.solicitacao.fornecedor_escolhido)
        self.assertEqual(self.solicitacao.justificativa_gerencia, "Fornecedor fora do escopo")

    def test_aprovar_fornecedor_gerente_define_status_final_quando_diretoria_ja_aprovou(self):
        self.solicitacao.aprovacao_fornecedor_diretor = "aprovado"
        self.solicitacao.save(update_fields=["aprovacao_fornecedor_diretor"])
        self.client.force_login(self.gerente_contrato)

        response = self.client.post(
            reverse("aprovar_fornecedor_gerente", args=[self.solicitacao.pk]),
            {"acao": "aprovar"},
            follow=True,
        )

        self.assertRedirects(response, reverse("lista_solicitacoes"))
        self.solicitacao.refresh_from_db()
        self.assertEqual(self.solicitacao.status, "Fornecedor aprovado")


class ApproveFornecedorDiretoriaTests(BaseUserTestCase):
    def setUp(self):
        self.centro = self.create_center()
        self.coordenador = self.create_user("coorddir", "coordenador")
        self.lider = self.create_user("liderdir", "lider_contrato")
        self.gerente_contrato = self.create_user("gcdir", "gerente_contrato")
        self.diretoria = self.create_user("dirdir", "diretoria")
        self.suprimento = self.create_user("supdir", "suprimento")
        self.coordenador.centros.add(self.centro)
        self.contrato = self.create_contract(
            codigo="PRJ-DIR-APR",
            coordenador=self.coordenador,
            lider_contrato=self.lider,
        )
        self.fornecedor = self.create_supplier("Fornecedor Diretoria", "77.777.777/0001-77")
        self.solicitacao = SolicitacaoProspeccao.objects.create(
            contrato=self.contrato,
            coordenador=self.coordenador,
            lider_contrato=self.lider,
            descricao="Solicitação diretoria",
            aprovado=True,
            triagem_realizada=True,
            fornecedor_escolhido=self.fornecedor,
            status="Fornecedor selecionado",
        )

    def test_aprovar_fornecedor_diretoria_define_status_final_quando_gerente_ja_aprovou(self):
        self.solicitacao.aprovacao_fornecedor_gerente = "aprovado"
        self.solicitacao.save(update_fields=["aprovacao_fornecedor_gerente"])
        self.client.force_login(self.diretoria)

        response = self.client.post(
            reverse("aprovar_fornecedor_diretor", args=[self.solicitacao.pk]),
            {"acao": "aprovar"},
            follow=True,
        )

        self.assertRedirects(response, reverse("lista_solicitacoes"))
        self.solicitacao.refresh_from_db()
        self.assertEqual(self.solicitacao.aprovacao_fornecedor_diretor, "aprovado")
        self.assertEqual(self.solicitacao.status, "Fornecedor aprovado")

    def test_reprovar_fornecedor_diretoria_limpa_escolha_e_exige_nova_triagem(self):
        self.client.force_login(self.diretoria)

        response = self.client.post(
            reverse("aprovar_fornecedor_diretor", args=[self.solicitacao.pk]),
            {"acao": "reprovar", "justificativa": "Fornecedor fora da estratégia."},
            follow=True,
        )

        self.assertRedirects(response, reverse("lista_solicitacoes"))
        self.solicitacao.refresh_from_db()
        self.assertEqual(self.solicitacao.aprovacao_fornecedor_diretor, "reprovado")
        self.assertEqual(self.solicitacao.status, "Fornecedor reprovado")
        self.assertFalse(self.solicitacao.triagem_realizada)
        self.assertIsNone(self.solicitacao.fornecedor_escolhido)
        self.assertEqual(self.solicitacao.justificativa_diretoria, "Fornecedor fora da estratégia.")
        self.assertGreaterEqual(len(mail.outbox), 1)


class DocumentoBMApprovalTests(BaseUserTestCase):
    def setUp(self):
        self.coordenador_group = Group.objects.create(name="Coordenador de Contrato")
        self.gerente_group = Group.objects.create(name="Gerente de Contrato")
        self.coordenador = self.create_user("coord", "coordenador")
        self.gerente_contrato = self.create_user("gc", "gerente_contrato")
        self.suprimento = self.create_user("supri", "suprimento")
        self.coordenador.groups.add(self.coordenador_group)
        self.gerente_contrato.groups.add(self.gerente_group)

        self.cliente = self.create_client("Cliente BM", "55.555.555/0001-55")
        self.contrato = self.create_contract(
            codigo="PRJ-BM",
            cliente=self.cliente,
            coordenador=self.coordenador,
            lider_contrato=self.gerente_contrato,
        )
        self.solicitacao = SolicitacaoProspeccao.objects.create(
            contrato=self.contrato,
            coordenador=self.coordenador,
            lider_contrato=self.gerente_contrato,
            descricao="Solicitação BM",
            status="Planejamento do contrato",
        )
        self.documento_bm = DocumentoBM.objects.create(solicitacao=self.solicitacao)

    def test_aprovar_bm_como_coordenador(self):
        self.client.force_login(self.coordenador)

        response = self.client.get(reverse("aprovar_bm", args=[self.documento_bm.pk, "coordenador"]))

        self.assertRedirects(
            response,
            reverse("detalhe_bm", kwargs={"pk": self.documento_bm.pk}),
            fetch_redirect_response=False,
        )
        self.documento_bm.refresh_from_db()
        self.assertEqual(self.documento_bm.status_coordenador, "aprovado")
        self.assertIsNotNone(self.documento_bm.data_aprovacao_coordenador)

    def test_aprovar_bm_como_gerente(self):
        self.client.force_login(self.gerente_contrato)

        response = self.client.get(reverse("aprovar_bm", args=[self.documento_bm.pk, "gerente"]))

        self.assertRedirects(response, reverse("detalhe_bm", kwargs={"pk": self.documento_bm.pk}))
        self.documento_bm.refresh_from_db()
        self.assertEqual(self.documento_bm.status_gerente, "aprovado")
        self.assertIsNotNone(self.documento_bm.data_aprovacao_gerente)

    def test_aprovar_bm_bloqueia_usuario_sem_group_do_django(self):
        self.client.force_login(self.suprimento)

        response = self.client.get(
            reverse("aprovar_bm", args=[self.documento_bm.pk, "gerente"]),
            follow=True,
        )

        self.assertRedirects(response, reverse("lista_solicitacoes"))
        self.documento_bm.refresh_from_db()
        self.assertEqual(self.documento_bm.status_gerente, "pendente")

    def test_reprovar_bm_como_gerente_envia_email_para_suprimento(self):
        self.client.force_login(self.gerente_contrato)

        response = self.client.get(reverse("reprovar_bm", args=[self.documento_bm.pk, "gerente"]))

        self.assertRedirects(response, reverse("detalhe_bm", kwargs={"pk": self.documento_bm.pk}))
        self.documento_bm.refresh_from_db()
        self.assertEqual(self.documento_bm.status_gerente, "reprovado")
        self.assertEqual(len(mail.outbox), 1)
        self.assertIn(self.suprimento.email, mail.outbox[0].to)

    def test_reprovar_bm_bloqueia_usuario_sem_permissao(self):
        self.client.force_login(self.suprimento)

        response = self.client.get(
            reverse("reprovar_bm", args=[self.documento_bm.pk, "coordenador"]),
            follow=True,
        )

        self.assertRedirects(response, reverse("lista_solicitacoes"))
        self.documento_bm.refresh_from_db()
        self.assertEqual(self.documento_bm.status_coordenador, "pendente")


class DiretoriaPagamentoBMTests(BaseUserTestCase):
    def setUp(self):
        self.diretoria = self.create_user("diretoria_bm", "diretoria")
        self.gerente_contrato = self.create_user("gerente_bm", "gerente_contrato")
        self.lider = self.create_user("lider_bm", "lider_contrato")
        self.suprimento = self.create_user("suprimento_bm", "suprimento")
        self.financeiro = self.create_user("financeiro_bm", "financeiro")
        self.financeiro_extra = self.create_user("financeiro_bm_extra", "financeiro")
        self.coordenador = self.create_user("coord_bm", "coordenador")
        self.centro = self.create_center()
        self.coordenador.centros.add(self.centro)

        self.contrato_base = self.create_contract(
            codigo="PRJ-DIR-BM",
            coordenador=self.coordenador,
            lider_contrato=self.lider,
        )
        self.contrato_terceiro = self.create_supplier_contract(
            cod_projeto=self.contrato_base,
            coordenador=self.coordenador,
            lider_contrato=self.lider,
            num_contrato="CT-DIR-BM",
        )
        self.evento = self.create_event(contrato=self.contrato_terceiro)
        self.bm = BM.objects.create(
            contrato=self.contrato_terceiro,
            evento=self.evento,
            numero_bm=1,
            parcela_paga=1,
            valor_pago=Decimal("1234.56"),
            status_coordenador="aprovado",
            status_gerente="aprovado",
        )

    def test_diretoria_nao_aprova_pagamento_sem_aprovacoes_previas(self):
        self.bm.status_coordenador = "pendente"
        self.bm.status_gerente = "pendente"
        self.bm.save(update_fields=["status_coordenador", "status_gerente"])
        self.client.force_login(self.diretoria)

        response = self.client.post(
            reverse("avaliar_bm", args=[self.bm.pk]),
            {"acao": "aprovar_pagamento"},
        )

        self.assertEqual(response.status_code, 400)
        self.bm.refresh_from_db()
        self.assertEqual(self.bm.aprovacao_pagamento, "pendente")

    def test_diretoria_aprova_pagamento_e_notifica_todos_os_membros_de_financeiro(self):
        self.client.force_login(self.diretoria)

        response = self.client.post(
            reverse("avaliar_bm", args=[self.bm.pk]),
            {"acao": "aprovar_pagamento"},
        )

        self.assertEqual(response.status_code, 200)
        self.bm.refresh_from_db()
        self.assertEqual(self.bm.aprovacao_pagamento, "aprovado")
        self.assertIsNotNone(self.bm.data_aprovacao_diretor)
        self.assertEqual(len(mail.outbox), 1)
        self.assertCountEqual(
            mail.outbox[0].to,
            [
                self.suprimento.email,
                self.financeiro.email,
                self.financeiro_extra.email,
            ],
        )

    def test_diretoria_reprova_pagamento_registra_justificativa(self):
        self.client.force_login(self.diretoria)

        response = self.client.post(
            reverse("avaliar_bm", args=[self.bm.pk]),
            {"acao": "reprovar_pagamento", "justificativa": "Valor divergente."},
        )

        self.assertEqual(response.status_code, 200)
        self.bm.refresh_from_db()
        self.assertEqual(self.bm.aprovacao_pagamento, "reprovado")
        self.assertEqual(self.bm.justificativa_reprovacao_diretor, "Valor divergente.")
        self.assertIsNotNone(self.bm.data_aprovacao_diretor)


class EventBMApprovalFlowTests(BaseUserTestCase):
    def setUp(self):
        self.centro = self.create_center()
        self.lider = self.create_user("lider_bm_fluxo", "lider_contrato")
        self.gerente_lider = self.create_user("gl_bm_fluxo", "gerente_lider")
        self.gerente_contrato = self.create_user("gc_bm_fluxo", "gerente_contrato")
        self.diretoria = self.create_user("dir_bm_fluxo", "diretoria")
        self.coordenador = self.create_user("coord_bm_fluxo", "coordenador")
        self.suprimento = self.create_user("sup_bm_fluxo", "suprimento")
        self.financeiro = self.create_user("fin_bm_fluxo", "financeiro")
        self.coordenador.centros.add(self.centro)
        self.gerente_lider.centros.add(self.centro)

        self.contrato_base = self.create_contract(
            codigo="PRJ-BM-FLUXO",
            coordenador=self.coordenador,
            lider_contrato=self.lider,
        )
        self.contrato_terceiro = self.create_supplier_contract(
            cod_projeto=self.contrato_base,
            coordenador=self.coordenador,
            lider_contrato=self.lider,
            num_contrato="CT-BM-FLUXO",
        )
        self.evento = self.create_event(contrato=self.contrato_terceiro)

    def create_bm_evento(self):
        return BM.objects.create(
            contrato=self.contrato_terceiro,
            evento=self.evento,
            numero_bm=1,
            parcela_paga=1,
            valor_pago=Decimal("500.00"),
        )

    def test_lider_aprova_bm_e_notifica_gerente_contrato(self):
        bm = self.create_bm_evento()
        self.client.force_login(self.lider)

        response = self.client.post(
            reverse("avaliar_bm", args=[bm.pk]),
            {"acao": "aprovar"},
        )

        self.assertEqual(response.status_code, 200)
        bm.refresh_from_db()
        self.assertEqual(bm.status_coordenador, "aprovado")
        self.assertEqual(bm.status_gerente, "pendente")
        self.assertEqual(len(mail.outbox), 1)
        self.assertIn(self.gerente_contrato.email, mail.outbox[0].to)

    def test_gerente_contrato_aprova_bm_sem_precisar_de_aprovacao_do_lider(self):
        bm = self.create_bm_evento()
        self.client.force_login(self.gerente_contrato)

        response = self.client.post(
            reverse("avaliar_bm", args=[bm.pk]),
            {"acao": "aprovar"},
        )

        self.assertEqual(response.status_code, 200)
        bm.refresh_from_db()
        self.assertEqual(bm.status_coordenador, "pendente")
        self.assertEqual(bm.status_gerente, "aprovado")

    def test_diretoria_pode_aprovar_pagamento_quando_apenas_o_lider_aprovou(self):
        bm = self.create_bm_evento()
        self.client.force_login(self.lider)
        self.client.post(
            reverse("avaliar_bm", args=[bm.pk]),
            {"acao": "aprovar"},
        )
        self.client.force_login(self.diretoria)

        response = self.client.post(
            reverse("avaliar_bm", args=[bm.pk]),
            {"acao": "aprovar_pagamento"},
        )

        self.assertEqual(response.status_code, 200)
        bm.refresh_from_db()
        self.assertEqual(bm.aprovacao_pagamento, "aprovado")

    def test_segunda_aprovacao_operacional_e_bloqueada_apos_aprovacao_inicial(self):
        bm = self.create_bm_evento()
        self.client.force_login(self.lider)
        self.client.post(
            reverse("avaliar_bm", args=[bm.pk]),
            {"acao": "aprovar"},
        )
        self.client.force_login(self.gerente_contrato)

        response = self.client.post(
            reverse("avaliar_bm", args=[bm.pk]),
            {"acao": "aprovar"},
        )

        self.assertEqual(response.status_code, 400)
        bm.refresh_from_db()
        self.assertEqual(bm.status_gerente, "pendente")


class DiretoriaHomeDashboardTests(BaseUserTestCase):
    def setUp(self):
        self.diretoria = self.create_user("diretoria_home", "diretoria")
        self.coordenador = self.create_user("coord_home", "coordenador")
        self.lider = self.create_user("lider_home", "lider_contrato")
        self.centro = self.create_center()
        self.coordenador.centros.add(self.centro)

        self.contrato_base = self.create_contract(
            codigo="PRJ-HOME-DIR",
            coordenador=self.coordenador,
            lider_contrato=self.lider,
        )
        self.contrato_terceiro = self.create_supplier_contract(
            cod_projeto=self.contrato_base,
            coordenador=self.coordenador,
            lider_contrato=self.lider,
            num_contrato="CT-HOME-DIR",
        )
        self.evento = self.create_event(contrato=self.contrato_terceiro)

    def test_home_diretoria_lista_bms_com_aprovacao_do_lider_ou_do_gerente(self):
        bm_aprovado_por_ambos = BM.objects.create(
            contrato=self.contrato_terceiro,
            evento=self.evento,
            numero_bm=1,
            parcela_paga=1,
            valor_pago=Decimal("100.00"),
            status_coordenador="aprovado",
            status_gerente="aprovado",
        )
        bm_aprovado_pelo_gerente = BM.objects.create(
            contrato=self.contrato_terceiro,
            evento=self.evento,
            numero_bm=2,
            parcela_paga=2,
            valor_pago=Decimal("200.00"),
            status_coordenador="pendente",
            status_gerente="aprovado",
        )
        bm_aprovado_pelo_lider = BM.objects.create(
            contrato=self.contrato_terceiro,
            evento=self.evento,
            numero_bm=3,
            parcela_paga=3,
            valor_pago=Decimal("300.00"),
            status_coordenador="aprovado",
            status_gerente="pendente",
        )
        BM.objects.create(
            contrato=self.contrato_terceiro,
            evento=self.evento,
            numero_bm=4,
            parcela_paga=4,
            valor_pago=Decimal("400.00"),
            status_coordenador="pendente",
            status_gerente="pendente",
        )

        self.client.force_login(self.diretoria)
        response = self.client.get(reverse("home"))

        self.assertEqual(response.status_code, 200)
        bms_pendentes = list(response.context["bms_pendentes"])
        self.assertCountEqual(
            [bm.id for bm in bms_pendentes],
            [bm_aprovado_por_ambos.id, bm_aprovado_pelo_gerente.id, bm_aprovado_pelo_lider.id],
        )

    def test_home_diretoria_lista_aditivos_pendentes_de_aprovacao_final(self):
        aditivo_visivel = AditivoContratoTerceiro.objects.create(
            contrato=self.contrato_terceiro,
            solicitado_por=self.lider,
            motivo="Prorrogacao",
            novo_valor_total=Decimal("2000.00"),
            nova_data_fim=date(2026, 7, 31),
            status_lider="aprovado",
        )
        AditivoContratoTerceiro.objects.create(
            contrato=self.contrato_terceiro,
            solicitado_por=self.lider,
            motivo="Ainda sem aprovacao operacional",
            novo_valor_total=Decimal("2100.00"),
            nova_data_fim=date(2026, 8, 15),
        )
        AditivoContratoTerceiro.objects.create(
            contrato=self.contrato_terceiro,
            solicitado_por=self.lider,
            motivo="Reprovado operacionalmente",
            novo_valor_total=Decimal("2200.00"),
            nova_data_fim=date(2026, 8, 20),
            status_gerente="reprovado",
        )

        self.client.force_login(self.diretoria)
        response = self.client.get(reverse("home"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(list(response.context["aditivos_pendentes"]), [aditivo_visivel])
        self.assertContains(response, f"Aditivo #{aditivo_visivel.id}")


class SuprimentoHomeDashboardTests(BaseUserTestCase):
    def setUp(self):
        self.suprimento = self.create_user("sup_home_aditivo", "suprimento")
        self.coordenador = self.create_user("coord_home_sup", "coordenador")
        self.lider = self.create_user("lider_home_sup", "lider_contrato")
        self.centro = self.create_center()
        self.coordenador.centros.add(self.centro)

        self.contrato_base = self.create_contract(
            codigo="PRJ-HOME-SUP",
            coordenador=self.coordenador,
            lider_contrato=self.lider,
        )
        self.contrato_terceiro = self.create_supplier_contract(
            cod_projeto=self.contrato_base,
            coordenador=self.coordenador,
            lider_contrato=self.lider,
            num_contrato="CT-HOME-SUP",
        )

    def test_home_suprimento_lista_aditivos_sem_documento_ou_reprovados(self):
        aditivo_sem_documento = AditivoContratoTerceiro.objects.create(
            contrato=self.contrato_terceiro,
            solicitado_por=self.lider,
            motivo="Sem documento",
            novo_valor_total=Decimal("1500.00"),
            nova_data_fim=date(2026, 6, 30),
        )
        aditivo_reprovado = AditivoContratoTerceiro.objects.create(
            contrato=self.contrato_terceiro,
            solicitado_por=self.lider,
            motivo="Reprovado pela diretoria",
            novo_valor_total=Decimal("1700.00"),
            nova_data_fim=date(2026, 7, 30),
            arquivo_aditivo=SimpleUploadedFile("aditivo.pdf", b"conteudo", content_type="application/pdf"),
            status_diretoria="reprovado",
        )
        AditivoContratoTerceiro.objects.create(
            contrato=self.contrato_terceiro,
            solicitado_por=self.lider,
            motivo="Ja aprovado",
            novo_valor_total=Decimal("1800.00"),
            nova_data_fim=date(2026, 8, 30),
            arquivo_aditivo=SimpleUploadedFile("aditivo_ok.pdf", b"conteudo", content_type="application/pdf"),
            status_lider="aprovado",
            status_diretoria="aprovado",
        )

        self.client.force_login(self.suprimento)
        response = self.client.get(reverse("home"))

        self.assertEqual(response.status_code, 200)
        self.assertCountEqual(
            [a.id for a in response.context["aditivos_pendentes"]],
            [aditivo_sem_documento.id, aditivo_reprovado.id],
        )
        self.assertContains(response, f"Aditivo #{aditivo_sem_documento.id}")
        self.assertContains(response, f"Aditivo #{aditivo_reprovado.id}")


class PrevisaoPagamentosTests(BaseUserTestCase):
    def setUp(self):
        self.centro = self.create_center()
        self.outro_centro = self.create_center("CT2", "Centro 2")
        self.suprimento = self.create_user("suprimento_prev", "suprimento")
        self.gerente_lider = self.create_user("gerente_lider_prev", "gerente_lider")
        self.gerente_contrato = self.create_user("gerente_contrato_prev", "gerente_contrato")
        self.outro_gerente_contrato = self.create_user("gerente_contrato_outro_prev", "gerente_contrato")
        self.coordenador = self.create_user("coord_prev", "coordenador")
        self.outro_coordenador = self.create_user("coord_outro_prev", "coordenador")
        self.gerente_lider.centros.add(self.centro)
        self.coordenador.centros.add(self.centro)
        self.outro_coordenador.centros.add(self.outro_centro)
        self.fornecedor = self.create_supplier("Fornecedor Previsao", "88.888.888/0001-88")
        self.fornecedor_outro = self.create_supplier("Fornecedor Fora Escopo", "99.999.999/0001-99")
        self.contrato_base = self.create_contract(
            codigo="PRJ-PREV",
            coordenador=self.coordenador,
            lider_contrato=self.gerente_contrato,
        )
        self.contrato_terceiro = self.create_supplier_contract(
            cod_projeto=self.contrato_base,
            empresa_terceira=self.fornecedor,
            coordenador=self.coordenador,
            lider_contrato=self.gerente_contrato,
            guarda_chuva=True,
            num_contrato="CT-PREV",
        )
        self.evento = self.create_event(
            contrato=self.contrato_terceiro,
            empresa_terceira=self.fornecedor,
            data_prevista=timezone.localdate(),
        )
        self.evento.data_prevista_pagamento = timezone.localdate()
        self.evento.valor_previsto = Decimal("1500.00")
        self.evento.save(update_fields=["data_prevista_pagamento", "valor_previsto"])
        self.os_visivel = self.create_os(
            contrato=self.contrato_terceiro,
            coordenador=self.coordenador,
            lider_contrato=self.gerente_contrato,
            titulo="OS Visível Prev",
        )
        self.os_visivel.data_pagamento = timezone.localdate()
        self.os_visivel.valor = Decimal("700.00")
        self.os_visivel.valor_pago = Decimal("650.00")
        self.os_visivel.save(update_fields=["data_pagamento", "valor", "valor_pago"])
        self.outro_contrato_base = self.create_contract(
            codigo="PRJ-OUT",
            coordenador=self.outro_coordenador,
            lider_contrato=self.outro_gerente_contrato,
        )
        self.outro_contrato_terceiro = self.create_supplier_contract(
            cod_projeto=self.outro_contrato_base,
            empresa_terceira=self.fornecedor_outro,
            coordenador=self.outro_coordenador,
            lider_contrato=self.outro_gerente_contrato,
            guarda_chuva=True,
            num_contrato="CT-OUT",
        )
        self.outro_evento = self.create_event(
            contrato=self.outro_contrato_terceiro,
            empresa_terceira=self.fornecedor_outro,
            data_prevista=timezone.localdate(),
        )
        self.outro_evento.data_prevista_pagamento = timezone.localdate()
        self.outro_evento.valor_previsto = Decimal("900.00")
        self.outro_evento.save(update_fields=["data_prevista_pagamento", "valor_previsto"])
        self.os_fora_escopo = self.create_os(
            contrato=self.outro_contrato_terceiro,
            coordenador=self.outro_coordenador,
            lider_contrato=self.outro_gerente_contrato,
            titulo="OS Fora Escopo Prev",
        )
        self.os_fora_escopo.data_pagamento = timezone.localdate()
        self.os_fora_escopo.valor = Decimal("400.00")
        self.os_fora_escopo.valor_pago = Decimal("350.00")
        self.os_fora_escopo.save(update_fields=["data_pagamento", "valor", "valor_pago"])

    def create_bm_for_previsao(self, **overrides):
        defaults = {
            "contrato": self.contrato_terceiro,
            "evento": self.evento,
            "numero_bm": 10,
            "parcela_paga": 1,
            "valor_pago": Decimal("1200.00"),
            "data_pagamento": timezone.localdate(),
            "status_coordenador": "aprovado",
            "status_gerente": "pendente",
            "data_aprovacao_coordenador": timezone.now(),
        }
        defaults.update(overrides)
        return BM.objects.create(**defaults)

    def create_out_of_scope_bm(self, **overrides):
        defaults = {
            "contrato": self.outro_contrato_terceiro,
            "evento": self.outro_evento,
            "numero_bm": 99,
            "parcela_paga": 1,
            "valor_pago": Decimal("800.00"),
            "data_pagamento": timezone.localdate(),
            "status_coordenador": "aprovado",
            "status_gerente": "pendente",
            "data_aprovacao_coordenador": timezone.now(),
        }
        defaults.update(overrides)
        return BM.objects.create(**defaults)

    def test_previsao_pagamentos_lista_bm_aprovado_pelo_lider_sem_aprovacao_do_gerente(self):
        bm = self.create_bm_for_previsao()
        self.client.force_login(self.suprimento)

        response = self.client.get(
            reverse("previsao_pagamentos"),
            {
                "data_inicial": timezone.localdate().strftime("%Y-%m-%d"),
                "data_limite": timezone.localdate().strftime("%Y-%m-%d"),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Boletins de Medição")
        self.assertContains(response, str(bm.numero_bm))
        self.assertContains(response, "Líder / Gerente-Líder")

    def test_download_bms_aprovados_inclui_bm_aprovado_pelo_lider(self):
        bm = self.create_bm_for_previsao(
            arquivo_bm=SimpleUploadedFile("bm_teste.pdf", b"conteudo pdf", content_type="application/pdf")
        )
        self.client.force_login(self.suprimento)

        response = self.client.get(
            reverse("download_bms_aprovados"),
            {
                "data_inicial": timezone.localdate().strftime("%Y-%m-%d"),
                "data_limite": timezone.localdate().strftime("%Y-%m-%d"),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/zip")
        zip_buffer = BytesIO(response.content)
        with zipfile.ZipFile(zip_buffer) as zip_file:
            self.assertEqual(len(zip_file.namelist()), 1)
            self.assertIn(f"{self.contrato_base.cod_projeto}_BM{bm.numero_bm}_", zip_file.namelist()[0])

    def test_exportar_previsao_pagamentos_permite_gerente_lider(self):
        self.client.force_login(self.gerente_lider)

        response = self.client.get(
            reverse("exportar_previsao_pagamentos_excel"),
            {
                "data_inicial": timezone.localdate().strftime("%Y-%m-%d"),
                "data_limite": timezone.localdate().strftime("%Y-%m-%d"),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_exportar_previsao_pagamentos_permite_gerente_contrato(self):
        self.client.force_login(self.gerente_contrato)

        response = self.client.get(
            reverse("exportar_previsao_pagamentos_excel"),
            {
                "data_inicial": timezone.localdate().strftime("%Y-%m-%d"),
                "data_limite": timezone.localdate().strftime("%Y-%m-%d"),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_previsao_pagamentos_mostra_atalho_para_gerente_contrato(self):
        bm = self.create_bm_for_previsao()
        self.client.force_login(self.gerente_contrato)

        response = self.client.get(
            reverse("previsao_pagamentos"),
            {
                "data_inicial": timezone.localdate().strftime("%Y-%m-%d"),
                "data_limite": timezone.localdate().strftime("%Y-%m-%d"),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse("detalhes_entrega", args=[bm.evento.id]))
        self.assertContains(response, "Verificar")

    def test_previsao_pagamentos_restringe_gerente_lider_ao_mesmo_centro(self):
        bm_visivel = self.create_bm_for_previsao(numero_bm=11)
        self.create_out_of_scope_bm(numero_bm=77)
        self.client.force_login(self.gerente_lider)

        response = self.client.get(
            reverse("previsao_pagamentos"),
            {
                "data_inicial": timezone.localdate().strftime("%Y-%m-%d"),
                "data_limite": timezone.localdate().strftime("%Y-%m-%d"),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertCountEqual(
            [bm.id for bm in response.context["bms"]],
            [bm_visivel.id],
        )
        self.assertContains(response, str(bm_visivel.numero_bm))
        self.assertNotContains(response, "Fornecedor Fora Escopo")

    def test_previsao_pagamentos_totais_restringem_gerente_lider_em_eventos_e_os(self):
        self.client.force_login(self.gerente_lider)

        response = self.client.get(
            reverse("previsao_pagamentos"),
            {
                "data_inicial": timezone.localdate().strftime("%Y-%m-%d"),
                "data_limite": timezone.localdate().strftime("%Y-%m-%d"),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["total_previsto"], Decimal("2200.00"))
        self.assertEqual(response.context["total_pago"], Decimal("650.00"))
        pagamentos = list(response.context["pagamentos"])
        self.assertEqual(len(pagamentos), 2)
        self.assertCountEqual([item["tipo"] for item in pagamentos], ["Evento", "OS"])
        self.assertTrue(all(item["projeto"] == "PRJ-PREV" for item in pagamentos))

    def test_previsao_pagamentos_restringe_gerente_contrato_ao_proprio_vinculo(self):
        bm_visivel = self.create_bm_for_previsao(numero_bm=12)
        self.create_out_of_scope_bm(numero_bm=78, status_gerente="aprovado", data_aprovacao_gerente=timezone.now())
        self.client.force_login(self.gerente_contrato)

        response = self.client.get(
            reverse("previsao_pagamentos"),
            {
                "data_inicial": timezone.localdate().strftime("%Y-%m-%d"),
                "data_limite": timezone.localdate().strftime("%Y-%m-%d"),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertCountEqual(
            [bm.id for bm in response.context["bms"]],
            [bm_visivel.id],
        )
        self.assertContains(response, str(bm_visivel.numero_bm))
        self.assertNotContains(response, "Fornecedor Fora Escopo")

    def test_previsao_pagamentos_totais_restringem_gerente_contrato_em_eventos_e_os(self):
        self.client.force_login(self.gerente_contrato)

        response = self.client.get(
            reverse("previsao_pagamentos"),
            {
                "data_inicial": timezone.localdate().strftime("%Y-%m-%d"),
                "data_limite": timezone.localdate().strftime("%Y-%m-%d"),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["total_previsto"], Decimal("2200.00"))
        self.assertEqual(response.context["total_pago"], Decimal("650.00"))
        pagamentos = list(response.context["pagamentos"])
        self.assertEqual(len(pagamentos), 2)
        self.assertCountEqual([item["tipo"] for item in pagamentos], ["Evento", "OS"])
        self.assertTrue(all(item["projeto"] == "PRJ-PREV" for item in pagamentos))

    def test_previsao_pagamentos_tabela_mostra_evento_e_os_no_escopo(self):
        self.client.force_login(self.gerente_lider)

        response = self.client.get(
            reverse("previsao_pagamentos"),
            {
                "data_inicial": timezone.localdate().strftime("%Y-%m-%d"),
                "data_limite": timezone.localdate().strftime("%Y-%m-%d"),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "<td>Evento</td>", html=True)
        self.assertContains(response, "<td>OS</td>", html=True)
        self.assertContains(response, "PRJ-PREV")
        pagamentos = list(response.context["pagamentos"])
        self.assertTrue(all(item["projeto"] == "PRJ-PREV" for item in pagamentos))

    def test_previsao_pagamentos_nao_duplica_evento_ou_os_por_multiplos_centros_compartilhados(self):
        centro_extra = self.create_center("CT3", "Centro 3")
        self.gerente_lider.centros.add(centro_extra)
        self.coordenador.centros.add(centro_extra)

        self.client.force_login(self.gerente_lider)

        response = self.client.get(
            reverse("previsao_pagamentos"),
            {
                "data_inicial": timezone.localdate().strftime("%Y-%m-%d"),
                "data_limite": timezone.localdate().strftime("%Y-%m-%d"),
            },
        )

        self.assertEqual(response.status_code, 200)
        pagamentos = list(response.context["pagamentos"])
        self.assertEqual(len(pagamentos), 2)
        self.assertEqual(response.context["total_previsto"], Decimal("2200.00"))
        self.assertEqual(response.context["total_pago"], Decimal("650.00"))
        self.assertEqual(
            sum(1 for item in pagamentos if item["tipo"] == "Evento" and item["projeto"] == "PRJ-PREV"),
            1,
        )
        self.assertEqual(
            sum(1 for item in pagamentos if item["tipo"] == "OS" and item["projeto"] == "PRJ-PREV"),
            1,
        )

    def test_download_bms_aprovados_respeita_escopo_do_gerente_lider(self):
        bm_visivel = self.create_bm_for_previsao(
            numero_bm=15,
            arquivo_bm=SimpleUploadedFile("bm_gl_visivel.pdf", b"conteudo visivel", content_type="application/pdf"),
        )
        self.create_out_of_scope_bm(
            numero_bm=81,
            arquivo_bm=SimpleUploadedFile("bm_gl_oculto.pdf", b"conteudo oculto", content_type="application/pdf"),
        )
        self.client.force_login(self.gerente_lider)

        response = self.client.get(
            reverse("download_bms_aprovados"),
            {
                "data_inicial": timezone.localdate().strftime("%Y-%m-%d"),
                "data_limite": timezone.localdate().strftime("%Y-%m-%d"),
            },
        )

        self.assertEqual(response.status_code, 200)
        zip_buffer = BytesIO(response.content)
        with zipfile.ZipFile(zip_buffer) as zip_file:
            names = zip_file.namelist()
            self.assertEqual(len(names), 1)
            self.assertIn(f"{self.contrato_base.cod_projeto}_BM{bm_visivel.numero_bm}_", names[0])
            self.assertNotIn("PRJ-OUT", names[0])

    def test_download_bms_aprovados_respeita_escopo_do_gerente_contrato(self):
        bm_visivel = self.create_bm_for_previsao(
            numero_bm=13,
            arquivo_bm=SimpleUploadedFile("bm_visivel.pdf", b"conteudo visivel", content_type="application/pdf"),
        )
        self.create_out_of_scope_bm(
            numero_bm=79,
            arquivo_bm=SimpleUploadedFile("bm_oculto.pdf", b"conteudo oculto", content_type="application/pdf"),
        )
        self.client.force_login(self.gerente_contrato)

        response = self.client.get(
            reverse("download_bms_aprovados"),
            {
                "data_inicial": timezone.localdate().strftime("%Y-%m-%d"),
                "data_limite": timezone.localdate().strftime("%Y-%m-%d"),
            },
        )

        self.assertEqual(response.status_code, 200)
        zip_buffer = BytesIO(response.content)
        with zipfile.ZipFile(zip_buffer) as zip_file:
            names = zip_file.namelist()
            self.assertEqual(len(names), 1)
            self.assertIn(f"{self.contrato_base.cod_projeto}_BM{bm_visivel.numero_bm}_", names[0])
            self.assertNotIn("PRJ-OUT", names[0])

    def test_exportar_previsao_pagamentos_restringe_gerente_lider_ao_mesmo_centro(self):
        self.create_bm_for_previsao(numero_bm=14)
        self.create_out_of_scope_bm(numero_bm=80)
        self.client.force_login(self.gerente_lider)

        response = self.client.get(
            reverse("exportar_previsao_pagamentos_excel"),
            {
                "data_inicial": timezone.localdate().strftime("%Y-%m-%d"),
                "data_limite": timezone.localdate().strftime("%Y-%m-%d"),
            },
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        exported_values = " ".join(str(cell) for row in worksheet.iter_rows(values_only=True) for cell in row if cell)
        self.assertIn("PRJ-PREV", exported_values)
        self.assertNotIn("PRJ-OUT", exported_values)
        self.assertEqual(worksheet.max_row, 3)

    def test_exportar_previsao_pagamentos_restringe_gerente_contrato_ao_proprio_vinculo(self):
        self.create_bm_for_previsao(numero_bm=16)
        self.create_out_of_scope_bm(numero_bm=82)
        self.client.force_login(self.gerente_contrato)

        response = self.client.get(
            reverse("exportar_previsao_pagamentos_excel"),
            {
                "data_inicial": timezone.localdate().strftime("%Y-%m-%d"),
                "data_limite": timezone.localdate().strftime("%Y-%m-%d"),
            },
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        exported_values = " ".join(str(cell) for row in worksheet.iter_rows(values_only=True) for cell in row if cell)
        self.assertIn("PRJ-PREV", exported_values)
        self.assertNotIn("PRJ-OUT", exported_values)
        self.assertEqual(worksheet.max_row, 3)


class SolicitarOSViewTests(BaseUserTestCase):
    def setUp(self):
        self.centro = self.create_center()
        self.lider = self.create_user("lideros", "lider_contrato")
        self.gerente_lider = self.create_user("glos", "gerente_lider")
        self.coordenador = self.create_user("coordos", "coordenador")
        self.suprimento = self.create_user("suprios", "suprimento")
        self.coordenador.centros.add(self.centro)
        self.gerente_lider.centros.add(self.centro)
        self.contrato_base = self.create_contract(
            codigo="PRJ-OS",
            coordenador=self.coordenador,
            lider_contrato=self.lider,
        )
        self.contrato_terceiro = self.create_supplier_contract(
            cod_projeto=self.contrato_base,
            coordenador=self.coordenador,
            lider_contrato=self.lider,
            guarda_chuva=True,
            num_contrato="CT-OS-001",
        )

    def test_solicitar_os_com_contrato_bloqueia_usuario_sem_permissao(self):
        self.client.force_login(self.suprimento)

        response = self.client.get(reverse("solicitar_os_com_contrato"), follow=True)

        self.assertRedirects(response, reverse("home"))
        self.assertEqual(SolicitacaoOrdemServico.objects.count(), 0)

    def test_lider_contrato_pode_criar_solicitacao_os_com_contrato(self):
        self.client.force_login(self.lider)

        response = self.client.post(
            reverse("solicitar_os_com_contrato"),
            {
                "contrato": self.contrato_terceiro.pk,
                "cod_projeto": self.contrato_base.pk,
                "titulo": "Nova OS do líder",
                "descricao": "Precisa executar o serviço",
                "valor_previsto": "R$ 1.234,56",
                "prazo_execucao": "2026-04-30",
                "coordenador": self.coordenador.pk,
            },
            follow=False,
        )

        os_request = SolicitacaoOrdemServico.objects.get()
        self.assertRedirects(response, reverse("detalhe_ordem_servico", kwargs={"pk": os_request.pk}), fetch_redirect_response=False)
        self.assertEqual(os_request.solicitante, self.lider)
        self.assertEqual(os_request.lider_contrato, self.lider)
        self.assertEqual(os_request.status, "pendente_lider")
        self.assertEqual(os_request.contrato, self.contrato_terceiro)
        self.assertEqual(os_request.valor_previsto, Decimal("1234.56"))
        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].to, [self.lider.email])

    def test_gerente_lider_ve_apenas_contratos_do_mesmo_centro_no_form(self):
        outro_centro = self.create_center("CT2", "Centro 2")
        outro_coord = self.create_user("coordfora", "coordenador")
        outro_coord.centros.add(outro_centro)
        outro_base = self.create_contract(codigo="PRJ-FORA", coordenador=outro_coord, lider_contrato=self.lider)
        self.create_supplier_contract(
            cod_projeto=outro_base,
            coordenador=outro_coord,
            lider_contrato=self.lider,
            guarda_chuva=True,
            num_contrato="CT-OS-002",
        )

        self.client.force_login(self.gerente_lider)
        response = self.client.get(reverse("solicitar_os_com_contrato"))

        self.assertEqual(response.status_code, 200)
        contratos = list(response.context["form"].fields["contrato"].queryset)
        coordenadores = list(response.context["form"].fields["coordenador"].queryset)
        self.assertEqual(contratos, [self.contrato_terceiro])
        self.assertEqual(coordenadores, [self.coordenador])


class CenterScopedViewTests(BaseUserTestCase):
    def setUp(self):
        self.centro_compartilhado = self.create_center()
        self.centro_outro = self.create_center("CT2", "Centro 2")
        self.gerente_lider = self.create_user("glscope", "gerente_lider")
        self.lider = self.create_user("liderscope", "lider_contrato")
        self.gerente_lider.centros.add(self.centro_compartilhado)
        self.coord_ok = self.create_user("coordok", "coordenador")
        self.coord_ok.centros.add(self.centro_compartilhado)
        self.coord_fora = self.create_user("coordfora2", "coordenador")
        self.coord_fora.centros.add(self.centro_outro)

        self.base_ok = self.create_contract(codigo="PRJ-OK", coordenador=self.coord_ok, lider_contrato=self.lider)
        self.base_fora = self.create_contract(codigo="PRJ-FORA2", coordenador=self.coord_fora, lider_contrato=self.lider)
        self.ct_ok = self.create_supplier_contract(
            cod_projeto=self.base_ok,
            coordenador=self.coord_ok,
            lider_contrato=self.lider,
            guarda_chuva=True,
            num_contrato="CT-SCOPE-1",
        )
        self.ct_fora = self.create_supplier_contract(
            cod_projeto=self.base_fora,
            coordenador=self.coord_fora,
            lider_contrato=self.lider,
            guarda_chuva=True,
            num_contrato="CT-SCOPE-2",
        )
        self.solic_ok = SolicitacaoProspeccao.objects.create(
            contrato=self.base_ok,
            coordenador=self.coord_ok,
            lider_contrato=self.lider,
            descricao="Solicitação visível",
            status="Solicitação de prospecção",
        )
        self.solic_fora = SolicitacaoProspeccao.objects.create(
            contrato=self.base_fora,
            coordenador=self.coord_fora,
            lider_contrato=self.lider,
            descricao="Solicitação oculta",
            status="Solicitação de prospecção",
        )
        self.os_visivel = self.create_os(
            contrato=self.ct_ok,
            coordenador=self.coord_ok,
            lider_contrato=self.lider,
            titulo="OS visível",
        )
        self.os_oculta = self.create_os(
            contrato=self.ct_fora,
            coordenador=self.coord_fora,
            lider_contrato=self.lider,
            titulo="OS oculta",
        )
        self.os_request_visivel = self.create_os_request(
            contrato=self.ct_ok,
            solicitante=self.lider,
            lider_contrato=self.lider,
            coordenador=self.coord_ok,
            titulo="Solicitação OS visível",
        )
        self.os_request_oculta = self.create_os_request(
            contrato=self.ct_fora,
            solicitante=self.lider,
            lider_contrato=self.lider,
            coordenador=self.coord_fora,
            titulo="Solicitação OS oculta",
        )

    def test_lista_solicitacoes_do_gerente_lider_filtra_por_centro(self):
        self.client.force_login(self.gerente_lider)

        response = self.client.get(reverse("lista_solicitacoes"))

        self.assertEqual(response.status_code, 200)
        lista_ids = {item["solicitacao"].pk for item in response.context["lista_solicitacoes"]}
        os_ids = {item.pk for item in response.context["ordens_servico_page"].object_list}
        self.assertIn(self.solic_ok.pk, lista_ids)
        self.assertNotIn(self.solic_fora.pk, lista_ids)
        self.assertIn(self.os_visivel.pk, os_ids)
        self.assertNotIn(self.os_oculta.pk, os_ids)

    def test_detalhes_solicitacao_bloqueia_gerente_lider_fora_do_centro(self):
        self.client.force_login(self.gerente_lider)

        response = self.client.get(reverse("detalhes_solicitacao", args=[self.solic_fora.pk]), follow=True)

        self.assertRedirects(response, reverse("home"))

    def test_detalhe_os_bloqueia_gerente_lider_fora_do_centro(self):
        self.client.force_login(self.gerente_lider)

        response = self.client.get(reverse("detalhe_ordem_servico", args=[self.os_request_oculta.pk]), follow=True)

        self.assertRedirects(response, reverse("lista_solicitacoes"))

    def test_lista_ordens_servico_filtra_por_centro_e_busca_por_titulo(self):
        self.client.force_login(self.gerente_lider)

        response = self.client.get(reverse("lista_ordens_servico"), {"search": "visível"})

        self.assertEqual(response.status_code, 200)
        os_ids = {item.pk for item in response.context["page_obj"].object_list}
        self.assertEqual(os_ids, {self.os_visivel.pk})


class DeliveryRegistrationTests(BaseUserTestCase):
    def setUp(self):
        self.centro = self.create_center()
        self.outro_centro = self.create_center("CT2", "Centro 2")
        self.lider = self.create_user("liderent", "lider_contrato")
        self.gerente_lider = self.create_user("glent", "gerente_lider")
        self.gerente_contrato = self.create_user("gcent", "gerente_contrato")
        self.suprimento = self.create_user("suprent", "suprimento")
        self.coordenador = self.create_user("coordent", "coordenador")
        self.coordenador.centros.add(self.centro)
        self.gerente_lider.centros.add(self.centro)
        self.contrato_base = self.create_contract(
            codigo="PRJ-ENT",
            coordenador=self.coordenador,
            lider_contrato=self.lider,
        )
        self.contrato_terceiro = self.create_supplier_contract(
            cod_projeto=self.contrato_base,
            coordenador=self.coordenador,
            lider_contrato=self.lider,
            guarda_chuva=True,
            num_contrato="CT-ENT-1",
        )
        self.evento = self.create_event(contrato=self.contrato_terceiro, data_prevista=date(2026, 4, 18))
        self.os = self.create_os(
            contrato=self.contrato_terceiro,
            coordenador=self.coordenador,
            lider_contrato=self.lider,
            prazo_execucao=date(2026, 4, 20),
        )

    def test_registrar_entrega_evento_bloqueia_lider_sem_vinculo(self):
        outro_lider = self.create_user("liderfora", "lider_contrato")
        self.client.force_login(outro_lider)

        response = self.client.get(reverse("registrar_entrega", args=[self.evento.pk]), follow=True)

        self.assertRedirects(response, reverse("home"))

    def test_registrar_entrega_evento_como_suprimento_abre_tela(self):
        self.client.force_login(self.suprimento)

        response = self.client.get(reverse("registrar_entrega", args=[self.evento.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["evento"], self.evento)

    def test_registrar_entrega_evento_como_gerente_contrato_atualiza_campos(self):
        self.client.force_login(self.gerente_contrato)

        response = self.client.post(
            reverse("registrar_entrega", args=[self.evento.pk]),
            {
                "observacao": "Entrega concluída",
                "caminho_evidencia": "C:/evidencias/entrega.pdf",
                "justificativa": "",
                "avaliacao": "Aprovado",
                "data_entrega": "2026-04-21",
                "realizado": "on",
                "valor_pago": "480.00",
                "data_pagamento": "2026-04-22",
            },
            follow=False,
        )

        self.assertRedirects(
            response,
            reverse("contrato_fornecedor_detalhe", kwargs={"pk": self.contrato_terceiro.pk}),
            fetch_redirect_response=False,
        )
        self.evento.refresh_from_db()
        self.assertTrue(self.evento.realizado)
        self.assertEqual(self.evento.avaliacao, "Aprovado")
        self.assertEqual(self.evento.data_entrega, date(2026, 4, 21))
        self.assertEqual(self.evento.valor_pago, Decimal("480.00"))

    def test_registrar_entrega_evento_como_suprimento_atualiza_campos(self):
        self.client.force_login(self.suprimento)

        response = self.client.post(
            reverse("registrar_entrega", args=[self.evento.pk]),
            {
                "observacao": "Entrega registrada pelo suprimento",
                "caminho_evidencia": "C:/evidencias/suprimento.pdf",
                "justificativa": "",
                "avaliacao": "Aprovado",
                "data_entrega": "2026-04-19",
                "realizado": "on",
                "valor_pago": "500.00",
                "data_pagamento": "2026-04-22",
            },
            follow=False,
        )

        self.assertRedirects(
            response,
            reverse("contrato_fornecedor_detalhe", kwargs={"pk": self.contrato_terceiro.pk}),
            fetch_redirect_response=False,
        )
        self.evento.refresh_from_db()
        self.assertTrue(self.evento.realizado)
        self.assertEqual(self.evento.observacao, "Entrega registrada pelo suprimento")
        self.assertEqual(self.evento.data_entrega, date(2026, 4, 19))
        self.assertEqual(self.evento.valor_pago, Decimal("500.00"))

    def test_registrar_entrega_os_como_gerente_lider_finaliza_com_atraso(self):
        self.client.force_login(self.gerente_lider)

        response = self.client.post(
            reverse("registrar_entrega_os", args=[self.os.pk]),
            {
                "caminho_evidencia": "C:/evidencias/os.pdf",
                "avaliacao": "Aprovado",
                "data_entrega": "2026-04-22",
                "realizado": "on",
                "valor_pago": "350.00",
                "data_pagamento": "2026-04-23",
                "observacao": "Finalizada com atraso",
            },
            follow=False,
        )

        self.assertRedirects(
            response,
            reverse("detalhes_os", kwargs={"pk": self.os.pk}),
            fetch_redirect_response=False,
        )


class OrdemServicoApprovalFlowTests(BaseUserTestCase):
    def setUp(self):
        self.centro = self.create_center()
        self.outro_centro = self.create_center("CT2", "Centro 2")
        self.lider = self.create_user("liderflow", "lider_contrato")
        self.gerente_lider = self.create_user("glflow", "gerente_lider")
        self.gerente_contrato = self.create_user("gcflow", "gerente_contrato")
        self.suprimento = self.create_user("supriflow", "suprimento")
        self.coordenador = self.create_user("coordflow", "coordenador")
        self.coordenador.centros.add(self.centro)
        self.gerente_lider.centros.add(self.centro)
        self.contrato_base = self.create_contract(
            codigo="PRJ-FLOW",
            coordenador=self.coordenador,
            lider_contrato=self.lider,
        )
        self.contrato_terceiro = self.create_supplier_contract(
            cod_projeto=self.contrato_base,
            coordenador=self.coordenador,
            lider_contrato=self.lider,
            guarda_chuva=True,
            num_contrato="CT-FLOW-1",
        )

    def test_aprovar_os_lider_como_lider_envia_para_suprimento(self):
        os_request = self.create_os_request(
            contrato=self.contrato_terceiro,
            solicitante=self.lider,
            lider_contrato=self.lider,
            coordenador=self.coordenador,
            status="pendente_lider",
        )
        self.client.force_login(self.lider)

        response = self.client.get(reverse("aprovar_os_lider", args=[os_request.pk, "aprovar"]), follow=False)

        self.assertRedirects(
            response,
            reverse("detalhe_ordem_servico", kwargs={"pk": os_request.pk}),
            fetch_redirect_response=False,
        )
        os_request.refresh_from_db()
        self.assertEqual(os_request.status, "pendente_suprimento")
        self.assertEqual(os_request.aprovacao_lider, self.lider.username)
        self.assertIsNotNone(os_request.aprovado_lider_em)
        self.assertGreaterEqual(len(mail.outbox), 1)

    def test_aprovar_os_lider_bloqueia_gerente_lider_fora_do_centro(self):
        gerente_lider_fora = self.create_user("glsemfluxo", "gerente_lider")
        gerente_lider_fora.centros.add(self.outro_centro)
        os_request = self.create_os_request(
            contrato=self.contrato_terceiro,
            solicitante=self.lider,
            lider_contrato=self.lider,
            coordenador=self.coordenador,
            status="pendente_lider",
        )
        self.client.force_login(gerente_lider_fora)

        response = self.client.get(reverse("aprovar_os_lider", args=[os_request.pk, "aprovar"]), follow=False)

        self.assertRedirects(
            response,
            reverse("detalhe_ordem_servico", kwargs={"pk": os_request.pk}),
            fetch_redirect_response=False,
        )
        os_request.refresh_from_db()
        self.assertEqual(os_request.status, "pendente_lider")

    def test_aprovar_os_gerente_contrato_cria_os(self):
        os_request = self.create_os_request(
            contrato=self.contrato_terceiro,
            solicitante=self.lider,
            lider_contrato=self.lider,
            coordenador=self.coordenador,
            status="pendente_gerente",
            titulo="OS para aprovação final",
        )
        self.client.force_login(self.gerente_contrato)

        response = self.client.get(reverse("aprovar_os_gerente_contrato", args=[os_request.pk, "aprovar"]), follow=False)

        self.assertRedirects(
            response,
            reverse("detalhe_ordem_servico", kwargs={"pk": os_request.pk}),
            fetch_redirect_response=False,
        )
        os_request.refresh_from_db()
        self.assertEqual(os_request.status, "aprovada")
        os_cadastrada = OS.objects.get(solicitacao=os_request)
        self.assertEqual(os_cadastrada.contrato, self.contrato_terceiro)
        self.assertEqual(os_cadastrada.titulo, "OS para aprovação final")
        self.assertEqual(os_cadastrada.coordenador, self.lider)
        self.assertEqual(os_cadastrada.valor, os_request.valor_previsto)

    def test_aprovar_os_gerente_contrato_reprova_sem_criar_os(self):
        os_request = self.create_os_request(
            contrato=self.contrato_terceiro,
            solicitante=self.lider,
            lider_contrato=self.lider,
            coordenador=self.coordenador,
            status="pendente_gerente",
        )
        self.client.force_login(self.gerente_contrato)

        response = self.client.get(reverse("aprovar_os_gerente_contrato", args=[os_request.pk, "reprovar"]), follow=False)

        self.assertRedirects(
            response,
            reverse("detalhe_ordem_servico", kwargs={"pk": os_request.pk}),
            fetch_redirect_response=False,
        )
        os_request.refresh_from_db()
        self.assertEqual(os_request.status, "reprovada")
        self.assertFalse(OS.objects.filter(solicitacao=os_request).exists())


class AuditoriaTests(BaseUserTestCase):
    def setUp(self):
        self.centro = self.create_center()
        self.lider = self.create_user("lideraudit", "lider_contrato")
        self.gerente_lider = self.create_user("glaudit", "gerente_lider")
        self.coordenador = self.create_user("coordaudit", "coordenador")
        self.coordenador.centros.add(self.centro)
        self.gerente_lider.centros.add(self.centro)
        self.contrato_base = self.create_contract(
            codigo="PRJ-AUD",
            coordenador=self.coordenador,
            lider_contrato=self.lider,
        )
        self.contrato_terceiro = self.create_supplier_contract(
            cod_projeto=self.contrato_base,
            coordenador=self.coordenador,
            lider_contrato=self.lider,
            guarda_chuva=True,
            num_contrato="CT-AUD-1",
        )

    def xest_criacao_via_view_gera_registro_auditoria(self):
        self.client.force_login(self.lider)

        response = self.client.post(
            reverse("solicitar_os_com_contrato"),
            {
                "contrato": self.contrato_terceiro.pk,
                "cod_projeto": self.contrato_base.pk,
                "titulo": "OS auditada",
                "descricao": "Criada com auditoria",
                "valor_previsto": "R$ 999,99",
                "prazo_execucao": "2026-04-30",
                "coordenador": self.coordenador.pk,
            },
            follow=False,
        )

        self.assertEqual(response.status_code, 302)
        os_request = SolicitacaoOrdemServico.objects.get(titulo="OS auditada")
        auditoria = RegistroAuditoria.objects.filter(
            object_id=os_request.pk,
            acao="criado",
            modelo="Solicitação Ordem Servico",
        ).first()
        self.assertIsNotNone(auditoria)
        self.assertEqual(auditoria.usuario, self.lider)
        self.assertEqual(auditoria.detalhes, "POST /solicitar-os/")

    def xest_atualizacao_via_view_gera_registro_auditoria(self):
        os_cadastrada = self.create_os(
            contrato=self.contrato_terceiro,
            coordenador=self.coordenador,
            lider_contrato=self.lider,
            prazo_execucao=date(2026, 4, 20),
        )
        self.client.force_login(self.gerente_lider)

        response = self.client.post(
            reverse("registrar_entrega_os", args=[os_cadastrada.pk]),
            {
                "caminho_evidencia": "C:/audit/os.pdf",
                "avaliacao": "Aprovado",
                "data_entrega": "2026-04-21",
                "realizado": "on",
                "valor_pago": "350.00",
                "data_pagamento": "2026-04-22",
                "observacao": "Atualização auditada",
            },
            follow=False,
        )

        self.assertEqual(response.status_code, 302)
        auditoria = RegistroAuditoria.objects.filter(
            object_id=os_cadastrada.pk,
            acao="atualizado",
            modelo="Os",
        ).first()
        self.assertIsNotNone(auditoria)
        self.assertEqual(auditoria.usuario, self.gerente_lider)
        self.assertEqual(auditoria.detalhes, f"POST /ordem-servico/{os_cadastrada.pk}/registrar-entrega/")
        self.os.refresh_from_db()
        self.assertTrue(self.os.realizado)
        self.assertTrue(self.os.com_atraso)
        self.assertEqual(self.os.status, "finalizada")
        self.assertEqual(self.os.data_entrega, date(2026, 4, 22))

    def xest_registrar_entrega_os_bloqueia_gerente_lider_sem_centro_compartilhado(self):
        gerente_lider_fora = self.create_user("glfora", "gerente_lider")
        gerente_lider_fora.centros.add(self.outro_centro)
        self.client.force_login(gerente_lider_fora)

        response = self.client.get(reverse("registrar_entrega_os", args=[self.os.pk]), follow=False)

        self.assertRedirects(
            response,
            reverse("detalhes_os", kwargs={"pk": self.os.pk}),
            fetch_redirect_response=False,
        )


class AuditoriaViewTests(BaseUserTestCase):
    def setUp(self):
        self.centro = self.create_center()
        self.lider = self.create_user("lideraudit2", "lider_contrato")
        self.gerente_lider = self.create_user("glaudit2", "gerente_lider")
        self.coordenador = self.create_user("coordaudit2", "coordenador")
        self.coordenador.centros.add(self.centro)
        self.gerente_lider.centros.add(self.centro)
        self.contrato_base = self.create_contract(
            codigo="PRJ-AUD2",
            coordenador=self.coordenador,
            lider_contrato=self.lider,
        )
        self.contrato_terceiro = self.create_supplier_contract(
            cod_projeto=self.contrato_base,
            coordenador=self.coordenador,
            lider_contrato=self.lider,
            guarda_chuva=True,
            num_contrato="CT-AUD-2",
        )

    def test_criacao_via_view_gera_registro_auditoria(self):
        self.client.force_login(self.lider)

        response = self.client.post(
            reverse("solicitar_os_com_contrato"),
            {
                "contrato": self.contrato_terceiro.pk,
                "cod_projeto": self.contrato_base.pk,
                "titulo": "OS auditada 2",
                "descricao": "Criada com auditoria",
                "valor_previsto": "R$ 999,99",
                "prazo_execucao": "2026-04-30",
                "coordenador": self.coordenador.pk,
            },
            follow=False,
        )

        self.assertEqual(response.status_code, 302)
        os_request = SolicitacaoOrdemServico.objects.get(titulo="OS auditada 2")
        auditoria = RegistroAuditoria.objects.filter(
            object_id=os_request.pk,
            acao="criado",
            usuario=self.lider,
            detalhes="POST /solicitar-os/",
        ).first()
        self.assertIsNotNone(auditoria)

    def test_atualizacao_via_view_gera_registro_auditoria(self):
        os_cadastrada = self.create_os(
            contrato=self.contrato_terceiro,
            coordenador=self.coordenador,
            lider_contrato=self.lider,
            prazo_execucao=date(2026, 4, 20),
        )
        self.client.force_login(self.gerente_lider)

        response = self.client.post(
            reverse("registrar_entrega_os", args=[os_cadastrada.pk]),
            {
                "caminho_evidencia": "C:/audit/os.pdf",
                "avaliacao": "Aprovado",
                "data_entrega": "2026-04-21",
                "realizado": "on",
                "valor_pago": "350.00",
                "data_pagamento": "2026-04-22",
                "observacao": "Atualização auditada",
            },
            follow=False,
        )

        self.assertEqual(response.status_code, 302)
        auditoria = RegistroAuditoria.objects.filter(
            object_id=os_cadastrada.pk,
            acao="atualizado",
            usuario=self.gerente_lider,
            detalhes=f"POST /ordem-servico/{os_cadastrada.pk}/registrar-entrega/",
        ).first()
        self.assertIsNotNone(auditoria)
        os_cadastrada.refresh_from_db()
        self.assertTrue(os_cadastrada.realizado)
        self.assertTrue(os_cadastrada.com_atraso)
        self.assertEqual(os_cadastrada.status, "finalizada")


class ContratacaoFlowTests(BaseUserTestCase):
    def setUp(self):
        self.centro = self.create_center()
        self.lider = self.create_user("lidercontr", "lider_contrato")
        self.gerente_lider = self.create_user("glcontr", "gerente_lider")
        self.gerente_contrato = self.create_user("gccontr", "gerente_contrato")
        self.diretoria = self.create_user("dircontr", "diretoria")
        self.suprimento = self.create_user("supcontr", "suprimento")
        self.coordenador = self.create_user("coordcontr", "coordenador")
        self.coordenador.centros.add(self.centro)
        self.gerente_lider.centros.add(self.centro)
        self.contrato_base = self.create_contract(
            codigo="PRJ-CONTR",
            coordenador=self.coordenador,
            lider_contrato=self.lider,
        )
        self.fornecedor = self.create_supplier(nome="Fornecedor Contratação")
        self.solicitacao_contrato = SolicitacaoContrato.objects.create(
            contrato=self.contrato_base,
            coordenador=self.coordenador,
            lider_contrato=self.lider,
            descricao="Solicitação de contratação",
            fornecedor_escolhido=self.fornecedor,
            data_inicio=date(2026, 5, 1),
            data_fim=date(2026, 6, 1),
            valor_provisionado=Decimal("1500.00"),
            status="Solicitação de contratação",
        )
        self.solicitacao_prospeccao = SolicitacaoProspeccao.objects.create(
            contrato=self.contrato_base,
            coordenador=self.coordenador,
            lider_contrato=self.lider,
            descricao="Solicitação de prospecção",
            fornecedor_escolhido=self.fornecedor,
            data_inicio=date(2026, 5, 1),
            data_fim=date(2026, 6, 1),
            status="Fornecedor aprovado",
        )

    def test_detalhes_solicitacao_contrato_aprovacao_completa_define_status_final(self):
        self.client.force_login(self.gerente_contrato)
        response_gerente = self.client.post(
            reverse("detalhes_solicitacao_contrato", args=[self.solicitacao_contrato.pk]),
            {"acao": "aprovar"},
            follow=False,
        )

        self.assertRedirects(
            response_gerente,
            reverse("detalhes_solicitacao_contrato", kwargs={"pk": self.solicitacao_contrato.pk}),
            fetch_redirect_response=False,
        )
        self.solicitacao_contrato.refresh_from_db()
        self.assertEqual(self.solicitacao_contrato.aprovacao_fornecedor_gerente, "aprovado")

        self.client.force_login(self.diretoria)
        response_diretoria = self.client.post(
            reverse("detalhes_solicitacao_contrato", args=[self.solicitacao_contrato.pk]),
            {"acao": "aprovar"},
            follow=False,
        )

        self.assertRedirects(
            response_diretoria,
            reverse("detalhes_solicitacao_contrato", kwargs={"pk": self.solicitacao_contrato.pk}),
            fetch_redirect_response=False,
        )
        self.solicitacao_contrato.refresh_from_db()
        self.assertEqual(self.solicitacao_contrato.aprovacao_fornecedor_diretor, "aprovado")
        self.assertEqual(self.solicitacao_contrato.status, "Fornecedor aprovado")

    def test_detalhes_solicitacao_contrato_reprovacao_sem_justificativa_redireciona_para_mesma_tela(self):
        self.client.force_login(self.diretoria)

        response = self.client.post(
            reverse("detalhes_solicitacao_contrato", args=[self.solicitacao_contrato.pk]),
            {"acao": "reprovar", "justificativa": ""},
            follow=False,
        )

        self.assertRedirects(
            response,
            reverse("detalhes_solicitacao_contrato", kwargs={"pk": self.solicitacao_contrato.pk}),
            fetch_redirect_response=False,
        )

    def test_cadastrar_minuta_contrato_cria_documento_para_solicitacao_contrato(self):
        self.client.force_login(self.suprimento)

        response = self.client.post(
            reverse("cadastrar_minuta_contrato", args=[self.solicitacao_contrato.pk]),
            {
                "numero_contrato": "MIN-CONTR-001",
                "objeto": "Objeto da contratação",
                "valor_total": "1.500,00",
                "observacao": "Minuta inicial",
            },
            follow=False,
        )

        self.assertRedirects(response, reverse("lista_solicitacoes"), fetch_redirect_response=False)
        documento = self.solicitacao_contrato.minuta_contrato
        self.assertEqual(documento.numero_contrato, "MIN-CONTR-001")
        self.assertEqual(documento.prazo_inicio, date(2026, 5, 1))
        self.assertEqual(documento.prazo_fim, date(2026, 6, 1))
        self.assertEqual(documento.valor_total, Decimal("1500.00"))
        self.solicitacao_contrato.refresh_from_db()
        self.assertEqual(self.solicitacao_contrato.status, "Planejamento do contrato")

    def test_cadastrar_contrato_cria_documento_para_solicitacao_prospeccao(self):
        self.client.force_login(self.suprimento)

        response = self.client.post(
            reverse("cadastrar_contrato", args=[self.solicitacao_prospeccao.pk]),
            {
                "numero_contrato": "MIN-PROS-001",
                "objeto": "Objeto da prospecção",
                "valor_total": "2.345,67",
                "observacao": "Contrato de prospecção",
            },
            follow=False,
        )

        self.assertRedirects(
            response,
            reverse("detalhes_solicitacao", kwargs={"pk": self.solicitacao_prospeccao.pk}),
            fetch_redirect_response=False,
        )
        documento = self.solicitacao_prospeccao.contrato_relacionado
        self.assertEqual(documento.numero_contrato, "MIN-PROS-001")
        self.assertEqual(documento.prazo_inicio, date(2026, 5, 1))
        self.assertEqual(documento.prazo_fim, date(2026, 6, 1))
        self.assertEqual(documento.valor_total, Decimal("2345.67"))


class AutomaticContractCreationTests(BaseUserTestCase):
    def setUp(self):
        self.centro = self.create_center()
        self.lider = self.create_user("liderauto", "lider_contrato")
        self.gerente_contrato = self.create_user("gcauto", "gerente_contrato")
        self.suprimento = self.create_user("supauto", "suprimento")
        self.coordenador = self.create_user("coordauto", "coordenador")
        self.coordenador.centros.add(self.centro)
        self.contrato_base = self.create_contract(
            codigo="PRJ-AUTO",
            coordenador=self.coordenador,
            lider_contrato=self.lider,
        )
        self.fornecedor = self.create_supplier(nome="Fornecedor Auto")

    def test_criar_contrato_se_aprovado_gera_contrato_para_prospeccao(self):
        solicitacao = SolicitacaoProspeccao.objects.create(
            contrato=self.contrato_base,
            coordenador=self.coordenador,
            lider_contrato=self.lider,
            fornecedor_escolhido=self.fornecedor,
            descricao="Prospecção aprovada",
            status="Aprovação Final",
            aprovacao_gerencia=True,
        )
        DocumentoBM.objects.create(
            solicitacao=solicitacao,
            status_gerente="aprovado",
        )
        DocumentoContratoTerceiro.objects.create(
            solicitacao=solicitacao,
            numero_contrato="AUTO-PROS-001",
            objeto="Objeto automático",
            prazo_inicio=date(2026, 5, 1),
            prazo_fim=date(2026, 6, 1),
            valor_total=Decimal("3210.00"),
        )
        PropostaFornecedor.objects.create(
            solicitacao=solicitacao,
            fornecedor=self.fornecedor,
            condicao_pagamento="30",
        )
        evento = Evento.objects.create(
            prospeccao=solicitacao,
            empresa_terceira=self.fornecedor,
            descricao="Entrega prospecção",
            data_prevista=date(2026, 5, 10),
            valor_previsto=Decimal("100.00"),
        )

        contrato = criar_contrato_se_aprovado(solicitacao)

        self.assertIsNotNone(contrato)
        self.assertEqual(contrato.prospeccao, solicitacao)
        self.assertEqual(contrato.empresa_terceira, self.fornecedor)
        self.assertEqual(contrato.num_contrato, "AUTO-PROS-001")
        self.assertEqual(contrato.valor_total, Decimal("3210.00"))
        self.assertEqual(contrato.status, "ativo")
        solicitacao.refresh_from_db()
        self.assertEqual(solicitacao.status, "Onboarding")
        evento.refresh_from_db()
        self.assertEqual(evento.contrato_terceiro, contrato)
        self.assertEqual(len(mail.outbox), 1)

    def test_criar_contrato_se_aprovado_minuta_gera_contrato_para_contratacao(self):
        solicitacao = SolicitacaoContrato.objects.create(
            contrato=self.contrato_base,
            coordenador=self.coordenador,
            lider_contrato=self.lider,
            fornecedor_escolhido=self.fornecedor,
            descricao="Contratação aprovada",
            status="Aprovação Final",
            aprovacao_gerencia=True,
            guarda_chuva=True,
            valor_provisionado=Decimal("4567.00"),
        )
        DocumentoBM.objects.create(
            solicitacao_contrato=solicitacao,
            status_gerente="aprovado",
        )
        DocumentoContratoTerceiro.objects.create(
            solicitacao_contrato=solicitacao,
            numero_contrato="AUTO-CONTR-001",
            objeto="Objeto contratação",
            prazo_inicio=date(2026, 7, 1),
            prazo_fim=date(2026, 8, 1),
            valor_total=Decimal("4567.00"),
        )
        PropostaFornecedor.objects.create(
            solicitacao_contrato=solicitacao,
            fornecedor=self.fornecedor,
            condicao_pagamento="60",
        )
        evento = Evento.objects.create(
            solicitacao_contrato=solicitacao,
            empresa_terceira=self.fornecedor,
            descricao="Entrega contratação",
            data_prevista=date(2026, 7, 10),
            valor_previsto=Decimal("100.00"),
        )

        contrato = criar_contrato_se_aprovado_minuta(solicitacao)

        self.assertIsNotNone(contrato)
        self.assertEqual(contrato.solicitacao, solicitacao)
        self.assertEqual(contrato.empresa_terceira, self.fornecedor)
        self.assertEqual(contrato.num_contrato, "AUTO-CONTR-001")
        self.assertTrue(contrato.guarda_chuva)
        self.assertEqual(contrato.valor_total, Decimal("4567.00"))
        solicitacao.refresh_from_db()
        self.assertEqual(solicitacao.status, "Onboarding")
        evento.refresh_from_db()
        self.assertEqual(evento.contrato_terceiro, contrato)
        self.assertEqual(len(mail.outbox), 1)


class TemplateVisibilityTests(BaseUserTestCase):
    def setUp(self):
        self.suprimento = self.create_user("supritemplate", "suprimento")
        self.lider = self.create_user("lidertemplate", "lider_contrato")

    def test_base_mostra_report_apenas_para_suprimento(self):
        self.client.force_login(self.suprimento)
        response = self.client.get(reverse("guia_permissoes"))
        self.assertContains(response, "Report")

        self.client.force_login(self.lider)
        response = self.client.get(reverse("guia_permissoes"))
        self.assertNotContains(response, "Report")

    def test_lista_ordens_servico_mostra_botao_solicitar_os_para_lider(self):
        self.client.force_login(self.lider)
        response = self.client.get(reverse("lista_ordens_servico"))
        self.assertContains(response, "Solicitar Nova OS")


class AditivoContratoTerceiroTests(BaseUserTestCase):
    def setUp(self):
        self.centro = self.create_center()
        self.lider = self.create_user("lider_aditivo", "lider_contrato")
        self.gerente_lider = self.create_user("gl_aditivo", "gerente_lider")
        self.gerente_contrato = self.create_user("gc_aditivo", "gerente_contrato")
        self.diretoria = self.create_user("dir_aditivo", "diretoria")
        self.suprimento = self.create_user("sup_aditivo", "suprimento")
        self.coordenador = self.create_user("coord_aditivo", "coordenador")
        self.coordenador.centros.add(self.centro)
        self.gerente_lider.centros.add(self.centro)

        self.contrato_base = self.create_contract(
            codigo="PRJ-ADITIVO",
            coordenador=self.coordenador,
            lider_contrato=self.lider,
            valor_total=Decimal("1000.00"),
        )
        self.contrato_terceiro = self.create_supplier_contract(
            cod_projeto=self.contrato_base,
            coordenador=self.coordenador,
            lider_contrato=self.lider,
            num_contrato="CT-ADITIVO",
        )
        self.contrato_terceiro.valor_total = Decimal("1000.00")
        self.contrato_terceiro.data_fim = date(2026, 5, 31)
        self.contrato_terceiro.save(update_fields=["valor_total", "data_fim"])

    def create_aditivo(self):
        return AditivoContratoTerceiro.objects.create(
            contrato=self.contrato_terceiro,
            solicitado_por=self.lider,
            motivo="Aumentar escopo do contrato",
            valor_total_anterior=Decimal("1000.00"),
            novo_valor_total=Decimal("1500.00"),
            data_fim_anterior=date(2026, 5, 31),
            nova_data_fim=date(2026, 6, 30),
        )

    def test_helper_permite_solicitacao_para_grupos_esperados(self):
        self.assertTrue(can_user_request_contract_addendum(self.lider, self.contrato_terceiro))
        self.assertTrue(can_user_request_contract_addendum(self.gerente_lider, self.contrato_terceiro))
        self.assertTrue(can_user_request_contract_addendum(self.gerente_contrato, self.contrato_terceiro))
        self.assertFalse(can_user_request_contract_addendum(self.suprimento, self.contrato_terceiro))

    def test_detalhe_contrato_exibe_botao_pedir_aditivo_para_lider(self):
        self.client.force_login(self.lider)
        response = self.client.get(reverse("contrato_fornecedor_detalhe", kwargs={"pk": self.contrato_terceiro.pk}))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse("solicitar_aditivo_contrato", args=[self.contrato_terceiro.pk]))

    def test_gerente_lider_do_mesmo_centro_pode_aprovar_como_lideranca(self):
        aditivo = self.create_aditivo()
        aditivo.arquivo_aditivo = SimpleUploadedFile("aditivo.pdf", b"conteudo", content_type="application/pdf")
        aditivo.save()
        self.client.force_login(self.gerente_lider)

        response = self.client.post(
            reverse("avaliar_aditivo_contrato", args=[aditivo.pk]),
            {"acao": "aprovar_lider"},
            follow=False,
        )

        self.assertEqual(response.status_code, 302)
        aditivo.refresh_from_db()
        self.assertEqual(aditivo.status_lider, "aprovado")

    def test_gerente_lider_fora_do_centro_nao_pode_solicitar_aditivo(self):
        gerente_lider_fora = self.create_user("gl_aditivo_fora", "gerente_lider")
        outro_centro = self.create_center(codigo="CT2", nome="Centro 2")
        gerente_lider_fora.centros.add(outro_centro)
        self.client.force_login(gerente_lider_fora)

        response = self.client.get(
            reverse("solicitar_aditivo_contrato", args=[self.contrato_terceiro.pk]),
            follow=False,
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response.url,
            reverse("contrato_fornecedor_detalhe", args=[self.contrato_terceiro.pk]),
        )

    def test_solicitar_aditivo_notifica_suprimento(self):
        self.client.force_login(self.lider)
        response = self.client.post(
            reverse("solicitar_aditivo_contrato", args=[self.contrato_terceiro.pk]),
            {
                "motivo": "Prorrogar prazo por demanda adicional",
                "novo_valor_total": "1800.00",
                "nova_data_fim": "2026-07-15",
            },
            follow=False,
        )

        self.assertEqual(response.status_code, 302)
        aditivo = AditivoContratoTerceiro.objects.get(contrato=self.contrato_terceiro)
        self.assertEqual(aditivo.solicitado_por, self.lider)
        self.assertEqual(len(mail.outbox), 1)
        self.assertIn(self.suprimento.email, mail.outbox[0].to)

    def test_suprimento_envia_documento_e_reseta_fluxo(self):
        aditivo = self.create_aditivo()
        aditivo.status_lider = "reprovado"
        aditivo.justificativa_reprovacao_lider = "Arquivo ilegivel"
        aditivo.save(update_fields=["status_lider", "justificativa_reprovacao_lider"])

        self.client.force_login(self.suprimento)
        response = self.client.post(
            reverse("enviar_documento_aditivo_contrato", args=[aditivo.pk]),
            {
                "arquivo_aditivo": SimpleUploadedFile("aditivo.pdf", b"conteudo", content_type="application/pdf"),
            },
            follow=False,
        )

        self.assertEqual(response.status_code, 302)
        aditivo.refresh_from_db()
        self.assertTrue(bool(aditivo.arquivo_aditivo))
        self.assertEqual(aditivo.status_lider, "pendente")
        self.assertEqual(aditivo.status_gerente, "pendente")
        self.assertEqual(aditivo.status_diretoria, "pendente")

    def test_reprovacao_como_lider_exige_justificativa(self):
        aditivo = self.create_aditivo()
        aditivo.arquivo_aditivo = SimpleUploadedFile("aditivo.pdf", b"conteudo", content_type="application/pdf")
        aditivo.save()
        self.client.force_login(self.lider)

        response = self.client.post(
            reverse("avaliar_aditivo_contrato", args=[aditivo.pk]),
            {"acao": "reprovar_lider", "justificativa": ""},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        aditivo.refresh_from_db()
        self.assertEqual(aditivo.status_lider, "pendente")
        self.assertContains(response, "justificativa")

    def test_aprovacao_do_lider_dispensa_gerente(self):
        aditivo = self.create_aditivo()
        aditivo.arquivo_aditivo = SimpleUploadedFile("aditivo.pdf", b"conteudo", content_type="application/pdf")
        aditivo.save()

        self.client.force_login(self.lider)
        response = self.client.post(
            reverse("avaliar_aditivo_contrato", args=[aditivo.pk]),
            {"acao": "aprovar_lider"},
            follow=False,
        )

        self.assertEqual(response.status_code, 302)
        aditivo.refresh_from_db()
        self.assertEqual(aditivo.status_lider, "aprovado")
        self.assertEqual(aditivo.status_gerente, "pendente")

        self.client.force_login(self.gerente_contrato)
        response = self.client.post(
            reverse("avaliar_aditivo_contrato", args=[aditivo.pk]),
            {"acao": "aprovar_gerente"},
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        aditivo.refresh_from_db()
        self.assertEqual(aditivo.status_gerente, "pendente")

    def test_aprovacao_do_gerente_dispensa_lider(self):
        aditivo = self.create_aditivo()
        aditivo.arquivo_aditivo = SimpleUploadedFile("aditivo.pdf", b"conteudo", content_type="application/pdf")
        aditivo.save()
        self.client.force_login(self.gerente_contrato)

        response = self.client.post(
            reverse("avaliar_aditivo_contrato", args=[aditivo.pk]),
            {"acao": "aprovar_gerente"},
            follow=False,
        )

        self.assertEqual(response.status_code, 302)
        aditivo.refresh_from_db()
        self.assertEqual(aditivo.status_lider, "pendente")
        self.assertEqual(aditivo.status_gerente, "aprovado")

    def test_diretoria_aprova_e_atualiza_contrato(self):
        aditivo = self.create_aditivo()
        aditivo.arquivo_aditivo = SimpleUploadedFile("aditivo.pdf", b"conteudo", content_type="application/pdf")
        aditivo.status_lider = "aprovado"
        aditivo.save()
        self.client.force_login(self.diretoria)

        response = self.client.post(
            reverse("avaliar_aditivo_contrato", args=[aditivo.pk]),
            {"acao": "aprovar_diretoria"},
            follow=False,
        )

        self.assertEqual(response.status_code, 302)
        aditivo.refresh_from_db()
        self.contrato_terceiro.refresh_from_db()
        self.assertEqual(aditivo.status_diretoria, "aprovado")
        self.assertEqual(self.contrato_terceiro.valor_total, Decimal("1500.00"))
        self.assertEqual(self.contrato_terceiro.data_fim, date(2026, 6, 30))

    def test_detalhe_contrato_exibe_download_do_aditivo_aprovado(self):
        aditivo = self.create_aditivo()
        aditivo.arquivo_aditivo = SimpleUploadedFile("aditivo.pdf", b"conteudo", content_type="application/pdf")
        aditivo.status_lider = "aprovado"
        aditivo.status_diretoria = "aprovado"
        aditivo.documento_enviado_em = timezone.now()
        aditivo.save()

        self.client.force_login(self.suprimento)
        response = self.client.get(reverse("contrato_fornecedor_detalhe", kwargs={"pk": self.contrato_terceiro.pk}))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Documento do aditivo")
        self.assertContains(response, aditivo.arquivo_aditivo.url)


class DocmGenerationTests(BaseUserTestCase):
    def create_docm_template(self, document_text, header_text=""):
        temp_file = tempfile.NamedTemporaryFile(suffix=".docm", delete=False)
        temp_file.close()
        document_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
            '<w:body>'
            f'<w:p><w:r><w:t>{document_text}</w:t></w:r></w:p>'
            '</w:body>'
            '</w:document>'
        )
        header_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<w:hdr xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
            f'<w:p><w:r><w:t>{header_text}</w:t></w:r></w:p>'
            '</w:hdr>'
        )
        with zipfile.ZipFile(temp_file.name, "w") as docm_file:
            docm_file.writestr("word/document.xml", document_xml)
            docm_file.writestr("word/header1.xml", header_xml)
        return temp_file.name

    def read_docm_xml(self, response, filename):
        with zipfile.ZipFile(BytesIO(response.content)) as docm_file:
            return docm_file.read(filename).decode("utf-8")

    def setUp(self):
        self.suprimento = self.create_user("supr_docm", "suprimento")
        self.coordenador = self.create_user("coord_docm", "coordenador")
        self.lider = self.create_user("lider_docm", "lider_contrato")
        self.cliente = self.create_client("Cliente Docm", "00.000.000/0001-55")
        self.fornecedor = self.create_supplier("Fornecedor Docm", "11.111.111/0001-55")
        self.fornecedor.endereco = "Rua das Flores"
        self.fornecedor.numero = "123"
        self.fornecedor.bairro = "Centro"
        self.fornecedor.municipio = "Belo Horizonte"
        self.fornecedor.estado = "MG"
        self.fornecedor.cep = "30000-000"
        self.fornecedor.telefone = "31 99999-9999"
        self.fornecedor.email = "fornecedor.docm@example.com"
        self.fornecedor.ponto_focal = "Maria Silva"
        self.fornecedor.telefone_focal = "31 98888-8888"
        self.fornecedor.email_focal = "maria.silva@example.com"
        self.fornecedor.informacoes_bancarias = "Banco XPTO Ag 1234 Conta 56789"
        self.fornecedor.save()

        self.contrato_base = self.create_contract(
            codigo="PRJ-DOCM",
            cliente=self.cliente,
            coordenador=self.coordenador,
            lider_contrato=self.lider,
            valor_total=Decimal("1500.00"),
        )
        self.solicitacao_prospeccao = SolicitacaoProspeccao.objects.create(
            contrato=self.contrato_base,
            coordenador=self.coordenador,
            lider_contrato=self.lider,
            descricao="Servico especializado",
            data_inicio=date(2026, 5, 1),
            data_fim=date(2026, 5, 31),
            fornecedor_escolhido=self.fornecedor,
            status="Planejamento do contrato",
        )
        PropostaFornecedor.objects.create(
            solicitacao=self.solicitacao_prospeccao,
            fornecedor=self.fornecedor,
            valor_proposta=Decimal("1500.00"),
            arquivo_proposta=SimpleUploadedFile("proposta_tecnica.pdf", b"pdf"),
        )
        self.solicitacao_contrato = SolicitacaoContrato.objects.create(
            contrato=self.contrato_base,
            coordenador=self.coordenador,
            lider_contrato=self.lider,
            descricao="Servico contratado",
            data_inicio=date(2026, 6, 1),
            data_fim=date(2026, 6, 30),
            fornecedor_escolhido=self.fornecedor,
            valor_provisionado=Decimal("2500.00"),
            status="Planejamento do contrato",
        )
        PropostaFornecedor.objects.create(
            solicitacao_contrato=self.solicitacao_contrato,
            fornecedor=self.fornecedor,
            valor_proposta=Decimal("2500.00"),
            arquivo_proposta=SimpleUploadedFile("proposta_contrato.pdf", b"pdf"),
        )
        self.contrato_fornecedor = self.create_supplier_contract(
            cod_projeto=self.contrato_base,
            empresa_terceira=self.fornecedor,
            coordenador=self.coordenador,
            lider_contrato=self.lider,
            status="ativo",
            num_contrato="CT-DOCM-01",
        )
        self.contrato_fornecedor.data_inicio = date(2026, 5, 1)
        self.contrato_fornecedor.data_fim = date(2026, 5, 31)
        self.contrato_fornecedor.objeto = "Servico especializado"
        self.contrato_fornecedor.valor_total = Decimal("1500.00")
        self.contrato_fornecedor.save()
        self.aditivo = AditivoContratoTerceiro.objects.create(
            contrato=self.contrato_fornecedor,
            solicitado_por=self.lider,
            motivo="Prorrogacao contratual",
            valor_total_anterior=Decimal("1500.00"),
            novo_valor_total=Decimal("2500.00"),
            data_fim_anterior=date(2026, 5, 31),
            nova_data_fim=date(2026, 6, 30),
        )

    def test_gerar_minuta_contrato_docm_substitui_placeholders(self):
        template_path = self.create_docm_template(
            "__nome_empresa_terceira__ __descricao__ __ valor_proposta__ __valor_proposta_extenso__ __contrato__ __dias_totais__ __data_inicio__ __data_fim__ __documento_proposta__ ☒ ESPECÍFICO    ☐ GUARDA-CHUVA",
            "__numero_revisao__ __data_hoje__ __cod_contrato__",
        )
        self.client.force_login(self.suprimento)

        with patch("gestao_contratos.views.CONTRACT_TEMPLATE_DOCM_PATH", Path(template_path)):
            response = self.client.post(
                reverse("gerar_minuta_contrato_docm", args=[self.solicitacao_prospeccao.id]),
                {
                    "numero_contrato": "CONT-DOCM-001",
                    "objeto": "Servico especializado",
                    "valor_total": "1.500,00",
                    "observacao": "Observacao teste",
                },
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/vnd.ms-word.document.macroEnabled.12")
        self.assertIn("Contrato_CONT-DOCM-001.docm", response["Content-Disposition"])
        document_xml = self.read_docm_xml(response, "word/document.xml")
        header_xml = self.read_docm_xml(response, "word/header1.xml")
        self.assertIn("FORNECEDOR DOCM", document_xml)
        self.assertIn("Servico especializado", document_xml)
        self.assertIn("R$ 1.500,00", document_xml)
        self.assertIn("mil e quinhentos reais", document_xml)
        self.assertIn("PRJ-DOCM", document_xml)
        self.assertIn("31 dias", document_xml)
        self.assertIn("01/05/2026", document_xml)
        self.assertIn("31/05/2026", document_xml)
        self.assertIn("proposta_tecnica", document_xml)
        self.assertIn(".pdf", document_xml)
        self.assertIn("☒ ESPECÍFICO    ☐ GUARDA-CHUVA", document_xml)
        self.assertIn("CONT-DOCM-001", header_xml)

    def test_gerar_minuta_contrato_docm_marca_guarda_chuva_quando_for_o_caso(self):
        self.solicitacao_prospeccao.guarda_chuva = True
        self.solicitacao_prospeccao.save(update_fields=["guarda_chuva"])
        template_path = self.create_docm_template(
            "☒ ESPECÍFICO    ☐ GUARDA-CHUVA",
            "__cod_contrato__",
        )
        self.client.force_login(self.suprimento)

        with patch("gestao_contratos.views.CONTRACT_TEMPLATE_DOCM_PATH", Path(template_path)):
            response = self.client.post(
                reverse("gerar_minuta_contrato_docm", args=[self.solicitacao_prospeccao.id]),
                {
                    "numero_contrato": "CONT-DOCM-GC",
                    "objeto": "Servico especializado",
                    "valor_total": "1.500,00",
                    "observacao": "Observacao teste",
                },
            )

        self.assertEqual(response.status_code, 200)
        document_xml = self.read_docm_xml(response, "word/document.xml")
        self.assertIn("☐ ESPECÍFICO    ☒ GUARDA-CHUVA", document_xml)

    def test_gerar_minuta_contrato_contratacao_docm_substitui_placeholders(self):
        template_path = self.create_docm_template(
            "__nome_empresa_terceira__ __descricao__ __ valor_proposta__ __contrato__ __data_inicio__ __data_fim__",
            "__cod_contrato__",
        )
        self.client.force_login(self.suprimento)

        with patch("gestao_contratos.views.CONTRACT_TEMPLATE_DOCM_PATH", Path(template_path)):
            response = self.client.post(
                reverse("gerar_minuta_contrato_contratacao_docm", args=[self.solicitacao_contrato.id]),
                {
                    "numero_contrato": "CONT-DOCM-002",
                    "objeto": "Servico contratado",
                    "valor_total": "2.500,00",
                    "observacao": "Observacao teste",
                },
            )

        self.assertEqual(response.status_code, 200)
        document_xml = self.read_docm_xml(response, "word/document.xml")
        self.assertIn("FORNECEDOR DOCM", document_xml)
        self.assertIn("R$ 2.500,00", document_xml)
        self.assertIn("PRJ-DOCM", document_xml)
        self.assertIn("01/06/2026", document_xml)
        self.assertIn("30/06/2026", document_xml)

    def test_gerar_aditivo_contrato_docm_substitui_placeholders(self):
        template_path = self.create_docm_template(
            "__ordem_aditivo__ __numero_contrato__ __data_fim__ __data_fim_aditivo__ __descricao__ __novo_valor _total__ __novo_valor _total_extenso__ __contrato__ __informacoes_bancarias__ __dias_totais_novo__ __data_inicio__ __nova_data_fim__ __data_hoje_completo__ __nome_empresa__",
            "__ordem_aditivo__ __numero_contrato__ __numero_revisao__ __data_hoje__",
        )
        self.client.force_login(self.suprimento)

        with patch("gestao_contratos.views.ADDENDUM_TEMPLATE_DOCM_PATH", Path(template_path)):
            response = self.client.get(reverse("gerar_aditivo_contrato_docm", args=[self.aditivo.id]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/vnd.ms-word.document.macroEnabled.12")
        document_xml = self.read_docm_xml(response, "word/document.xml")
        header_xml = self.read_docm_xml(response, "word/header1.xml")
        self.assertIn("1º", document_xml)
        self.assertIn("CT-DOCM-01", document_xml)
        self.assertIn("31/05/2026", document_xml)
        self.assertIn("30/06/2026", document_xml)
        self.assertIn("Servico especializado", document_xml)
        self.assertIn("R$ 2.500,00", document_xml)
        self.assertIn("dois mil e quinhentos reais", document_xml)
        self.assertIn("PRJ-DOCM", document_xml)
        self.assertIn("BANCO XPTO AG 1234 CONTA 56789", document_xml)
        self.assertIn("61 dias", document_xml)
        self.assertIn("FORNECEDOR DOCM", document_xml)
        self.assertIn("CT-DOCM-01", header_xml)

