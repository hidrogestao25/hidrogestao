from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib import messages
from django.conf import settings
from django.contrib.contenttypes.models import ContentType
from django.core.mail import send_mail, EmailMultiAlternatives
from django.core.paginator import Paginator
from django.db.models import Sum, Q, DecimalField, Avg, Prefetch, Count, Max, Min
from decimal import Decimal
from django.db.models.functions import Coalesce, Greatest
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.utils import timezone
from django.utils.html import escape, strip_tags
from django.views.generic.edit import CreateView
from .models import Contrato, Cliente, EmpresaTerceira, ContratoTerceiros, SolicitacaoProspeccao, Indicadores, PropostaFornecedor, DocumentoContratoTerceiro, DocumentoBM, Evento, CalendarioPagamento, BM, NF, AvaliacaoFornecedor, NFCliente, SolicitacaoOrdemServico, OS, SolicitacaoContrato, AditivoContratoTerceiro, RegistroAuditoria
from .forms import ContratoForm, ClienteForm, FornecedorForm, ContratoFornecedorForm, SolicitacaoProspeccaoForm, DocumentoContratoTerceiroForm, DocumentoBMForm, EventoPrevisaoForm, EventoEntregaForm, FiltroPrevisaoForm, BMForm, NFForm, NFClienteForm, SolicitacaoOrdemServicoForm, UploadContratoOSForm, RegistroEntregaOSForm, OrdemServicoForm, SolicitacaoContratoForm, ContratoModalForm, SolicitacaoGuardaChuvaForm, SolicitacaoAditivoContratoTerceiroForm, DocumentoAditivoContratoTerceiroForm, DocumentoAditivoAssinadoContratoTerceiroForm

import plotly.graph_objs as go
import plotly.express as px
import pandas as pd
from plotly.offline import plot
import plotly.colors as pc

import os
from datetime import date
import zipfile
import openpyxl
from django.http import HttpResponse, JsonResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.styles import Font
from io import BytesIO
from datetime import datetime, timedelta
from pathlib import Path
import xml.etree.ElementTree as ET
from collections import Counter

User = get_user_model()
FROM_EMAIL = settings.DEFAULT_FROM_EMAIL
CONTRACT_TEMPLATE_DOCM_PATH = Path(settings.MEDIA_ROOT) / "modelos_word" / "Modelo Contrato.docm"
ADDENDUM_TEMPLATE_DOCM_PATH = Path(settings.MEDIA_ROOT) / "modelos_word" / "Modelo Aditivo.docm"
SIGNED_FILES_PENDING_STATUS = "Aguardando Arquivos Assinados"


def format_date_br(value):
    if not value:
        return "-"
    return value.strftime("%d/%m/%Y")


def format_date_long_br(value):
    if not value:
        return "-"
    meses = [
        "janeiro", "fevereiro", "marco", "abril", "maio", "junho",
        "julho", "agosto", "setembro", "outubro", "novembro", "dezembro",
    ]
    return f"{value.day} de {meses[value.month - 1]} de {value.year}"


def format_currency_br(value, with_symbol=False):
    if value in [None, ""]:
        value = Decimal("0.00")
    if not isinstance(value, Decimal):
        value = Decimal(value)
    formatted = f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {formatted}" if with_symbol else formatted


def get_signed_files_pending_status():
    return SIGNED_FILES_PENDING_STATUS


def notify_supply_about_signed_files_request(solicitacao):
    suprimentos = User.objects.filter(grupo="suprimento").exclude(email__isnull=True).exclude(email__exact="")
    lista_emails = [u.email for u in suprimentos]
    if not lista_emails:
        return

    contrato_referencia = getattr(solicitacao, "contrato", None)
    fornecedor = getattr(solicitacao, "fornecedor_escolhido", None)
    assunto = f"Solicitacao #{solicitacao.id} - Inserir arquivos assinados"
    etapa_aprovada = (
        "A minuta do contrato já foi aprovada pela gerência de contrato.\n"
        if getattr(solicitacao, "guarda_chuva", False)
        else "A minuta do contrato e a minuta do BM já foram aprovadas pela gerência de contrato.\n"
    )
    mensagem = (
        "Olá, equipe de Suprimento!\n\n"
        f"{etapa_aprovada}"
        "Agora é necessário inserir o contrato assinado para concluir a geração do contrato.\n\n"
        f"Solicitação: {solicitacao.id}\n"
        f"Projeto: {contrato_referencia or '-'}\n"
        f"Fornecedor: {fornecedor or '-'}\n\n"
        "Acesse o sistema HIDROGestão para anexar o contrato assinado.\n"
        "https://hidrogestao.pythonanywhere.com/"
    )
    try:
        send_mail(assunto, mensagem, FROM_EMAIL, lista_emails, fail_silently=False)
    except Exception:
        pass


def get_contract_completion_notification_emails(solicitacao):
    emails = set()

    def add_email(value):
        if value:
            emails.add(value)

    for grupo in ["suprimento", "gerente_contrato", "diretoria"]:
        for email in User.objects.filter(grupo=grupo).exclude(email__isnull=True).exclude(email__exact="").values_list("email", flat=True):
            add_email(email)

    coordenador = getattr(solicitacao, "coordenador", None)
    lider_contrato = getattr(solicitacao, "lider_contrato", None)

    if coordenador and coordenador.email:
        add_email(coordenador.email)
        centros_ids = list(coordenador.centros.values_list("id", flat=True))
        if centros_ids:
            gerente_tecnico_qs = User.objects.filter(
                grupo__in=["gerente", "gerente_lider"],
                centros__in=centros_ids,
            ).exclude(email__isnull=True).exclude(email__exact="").distinct()
            for email in gerente_tecnico_qs.values_list("email", flat=True):
                add_email(email)

    if lider_contrato and lider_contrato.email:
        add_email(lider_contrato.email)

    return sorted(emails)


def get_addendum_completion_notification_emails(aditivo):
    emails = set()

    def add_email(value):
        if value:
            emails.add(value)

    for grupo in ["suprimento", "gerente_contrato", "diretoria"]:
        for email in User.objects.filter(grupo=grupo).exclude(email__isnull=True).exclude(email__exact="").values_list("email", flat=True):
            add_email(email)

    contrato = aditivo.contrato
    if contrato.lider_contrato:
        add_email(contrato.lider_contrato.email)
    if aditivo.solicitado_por:
        add_email(aditivo.solicitado_por.email)

    gerente_lider_qs = User.objects.filter(grupo="gerente_lider", is_active=True)
    if contrato.coordenador_id:
        gerente_lider_qs = gerente_lider_qs.filter(
            Q(centros__in=contrato.coordenador.centros.all())
            | Q(centros__in=contrato.coordenadores.values_list("centros__id", flat=True))
        ).distinct()
    for email in gerente_lider_qs.exclude(email__isnull=True).exclude(email__exact="").values_list("email", flat=True):
        add_email(email)

    return sorted(emails)


def notify_contract_process_completed(solicitacao, contrato):
    lista_emails = get_contract_completion_notification_emails(solicitacao)
    if not lista_emails:
        return

    contrato_path = reverse_lazy("contrato_fornecedor_detalhe", kwargs={"pk": contrato.pk})
    contrato_url = f"https://hidrogestao.pythonanywhere.com{contrato_path}"
    projeto_label = "-"
    if getattr(contrato, "cod_projeto", None):
        projeto_label = getattr(contrato.cod_projeto, "cod_projeto", None) or str(contrato.cod_projeto)
    elif getattr(solicitacao, "contrato", None):
        projeto_label = getattr(solicitacao.contrato, "cod_projeto", None) or str(solicitacao.contrato)
    elif contrato.num_contrato:
        projeto_label = contrato.num_contrato
    elif getattr(contrato, "guarda_chuva", False):
        projeto_label = "Guarda-chuva"

    assunto = f"Processo concluido - Contrato gerado para {projeto_label}"
    mensagem = (
        "Prezados,\n\n"
        "Informamos que o processo de solicitacao foi concluido com sucesso, "
        "e o contrato foi gerado apos a insercao dos arquivos assinados.\n\n"
        f"Solicitação: {solicitacao.id}\n"
        f"Código do Projeto: {projeto_label}\n"
        f"Fornecedor: {contrato.empresa_terceira}\n"
        f"Número do Contrato: {contrato.num_contrato or '-'}\n"
        f"Valor Total: R$ {contrato.valor_total:,.2f}\n"
        f"Vigência: {contrato.data_inicio.strftime('%d/%m/%Y') if contrato.data_inicio else 'Não definida'} "
        f"a {contrato.data_fim.strftime('%d/%m/%Y') if contrato.data_fim else 'Não definida'}\n\n"
        "O processo esta finalizado no sistema e o contrato encontra-se disponivel para consulta e prosseguimento dos encaminhamentos.\n\n"
        f"Link direto para o contrato gerado:\n{contrato_url}\n\n"
        "Atenciosamente,\n"
        "Sistema de Gestão de Terceiros - HIDROGestão"
    )
    try:
        send_mail(assunto, mensagem, FROM_EMAIL, lista_emails, fail_silently=False)
    except Exception:
        pass


def update_signed_files_pending_status(solicitacao, bm=None, documento=None):
    if bm is None:
        bm = getattr(solicitacao, "minuta_boletins_medicao", None)
        if bm is None:
            bm = getattr(solicitacao, "minuta_boletins_medicao_contrato", None)
    if documento is None:
        documento = getattr(solicitacao, "contrato_relacionado", None)
        if documento is None:
            documento = getattr(solicitacao, "minuta_contrato", None)

    if not documento:
        return False

    if getattr(solicitacao, "guarda_chuva", False):
        fluxo_aprovado = solicitacao.aprovacao_gerencia is True
    else:
        fluxo_aprovado = bool(bm and bm.status_gerente == "aprovado" and solicitacao.aprovacao_gerencia is True)

    if fluxo_aprovado:
        if documento.arquivo_contrato_assinado:
            return False
        if solicitacao.status != SIGNED_FILES_PENDING_STATUS:
            solicitacao.status = SIGNED_FILES_PENDING_STATUS
            solicitacao.save(update_fields=["status"])
            notify_supply_about_signed_files_request(solicitacao)
        return True
    return False


def number_to_words_pt_br(number):
    units = ["zero", "um", "dois", "tres", "quatro", "cinco", "seis", "sete", "oito", "nove"]
    teens = ["dez", "onze", "doze", "treze", "quatorze", "quinze", "dezesseis", "dezessete", "dezoito", "dezenove"]
    tens = ["", "", "vinte", "trinta", "quarenta", "cinquenta", "sessenta", "setenta", "oitenta", "noventa"]
    hundreds = ["", "cento", "duzentos", "trezentos", "quatrocentos", "quinhentos", "seiscentos", "setecentos", "oitocentos", "novecentos"]

    def under_thousand(n):
        if n == 0:
            return ""
        if n == 100:
            return "cem"
        parts = []
        c = n // 100
        rem = n % 100
        if c:
            parts.append(hundreds[c])
        if rem:
            if rem < 10:
                parts.append(units[rem])
            elif rem < 20:
                parts.append(teens[rem - 10])
            else:
                d = rem // 10
                u = rem % 10
                parts.append(tens[d] if u == 0 else f"{tens[d]} e {units[u]}")
        return " e ".join([part for part in parts if part])

    def as_words(n):
        if n == 0:
            return "zero"
        if n < 1000:
            return under_thousand(n)
        scales = [
            (1_000_000_000, "bilhao", "bilhoes"),
            (1_000_000, "milhao", "milhoes"),
            (1000, "mil", "mil"),
        ]
        parts = []
        remainder = n
        for value, singular, plural in scales:
            amount = remainder // value
            remainder %= value
            if not amount:
                continue
            if value == 1000:
                parts.append("mil" if amount == 1 else f"{under_thousand(amount)} mil")
            else:
                label = singular if amount == 1 else plural
                parts.append(f"{under_thousand(amount) if amount < 1000 else as_words(amount)} {label}")
        if remainder:
            parts.append(under_thousand(remainder))
        if len(parts) == 1:
            return parts[0]
        return ", ".join(parts[:-1]) + " e " + parts[-1]

    return as_words(int(number))


def decimal_to_money_words_pt_br(value):
    if value in [None, ""]:
        value = Decimal("0.00")
    if not isinstance(value, Decimal):
        value = Decimal(value)
    quantized = value.quantize(Decimal("0.01"))
    inteiro = int(quantized)
    centavos = int((quantized - Decimal(inteiro)) * 100)

    partes = []
    if inteiro:
        partes.append(f"{number_to_words_pt_br(inteiro)} {'real' if inteiro == 1 else 'reais'}")
    else:
        partes.append("zero real")
    if centavos:
        partes.append(f"{number_to_words_pt_br(centavos)} {'centavo' if centavos == 1 else 'centavos'}")
    return " e ".join(partes)


def calculate_inclusive_days(start_date, end_date):
    if not start_date or not end_date:
        return "-"
    return f"{(end_date - start_date).days + 1} dias"


def get_selected_supplier_proposal(solicitacao=None, solicitacao_contrato=None, fornecedor=None):
    filtros = {"fornecedor": fornecedor}
    if solicitacao is not None:
        filtros["solicitacao"] = solicitacao
    if solicitacao_contrato is not None:
        filtros["solicitacao_contrato"] = solicitacao_contrato
    return PropostaFornecedor.objects.filter(**filtros).first()


def build_contract_document_preview(request_post, existing_document=None):
    documento = existing_document or DocumentoContratoTerceiro()
    documento.numero_contrato = (
        request_post.get("numero_contrato")
        or getattr(existing_document, "numero_contrato", "")
        or ""
    )
    documento.objeto = (
        request_post.get("objeto")
        or getattr(existing_document, "objeto", "")
        or ""
    )
    documento.observacao = request_post.get("observacao", getattr(existing_document, "observacao", "") or "")

    valor_total_raw = request_post.get("valor_total")
    if valor_total_raw not in [None, ""]:
        try:
            documento.valor_total = Decimal(
                str(valor_total_raw).replace(".", "").replace(",", ".")
            )
        except Exception:
            documento.valor_total = getattr(existing_document, "valor_total", Decimal("0.00")) or Decimal("0.00")
    else:
        documento.valor_total = getattr(existing_document, "valor_total", Decimal("0.00")) or Decimal("0.00")

    return documento


def replace_placeholders_in_docm(template_path, replacements):
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    output_buffer = BytesIO()
    with zipfile.ZipFile(template_path, "r") as source_docm, zipfile.ZipFile(output_buffer, "w") as target_docm:
        for file_info in source_docm.infolist():
            content = source_docm.read(file_info.filename)
            if file_info.filename.startswith("word/") and file_info.filename.endswith(".xml"):
                root = ET.fromstring(content)
                changed = False
                for paragraph in root.findall(".//w:p", namespace):
                    text_nodes = paragraph.findall(".//w:t", namespace)
                    if not text_nodes:
                        continue
                    original_text = "".join(node.text or "" for node in text_nodes)
                    updated_text = original_text
                    for placeholder, replacement in replacements.items():
                        updated_text = updated_text.replace(placeholder, replacement)
                    if updated_text != original_text:
                        text_nodes[0].text = updated_text
                        for node in text_nodes[1:]:
                            node.text = ""
                        changed = True
                if changed:
                    content = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            target_docm.writestr(file_info, content)
    output_buffer.seek(0)
    return output_buffer.getvalue()

GROUP_GUIDE = {
    "coordenador": {
        "label": "Coordenador de Contrato",
        "summary": "Acompanha a execução técnica dos contratos sob sua responsabilidade e consulta os registros ligados ao seu centro/contrato.",
        "must_do": [
            "Acompanhar informações técnicas do contrato e dos fornecedores vinculados.",
            "Apoiar o fluxo de entregas, eventos e ordens de serviço quando for o responsável direto.",
            "Consultar o andamento das solicitações e dos contratos ligados à sua operação.",
        ],
        "can_do": [
            "Visualizar contratos, fornecedores, solicitações, eventos e ordens de serviço dentro do seu escopo.",
            "Cadastrar e editar eventos em fluxos permitidos pelo sistema.",
            "Consultar status, documentos e históricos das solicitações relacionadas aos seus contratos.",
        ],
        "permissions": [
            "Acesso restrito ao que estiver ligado ao contrato/registro em que atua.",
            "Não aprova fornecedor, BM ou minuta como papel gerencial.",
        ],
    },
    "lider_contrato": {
        "label": "Líder de Contrato",
        "summary": "Conduz o fluxo operacional das contratações e da execução dos contratos sob sua liderança.",
        "must_do": [
            "Abrir solicitações de prospecção, contratação e ordem de serviço quando necessário.",
            "Escolher o fornecedor indicado para a solicitação.",
            "Registrar entrega de evento/OS, avaliar fornecedor e aprovar ou reprovar BM no papel de líder.",
        ],
        "can_do": [
            "Acompanhar solicitações, contratos, eventos, BMs e OS vinculados aos contratos em que é líder.",
            "Atuar em renegociação de prazo/valor e no acompanhamento das minutas.",
            "Visualizar indicadores e documentos do fornecedor/contrato sob sua gestão.",
        ],
        "permissions": [
            "Atua apenas nos registros do próprio contrato/liderança.",
            "Pode aprovar fornecedor e BM como papel de liderança, mas não substitui a diretoria.",
        ],
    },
    "gerente": {
        "label": "Gerente Técnico",
        "summary": "Acompanha a operação gerencial e a carteira sob sua responsabilidade.",
        "must_do": [
            "Monitorar solicitações, contratos, eventos e ordens de serviço do seu escopo.",
            "Acompanhar atrasos, entregas e indicadores operacionais.",
        ],
        "can_do": [
            "Visualizar registros vinculados à sua estrutura gerencial.",
            "Acompanhar dashboards, pagamentos e informações consolidadas disponíveis ao grupo.",
        ],
        "permissions": [
            "Escopo limitado ao que a regra de negócio libera para o grupo gerente.",
            "Não assume automaticamente o papel de líder de contrato ou gerente de contrato.",
        ],
    },
    "gerente_lider": {
        "label": "Gerente e Líder de Contratos",
        "summary": "Atua como liderança contratual e gerencial, mas somente sobre registros de coordenadores que compartilham centro com ele.",
        "must_do": [
            "Abrir solicitações de prospecção, contratação e ordem de serviço do seu escopo.",
            "Escolher fornecedor, registrar entregas e avaliar fornecedor quando o líder técnico estiver em centro compartilhado.",
            "Aprovar ou reprovar BM e atuar como líder no fluxo de aprovação do fornecedor dentro do seu escopo por centro.",
        ],
        "can_do": [
            "Consultar e atuar em contratos, solicitações, eventos, OS e BMs ligados a líderes técnicos com centro em comum.",
            "Substituir o líder de contrato na operação dentro desse mesmo recorte de centros.",
            "Acompanhar documentos, minutas, indicadores e andamento das entregas do seu escopo.",
        ],
        "permissions": [
            "Pode ter mais de um centro.",
            "Só atua quando há interseção entre os centros dele e o centro do líder técnico responsável pelo registro.",
        ],
    },
    "gerente_contrato": {
        "label": "Gerente de Contratos",
        "summary": "Atua como camada gerencial do fluxo contratual e pode substituir o líder de contrato nas etapas previstas.",
        "must_do": [
            "Avaliar e aprovar fornecedor como papel gerencial do contrato.",
            "Atuar na aprovação/reprovação de BM e nas minutas quando o fluxo exigir o papel de gerente de contrato.",
            "Dar suporte ao líder de contratos em entregas, avaliações e decisões contratuais.",
        ],
        "can_do": [
            "Abrir solicitações de prospecção, contratação e ordem de serviço.",
            "Escolher fornecedor e registrar entrega/avaliação quando estiver atuando em substituição ao líder.",
            "Receber notificações e acompanhar o fluxo geral das contratações.",
        ],
        "permissions": [
            "Substitui o líder de contrato nas etapas em que o sistema prevê essa atuação.",
            "Compartilha aprovações com diretoria nas etapas finais do fluxo de fornecedor e contrato.",
        ],
    },
    "diretoria": {
        "label": "Diretoria",
        "summary": "Executa aprovações estratégicas e acompanha a carteira consolidada do sistema.",
        "must_do": [
            "Aprovar ou reprovar fornecedor nas etapas finais que exigem diretoria.",
            "Acompanhar solicitações, contratos e indicadores estratégicos.",
        ],
        "can_do": [
            "Visualizar registros amplos do sistema liberados ao grupo.",
            "Receber notificações das novas solicitações e das aprovações finais.",
        ],
        "permissions": [
            "Atua principalmente como aprovador final e patrocinador da operação.",
            "Não substitui automaticamente os papéis operacionais do dia a dia.",
        ],
    },
    "financeiro": {
        "label": "Financeiro",
        "summary": "Controla documentos financeiros, pagamentos e acompanhamento de notas fiscais.",
        "must_do": [
            "Acompanhar BMs aprovados, notas fiscais e pagamentos.",
            "Executar os registros financeiros ligados às entregas e contratos.",
        ],
        "can_do": [
            "Visualizar previsões de pagamento, contratos e registros financeiros permitidos ao grupo.",
            "Cadastrar e editar dados financeiros conforme o fluxo do sistema.",
        ],
        "permissions": [
            "Foco no fluxo financeiro; não conduz as aprovações técnicas/contratuais do líder.",
        ],
    },
    "suprimento": {
        "label": "Suprimento",
        "summary": "Conduz o processo de prospecção, triagem, cadastro e documentação contratual.",
        "must_do": [
            "Receber as solicitações abertas pelos grupos responsáveis e realizar a triagem/cadastro necessário.",
            "Cadastrar propostas, contratos, minutas, BMs e fornecedores conforme o fluxo.",
            "Enviar e acompanhar o report operacional semanal.",
        ],
        "can_do": [
            "Acessar fornecedores, clientes, contratos, solicitações e documentos de suporte.",
            "Avançar o fluxo operacional após as aprovações de liderança/gerência/diretoria.",
            "Receber notificações automáticas das novas solicitações.",
        ],
        "permissions": [
            "É o grupo operacional central do processo de contratação e documentação.",
            "Não substitui as aprovações formais de líder, gerente de contrato ou diretoria.",
        ],
    },
    "fornecedor": {
        "label": "Fornecedor",
        "summary": "Participa do processo apenas no escopo externo liberado pela aplicação.",
        "must_do": [
            "Consultar e responder ao que for disponibilizado no fluxo do fornecedor.",
        ],
        "can_do": [
            "Acessar somente o que o sistema expuser para o perfil fornecedor.",
        ],
        "permissions": [
            "Escopo restrito e sem poderes internos de aprovação.",
        ],
    },
}


def user_shares_center_with_coordinator(user, coordinator):
    if not user or not coordinator:
        return False

    user_centros = getattr(user, "centros", None)
    coordinator_centros = getattr(coordinator, "centros", None)
    if user_centros is None or coordinator_centros is None:
        return False

    return coordinator_centros.filter(
        id__in=user_centros.values_list("id", flat=True)
    ).exists()


def user_is_contract_coordinator(user, contract):
    if not user or not contract:
        return False
    if getattr(contract, "coordenador_id", None) == user.id:
        return True
    coordenadores = getattr(contract, "coordenadores", None)
    if coordenadores is None:
        return False
    return coordenadores.filter(pk=user.pk).exists()


def user_shares_center_with_contract_coordinators(user, contract):
    if not user or not contract:
        return False
    if getattr(contract, "coordenador", None) and user_shares_center_with_coordinator(user, contract.coordenador):
        return True
    coordenadores = getattr(contract, "coordenadores", None)
    if coordenadores is None:
        return False
    user_centros = getattr(user, "centros", None)
    if user_centros is None:
        return False
    return coordenadores.filter(
        centros__in=user_centros.all()
    ).exists()


def send_request_notification_to_management(subject, message):
    recipients = list(
        User.objects.filter(grupo__in=["diretoria", "gerente_contrato"])
        .exclude(email__isnull=True)
        .exclude(email__exact="")
        .values_list("email", flat=True)
        .distinct()
    )

    if recipients:
        send_mail(subject, message, FROM_EMAIL, recipients, fail_silently=False)


def build_latest_audit_map(model_class, object_ids):
    if not object_ids:
        return {}

    content_type = ContentType.objects.get_for_model(model_class)
    audits = RegistroAuditoria.objects.filter(
        content_type=content_type,
        object_id__in=object_ids,
    ).order_by("object_id", "-data_hora")

    latest_audits = {}
    for audit in audits:
        latest_audits.setdefault(audit.object_id, audit)
    return latest_audits


def build_bm_approval_audit_map(bms):
    if not bms:
        return {}

    content_type = ContentType.objects.get_for_model(BM)
    bm_ids = [bm.id for bm in bms]
    audits = (
        RegistroAuditoria.objects.filter(
            content_type=content_type,
            object_id__in=bm_ids,
            usuario__isnull=False,
        )
        .select_related("usuario")
        .order_by("object_id", "data_hora")
    )

    audits_by_bm = {}
    for audit in audits:
        audits_by_bm.setdefault(audit.object_id, []).append(audit)

    approval_map = {}

    def resolve_nearest_audit(audit_list, approval_datetime):
        if not audit_list or not approval_datetime:
            return None

        matching_audits = [
            audit
            for audit in audit_list
            if abs((audit.data_hora - approval_datetime).total_seconds()) <= 300
        ]
        if not matching_audits:
            return None

        return min(
            matching_audits,
            key=lambda audit: abs((audit.data_hora - approval_datetime).total_seconds()),
        )

    for bm in bms:
        audit_list = audits_by_bm.get(bm.id, [])
        approval_map[bm.id] = {
            "coordenador": resolve_nearest_audit(audit_list, bm.data_aprovacao_coordenador),
            "gerente": resolve_nearest_audit(audit_list, bm.data_aprovacao_gerente),
        }

    return approval_map


def average_days_from_pairs(pairs):
    valid_deltas = []
    for start, end in pairs:
        if not start or not end:
            continue
        delta = end - start
        valid_deltas.append(delta.total_seconds() / 86400)
    if not valid_deltas:
        return None
    return round(sum(valid_deltas) / len(valid_deltas), 1)


def build_supply_retrabalho_queryset():
    retrabalho_prospeccao = SolicitacaoProspeccao.objects.filter(
        Q(solicitacao_origem__isnull=False)
        | Q(nenhum_fornecedor_ideal=True)
        | Q(aprovacao_fornecedor_gerente="reprovado")
        | Q(aprovacao_fornecedor_diretor="reprovado")
        | Q(reprovacao_gerencia=True)
    )
    retrabalho_contratacao = SolicitacaoContrato.objects.filter(
        Q(aprovacao_fornecedor_gerente="reprovado")
        | Q(aprovacao_fornecedor_diretor="reprovado")
        | Q(reprovacao_gerencia=True)
    )
    return retrabalho_prospeccao, retrabalho_contratacao


def is_supply_retrabalho_request(solicitacao):
    if isinstance(solicitacao, SolicitacaoProspeccao):
        return any(
            [
                bool(solicitacao.solicitacao_origem_id),
                bool(solicitacao.nenhum_fornecedor_ideal),
                solicitacao.aprovacao_fornecedor_gerente == "reprovado",
                solicitacao.aprovacao_fornecedor_diretor == "reprovado",
                bool(solicitacao.reprovacao_gerencia),
            ]
        )

    if isinstance(solicitacao, SolicitacaoContrato):
        return any(
            [
                solicitacao.aprovacao_fornecedor_gerente == "reprovado",
                solicitacao.aprovacao_fornecedor_diretor == "reprovado",
                bool(solicitacao.reprovacao_gerencia),
            ]
        )

    return False


def build_supply_user_indicator_rows(
    solicitacoes_prospeccao,
    solicitacoes_contratacao,
    prospeccao_onboarding_map,
    contratacao_onboarding_map,
):
    rows = {}

    def ensure_row(user):
        row = rows.get(user.pk)
        if row is None:
            row = {
                "usuario_id": user.pk,
                "usuario_nome": user.get_full_name() or user.username,
                "grupo": user.get_grupo_display() if user.grupo else "-",
                "papeis": set(),
                "solicitacoes_vinculadas": 0,
                "em_aberto": 0,
                "concluidas": 0,
                "retrabalho": 0,
                "lead_times": [],
            }
            rows[user.pk] = row
        return row

    def register_request(solicitacao, onboarding_map):
        if not solicitacao.data_solicitacao:
            return

        if solicitacao.pk not in onboarding_map:
            return

        participants = {}

        if solicitacao.coordenador_id:
            participants.setdefault(
                solicitacao.coordenador_id,
                {"user": solicitacao.coordenador, "roles": set()},
            )["roles"].add("Coordenador")

        if solicitacao.lider_contrato_id:
            participants.setdefault(
                solicitacao.lider_contrato_id,
                {"user": solicitacao.lider_contrato, "roles": set()},
            )["roles"].add("Lider do fluxo")

        if not participants:
            return

        onboarding_audit = onboarding_map.get(solicitacao.pk) if solicitacao.status == "Onboarding" else None
        lead_time_days = None
        if onboarding_audit:
            lead_time_days = round((onboarding_audit.data_hora - solicitacao.data_solicitacao).total_seconds() / 86400, 1)

        retrabalho = is_supply_retrabalho_request(solicitacao)
        concluida = solicitacao.status == "Onboarding"

        for participant in participants.values():
            row = ensure_row(participant["user"])
            row["papeis"].update(participant["roles"])
            row["solicitacoes_vinculadas"] += 1
            if concluida:
                row["concluidas"] += 1
                if lead_time_days is not None:
                    row["lead_times"].append(lead_time_days)
            else:
                row["em_aberto"] += 1

            if retrabalho:
                row["retrabalho"] += 1

    for solicitacao in solicitacoes_prospeccao:
        register_request(solicitacao, prospeccao_onboarding_map)

    for solicitacao in solicitacoes_contratacao:
        register_request(solicitacao, contratacao_onboarding_map)

    finalized_rows = []
    for row in rows.values():
        finalized_rows.append(
            {
                "usuario_id": row["usuario_id"],
                "usuario_nome": row["usuario_nome"],
                "grupo": row["grupo"],
                "papeis": ", ".join(sorted(row["papeis"])),
                "solicitacoes_vinculadas": row["solicitacoes_vinculadas"],
                "em_aberto": row["em_aberto"],
                "concluidas": row["concluidas"],
                "retrabalho": row["retrabalho"],
                "media_prazo": round(sum(row["lead_times"]) / len(row["lead_times"]), 1) if row["lead_times"] else None,
            }
        )

    return sorted(
        finalized_rows,
        key=lambda item: (
            -item["solicitacoes_vinculadas"],
            -item["em_aberto"],
            item["usuario_nome"].lower(),
        ),
    )


def filter_records_with_date_and_audit(records, audit_map, date_attr):
    filtered_records = []
    for record in records:
        if not getattr(record, date_attr, None):
            continue
        if record.pk not in audit_map:
            continue
        filtered_records.append(record)
    return filtered_records


def build_os_average_pairs(os_records, os_audit_map):
    pairs = []
    for solicitacao_os in os_records:
        if not solicitacao_os.criado_em:
            continue
        if solicitacao_os.pk not in os_audit_map:
            continue
        if not is_request_concluded(solicitacao_os):
            continue
        pairs.append((solicitacao_os.criado_em, os_audit_map[solicitacao_os.pk].data_hora))
    return pairs


@login_required
def guia_permissoes(request):
    current_group = getattr(request.user, "grupo", None)
    current_group_guide = GROUP_GUIDE.get(current_group)
    can_view_all_groups = current_group == "suprimento"
    current_group_item = None
    all_group_items = []

    for key, label in User.GRUPOS_CHOICES:
        guide = GROUP_GUIDE.get(key)
        if not guide:
            continue

        item = {
            "key": key,
            "label": label,
            "guide": guide,
            "is_current": key == current_group,
        }

        if key == current_group:
            current_group_item = item

        if can_view_all_groups:
            all_group_items.append(item)

    return render(
        request,
        "gestao_contratos/guia_permissoes.html",
        {
            "current_group": current_group,
            "current_group_guide": current_group_guide,
            "current_group_item": current_group_item,
            "can_view_all_groups": can_view_all_groups,
            "group_guides": all_group_items,
        },
    )


@login_required
def indicadores_suprimento(request):
    if getattr(request.user, "grupo", None) != "suprimento":
        messages.error(request, "Voce nao tem permissao para acessar os indicadores de suprimento.")
        return redirect("home")

    solicitacoes_prospeccao = SolicitacaoProspeccao.objects.select_related(
        "contrato",
        "fornecedor_escolhido",
    )
    solicitacoes_contratacao = SolicitacaoContrato.objects.select_related(
        "contrato",
        "fornecedor_escolhido",
    )
    contratos_fornecedor = list(
        ContratoTerceiros.objects.select_related(
            "cod_projeto",
            "empresa_terceira",
            "prospeccao",
            "solicitacao",
        )
    )

    prospeccoes_onboarding = list(solicitacoes_prospeccao.filter(status="Onboarding"))
    contratacoes_onboarding = list(solicitacoes_contratacao.filter(status="Onboarding"))
    prospeccao_audit_map = build_latest_audit_map(
        SolicitacaoProspeccao,
        list(solicitacoes_prospeccao.values_list("pk", flat=True)),
    )
    contratacao_audit_map = build_latest_audit_map(
        SolicitacaoContrato,
        list(solicitacoes_contratacao.values_list("pk", flat=True)),
    )
    os_queryset = SolicitacaoOrdemServico.objects.all()
    os_audit_map = build_latest_audit_map(
        SolicitacaoOrdemServico,
        list(os_queryset.values_list("pk", flat=True)),
    )
    prospeccao_onboarding_map = build_latest_audit_map(
        SolicitacaoProspeccao,
        [solicitacao.pk for solicitacao in prospeccoes_onboarding],
    )
    contratacao_onboarding_map = build_latest_audit_map(
        SolicitacaoContrato,
        [solicitacao.pk for solicitacao in contratacoes_onboarding],
    )

    prospeccao_pairs = []
    contratacao_pairs = []
    latest_completed = []

    for contrato in contratos_fornecedor:
        if contrato.prospeccao and contrato.prospeccao.data_solicitacao:
            onboarding_audit = prospeccao_onboarding_map.get(contrato.prospeccao.pk)
            if onboarding_audit:
                prospeccao_pairs.append((contrato.prospeccao.data_solicitacao, onboarding_audit.data_hora))
                latest_completed.append(
                    {
                        "tipo": "Prospeccao",
                        "solicitacao_id": contrato.prospeccao.pk,
                        "projeto": str(contrato.cod_projeto) if contrato.cod_projeto else "-",
                        "fornecedor": str(contrato.empresa_terceira),
                        "dias": round((onboarding_audit.data_hora - contrato.prospeccao.data_solicitacao).total_seconds() / 86400, 1),
                        "concluido_em": onboarding_audit.data_hora,
                    }
                )

        if contrato.solicitacao and contrato.solicitacao.data_solicitacao:
            onboarding_audit = contratacao_onboarding_map.get(contrato.solicitacao.pk)
            if onboarding_audit:
                contratacao_pairs.append((contrato.solicitacao.data_solicitacao, onboarding_audit.data_hora))
                latest_completed.append(
                    {
                        "tipo": "Contratacao",
                        "solicitacao_id": contrato.solicitacao.pk,
                        "projeto": str(contrato.cod_projeto) if contrato.cod_projeto else "-",
                        "fornecedor": str(contrato.empresa_terceira),
                        "dias": round((onboarding_audit.data_hora - contrato.solicitacao.data_solicitacao).total_seconds() / 86400, 1),
                        "concluido_em": onboarding_audit.data_hora,
                    }
                )

    media_prazo_prospeccao = average_days_from_pairs(prospeccao_pairs)
    media_prazo_contratacao = average_days_from_pairs(contratacao_pairs)
    media_prazo_geral = average_days_from_pairs(prospeccao_pairs + contratacao_pairs)

    backlog_status_excluded = ["Onboarding", "Reprovada pelo suprimento"]
    backlog_prospeccao = solicitacoes_prospeccao.exclude(status__in=backlog_status_excluded)
    backlog_contratacao = solicitacoes_contratacao.exclude(status__in=backlog_status_excluded)
    backlog_os = SolicitacaoOrdemServico.objects.exclude(status__in=["aprovada", "reprovada", "finalizada"])
    backlog_guarda_chuva = backlog_prospeccao.filter(guarda_chuva=True).count() + backlog_contratacao.filter(guarda_chuva=True).count()
    backlog_total = backlog_prospeccao.count() + backlog_contratacao.count() + backlog_os.count()

    guarda_chuva_pairs = []
    for solicitacao in prospeccoes_onboarding:
        if solicitacao.guarda_chuva:
            onboarding_audit = prospeccao_onboarding_map.get(solicitacao.pk)
            if onboarding_audit and solicitacao.data_solicitacao:
                guarda_chuva_pairs.append((solicitacao.data_solicitacao, onboarding_audit.data_hora))
    for solicitacao in contratacoes_onboarding:
        if solicitacao.guarda_chuva:
            onboarding_audit = contratacao_onboarding_map.get(solicitacao.pk)
            if onboarding_audit and solicitacao.data_solicitacao:
                guarda_chuva_pairs.append((solicitacao.data_solicitacao, onboarding_audit.data_hora))

    media_prazo_guarda_chuva = average_days_from_pairs(guarda_chuva_pairs)
    media_prazo_os = average_days_from_pairs(
        build_os_average_pairs(list(os_queryset), os_audit_map)
    )

    retrabalho_prospeccao, retrabalho_contratacao = build_supply_retrabalho_queryset()
    total_solicitacoes = solicitacoes_prospeccao.count() + solicitacoes_contratacao.count()
    total_retrabalho = retrabalho_prospeccao.count() + retrabalho_contratacao.count()
    percentual_retrabalho = round((total_retrabalho / total_solicitacoes) * 100, 1) if total_solicitacoes else 0.0

    contratos_ativos_qs = ContratoTerceiros.objects.exclude(status="encerrado")
    valor_total_contratos_ativos = contratos_ativos_qs.aggregate(total=Coalesce(Sum("valor_total"), Decimal("0.00")))["total"]

    eventos_sem_nf_count = Evento.objects.filter(
        boletins_medicao__aprovacao_pagamento="aprovado",
        nota_fiscal__isnull=True,
    ).filter(
        Q(boletins_medicao__status_coordenador="aprovado")
        | Q(boletins_medicao__status_gerente="aprovado")
    ).distinct().count()

    bms = list(BM.objects.all())
    bms_pendentes_operacionais = sum(1 for bm in bms if bm_is_operationally_pending(bm))
    bms_pendentes_diretoria = sum(
        1
        for bm in bms
        if bm.aprovacao_pagamento == "pendente" and bm_has_operational_approval(bm)
    )

    prospeccao_status_records = filter_records_with_date_and_audit(
        list(solicitacoes_prospeccao),
        prospeccao_audit_map,
        "data_solicitacao",
    )
    contratacao_status_records = filter_records_with_date_and_audit(
        list(solicitacoes_contratacao),
        contratacao_audit_map,
        "data_solicitacao",
    )
    os_status_records = filter_records_with_date_and_audit(
        list(backlog_os),
        os_audit_map,
        "criado_em",
    )

    prospeccao_status_rows = sorted(Counter(record.status for record in prospeccao_status_records).items(), key=lambda item: (-item[1], item[0]))
    contratacao_status_rows = sorted(Counter(record.status for record in contratacao_status_records).items(), key=lambda item: (-item[1], item[0]))
    guarda_chuva_status_counter = Counter()
    for record in prospeccao_status_records:
        if record.guarda_chuva:
            guarda_chuva_status_counter[record.status] += 1
    for record in contratacao_status_records:
        if record.guarda_chuva:
            guarda_chuva_status_counter[record.status] += 1
    guarda_chuva_status_rows = sorted(guarda_chuva_status_counter.items(), key=lambda item: (-item[1], item[0]))
    os_status_rows = sorted(Counter(record.status for record in os_status_records).items(), key=lambda item: (-item[1], item[0]))
    user_indicator_rows = build_supply_user_indicator_rows(
        solicitacoes_prospeccao,
        solicitacoes_contratacao,
        prospeccao_onboarding_map,
        contratacao_onboarding_map,
    )

    latest_completed = sorted(latest_completed, key=lambda item: item["concluido_em"], reverse=True)[:10]

    context = {
        "media_prazo_geral": media_prazo_geral,
        "media_prazo_prospeccao": media_prazo_prospeccao,
        "media_prazo_contratacao": media_prazo_contratacao,
        "total_solicitacoes_prospeccao": len(prospeccao_status_records),
        "total_solicitacoes_contratacao": len(contratacao_status_records),
        "total_solicitacoes_os": len(os_status_records),
        "backlog_total": backlog_total,
        "backlog_os_total": backlog_os.count(),
        "backlog_guarda_chuva_total": backlog_guarda_chuva,
        "media_prazo_os": media_prazo_os,
        "media_prazo_guarda_chuva": media_prazo_guarda_chuva,
        "total_retrabalho": total_retrabalho,
        "percentual_retrabalho": percentual_retrabalho,
        "valor_total_contratos_ativos": valor_total_contratos_ativos,
        "contratos_ativos_total": contratos_ativos_qs.count(),
        "contratos_gerados_total": len(latest_completed),
        "eventos_sem_nf_count": eventos_sem_nf_count,
        "bms_pendentes_operacionais": bms_pendentes_operacionais,
        "bms_pendentes_diretoria": bms_pendentes_diretoria,
        "prospeccao_status_rows": prospeccao_status_rows,
        "contratacao_status_rows": contratacao_status_rows,
        "guarda_chuva_status_rows": guarda_chuva_status_rows,
        "os_status_rows": os_status_rows,
        "user_indicator_rows": user_indicator_rows,
        "latest_completed": latest_completed,
        "total_solicitacoes_guarda_chuva": guarda_chuva_status_counter.total() if hasattr(guarda_chuva_status_counter, "total") else sum(guarda_chuva_status_counter.values()),
    }
    return render(request, "gestao_contratos/indicadores_suprimento.html", context)


def get_week_ranges(reference_date):
    current_week_start = reference_date - timedelta(days=reference_date.weekday())
    current_week_end = current_week_start + timedelta(days=6)
    previous_week_start = current_week_start - timedelta(days=7)
    previous_week_end = current_week_start - timedelta(days=1)
    next_week_start = current_week_start + timedelta(days=7)
    next_week_end = current_week_start + timedelta(days=13)
    return {
        "previous_week_start": previous_week_start,
        "previous_week_end": previous_week_end,
        "current_week_start": current_week_start,
        "current_week_end": current_week_end,
        "next_week_start": next_week_start,
        "next_week_end": next_week_end,
    }


def is_request_concluded(solicitacao):
    if isinstance(solicitacao, SolicitacaoProspeccao):
        return hasattr(solicitacao, "contratoterceiros") or solicitacao.status == "Onboarding"
    if isinstance(solicitacao, SolicitacaoContrato):
        return hasattr(solicitacao, "contratoterceiros") or solicitacao.status == "Onboarding"
    if isinstance(solicitacao, SolicitacaoOrdemServico):
        return solicitacao.status in ["aprovada", "reprovada", "finalizada"] or hasattr(solicitacao, "os")
    return False


def format_request_line(tipo, solicitacao):
    titulo = getattr(solicitacao, "titulo", None) or getattr(solicitacao, "contrato", None) or getattr(solicitacao, "descricao", None) or "-"
    return f"- {tipo} #{solicitacao.id}: {titulo} | status: {solicitacao.status}"


def br_date(value):
    return value.strftime("%d/%m/%Y") if value else "-"


def render_empty(message):
    return f"""
        <tr>
            <td colspan="4" style="padding:12px;color:#777;text-align:center;">
                {escape(message)}
            </td>
        </tr>
    """


def render_section(title, rows):
    return f"""
        <h2 style="font-size:18px;margin:28px 0 12px;color:#0f172a;">
            {escape(title)}
        </h2>

        <table width="100%" cellpadding="0" cellspacing="0"
            style="border-collapse:collapse;border:1px solid #e2e8f0;border-radius:10px;overflow:hidden;margin-bottom:18px;">
            <thead>
                <tr style="background:#f1f5f9;">
                    <th align="left" style="padding:10px;border-bottom:1px solid #e2e8f0;">Tipo / ID</th>
                    <th align="left" style="padding:10px;border-bottom:1px solid #e2e8f0;">Descrição</th>
                    <th align="left" style="padding:10px;border-bottom:1px solid #e2e8f0;">Fornecedor / Info</th>
                    <th align="left" style="padding:10px;border-bottom:1px solid #e2e8f0;">Data / Status</th>
                </tr>
            </thead>
            <tbody>
                {rows}
            </tbody>
        </table>
    """

def build_weekly_supply_report(user):
    today = timezone.localdate()
    weeks = get_week_ranges(today)

    previous_week_requests = []

    for tipo, queryset in [
        (
            "Prospecção",
            SolicitacaoProspeccao.objects.filter(
                data_solicitacao__date__range=(
                    weeks["previous_week_start"],
                    weeks["previous_week_end"],
                )
            ).order_by("data_solicitacao"),
        ),
        (
            "Contratação",
            SolicitacaoContrato.objects.filter(
                data_solicitacao__date__range=(
                    weeks["previous_week_start"],
                    weeks["previous_week_end"],
                )
            ).order_by("data_solicitacao"),
        ),
        (
            "Ordem de Serviço",
            SolicitacaoOrdemServico.objects.filter(
                criado_em__date__range=(
                    weeks["previous_week_start"],
                    weeks["previous_week_end"],
                )
            ).order_by("criado_em"),
        ),
    ]:
        for solicitacao in queryset:
            previous_week_requests.append((tipo, solicitacao))

    concluded_requests = [
        (tipo, s) for tipo, s in previous_week_requests if is_request_concluded(s)
    ]

    pending_requests = [
        (tipo, s) for tipo, s in previous_week_requests if not is_request_concluded(s)
    ]

    all_pending_requests = []

    for tipo, queryset in [
        (
            "Prospecção",
            SolicitacaoProspeccao.objects.all().order_by("data_solicitacao"),
        ),
        (
            "Contratação",
            SolicitacaoContrato.objects.all().order_by("data_solicitacao"),
        ),
        (
            "Ordem de Serviço",
            SolicitacaoOrdemServico.objects.all().order_by("criado_em"),
        ),
    ]:
        for solicitacao in queryset:
            if not is_request_concluded(solicitacao):
                all_pending_requests.append((tipo, solicitacao))

    contratos_finalizados = ContratoTerceiros.objects.filter(
        status="encerrado",
        data_fim__range=(
            weeks["previous_week_start"],
            weeks["previous_week_end"],
        ),
    ).order_by("data_fim")

    eventos_previstos_semana_anterior = Evento.objects.filter(
        data_prevista__range=(
            weeks["previous_week_start"],
            weeks["previous_week_end"],
        )
    ).order_by("data_prevista")

    eventos_entregues_semana_anterior = Evento.objects.filter(
        data_entrega__range=(
            weeks["previous_week_start"],
            weeks["previous_week_end"],
        )
    ).order_by("data_entrega")

    eventos_previstos_semana_atual = Evento.objects.filter(
        data_prevista__range=(
            weeks["current_week_start"],
            weeks["current_week_end"],
        )
    ).order_by("data_prevista")

    contratos_previstos_semana_atual = ContratoTerceiros.objects.filter(
        data_fim__range=(
            weeks["current_week_start"],
            weeks["current_week_end"],
        ),
    ).exclude(status="encerrado").order_by("data_fim")

    eventos_previstos_proxima_semana = Evento.objects.filter(
        data_prevista__range=(
            weeks["next_week_start"],
            weeks["next_week_end"],
        )
    ).order_by("data_prevista")

    contratos_previstos_proxima_semana = ContratoTerceiros.objects.filter(
        data_fim__range=(
            weeks["next_week_start"],
            weeks["next_week_end"],
        ),
    ).exclude(status="encerrado").order_by("data_fim")

    eventos_previstos_nao_entregues = Evento.objects.filter(
        data_prevista__lt=today,
    ).filter(
        Q(realizado=False) | Q(data_entrega__isnull=True)
    ).order_by("data_prevista")

    contratos_atrasados_nao_encerrados = ContratoTerceiros.objects.filter(
        data_fim__lt=today,
    ).exclude(status="encerrado").order_by("data_fim")

    def br_date(value):
        return value.strftime("%d/%m/%Y") if value else "-"

    def render_empty(message, colspan=4):
        return f"""
            <tr>
                <td colspan="{colspan}" style="padding:14px;color:#64748b;text-align:center;border-bottom:1px solid #e2e8f0;">
                    {escape(message)}
                </td>
            </tr>
        """

    def render_table(headers, rows):
        header_html = "".join(
            f"""
            <th align="left" style="padding:10px;background:#f1f5f9;border-bottom:1px solid #e2e8f0;color:#334155;font-size:13px;">
                {escape(header)}
            </th>
            """
            for header in headers
        )

        return f"""
            <table width="100%" cellpadding="0" cellspacing="0"
                style="border-collapse:collapse;border:1px solid #e2e8f0;border-radius:10px;overflow:hidden;margin-top:10px;margin-bottom:18px;">
                <thead>
                    <tr>{header_html}</tr>
                </thead>
                <tbody>
                    {rows}
                </tbody>
            </table>
        """

    def render_period_block(title, period, content, bg_color="#ffffff", border_color="#0f766e"):
        return f"""
            <div style="background:{bg_color};border:1px solid #e2e8f0;border-left:6px solid {border_color};border-radius:14px;padding:22px;margin-bottom:28px;">
                <h2 style="margin:0;color:#0f172a;font-size:22px;">
                    {escape(title)}
                </h2>

                <p style="margin:6px 0 20px;color:#64748b;font-size:14px;">
                    Período: {period}
                </p>

                {content}
            </div>
        """

    def render_summary_cards(items):
        cards = ""

        for label, value, color in items:
            cards += f"""
                <td style="padding:8px;">
                    <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:12px;padding:14px;text-align:center;">
                        <div style="font-size:26px;font-weight:bold;color:{color};">
                            {value}
                        </div>
                        <div style="font-size:13px;color:#475569;">
                            {escape(label)}
                        </div>
                    </div>
                </td>
            """

        return f"""
            <table width="100%" cellpadding="0" cellspacing="0" style="margin-bottom:20px;">
                <tr>{cards}</tr>
            </table>
        """

    def request_rows(requests, include_current_status=False):
        if not requests:
            colspan = 5 if include_current_status else 4
            return render_empty("Nenhum registro encontrado.", colspan=colspan)

        rows = ""

        for tipo, solicitacao in requests:
            status = "Concluída" if is_request_concluded(solicitacao) else "Pendente"
            current_status = getattr(solicitacao, "status", "-")
            current_status_cell = ""

            if include_current_status:
                current_status_cell = (
                    '<td style="padding:10px;border-bottom:1px solid #e2e8f0;">'
                    f"{escape(str(current_status))}"
                    "</td>"
                )

            rows += f"""
                <tr>
                    <td style="padding:10px;border-bottom:1px solid #e2e8f0;">{escape(tipo)}</td>
                    <td style="padding:10px;border-bottom:1px solid #e2e8f0;">#{solicitacao.id}</td>
                    <td style="padding:10px;border-bottom:1px solid #e2e8f0;">{escape(str(solicitacao))}</td>
                    <td style="padding:10px;border-bottom:1px solid #e2e8f0;">{escape(status)}</td>
                    {current_status_cell}
                </tr>
            """

        return rows

    def contract_rows(contratos):
        if not contratos.exists():
            return render_empty("Nenhum contrato finalizado no período.")

        rows = ""

        for contrato in contratos:
            rows += f"""
                <tr>
                    <td style="padding:10px;border-bottom:1px solid #e2e8f0;">#{contrato.id}</td>
                    <td style="padding:10px;border-bottom:1px solid #e2e8f0;">{escape(str(contrato))}</td>
                    <td style="padding:10px;border-bottom:1px solid #e2e8f0;">{br_date(contrato.data_fim)}</td>
                    <td style="padding:10px;border-bottom:1px solid #e2e8f0;">Encerrado</td>
                </tr>
            """

        return rows

    def event_rows(eventos, field_name):
        if not eventos.exists():
            return render_empty("Nenhum evento encontrado.")

        rows = ""

        for evento in eventos:
            data_valor = getattr(evento, field_name)
            fornecedor = (
                evento.empresa_terceira
                or getattr(evento.contrato_terceiro, "empresa_terceira", None)
                or "-"
            )

            rows += f"""
                <tr>
                    <td style="padding:10px;border-bottom:1px solid #e2e8f0;">#{evento.id}</td>
                    <td style="padding:10px;border-bottom:1px solid #e2e8f0;">{escape(evento.descricao or "-")}</td>
                    <td style="padding:10px;border-bottom:1px solid #e2e8f0;">{escape(str(fornecedor))}</td>
                    <td style="padding:10px;border-bottom:1px solid #e2e8f0;">{br_date(data_valor)}</td>
                </tr>
            """

        return rows

    def overdue_event_rows(eventos):
        if not eventos.exists():
            return render_empty("Nenhum evento pendente de entrega encontrado.", colspan=5)

        rows = ""

        for evento in eventos:
            fornecedor = (
                evento.empresa_terceira
                or getattr(evento.contrato_terceiro, "empresa_terceira", None)
                or "-"
            )

            rows += f"""
                <tr>
                    <td style="padding:10px;border-bottom:1px solid #e2e8f0;">#{evento.id}</td>
                    <td style="padding:10px;border-bottom:1px solid #e2e8f0;">{escape(evento.descricao or "-")}</td>
                    <td style="padding:10px;border-bottom:1px solid #e2e8f0;">{escape(str(fornecedor))}</td>
                    <td style="padding:10px;border-bottom:1px solid #e2e8f0;">{br_date(evento.data_prevista)}</td>
                    <td style="padding:10px;border-bottom:1px solid #e2e8f0;">Não entregue</td>
                </tr>
            """

        return rows

    def overdue_contract_rows(contratos):
        if not contratos.exists():
            return render_empty("Nenhum contrato em atraso encontrado.")

        rows = ""

        for contrato in contratos:
            rows += f"""
                <tr>
                    <td style="padding:10px;border-bottom:1px solid #e2e8f0;">#{contrato.id}</td>
                    <td style="padding:10px;border-bottom:1px solid #e2e8f0;">{escape(str(contrato))}</td>
                    <td style="padding:10px;border-bottom:1px solid #e2e8f0;">{br_date(contrato.data_fim)}</td>
                    <td style="padding:10px;border-bottom:1px solid #e2e8f0;">{escape(contrato.get_status_display())}</td>
                </tr>
            """

        return rows

    def render_section(title, table_html):
        return f"""
            <h3 style="font-size:16px;margin:22px 0 8px;color:#0f172a;">
                {escape(title)}
            </h3>
            {table_html}
        """

    user_name = user.get_full_name() or user.username

    previous_week_period = (
        f'{br_date(weeks["previous_week_start"])} a '
        f'{br_date(weeks["previous_week_end"])}'
    )

    current_week_period = (
        f'{br_date(weeks["current_week_start"])} a '
        f'{br_date(weeks["current_week_end"])}'
    )

    next_week_period = (
        f'{br_date(weeks["next_week_start"])} a '
        f'{br_date(weeks["next_week_end"])}'
    )

    general_pending_period = f"Posição consolidada até {br_date(today)}"

    previous_week_content = f"""
        {render_summary_cards([
            ("Solicitações criadas", len(previous_week_requests), "#0f766e"),
            ("Concluídas", len(concluded_requests), "#15803d"),
            ("Pendentes", len(pending_requests), "#b45309"),
            ("Contratos finalizados", contratos_finalizados.count(), "#334155"),
            ("Eventos previstos", eventos_previstos_semana_anterior.count(), "#2563eb"),
            ("Eventos entregues", eventos_entregues_semana_anterior.count(), "#7c3aed"),
        ])}

        {render_section(
            "Solicitações criadas",
            render_table(
                ["Tipo", "ID", "Descrição", "Status"],
                request_rows(previous_week_requests),
            ),
        )}

        {render_section(
            "Solicitações concluídas",
            render_table(
                ["Tipo", "ID", "Descrição", "Status"],
                request_rows(concluded_requests),
            ),
        )}

        {render_section(
            "Solicitações não concluídas",
            render_table(
                ["Tipo", "ID", "Descrição", "Status"],
                request_rows(pending_requests),
            ),
        )}

        {render_section(
            "Contratos finalizados",
            render_table(
                ["ID", "Contrato", "Data de encerramento", "Status"],
                contract_rows(contratos_finalizados),
            ),
        )}

        {render_section(
            "Eventos previstos para entrega",
            render_table(
                ["ID", "Descrição", "Fornecedor", "Data prevista"],
                event_rows(eventos_previstos_semana_anterior, "data_prevista"),
            ),
        )}

        {render_section(
            "Eventos efetivamente entregues",
            render_table(
                ["ID", "Descrição", "Fornecedor", "Data de entrega"],
                event_rows(eventos_entregues_semana_anterior, "data_entrega"),
            ),
        )}
    """

    current_week_content = f"""
        {render_summary_cards([
            ("Eventos previstos para entrega", eventos_previstos_semana_atual.count(), "#2563eb"),
            ("Contratos previstos para finalizar", contratos_previstos_semana_atual.count(), "#b45309"),
        ])}

        {render_section(
            "Eventos previstos para entrega na semana atual",
            render_table(
                ["ID", "Descrição", "Fornecedor", "Data prevista"],
                event_rows(eventos_previstos_semana_atual, "data_prevista"),
            ),
        )}

        {render_section(
            "Contratos previstos para finalizar na semana atual",
            render_table(
                ["ID", "Contrato", "Data prevista de finalização", "Status"],
                contract_rows(contratos_previstos_semana_atual),
            ),
        )}
    """

    next_week_content = f"""
        {render_summary_cards([
            ("Eventos previstos para entrega", eventos_previstos_proxima_semana.count(), "#2563eb"),
            ("Contratos previstos para finalizar", contratos_previstos_proxima_semana.count(), "#b45309"),
        ])}

        {render_section(
            "Eventos previstos para entrega na próxima semana",
            render_table(
                ["ID", "Descrição", "Fornecedor", "Data prevista"],
                event_rows(eventos_previstos_proxima_semana, "data_prevista"),
            ),
        )}

        {render_section(
            "Contratos previstos para finalizar na próxima semana",
            render_table(
                ["ID", "Contrato", "Data prevista de finalização", "Status"],
                contract_rows(contratos_previstos_proxima_semana),
            ),
        )}
    """

    general_pending_content = f"""
        {render_summary_cards([
            ("Solicitações pendentes", len(all_pending_requests), "#b45309"),
            ("Eventos previstos não entregues", eventos_previstos_nao_entregues.count(), "#dc2626"),
            ("Contratos em atraso", contratos_atrasados_nao_encerrados.count(), "#7c2d12"),
        ])}

        {render_section(
            "Todas as solicitações pendentes",
            render_table(
                ["Tipo", "ID", "Descrição", "Situação", "Status atual"],
                request_rows(all_pending_requests, include_current_status=True),
            ),
        )}

        {render_section(
            "Eventos previstos que não foram entregues",
            render_table(
                ["ID", "Descrição", "Fornecedor", "Data prevista", "Situação"],
                overdue_event_rows(eventos_previstos_nao_entregues),
            ),
        )}

        {render_section(
            "Contratos com data fim vencida e ainda não encerrados",
            render_table(
                ["ID", "Contrato", "Data fim", "Status atual"],
                overdue_contract_rows(contratos_atrasados_nao_encerrados),
            ),
        )}
    """

    html = f"""
    <div style="margin:0;padding:0;background:#f4f6f8;font-family:Arial,Helvetica,sans-serif;color:#263238;">
        <div style="max-width:960px;margin:0 auto;padding:24px;">

            <div style="background:#0f766e;color:#fff;padding:28px;border-radius:14px 14px 0 0;">
                <h1 style="margin:0;font-size:26px;">Report Semanal de Suprimentos</h1>
                <p style="margin:8px 0 0;font-size:15px;">
                    Olá, {escape(user_name)}. Segue o resumo semanal organizado por período.
                </p>
            </div>

            <div style="background:#ffffff;padding:24px;border-radius:0 0 14px 14px;box-shadow:0 4px 18px rgba(0,0,0,.08);">

                {render_period_block(
                    "Semana anterior",
                    previous_week_period,
                    previous_week_content,
                    "#ffffff",
                    "#0f766e",
                )}

                {render_period_block(
                    "Semana atual",
                    current_week_period,
                    current_week_content,
                    "#ffffff",
                    "#2563eb",
                )}

                {render_period_block(
                    "Próxima semana",
                    next_week_period,
                    next_week_content,
                    "#ffffff",
                    "#7c3aed",
                )}

                {render_period_block(
                    "Pendências gerais",
                    general_pending_period,
                    general_pending_content,
                    "#ffffff",
                    "#b45309",
                )}

                <p style="font-size:12px;color:#94a3b8;margin-top:32px;text-align:center;">
                    Este é um e-mail automático. Por favor, não responda diretamente.
                </p>

            </div>
        </div>
    </div>
    """

    return html


def can_user_manage_supplier_choice(user, solicitacao):
    if user.grupo == "lider_contrato":
        return user == solicitacao.lider_contrato
    if user.grupo == "gerente_contrato":
        return True
    if user.grupo == "gerente_lider":
        return user_shares_center_with_coordinator(user, solicitacao.coordenador)
    return False


def can_user_manage_event_delivery(user, contrato):
    if not user or not contrato:
        return False
    if user.grupo == "lider_contrato":
        return user == contrato.lider_contrato
    if user.grupo == "gerente_contrato":
        return True
    if user.grupo == "suprimento":
        return True
    if user.grupo == "gerente_lider":
        return user_shares_center_with_coordinator(user, contrato.coordenador)
    return False


def can_user_manage_os_delivery(user, os):
    if not user or not os:
        return False
    if user.grupo == "lider_contrato":
        return user == os.lider_contrato
    if user.grupo == "gerente_contrato":
        return True
    if user.grupo == "gerente_lider":
        return user_shares_center_with_coordinator(user, os.coordenador)
    return False


def can_user_request_contract_addendum(user, contrato):
    if not user or not contrato:
        return False
    if user.grupo == "lider_contrato":
        return user == contrato.lider_contrato
    if user.grupo == "gerente_lider":
        return user_shares_center_with_coordinator(user, contrato.coordenador)
    if user.grupo == "gerente_contrato":
        return True
    return False


def can_user_view_addendum(user, aditivo):
    if not user or not aditivo:
        return False
    if user.grupo in ["suprimento", "diretoria", "gerente_contrato"]:
        return True
    if user.grupo == "lider_contrato":
        return user == aditivo.contrato.lider_contrato or user == aditivo.solicitado_por
    if user.grupo == "gerente_lider":
        return user_shares_center_with_contract_coordinators(user, aditivo.contrato)
    return False


def can_user_approve_addendum_request_as_gerente(user, aditivo):
    return bool(user and aditivo and user.grupo == "gerente_contrato" and aditivo.status_gerente == "pendente")


def can_user_approve_addendum_request_as_diretoria(user, aditivo):
    return bool(user and aditivo and user.grupo == "diretoria" and aditivo.status_diretoria == "pendente")


def can_user_upload_addendum_draft(user, aditivo):
    return bool(
        user
        and aditivo
        and user.grupo == "suprimento"
        and aditivo.solicitacao_aprovada_totalmente
        and not aditivo.aprovado_totalmente
        and not aditivo.minuta_aprovada
    )


def can_user_approve_addendum_draft(user, aditivo):
    return bool(
        user
        and aditivo
        and user.grupo == "gerente_contrato"
        and aditivo.tem_documento
        and aditivo.status_lider == "pendente"
        and not aditivo.aprovado_totalmente
    )


def can_user_upload_signed_addendum(user, aditivo):
    return bool(
        user
        and aditivo
        and user.grupo == "suprimento"
        and aditivo.minuta_aprovada
        and not aditivo.tem_documento_assinado
    )


def build_addendum_timeline(aditivo):
    labels = [
        "Solicitacao do Aditivo",
        "Aprovacao da Solicitacao",
        "Minuta do Aditivo",
        "Aprovacao da Minuta",
        "Documento Assinado",
        "Concluido",
    ]
    if aditivo.aprovado_totalmente:
        current_index = 5
    elif aditivo.minuta_aprovada:
        current_index = 4
    elif aditivo.tem_documento:
        current_index = 3
    elif aditivo.solicitacao_aprovada_totalmente:
        current_index = 2
    else:
        current_index = 1

    progress_percent = 0
    if len(labels) > 1:
        if current_index == 5:
            progress_percent = 100
        else:
            progress_percent = (current_index / (len(labels))) * 100

    return {
        "labels": labels,
        "current_index": current_index,
        "progress_percent": float(progress_percent),
    }


def get_visible_addendums_for_user(user):
    queryset = AditivoContratoTerceiro.objects.select_related(
        "contrato",
        "contrato__empresa_terceira",
        "contrato__cod_projeto",
        "contrato__lider_contrato",
        "contrato__coordenador",
        "solicitado_por",
        "documento_enviado_por",
        "documento_assinado_enviado_por",
    )
    if not user:
        return queryset.none()
    if user.grupo in ["suprimento", "diretoria", "gerente_contrato"]:
        return queryset
    if user.grupo == "lider_contrato":
        return queryset.filter(Q(contrato__lider_contrato=user) | Q(solicitado_por=user)).distinct()
    if user.grupo == "gerente_lider":
        return queryset.filter(contrato__coordenadores__centros__in=user.centros.all()).distinct()
    return queryset.none()


def bm_operational_approval_query():
    return (
        Q(status_coordenador="aprovado", status_gerente="pendente")
        | Q(status_coordenador="pendente", status_gerente="aprovado")
        | Q(status_coordenador="aprovado", status_gerente="aprovado")
    )


def bm_has_operational_approval(bm):
    return (
        bm.status_coordenador == "aprovado"
        or bm.status_gerente == "aprovado"
    ) and bm.status_coordenador != "reprovado" and bm.status_gerente != "reprovado"


def bm_is_operationally_pending(bm):
    return bm.status_coordenador == "pendente" and bm.status_gerente == "pendente"


def filter_payment_events_for_user(user, filtros_base):
    if user.grupo in ["suprimento", "diretoria", "financeiro"]:
        return Evento.objects.filter(filtros_base, contrato_terceiro__isnull=False)
    if user.grupo in ["gerente", "gerente_lider"]:
        return Evento.objects.filter(
            filtros_base,
            contrato_terceiro__isnull=False,
            contrato_terceiro__coordenador__centros__in=user.centros.all(),
        ).distinct()
    if user.grupo == "gerente_contrato":
        return Evento.objects.filter(
            filtros_base,
            contrato_terceiro__isnull=False,
            contrato_terceiro__lider_contrato=user,
        )
    return Evento.objects.none()


def filter_payment_os_for_user(user, filtros_os):
    if user.grupo in ["suprimento", "diretoria", "financeiro"]:
        return OS.objects.filter(filtros_os)
    if user.grupo in ["gerente", "gerente_lider"]:
        return OS.objects.filter(
            filtros_os,
            coordenador__centros__in=user.centros.all(),
        ).distinct()
    if user.grupo == "gerente_contrato":
        return OS.objects.filter(
            filtros_os,
            lider_contrato=user,
        )
    return OS.objects.none()


def filter_payment_bms_for_user(user, queryset):
    if user.grupo in ["suprimento", "diretoria", "financeiro"]:
        return queryset
    if user.grupo in ["gerente", "gerente_lider"]:
        return queryset.filter(
            contrato__coordenador__centros__in=user.centros.all(),
        ).distinct()
    if user.grupo == "gerente_contrato":
        return queryset.filter(contrato__lider_contrato=user)
    return queryset.none()


class ContratoCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Contrato
    form_class = ContratoForm
    template_name = 'forms/contrato_form.html'
    success_url = reverse_lazy('lista_contratos')

    def test_func(self):
        # só permite se for grupo suprimento
        return self.request.user.grupo == "suprimento"

    def handle_no_permission(self):
        # Redireciona para a home
        return redirect('home')


class ClienteCreateView(LoginRequiredMixin, UserPassesTestMixin,CreateView):
    model = Cliente
    form_class = ClienteForm
    template_name = 'forms/cliente_form.html'
    success_url = reverse_lazy('lista_clientes')

    def test_func(self):
        # só permite se for grupo suprimento
        return self.request.user.grupo == "suprimento"

    def handle_no_permission(self):
        # Redireciona para a home
        return redirect('home')


class FornecedorCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = EmpresaTerceira
    form_class = FornecedorForm
    template_name = 'forms/fornecedor_form.html'
    success_url = reverse_lazy('lista_fornecedores')

    def test_func(self):
        # só permite se for grupo suprimento
        return self.request.user.grupo in ["suprimento"]

    def handle_no_permission(self):
        # Redireciona para a home
        return redirect('home')


class ContratoFornecedorCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = ContratoTerceiros
    form_class = ContratoFornecedorForm
    template_name = 'forms/contrato_fornecedor_form.html'
    success_url = reverse_lazy('lista_contratos_fornecedores')

    def test_func(self):
        # só permite se for grupo suprimento
        return self.request.user.grupo == "suprimento"

    def handle_no_permission(self):
        # Redireciona para a home
        return redirect('home')

    def form_valid(self, form):
        print("==== DEBUG FORM_VALID ====")
        print("Arquivos recebidos:", self.request.FILES)
        print("Campos POST:", self.request.POST)

        response = super().form_valid(form)

        print("Objeto salvo:", self.object)
        print("Arquivo salvo:", self.object.num_contrato_arquivo)
        print("==========================")

        return response


class OSCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = OS
    form_class = OrdemServicoForm
    template_name = 'forms/os_form.html'
    success_url = reverse_lazy('lista_ordens_servico')

    def test_func(self):
        # só permite se for grupo suprimento
        return self.request.user.grupo == "suprimento"

    def handle_no_permission(self):
        # Redireciona para a home
        return redirect('home')


#def is_financeiro(user):
#    return user.is_authenticated and getattr(user, "grupo", None) == "financeiro"

def home(request):
    user = request.user
    hoje = timezone.now().date()
    limite = hoje + timedelta(days=10)
    limite_contrato = hoje + timedelta(days=30)

    grupo = getattr(user, "grupo", None)
    is_suprimento = grupo == "suprimento"
    is_coordenador = grupo == "coordenador"
    is_gerente = grupo == "gerente"
    is_gerente_lider = grupo =="gerente_lider"
    is_diretoria = grupo == "diretoria"
    is_financeiro = grupo == "financeiro"
    is_lider = grupo == "lider_contrato"
    is_gerente_contrato = grupo == "gerente_contrato"

    context = {
        "is_suprimento": is_suprimento,
        "is_coordenador": is_coordenador,
        "is_lider":is_lider,
        "is_gerente": is_gerente,
        "is_gerente_lider": is_gerente_lider,
        "is_gerente_contrato":is_gerente_contrato,
        "is_diretoria": is_diretoria,
        "is_financeiro": is_financeiro
    }

    # ==================== SUPRIMENTO ====================
    if is_suprimento:
        solicitacoes_prospeccao = SolicitacaoProspeccao.objects.filter(
            Q(aprovado__isnull=True)
            | (Q(aprovado=True) & Q(triagem_realizada=False))
            | (Q(aprovado=True) & Q(triagem_realizada=True) & Q(fornecedor_escolhido__isnull=True))
            | (Q(aprovacao_fornecedor_gerente="aprovado") & Q(aprovacao_gerencia=False))
            | Q(status__in=["Fornecedor aprovado", "Planejamento do Contrato", SIGNED_FILES_PENDING_STATUS])
        ).exclude(status__in=["Onboarding"]).distinct()

        solicitacoes_contratos = SolicitacaoContrato.objects.filter(
            Q(aprovacao_fornecedor_gerente="aprovado")
            | Q(aprovacao_fornecedor_diretor="aprovado")
            | Q(status__in=["Planejamento do Contrato", SIGNED_FILES_PENDING_STATUS])
        ).exclude(status__in=["Onboarding"]).distinct()

        solicitacoes_os = SolicitacaoOrdemServico.objects.filter(
            Q(aprovacao_lider="aprovado")
            #| Q(aprovacao_diretor="aprovado")
            | Q(status__in=["pendente_suprimento"])
        ).distinct()

        aditivos_pendentes = AditivoContratoTerceiro.objects.filter(
            Q(status_gerente="aprovado", status_diretoria="aprovado", arquivo_aditivo__isnull=True)
            | Q(status_gerente="aprovado", status_diretoria="aprovado", arquivo_aditivo="")
            | Q(status_lider="reprovado")
            | Q(status_lider="aprovado", arquivo_aditivo_assinado__isnull=True)
            | Q(status_lider="aprovado", arquivo_aditivo_assinado="")
        ).select_related(
            "contrato",
            "contrato__empresa_terceira",
            "solicitado_por",
        ).distinct()

        bms_pendentes = BM.objects.filter(
            Q(status_coordenador="pendente", status_gerente="pendente")
            | Q(aprovacao_pagamento="pendente", status_coordenador="aprovado")
            | Q(aprovacao_pagamento="pendente", status_gerente="aprovado")
        ).exclude(
            Q(status_coordenador="reprovado") | Q(status_gerente="reprovado")
        ).select_related(
            "contrato",
            "contrato__empresa_terceira"
        ).order_by("-data_pagamento").distinct()

        proxima_data_pagamento_bm = Evento.objects.filter(
            data_prevista_pagamento__isnull=False,
            data_prevista_pagamento__gte=hoje,
            contrato_terceiro__isnull=False,
        ).exclude(
            boletins_medicao__isnull=False
        ).aggregate(
            proxima_data=Min("data_prevista_pagamento")
        )["proxima_data"]

        eventos_bm_para_entrega = Evento.objects.none()
        if proxima_data_pagamento_bm:
            eventos_bm_para_entrega = Evento.objects.filter(
                data_prevista_pagamento__isnull=False,
                data_prevista_pagamento__gte=hoje,
                data_prevista_pagamento__lte=proxima_data_pagamento_bm,
                contrato_terceiro__isnull=False,
            ).exclude(
                boletins_medicao__isnull=False
            ).select_related(
                "contrato_terceiro",
                "empresa_terceira"
            ).order_by(
                "data_prevista_pagamento",
                "data_prevista",
            ).distinct()

        eventos_proximos = Evento.objects.filter(
            data_prevista__gte=hoje,
            data_prevista__lte=limite,
            data_entrega__isnull=True,
            contrato_terceiro__isnull=False
        ).select_related(
            "contrato_terceiro",
            "empresa_terceira"
        ).order_by("data_prevista")

        # Entregas atrasadas (todas)
        entregas_atrasadas = Evento.objects.filter(
            data_prevista__lt=hoje,
            data_entrega__isnull=True,
            contrato_terceiro__isnull=False
        ).select_related(
            "contrato_terceiro",
            "empresa_terceira"
        ).order_by("data_prevista")

        eventos_para_avaliar = Evento.objects.filter(
            data_entrega__isnull=False,
            contrato_terceiro__isnull=False
        ).annotate(
            total_avaliacoes=Count("avaliacoes")
        ).filter(
            total_avaliacoes=0
        ).select_related(
            "contrato_terceiro",
            "empresa_terceira"
        ).order_by("data_entrega")

        contratos_vencendo = ContratoTerceiros.objects.filter(
            data_fim__gte=hoje,
            data_fim__lte=limite_contrato,
            status='ativo'
        ).select_related(
            "empresa_terceira"
        ).order_by("data_fim")

        os_em_aberto = OS.objects.filter(
            status__in=["em_execucao", "paralizada"]
        ).select_related(
            "contrato",
            "contrato__empresa_terceira"
        ).order_by("prazo_execucao")

        context.update({
            "painel_titulo": "Painel de Suprimentos",
            "solicitacoes_prospeccao": solicitacoes_prospeccao,
            "solicitacoes_contrato": solicitacoes_contratos,
            "solicitacoes_os": solicitacoes_os,
            "aditivos_pendentes": aditivos_pendentes,
            "eventos_proximos": eventos_proximos,
            "eventos_para_avaliar": eventos_para_avaliar,
            "entregas_atrasadas": entregas_atrasadas,
            "is_suprimento": user.grupo == 'suprimento',
            "bms_pendentes": bms_pendentes,
            "eventos_bm_para_entrega": eventos_bm_para_entrega,
            "proxima_data_pagamento_bm": proxima_data_pagamento_bm,
            "contratos_vencendo":contratos_vencendo,
            "os_em_aberto": os_em_aberto,
            "today": hoje
        })

    # ==================== COORDENADOR ====================
    elif is_coordenador:
        solicitacoes_prospeccao = SolicitacaoProspeccao.objects.filter(
            coordenador=user
        ).exclude(status__in=["Onboarding"]).distinct()

        solicitacoes_contrato = SolicitacaoContrato.objects.filter(
            coordenador=user
        ).exclude(status__in=["Onboarding"]).distinct()

        solicitacoes_os = SolicitacaoOrdemServico.objects.filter(
            coordenador=user
        ).exclude(status__in=["finalizada","reprovada"]).distinct()

        eventos_atrasados = Evento.objects.filter(
            contrato_terceiro__coordenador=user,
            data_prevista__lt=hoje,
            realizado=False,
            contrato_terceiro__isnull=False,
        ).select_related(
            "contrato_terceiro",
            "empresa_terceira"
        ).order_by("data_prevista")

        bms_pendentes = BM.objects.none()

        eventos_proximos = Evento.objects.filter(
            data_prevista__range=[hoje, limite],
            realizado=False,
            contrato_terceiro__coordenador=user
        ).order_by("data_prevista")

        eventos_para_avaliar = Evento.objects.filter(
            contrato_terceiro__coordenador=user,
            realizado=True,
            avaliacoes__isnull=True
        ).select_related(
            "contrato_terceiro",
            "empresa_terceira"
        ).order_by("data_entrega")

        contratos_vencendo = ContratoTerceiros.objects.filter(
            coordenador=user,
            data_fim__gte=hoje,
            data_fim__lte=limite_contrato,
            status='ativo'
        ).select_related(
            "empresa_terceira"
        ).order_by("data_fim")

        os_em_aberto = OS.objects.filter(
            status__in=["em_execucao", "paralizada"],
            coordenador=user
        ).select_related(
            "contrato",
            "contrato__empresa_terceira"
        ).order_by("prazo_execucao")

        context.update({
            "painel_titulo": "Painel do Coordenador",
            "solicitacoes_prospeccao": solicitacoes_prospeccao,
            "solicitacoes_contrato": solicitacoes_contrato,
            "solicitacoes_os": solicitacoes_os,
            "bms_pendentes": bms_pendentes,
            "eventos_proximos": eventos_proximos,
            "eventos_atrasados": eventos_atrasados,
            "eventos_para_avaliar":eventos_para_avaliar,
            "contratos_vencendo": contratos_vencendo,
            "today": hoje,
        })


    # ==================== LIDER DE CONTRATO ==========

    elif is_lider:
        solicitacoes_prospeccao = SolicitacaoProspeccao.objects.filter(
            lider_contrato=user
        ).filter(
            Q(status="Solicitação de prospecção") |
            Q(triagem_realizada=True, status="Triagem realizada")
        ).distinct()
        solicitacoes_contrato = SolicitacaoContrato.objects.filter(
            lider_contrato=user,
            status="Solicitação de contratação"
        ).distinct()
        solicitacoes_os = SolicitacaoOrdemServico.objects.filter(
            lider_contrato=user,
            status="pendente_lider"
        ).distinct()

        eventos_proximos = Evento.objects.filter(
            data_prevista__gte=hoje,
            data_prevista__lte=limite,
            data_entrega__isnull=True,
            contrato_terceiro__isnull=False,
            contrato_terceiro__lider_contrato=user
        ).select_related(
            "contrato_terceiro",
            "empresa_terceira"
        ).order_by("data_prevista")

        # Entregas atrasadas
        entregas_atrasadas = Evento.objects.filter(
            data_prevista__lt=hoje,
            data_entrega__isnull=True,
            contrato_terceiro__isnull=False,
            contrato_terceiro__lider_contrato=user
        ).select_related(
            "contrato_terceiro",
            "empresa_terceira"
        ).order_by("data_prevista")

        contratos_vencendo = ContratoTerceiros.objects.filter(
            data_fim__gte=hoje,
            data_fim__lte=limite_contrato,
            status='ativo',
            lider_contrato=user
        ).select_related(
            "empresa_terceira"
        ).order_by("data_fim")

        os_em_aberto = OS.objects.filter(
            status__in=["em_execucao", "paralizada"],
            lider_contrato=user
        ).select_related(
            "contrato",
            "contrato__empresa_terceira"
        ).order_by("prazo_execucao")

        context.update({
            "painel_titulo": "Painel do Lider de Contratos",
            "solicitacoes_prospeccao": solicitacoes_prospeccao,
            "solicitacoes_contrato": solicitacoes_contrato,
            "solicitacoes_os": solicitacoes_os,
            "eventos_proximos": eventos_proximos,
            "entregas_atrasadas": entregas_atrasadas,
            "contratos_vencendo": contratos_vencendo,
            "is_lider": is_lider,
            "today": hoje,
        })

    # ==================== GERENTE ====================
    elif is_gerente:
        centros_gerente = getattr(user, "centros", None)
        centros_ids = centros_gerente.values_list("id", flat=True) if centros_gerente else []

        solicitacoes_prospeccao = SolicitacaoProspeccao.objects.filter(
            coordenador__centros__in=centros_ids
        ).exclude(status__in=["Onboarding"]).distinct()

        solicitacoes_contratos = SolicitacaoContrato.objects.filter(
            coordenador__centros__in=centros_ids
        ).exclude(status__in=["Onboarding"]).distinct()

        solicitacoes_os = SolicitacaoOrdemServico.objects.filter(
            coordenador__centros__in=centros_ids
        ).exclude(status__in=["finalizada","reprovada"]).distinct()

        bms_pendentes = BM.objects.filter(
            contrato__coordenador__centros__in=centros_ids,
            status_coordenador="pendente",
            status_gerente="pendente"
        ).select_related("contrato", "contrato__empresa_terceira").order_by("-data_pagamento")

        eventos_proximos = Evento.objects.filter(
            data_prevista__range=[hoje, limite],
            realizado=False,
            contrato_terceiro__coordenador__centros__in=centros_ids
        ).order_by("data_prevista").distinct()

        eventos_para_avaliar = Evento.objects.filter(
            contrato_terceiro__coordenador__centros__in=centros_ids,
            realizado=True,
            avaliacoes__isnull=True
        ).select_related(
            "contrato_terceiro",
            "empresa_terceira"
        ).order_by("data_entrega")

        os_em_aberto = OS.objects.filter(
            status__in=["em_execucao", "paralizada"],
            coordenador__centros__in=centros_ids
        ).select_related(
            "contrato",
            "contrato__empresa_terceira"
        ).order_by("prazo_execucao")

        contratos_vencendo = ContratoTerceiros.objects.filter(
            coordenador__centros__in=centros_ids,
            data_fim__gte=hoje,
            data_fim__lte=limite_contrato,
            status='ativo'
        ).select_related(
            "empresa_terceira"
        ).order_by("data_fim")

        # Entregas atrasadas dos centros do gerente
        eventos_atrasados = Evento.objects.filter(
            contrato_terceiro__coordenador__centros__in=centros_ids,
            data_prevista__lt=hoje,
            realizado=False,
            contrato_terceiro__isnull=False,
        ).select_related(
            "contrato_terceiro",
            "empresa_terceira"
        ).order_by("data_prevista").distinct()

        context.update({
            "painel_titulo": "Painel da Gerência",
            "solicitacoes_prospeccao": solicitacoes_prospeccao,
            "solicitacoes_contrato": solicitacoes_contratos,
            "solicitacoes_os": solicitacoes_os,
            "bms_pendentes": bms_pendentes,
            "eventos_proximos": eventos_proximos,
            "eventos_atrasados": eventos_atrasados,
            "eventos_proximos": eventos_proximos,
            "eventos_para_avaliar": eventos_para_avaliar,
            "contratos_vencendo":contratos_vencendo,
            "os_em_aberto": os_em_aberto,
            "today": hoje,
        })

    # ==================== GERENTE TÉCNICO E LIDER DE CONTRATO ====================
    elif is_gerente_lider:
        centros_gerente = getattr(user, "centros", None)
        centros_ids = centros_gerente.values_list("id", flat=True) if centros_gerente else []

        solicitacoes_prospeccao = SolicitacaoProspeccao.objects.filter(
            coordenador__centros__in=centros_ids
        ).exclude(status__in=["Onboarding"]).distinct()

        solicitacoes_contratos = SolicitacaoContrato.objects.filter(
            coordenador__centros__in=centros_ids
        ).exclude(status__in=["Onboarding"]).distinct()

        solicitacoes_os = SolicitacaoOrdemServico.objects.filter(
            coordenador__centros__in=centros_ids
        ).exclude(status__in=["finalizada","reprovada"]).distinct()

        bms_pendentes = BM.objects.filter(
            contrato__coordenador__centros__in=centros_ids
        ).filter(
            Q(status_coordenador="pendente", status_gerente="pendente")
            | (Q(aprovacao_pagamento="pendente") & bm_operational_approval_query())
        ).exclude(
            Q(status_coordenador="reprovado") | Q(status_gerente="reprovado")
        ).select_related(
            "contrato",
            "contrato__empresa_terceira"
        ).order_by("-data_pagamento").distinct()

        eventos_proximos = Evento.objects.filter(
            data_prevista__range=[hoje, limite],
            realizado=False,
            contrato_terceiro__coordenador__centros__in=centros_ids
        ).order_by("data_prevista").distinct()

        eventos_para_avaliar = Evento.objects.filter(
            contrato_terceiro__coordenador__centros__in=centros_ids,
            realizado=True,
            avaliacoes__isnull=True
        ).select_related(
            "contrato_terceiro",
            "empresa_terceira"
        ).order_by("data_entrega")

        os_em_aberto = OS.objects.filter(
            status__in=["em_execucao", "paralizada"],
            coordenador__centros__in=centros_ids
        ).select_related(
            "contrato",
            "contrato__empresa_terceira"
        ).order_by("prazo_execucao")

        contratos_vencendo = ContratoTerceiros.objects.filter(
            coordenador__centros__in=centros_ids,
            data_fim__gte=hoje,
            data_fim__lte=limite_contrato,
            status='ativo'
        ).select_related(
            "empresa_terceira"
        ).order_by("data_fim")

        # Entregas atrasadas dos centros do gerente
        entregas_atrasadas = Evento.objects.filter(
            contrato_terceiro__coordenador__centros__in=centros_ids,
            data_prevista__lt=hoje,
            data_entrega__isnull=True,
            contrato_terceiro__isnull=False,
        ).select_related(
            "contrato_terceiro",
            "empresa_terceira"
        ).order_by("data_prevista").distinct()

        context.update({
            "painel_titulo": "Painel da Gerência",
            "solicitacoes_prospeccao": solicitacoes_prospeccao,
            "solicitacoes_contrato": solicitacoes_contratos,
            "solicitacoes_os": solicitacoes_os,
            "bms_pendentes": bms_pendentes,
            "eventos_proximos": eventos_proximos,
            "entregas_atrasadas":entregas_atrasadas,
            "eventos_para_avaliar":eventos_para_avaliar,
            "contratos_vencendo": contratos_vencendo,
            "os_em_aberto": os_em_aberto,
            "today": hoje,
        })

    #===================== GERENTE DE CONTRATO ==========
    elif is_gerente_contrato:
        """centros_gerente = getattr(user, "centros", None)
        centros_ids = centros_gerente.values_list("id", flat=True) if centros_gerente else []"""

        solicitacoes_prospeccao = SolicitacaoProspeccao.objects.filter(
            lider_contrato__grupo__in=["lider_contrato", "gerente_contrato", "gerente_lider"]
        ).filter(
            Q(status="Solicitação de prospecção") |
            Q(triagem_realizada=True, status="Triagem realizada") |
            Q(status__in=["Aprovação Final", "Fornecedor selecionado"])
        ).distinct()
        solicitacoes_contrato = SolicitacaoContrato.objects.filter(
            lider_contrato__grupo__in=["lider_contrato", "gerente_contrato", "gerente_lider"]
        ).filter(
            Q(status="Solicitação de contratação")
            | Q(status="Aprovação Final")
            | (
                Q(status="Planejamento do Contrato")
                & Q(minuta_contrato__arquivo_contrato__isnull=False)
            )
        ).distinct()
        solicitacoes_os = SolicitacaoOrdemServico.objects.filter(
            lider_contrato__grupo__in=["lider_contrato", "gerente_contrato", "gerente_lider"],
            status__in=["pendente_lider", "pendente_gerente"]
        ).distinct()

        aditivos_pendentes = AditivoContratoTerceiro.objects.filter(
            Q(status_gerente="pendente")
            | (
                Q(status_gerente="aprovado")
                & Q(status_diretoria="aprovado")
                & Q(status_lider="pendente")
                & Q(arquivo_aditivo__isnull=False)
                & ~Q(arquivo_aditivo="")
            )
        ).select_related(
            "contrato",
            "contrato__empresa_terceira",
            "solicitado_por",
        ).distinct()

        eventos_proximos = Evento.objects.filter(
            data_prevista__gte=hoje,
            data_prevista__lte=limite,
            data_entrega__isnull=True,
            contrato_terceiro__isnull=False,
            contrato_terceiro__lider_contrato__grupo__in=["lider_contrato", "gerente_contrato", "gerente_lider"]
        ).select_related(
            "contrato_terceiro",
            "empresa_terceira"
        ).order_by("data_prevista")

        # Entregas atrasadas
        entregas_atrasadas = Evento.objects.filter(
            data_prevista__lt=hoje,
            data_entrega__isnull=True,
            contrato_terceiro__isnull=False,
            contrato_terceiro__lider_contrato__grupo__in=["lider_contrato", "gerente_contrato", "gerente_lider"]
        ).select_related(
            "contrato_terceiro",
            "empresa_terceira"
        ).order_by("data_prevista").distinct()

        contratos_vencendo = ContratoTerceiros.objects.filter(
            data_fim__gte=hoje,
            data_fim__lte=limite_contrato,
            status='ativo',
            lider_contrato__grupo__in=["lider_contrato", "gerente_contrato", "gerente_lider"]
        ).select_related(
            "empresa_terceira"
        ).order_by("data_fim")

        """bms_pendentes = BM.objects.filter(
            contrato__lider_contrato__grupo__in=["lider_contrato", "gerente_contrato", "gerente_lider"],
            status_gerente="pendente"
        ).select_related("contrato", "contrato__empresa_terceira").order_by("-data_pagamento").distinct()"""

        os_em_aberto = OS.objects.filter(
            status__in=["em_execucao", "paralizada"],
            lider_contrato__grupo__in=["lider_contrato", "gerente_contrato", "gerente_lider"]
        ).select_related(
            "contrato",
            "contrato__empresa_terceira"
        ).order_by("prazo_execucao")

        context.update({
            "painel_titulo": "Painel da Gerência de Contratos",
            "solicitacoes_prospeccao": solicitacoes_prospeccao,
            "solicitacoes_contrato": solicitacoes_contrato,
            "solicitacoes_os": solicitacoes_os,
            "aditivos_pendentes": aditivos_pendentes,
            #"bms_pendentes": bms_pendentes,
            "eventos_proximos": eventos_proximos,
            "entregas_atrasadas": entregas_atrasadas,
            "is_gerente_contrato": is_gerente_contrato,
            "contratos_vencendo": contratos_vencendo,
            "os_em_aberto": os_em_aberto,
            "today": hoje,
        })

    # ==================== DIRETORIA ====================
    elif is_diretoria:

        solicitacoes_prospeccao = SolicitacaoProspeccao.objects.filter(
            Q(aprovacao_fornecedor_diretor="pendente")
        ).exclude(status__in=["Onboarding"]).distinct()

        solicitacoes_contratos = SolicitacaoContrato.objects.filter(
            Q(aprovacao_fornecedor_diretor="pendente")
        ).exclude(status__in=["Onboarding"]).distinct()

        solicitacoes_os = SolicitacaoOrdemServico.objects.filter(
            Q(aprovacao_diretor="pendente")
        ).distinct()

        aditivos_pendentes = AditivoContratoTerceiro.objects.filter(
            status_diretoria="pendente"
        ).select_related(
            "contrato",
            "contrato__empresa_terceira",
            "solicitado_por",
        ).distinct()

        # BM aprovados por Coordenador e Gerente mas pendentes na Diretoria
        bms_pendentes = BM.objects.filter(
            Q(aprovacao_pagamento="pendente")
        ).filter(
            bm_operational_approval_query()
        ).exclude(
            Q(status_coordenador="reprovado") | Q(status_gerente="reprovado")
        ).select_related(
            "contrato",
            "contrato__empresa_terceira"
        ).order_by("-data_pagamento").distinct()

        # Entregas atrasadas (todas)
        entregas_atrasadas = Evento.objects.filter(
            data_prevista__lt=hoje,
            data_entrega__isnull=True,
            contrato_terceiro__isnull=False
        ).select_related(
            "contrato_terceiro",
            "empresa_terceira"
        ).order_by("data_prevista")

        contratos_vencendo = ContratoTerceiros.objects.filter(
            data_fim__gte=hoje,
            data_fim__lte=limite_contrato,
            status='ativo'
        ).select_related(
            "empresa_terceira"
        ).order_by("data_fim")

        # PrÃ³ximas entregas (10 dias)
        eventos_proximos = Evento.objects.filter(
            data_prevista__gte=hoje,
            data_prevista__lte=limite,
            data_entrega__isnull=True,
            contrato_terceiro__isnull=False
        ).select_related(
            "contrato_terceiro",
            "empresa_terceira"
        ).order_by("data_prevista")

        os_em_aberto = OS.objects.filter(
            status__in=["em_execucao", "paralizada"]
        ).select_related(
            "contrato",
            "contrato__empresa_terceira"
        ).order_by("prazo_execucao")

        context.update({
            "painel_titulo": "Painel da Diretoria",
            "solicitacoes_prospeccao": solicitacoes_prospeccao,
            "solicitacoes_contrato": solicitacoes_contratos,
            "solicitacoes_os": solicitacoes_os,
            "aditivos_pendentes": aditivos_pendentes,
            "bms_pendentes": bms_pendentes,
            "entregas_atrasadas": entregas_atrasadas,
            "eventos_proximos": eventos_proximos,
            "is_diretoria": is_diretoria,
            "contratos_vencendo": contratos_vencendo,
            "os_em_aberto": os_em_aberto,
            "today": hoje,
        })

    # ==================== FINANCEIRO ====================
    elif is_financeiro:
        eventos_sem_nf = Evento.objects.filter(
            boletins_medicao__aprovacao_pagamento="aprovado",
            nota_fiscal__isnull=True,
        ).filter(
            Q(boletins_medicao__status_coordenador="aprovado")
            | Q(boletins_medicao__status_gerente="aprovado")
        ).select_related(
            "contrato_terceiro",
            "contrato_terceiro__empresa_terceira",
            "empresa_terceira",
        ).annotate(
            ultima_data_pagamento_bm=Max("boletins_medicao__data_pagamento")
        ).order_by("-ultima_data_pagamento_bm").distinct()

        eventos_proximos = Evento.objects.filter(
            data_prevista__range=[hoje, limite]
        ).order_by("data_prevista")

        context.update({
            "painel_titulo": "Painel do Financeiro",
            "eventos_sem_nf": eventos_sem_nf,
            "eventos_proximos": eventos_proximos,
            "is_financeiro": is_financeiro,

        })


    return render(request, "home.html", context)


@login_required
def enviar_report_suprimento(request):
    if request.user.grupo != "suprimento":
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    if request.method != "POST":
        return redirect("home")

    if not request.user.email:
        messages.error(request, "Seu usuário não possui e-mail cadastrado.")
        return redirect("home")

    assunto = "Report semanal de suprimentos"
    html_content = build_weekly_supply_report(request.user)
    mensagem = strip_tags(html_content)

    try:
        email = EmailMultiAlternatives(
            subject="Report Semanal de Suprimentos",
            body=mensagem,
            from_email=FROM_EMAIL,
            to=[request.user.email],
        )
        email.attach_alternative(html_content, "text/html")
        email.send()
        messages.success(request, f"Report enviado para {request.user.email}.")
    except Exception as exc:
        messages.error(request, f"Não foi possível enviar o report: {exc}")

    return redirect("home")



def logout(request):
    return render(request, 'logged_out.html')


@login_required
def lista_contratos(request):
    if request.user.grupo in ['suprimento', 'financeiro', 'diretoria']:
        contratos = Contrato.objects.all()
    elif request.user.grupo == 'coordenador':
        contratos = Contrato.objects.filter(
            Q(coordenador=request.user) | Q(coordenadores=request.user)
        ).distinct()
    elif request.user.grupo == 'lider_contrato':
        contratos = Contrato.objects.filter(lider_contrato=request.user)
    elif request.user.grupo == 'gerente':
        contratos = Contrato.objects.filter(
            Q(coordenador__centros__in=request.user.centros.all()) |
            Q(coordenadores__centros__in=request.user.centros.all())
        ).distinct()
    elif request.user.grupo == 'gerente_lider':
        contratos = Contrato.objects.filter(
            Q(coordenador__centros__in=request.user.centros.all()) |
            Q(coordenadores__centros__in=request.user.centros.all())
        ).distinct()
    elif request.user.grupo == 'gerente_contrato':
        contratos = Contrato.objects.filter(lider_contrato__grupo='lider_contrato')
    else:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    search_query = request.GET.get('search', '').strip()
    if search_query:
        contratos = contratos.filter(
            Q(cod_projeto__icontains=search_query) |
            Q(coordenador__username__icontains=search_query) |
            Q(coordenadores__username__icontains=search_query) |
            Q(cliente__nome__icontains=search_query) |
            Q(status__icontains=search_query)
        ).distinct()

    contratos = contratos.order_by('-data_inicio')

    paginator = Paginator(contratos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {'page_obj': page_obj, 'search_query': search_query}
    return render(request, 'gestao_contratos/lista_contratos.html', context)



@login_required
def lista_clientes(request):
    # Filtro inicial por grupo
    if request.user.grupo in ['suprimento', 'financeiro', 'diretoria']:
        clientes = Cliente.objects.all()
        """elif request.user.grupo == 'coordenador':
            clientes = Cliente.objects.filter(contratos__coordenador=request.user).distinct()
        elif request.user.grupo == 'lider_contrato':
            clientes = Cliente.objects.filter(contratos__lider_contrato=request.user).distinct()
        elif request.user.grupo == 'gerente':
            clientes = Cliente.objects.filter(
                contratos__coordenador__centros__in=request.user.centros.all()
            ).distinct()
        elif request.user.grupo == 'gerente_contrato':
            clientes = Cliente.objects.filter(
                contratos__lider_contrato__grupo='lider_contrato'
            ).distinct()"""
    else:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    # Campo de busca
    search_query = request.GET.get('search', '').strip()
    if search_query:
        clientes = clientes.filter(
            Q(nome__icontains=search_query) |
            Q(cpf_cnpj__icontains=search_query) |
            Q(endereco__icontains=search_query)
        )

    # Ordenação e paginação
    clientes = clientes.order_by('nome')
    paginator = Paginator(clientes, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
    }
    return render(request, 'gestao_contratos/lista_clientes.html', context)


@login_required
def lista_contratos_fornecedor(request):
    # Filtragem base por grupo de usuário
    if request.user.grupo in ['suprimento', 'financeiro', 'diretoria']:
        contratos = ContratoTerceiros.objects.all()
    elif request.user.grupo == 'coordenador':
        contratos = ContratoTerceiros.objects.filter(
            Q(coordenador=request.user) | Q(coordenadores=request.user)
        ).distinct()
    elif request.user.grupo == 'lider_contrato':
        contratos = ContratoTerceiros.objects.filter(lider_contrato=request.user)
    elif request.user.grupo == 'gerente':
        contratos = ContratoTerceiros.objects.filter(
            Q(coordenador__centros__in=request.user.centros.all()) |
            Q(coordenadores__centros__in=request.user.centros.all())
        ).distinct()
    elif request.user.grupo == 'gerente_lider':
        contratos = ContratoTerceiros.objects.filter(
            Q(coordenador__centros__in=request.user.centros.all()) |
            Q(coordenadores__centros__in=request.user.centros.all())
        ).distinct()
    elif request.user.grupo == 'gerente_contrato':
        contratos = ContratoTerceiros.objects.filter(
            lider_contrato__grupo__in=['lider_contrato','gerente_contrato', 'gerente_lider']
        )
    else:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    # Campo de busca
    search_query = request.GET.get('search', '').strip()
    if search_query:
        contratos = contratos.filter(
            Q(cod_projeto__cod_projeto__icontains=search_query) |
            Q(num_contrato__icontains=search_query) |
            Q(cod_projeto__cliente__nome__icontains=search_query) |
            Q(empresa_terceira__nome__icontains=search_query) |
            Q(coordenador__username__icontains=search_query) |
            Q(coordenadores__username__icontains=search_query) |
            Q(status__icontains=search_query) |
            Q(valor_total__icontains=search_query)
        ).distinct()

    # Ordenar e paginar
    contratos = contratos.order_by('-data_inicio')
    paginator = Paginator(contratos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
    }

    return render(request, 'gestao_contratos/lista_contratos_fornecedores.html', context)


@login_required
def lista_contratos_guarda_chuva(request):
    # Filtragem base por grupo de usuário
    if request.user.grupo in ['suprimento', 'financeiro', 'diretoria', 'coordenador', 'lider_contrato', 'gerente', 'gerente_lider', 'gerente_contrato']:
        contratos = ContratoTerceiros.objects.filter(guarda_chuva=True)
    else:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    # Campo de busca
    search_query = request.GET.get('search', '').strip()
    if search_query:
        contratos = contratos.filter(
            Q(cod_projeto__cod_projeto__icontains=search_query) |
            Q(num_contrato__icontains=search_query) |
            Q(cod_projeto__cliente__nome__icontains=search_query) |
            Q(empresa_terceira__nome__icontains=search_query) |
            Q(coordenador__username__icontains=search_query) |
            Q(status__icontains=search_query) |
            Q(valor_total__icontains=search_query)
        )

    # Ordenar e paginar
    contratos = contratos.order_by('-data_inicio')
    paginator = Paginator(contratos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
    }

    return render(request, 'gestao_contratos/lista_contratos_guarda_chuva.html', context)


@login_required
def lista_fornecedores(request):
    # Filtro base por grupo de usuário
    if request.user.grupo in ['suprimento', 'financeiro', 'diretoria']:
        fornecedores = EmpresaTerceira.objects.all()
    elif request.user.grupo == 'coordenador':
        fornecedores = EmpresaTerceira.objects.filter(
            Q(contratos__coordenador=request.user) |
            Q(contratos__coordenadores=request.user) |
            Q(contratos__os_cadastrada__status__in=['em_execucao', 'finalizada', 'paralizada']) &
            Q(contratos__os_cadastrada__coordenador=request.user)
        ).distinct()
    elif request.user.grupo == 'lider_contrato':
        fornecedores = EmpresaTerceira.objects.filter(
            Q(contratos__lider_contrato=request.user) |
            Q(contratos__os_cadastrada__status__in=['em_execucao', 'finalizada', 'paralizada']) &
            Q(contratos__os_cadastrada__lider_contrato=request.user)
        ).distinct()
    elif request.user.grupo == 'gerente':
        fornecedores = EmpresaTerceira.objects.filter(
            Q(contratos__coordenador__centros__in=request.user.centros.all()) |
            Q(contratos__coordenadores__centros__in=request.user.centros.all()) |
            Q(contratos__os_cadastrada__status__in=['em_execucao', 'finalizada', 'paralizada']) &
            Q(contratos__os_cadastrada__coordenador__centros__in=request.user.centros.all())
        ).distinct()
    elif request.user.grupo == 'gerente_lider':
        fornecedores = EmpresaTerceira.objects.filter(
            Q(contratos__coordenador__centros__in=request.user.centros.all()) |
            Q(contratos__coordenadores__centros__in=request.user.centros.all()) |
            Q(contratos__os_cadastrada__status__in=['em_execucao', 'finalizada', 'paralizada']) &
            Q(contratos__os_cadastrada__coordenador__centros__in=request.user.centros.all())
        ).distinct()
    elif request.user.grupo == 'gerente_contrato':
        fornecedores = EmpresaTerceira.objects.filter(
            Q(contratos__lider_contrato__grupo='lider_contrato') |
            Q(contratos__os_cadastrada__status__in=['em_execucao', 'finalizada', 'paralizada']) &
            Q(contratos__os_cadastrada__lider_contrato__grupo='lider_contrato')
        ).distinct()
    else:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    # Campo de busca
    search_query = request.GET.get('search', '').strip()
    if search_query:
        fornecedores = fornecedores.filter(
            Q(nome__icontains=search_query) |
            Q(cpf_cnpj__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(municipio__icontains=search_query) |
            Q(estado__icontains=search_query)
        )

    guarda_chuva = request.GET.get('guarda_chuva')
    if guarda_chuva == '1':
        fornecedores = fornecedores.filter(contratos__guarda_chuva=True)

    fornecedores = fornecedores.order_by('nome')

    # Criar indicadores para cada fornecedor
    indicadores_dict = {
        ind.empresa_terceira_id: ind
        for ind in Indicadores.objects.filter(
            empresa_terceira__in=fornecedores
        )
    }

    # Ordenação e paginação
    paginator = Paginator(fornecedores, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'guarda_chuva': guarda_chuva,
        'indicadores': indicadores_dict,
    }

    return render(request, 'gestao_contratos/lista_fornecedores.html', context)

@login_required
def cadastrar_fornecedor(request):
    if request.user.grupo not in ["lider_contrato", "suprimento", "gerente_lider", "gerente_contrato"]:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")
    if request.method == "POST":
        form = FornecedorForm(request.POST)
        if form.is_valid():
            fornecedor = form.save()

            suprimentos = User.objects.filter(grupo="suprimento", is_active=True).values_list("email", flat=True)

            if suprimentos:
                assunto = "Novo fornecedor cadastrado"
                mensagem = (
                    f"Um novo fornecedor foi cadastrado no sistema.\n\n"
                    f"Fornecedor: {fornecedor.nome}\n"
                    f"CNPJ/CPF: {fornecedor.cpf_cnpj}\n"
                    f"Cadastrado por: {request.user.get_full_name() or request.user.username}\n"
                    "Acesse o sistema HIDROGestão e complete o cadastro.\n"
                    "https://hidrogestao.pythonanywhere.com/"
                )
                try:
                    send_mail(
                        assunto, mensagem,
                        FROM_EMAIL,
                        list(suprimentos),
                        fail_silently=False,
                    )
                except Exception as e:
                    messages.warning(request, f"Erro ao enviar e-mail para suprimentos: {e}")

            return JsonResponse({
                "id": fornecedor.id,
                "nome": fornecedor.nome
            })
        return JsonResponse({"errors": form.errors}, status=400)


@login_required
def cadastrar_cliente_ajax(request):
    if request.user.grupo not in ["lider_contrato", "suprimento", "gerente_lider", "gerente_contrato"]:
        return JsonResponse({"errors": {"__all__": ["Você não tem permissão para isso."]}}, status=403)

    if request.method == "POST":
        form = ClienteForm(request.POST)
        if form.is_valid():
            cliente = form.save()
            return JsonResponse({
                "id": cliente.id,
                "nome": cliente.nome,
            })
        return JsonResponse({"errors": form.errors}, status=400)

    return JsonResponse({"errors": {"__all__": ["Método inválido."]}}, status=405)


@login_required
def cadastrar_nf_cliente(request, pk):
    if request.user.grupo not in ["financeiro", "suprimento"]:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    contrato = get_object_or_404(Contrato, pk=pk)

    if request.method == "POST":
        form = NFClienteForm(request.POST, request.FILES)
        if form.is_valid():
            nf = form.save(commit=False)
            nf.contrato = contrato
            nf.inserido_por = request.user
            nf.save()

            messages.success(request, "Nota Fiscal cadastrada com sucesso!")
        else:
            messages.error(request, "Erro ao salvar Nota Fiscal. Verifique os campos.")

    return redirect("contrato_cliente_detalhe", pk=pk)


@login_required
def contrato_cliente_detalhe(request, pk):
    contrato = get_object_or_404(Contrato, pk=pk)

    # ---- Buscar eventos ligados a esse contrato ----
    eventos = Evento.objects.filter(contrato_terceiro__cod_projeto=contrato)

    total_contrato = contrato.valor_total or 0

    # --- Criar DataFrame para segmentação por fornecedor ---
    df = pd.DataFrame(list(eventos.values(
        'empresa_terceira__nome',
        'valor_previsto',
        'valor_pago'
    )))

    if df.empty:
        resumo = pd.DataFrame(columns=['empresa_terceira__nome', 'valor_previsto', 'valor_pago'])
        total_previsto = total_pago = 0
    else:
        df = df.fillna(0)
        resumo = df.groupby('empresa_terceira__nome')[['valor_previsto', 'valor_pago']].sum().reset_index()
        total_previsto = resumo['valor_previsto'].sum()
        total_pago = resumo['valor_pago'].sum()

    # ---- Calcular percentuais ----
    if total_contrato > 0:
        perc_previsto = (total_previsto / total_contrato) * 100
        perc_pago = (total_pago / total_contrato) * 100
        perc_restante = 100 - max(perc_previsto, perc_pago)
    else:
        perc_previsto = perc_pago = perc_restante = 0

    # ---- Cores diferentes para cada fornecedor ----
    fornecedores = resumo['empresa_terceira__nome'].tolist()
    cores = px.colors.qualitative.Plotly * (len(fornecedores) // 10 + 1)

    # ---- Gráfico de barras ----
    fig = go.Figure()

    # Barra 1: Valor total do contrato
    fig.add_trace(go.Bar(
        x=['Valor Total do Contrato'],
        y=[total_contrato],
        name='Total do Contrato',
        text=[f"R$ {total_contrato:,.2f}"],
        textposition='inside',
        marker_color='#007bff'
    ))

    # Barras empilhadas: Valor Previsto por fornecedor
    for i, row in enumerate(resumo.itertuples()):
        fig.add_trace(go.Bar(
            x=['Valor Previsto para Fornecedores'],
            y=[row.valor_previsto],
            name=f"{row.empresa_terceira__nome}",
            text=[f"R$ {row.valor_previsto:,.2f}"],
            textposition='inside',
            hovertemplate=f"<b>{row.empresa_terceira__nome}</b><br>Previsto: R$ %{{y:,.2f}}<extra></extra>",
            marker_color=cores[i]
        ))

    # Barras empilhadas: Valor Pago por fornecedor
    for i, row in enumerate(resumo.itertuples()):
        fig.add_trace(go.Bar(
            x=['Valor Pago a Fornecedores'],
            y=[row.valor_pago],
            name=f"PAGO - {row.empresa_terceira__nome}",
            text=[f"R$ {row.valor_pago:,.2f}"],
            textposition='inside',
            hovertemplate=f"<b>{row.empresa_terceira__nome}</b><br>Pago: R$ %{{y:,.2f}}<extra></extra>",
            marker_color=cores[i]
        ))

    # ---- Adicionar valor total no topo das barras empilhadas ----
    if not resumo.empty:
        fig.add_annotation(
            x='Valor Previsto para Fornecedores',
            y=total_previsto,
            text=f"R$ {total_previsto:,.2f}",
            showarrow=False,
            yshift=10,
            font=dict(size=12, color='black')
        )
        fig.add_annotation(
            x='Valor Pago a Fornecedores',
            y=total_pago,
            text=f"R$ {total_pago:,.2f}",
            showarrow=False,
            yshift=10,
            font=dict(size=12, color='black')
        )

    fig.update_layout(
        title=f'Resumo Financeiro do Contrato ({contrato.cod_projeto})',
        yaxis_title='Valor (R$)',
        xaxis_title='Categoria',
        template='plotly_white',
        barmode='stack',
        height=500,
        legend=dict(orientation="h", y=-0.3, x=0)
    )

    grafico_contrato = fig.to_html(full_html=False)

    # ---- Texto com resumo percentual ----
    resumo_percentual = {
        'pago': round(perc_pago, 1),
        'previsto': round(perc_previsto, 1),
        'restante': round(perc_restante, 1)
    }

    # ---- Listagem das NFs já cadastradas para esse contrato ----
    open_modal_nf = False
    form_nf = NFClienteForm()
    nf_list = contrato.nota_fiscal.all().order_by("-data_emissao")


    # ---- Controle de edição ----
    if request.user.grupo == "suprimento":
        forms_edit = {}
        for nf in nf_list:
            forms_edit[nf.id] = NFClienteForm(instance=nf, prefix=f"edit_{nf.id}")
        # POST vindo do modal de NF
        if request.method == "POST" and request.user.is_authenticated and "submit_nf" in request.POST:
            form_nf = NFClienteForm(request.POST, request.FILES)
            if form_nf.is_valid():
                nf = form_nf.save(commit=False)
                nf.contrato = contrato
                nf.inserido_por = request.user
                nf.save()
                messages.success(request, "Nota Fiscal cadastrada com sucesso!")
                return redirect("contrato_cliente_detalhe", pk=pk)
            else:
                open_modal_nf = True
                messages.error(request, "Erro ao cadastrar Nota Fiscal. Verifique os campos abaixo.")

        # POST do formulário do contrato
        if request.method == "POST" and request.user.is_authenticated and "submit_contrato" in request.POST:
            form = ContratoForm(request.POST, instance=contrato)
            if form.is_valid():
                form.save()
                messages.success(request, "Contrato atualizado com sucesso!")
                return redirect("lista_contratos")
            else:
                messages.error(request, "Erro ao atualizar contrato.")
        else:
            form = ContratoForm(instance=contrato)

        return render(request, 'contratos/contrato_detail_edit.html', {
            'form': form,
            'contrato': contrato,
            'grafico_contrato': grafico_contrato,
            'resumo_percentual': resumo_percentual,
            'form_nf': form_nf,
            'nf_list': nf_list,
            'open_modal_nf': open_modal_nf,
            'forms_edit': forms_edit,
        })

    if request.user.grupo in ["financeiro"]:
        forms_edit = {}
        for nf in nf_list:
            forms_edit[nf.id] = NFClienteForm(instance=nf, prefix=f"edit_{nf.id}")
        if request.method == "POST" and request.user.is_authenticated:
            # executa validação do form vindo do modal
            form_nf = NFClienteForm(request.POST, request.FILES)
            if form_nf.is_valid():
                nf = form_nf.save(commit=False)
                nf.contrato = contrato
                nf.inserido_por = request.user
                nf.save()
                messages.success(request, "Nota Fiscal cadastrada com sucesso!")
                return redirect("contrato_cliente_detalhe", pk=pk)
            else:
                # manter o modal aberto e mostrar erros
                open_modal_nf = True
                messages.error(request, "Erro ao cadastrar Nota Fiscal. Verifique os campos e corrija os erros abaixo.")

        return render(request, "contratos/contrato_detail.html", {
            'contrato': contrato,
            'grafico_contrato': grafico_contrato,
            'resumo_percentual': resumo_percentual,
            'form_nf': form_nf,
            'nf_list': nf_list,
            'open_modal_nf': open_modal_nf,
            'forms_edit': forms_edit,
        })

    return render(request, "contratos/contrato_detail.html", {
        'contrato': contrato,
        'grafico_contrato': grafico_contrato,
        'resumo_percentual': resumo_percentual,
        'nf_list': nf_list,
    })


@login_required
def editar_nf_cliente(request, pk):
    nf = get_object_or_404(NFCliente, pk=pk)

    if request.user.grupo not in ["suprimento","financeiro"]:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    if request.method == "POST":
        prefix = f"edit_{nf.id}"
        form = NFClienteForm(request.POST, request.FILES, instance=nf, prefix=prefix)

        if form.is_valid():
            form.save()
            messages.success(request, "NF atualizada com sucesso!")
        else:
            print("ERROS AO EDITAR NF:", form.errors)
            messages.error(request, "Erro ao atualizar a Nota Fiscal.")

    return redirect("contrato_cliente_detalhe", pk=nf.contrato.pk)




@login_required
def excluir_nf_cliente(request, pk):
    nf = get_object_or_404(NFCliente, pk=pk)

    if request.user.grupo not in ["financeiro", "suprimento"]:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    contrato_pk = nf.contrato.pk
    nf.delete()
    messages.success(request, "Nota Fiscal excluÃ­da com sucesso!")

    return redirect("contrato_cliente_detalhe", pk=contrato_pk)


@login_required
def contrato_fornecedor_detalhe(request, pk):
    contrato = get_object_or_404(ContratoTerceiros, pk=pk)
    fornecedor = contrato.empresa_terceira

    if request.user.grupo == "gerente_lider" and not user_shares_center_with_contract_coordinators(request.user, contrato):
        messages.error(request, "Você não tem permissão para isso.")
        return redirect("home")

    indicadores_geral, _ = Indicadores.objects.get_or_create(empresa_terceira=fornecedor)
    indicadores_contrato = contrato
    contratos_fornecedor = ContratoTerceiros.objects.filter(
        empresa_terceira=contrato.empresa_terceira,
        status='ativo'
    )

    ordens_servico = OS.objects.filter(
        contrato=contrato
    ).order_by('-criado_em')


    proposta_fornecedor = None
    if contrato.prospeccao and contrato.prospeccao.fornecedor_escolhido:
        proposta_fornecedor = contrato.prospeccao.propostas.filter(
            fornecedor=contrato.prospeccao.fornecedor_escolhido
        ).first()

    eventos = Evento.objects.filter(contrato_terceiro=contrato).order_by("data_prevista")
    aditivos = contrato.aditivos.select_related(
        "solicitado_por",
        "documento_enviado_por",
        "documento_assinado_enviado_por",
    ).all()
    aditivo_ativo = next(
        (
            item
            for item in aditivos
            if not item.aprovado_totalmente and not item.solicitacao_reprovada_por_alguem
        ),
        None,
    )
    ultimo_aditivo_aprovado = (
        aditivos.filter(arquivo_aditivo_assinado__isnull=False)
        .exclude(arquivo_aditivo_assinado="")
        .order_by("-documento_assinado_enviado_em", "-criado_em")
        .first()
    )
    valor_total_contrato_vigente = (
        ultimo_aditivo_aprovado.novo_valor_total
        if ultimo_aditivo_aprovado and ultimo_aditivo_aprovado.novo_valor_total is not None
        else (contrato.valor_total or Decimal("0.00"))
    )
    total_eventos_previstos = eventos.aggregate(total=Sum("valor_previsto"))["total"] or Decimal("0.00")
    exibir_warning_valor_eventos = total_eventos_previstos != valor_total_contrato_vigente

    df = pd.DataFrame(list(eventos.values("data_prevista_pagamento", "valor_previsto", "valor_pago", "data_pagamento")))

    plot_div = None
    if not df.empty:
        # Converte explicitamente para numérico antes de preencher nulos
        df["valor_previsto"] = pd.to_numeric(df["valor_previsto"], errors="coerce").fillna(0.0)
        df["valor_pago"] = pd.to_numeric(df["valor_pago"], errors="coerce").fillna(0.0)
        df["data_prevista_pagamento"] = pd.to_datetime(
            df["data_prevista_pagamento"], errors="coerce"
        ).dt.normalize()
        df["data_pagamento"] = pd.to_datetime(
            df["data_pagamento"], errors="coerce"
        ).dt.normalize()

        # Ordenar pelas datas
        df = df.sort_values("data_prevista_pagamento")

        # Calcular valores acumulados
        df["valor_previsto_acum"] = df["valor_previsto"].cumsum()
        df["valor_pago_acum"] = df["valor_pago"].cumsum()

        # Criar gráfico
        trace_previsto = go.Scatter(
            x=df["data_prevista_pagamento"],
            y=df["valor_previsto_acum"],
            mode="lines+markers",
            name="Previsto (Acumulado)",
            line=dict(color="orange"),
            hovertemplate="Data: %{x|%d/%m/%Y}<br>Valor acumulado: R$ %{y:,.2f}<extra></extra>",
        )

        trace_pago = go.Scatter(
            x=df["data_pagamento"],
            y=df["valor_pago_acum"],
            mode="lines+markers",
            name="Pago (Acumulado)",
            line=dict(color="green"),
            hovertemplate="Data: %{x|%d/%m/%Y}<br>Valor acumulado: R$ %{y:,.2f}<extra></extra>",
        )

        layout = go.Layout(
            title="Evolução Acumulada de Pagamentos",
            xaxis=dict(title="Data", tickformat="%d/%m/%Y", hoverformat="%d/%m/%Y"),
            yaxis=dict(title="Valor (R$)"),
            template="plotly_white"
        )

        fig = go.Figure(data=[trace_previsto, trace_pago], layout=layout)
        plot_div = plot(fig, auto_open=False, output_type="div")

    # --- GRÁFICO DE COMPARAÃ‡ÃƒO ---
    if contrato.guarda_chuva:
        valor_total_contrato = contrato.valor_total or 0

        total_os = OS.objects.filter(
            contrato=contrato
        ).aggregate(Sum("valor_pago"))["valor_pago__sum"] or 0

        saldo_contrato = valor_total_contrato - total_os

        fig_comp = go.Figure()

        fig_comp.add_trace(go.Bar(
            x=["Valor Total Contrato", "Total Pago em OS", "Saldo Disponí­vel"],
            y=[valor_total_contrato, total_os, saldo_contrato],
            text=[
                f"R$ {valor_total_contrato:,.2f}",
                f"R$ {total_os:,.2f}",
                f"R$ {saldo_contrato:,.2f}",
            ],
            textposition="auto",
            marker_color=["#007bff", "#ffc107", "#28a745"]
        ))

        fig_comp.update_layout(
            title="Controle Financeiro do Contrato Guarda-Chuva",
            yaxis_title="Valor (R$)",
            xaxis_title="Categoria",
            template="plotly_white",
            height=450
        )

    else:
        # GRÁFICO ANTIGO
        if contrato.cod_projeto:
            contrato_cliente = contrato.cod_projeto
            valor_cliente = contrato_cliente.valor_total or 0
        else:
            valor_cliente = 0

        total_previsto = eventos.aggregate(Sum("valor_previsto"))["valor_previsto__sum"] or 0
        total_pago = eventos.aggregate(Sum("valor_pago"))["valor_pago__sum"] or 0

        perc_previsto = (total_previsto / valor_cliente * 100) if valor_cliente else 0
        perc_pago = (total_pago / valor_cliente * 100) if valor_cliente else 0

        fig_comp = go.Figure()

        fig_comp.add_trace(go.Bar(
            x=["Valor Contrato Cliente", "Valor Previsto Fornecedor", "Valor Pago Fornecedor"],
            y=[valor_cliente, total_previsto, total_pago],
            text=[
                f"R$ {valor_cliente:,.2f}<br>100%",
                f"R$ {total_previsto:,.2f}<br>{perc_previsto:.1f}%",
                f"R$ {total_pago:,.2f}<br>{perc_pago:.1f}%"
            ],
            textposition="auto",
            marker_color=["#007bff", "#ffc107", "#28a745"]
        ))

        fig_comp.update_layout(
            title="Comparativo Financeiro: Cliente & Fornecedor",
            yaxis_title="Valor (R$)",
            xaxis_title="Categoria",
            template="plotly_white",
            height=450
        )

    comparativo_div = plot(fig_comp, auto_open=False, output_type="div")


    return render(
        request,
        "contratos/contrato_fornecedor_detail.html",
        {
            "contrato": contrato,
            "proposta_fornecedor": proposta_fornecedor,
            "eventos": eventos,
            "plot_div": plot_div,
            "comparativo_div": comparativo_div,
            "indicadores_geral": indicadores_geral,
            "indicadores_contrato": indicadores_contrato,
            "contratos_ativos": contratos_fornecedor.count(),
            "ordens_servico": ordens_servico,
            "aditivos": aditivos,
            "aditivo_ativo": aditivo_ativo,
            "can_request_addendum": can_user_request_contract_addendum(request.user, contrato),
            "valor_total_contrato_vigente": valor_total_contrato_vigente,
            "total_eventos_previstos": total_eventos_previstos,
            "exibir_warning_valor_eventos": exibir_warning_valor_eventos,
        },
    )


@login_required
def solicitar_aditivo_contrato(request, pk):
    contrato = get_object_or_404(ContratoTerceiros, pk=pk)

    if not can_user_request_contract_addendum(request.user, contrato):
        messages.error(request, "Você não tem permissão para solicitar aditivo neste contrato.")
        return redirect("contrato_fornecedor_detalhe", pk=pk)

    aditivo_em_aberto = next(
        (
            item
            for item in contrato.aditivos.all()
            if not item.aprovado_totalmente and not item.solicitacao_reprovada_por_alguem
        ),
        None,
    )
    if aditivo_em_aberto:
        messages.warning(request, "Já existe um aditivo em andamento para este contrato.")
        return redirect("detalhes_aditivo_contrato", pk=aditivo_em_aberto.pk)

    if request.method == "POST":
        form = SolicitacaoAditivoContratoTerceiroForm(request.POST)
        if form.is_valid():
            aditivo = form.save(commit=False)
            aditivo.contrato = contrato
            aditivo.solicitado_por = request.user
            aditivo.valor_total_anterior = contrato.valor_total
            aditivo.data_fim_anterior = contrato.data_fim
            aditivo.save()

            destinatarios = list(
                User.objects.filter(grupo__in=["gerente_contrato", "diretoria"], is_active=True)
                .exclude(email__isnull=True)
                .exclude(email__exact="")
                .values_list("email", flat=True)
                .distinct()
            )
            if destinatarios:
                send_mail(
                    f"Nova solicitação de aditivo - Contrato {contrato.num_contrato}",
                    (
                        f"Olá,\n\n"
                        f"Foi solicitada uma nova análise de aditivo para o contrato abaixo:\n\n"
                        f"Projeto: {contrato.cod_projeto}\n"
                        f"Contrato: {contrato.num_contrato} - {contrato.empresa_terceira}\n"
                        f"Solicitante: {request.user.get_full_name() or request.user.username}\n"
                        f"Novo valor total: R$ {aditivo.novo_valor_total or 0}\n"
                        f"Nova data fim: {aditivo.nova_data_fim.strftime('%d/%m/%Y') if aditivo.nova_data_fim else 'Não informada'}\n"
                        f"Motivo: {aditivo.motivo}\n\n"
                        f"A solicitação aguarda aprovação do gerente de contrato e da diretoria.\n\n"
                        f"Atenciosamente,\n"
                        f"Sistema HIDROGestão"
                    ),
                    FROM_EMAIL,
                    destinatarios,
                    fail_silently=False,
                )

            messages.success(request, "Solicitação de aditivo registrada e enviada para aprovação.")
            return redirect("detalhes_aditivo_contrato", pk=aditivo.pk)
    else:
        form = SolicitacaoAditivoContratoTerceiroForm(
            initial={
                "novo_valor_total": contrato.valor_total,
                "nova_data_fim": contrato.data_fim,
            }
        )

    return render(
        request,
        "contratos/solicitar_aditivo_contrato.html",
        {"contrato": contrato, "form": form, "timeline_steps": None},
    )


@login_required
def detalhes_aditivo_contrato(request, pk):
    aditivo = get_object_or_404(AditivoContratoTerceiro, pk=pk)

    if not can_user_view_addendum(request.user, aditivo):
        messages.error(request, "Você não tem permissão para visualizar este aditivo.")
        return redirect("home")

    return render(
        request,
        "contratos/detalhes_aditivo_contrato.html",
        {
            "aditivo": aditivo,
            "contrato": aditivo.contrato,
            "timeline_steps": build_addendum_timeline(aditivo),
            "can_approve_request_as_gerente": can_user_approve_addendum_request_as_gerente(request.user, aditivo),
            "can_approve_request_as_diretoria": can_user_approve_addendum_request_as_diretoria(request.user, aditivo),
            "can_upload_draft": can_user_upload_addendum_draft(request.user, aditivo),
            "can_approve_draft": can_user_approve_addendum_draft(request.user, aditivo),
            "can_upload_signed": can_user_upload_signed_addendum(request.user, aditivo),
        },
    )


@login_required
def avaliar_solicitacao_aditivo_contrato(request, pk):
    aditivo = get_object_or_404(AditivoContratoTerceiro, pk=pk)
    contrato = aditivo.contrato

    if not can_user_view_addendum(request.user, aditivo):
        messages.error(request, "Você não tem permissão para avaliar esta solicitação de aditivo.")
        return redirect("home")

    if request.method == "POST":
        acao = request.POST.get("acao")
        justificativa = request.POST.get("justificativa", "").strip()

        if acao in ["reprovar_gerente", "reprovar_diretoria"] and not justificativa:
            messages.error(request, "A justificativa é obrigatória para reprovar a solicitação de aditivo.")
            return redirect("avaliar_solicitacao_aditivo_contrato", pk=aditivo.pk)

        if request.user.grupo == "gerente_contrato":
            if acao not in ["aprovar_gerente", "reprovar_gerente"]:
                messages.error(request, "Ação inválida para a gerência de contrato.")
                return redirect("avaliar_solicitacao_aditivo_contrato", pk=aditivo.pk)
            if aditivo.status_gerente != "pendente":
                messages.warning(request, "A gerência de contrato já avaliou esta solicitação.")
                return redirect("detalhes_aditivo_contrato", pk=aditivo.pk)
            aditivo.status_gerente = "aprovado" if acao == "aprovar_gerente" else "reprovado"
            aditivo.data_aprovacao_gerente = timezone.now()
            aditivo.justificativa_reprovacao_gerente = justificativa if acao == "reprovar_gerente" else None
        elif request.user.grupo == "diretoria":
            if acao not in ["aprovar_diretoria", "reprovar_diretoria"]:
                messages.error(request, "Ação inválida para a diretoria.")
                return redirect("avaliar_solicitacao_aditivo_contrato", pk=aditivo.pk)
            if aditivo.status_diretoria != "pendente":
                messages.warning(request, "A diretoria já avaliou esta solicitação.")
                return redirect("detalhes_aditivo_contrato", pk=aditivo.pk)
            aditivo.status_diretoria = "aprovado" if acao == "aprovar_diretoria" else "reprovado"
            aditivo.data_aprovacao_diretoria = timezone.now()
            aditivo.justificativa_reprovacao_diretoria = justificativa if acao == "reprovar_diretoria" else None
        else:
            messages.error(request, "Você não tem permissão para avaliar esta solicitação de aditivo.")
            return redirect("detalhes_aditivo_contrato", pk=aditivo.pk)

        aditivo.save()

        if aditivo.solicitacao_reprovada_por_alguem:
            destinatarios = []
            if aditivo.solicitado_por and aditivo.solicitado_por.email:
                destinatarios.append(aditivo.solicitado_por.email)
            if destinatarios:
                send_mail(
                    f"Solicitação de aditivo reprovada - Contrato {contrato.num_contrato}",
                    (
                        f"Olá,\n\n"
                        f"A solicitação de aditivo do contrato {contrato.num_contrato} foi reprovada.\n\n"
                        f"Justificativa: {justificativa}\n\n"
                        f"Atenciosamente,\n"
                        f"Sistema HIDROGestão"
                    ),
                    FROM_EMAIL,
                    list(dict.fromkeys(destinatarios)),
                    fail_silently=False,
                )
        elif aditivo.solicitacao_aprovada_totalmente:
            destinatarios = list(
                User.objects.filter(grupo="suprimento", is_active=True)
                .exclude(email__isnull=True)
                .exclude(email__exact="")
                .values_list("email", flat=True)
                .distinct()
            )
            if destinatarios:
                send_mail(
                    f"Solicitação de aditivo aprovada - Contrato {contrato.num_contrato}",
                    (
                        f"Olá, equipe de Suprimento!\n\n"
                        f"A solicitação de aditivo do contrato {contrato.num_contrato} foi aprovada "
                        f"pela gerência de contrato e pela diretoria.\n\n"
                        f"Favor inserir a minuta do aditivo no sistema.\n\n"
                        f"Atenciosamente,\n"
                        f"Sistema HIDROGestão"
                    ),
                    FROM_EMAIL,
                    destinatarios,
                    fail_silently=False,
                )

        messages.success(request, "Ação registrada com sucesso.")
        return redirect("detalhes_aditivo_contrato", pk=aditivo.pk)

    return render(
        request,
        "contratos/avaliar_solicitacao_aditivo_contrato.html",
        {
            "aditivo": aditivo,
            "contrato": contrato,
            "timeline_steps": build_addendum_timeline(aditivo),
            "can_approve_request_as_gerente": can_user_approve_addendum_request_as_gerente(request.user, aditivo),
            "can_approve_request_as_diretoria": can_user_approve_addendum_request_as_diretoria(request.user, aditivo),
        },
    )


@login_required
def enviar_documento_aditivo_contrato(request, pk):
    aditivo = get_object_or_404(AditivoContratoTerceiro, pk=pk)

    if request.user.grupo != "suprimento":
        messages.error(request, "Somente o Suprimento pode inserir a minuta do aditivo.")
        return redirect("detalhes_aditivo_contrato", pk=aditivo.pk)

    if not aditivo.solicitacao_aprovada_totalmente:
        messages.error(request, "A solicitação do aditivo ainda não foi aprovada por completo.")
        return redirect("detalhes_aditivo_contrato", pk=aditivo.pk)

    if aditivo.aprovado_totalmente:
        messages.warning(request, "Este aditivo já foi concluído e não pode mais ser alterado.")
        return redirect("detalhes_aditivo_contrato", pk=aditivo.pk)

    if request.method == "POST":
        form = DocumentoAditivoContratoTerceiroForm(request.POST, request.FILES, instance=aditivo)
        if form.is_valid():
            aditivo = form.save(commit=False)
            aditivo.documento_enviado_por = request.user
            aditivo.documento_enviado_em = timezone.now()
            aditivo.status_lider = "pendente"
            aditivo.data_aprovacao_lider = None
            aditivo.justificativa_reprovacao_lider = None
            aditivo.save()

            destinatarios = list(
                User.objects.filter(grupo="gerente_contrato", is_active=True)
                .exclude(email__isnull=True)
                .exclude(email__exact="")
                .values_list("email", flat=True)
                .distinct()
            )
            if destinatarios:
                send_mail(
                    f"Minuta de aditivo disponível - Contrato {aditivo.contrato.num_contrato}",
                    (
                        f"Olá,\n\n"
                        f"O Suprimento anexou a minuta do aditivo do contrato abaixo:\n\n"
                        f"Projeto: {aditivo.contrato.cod_projeto}\n"
                        f"Contrato: {aditivo.contrato.num_contrato} - {aditivo.contrato.empresa_terceira}\n"
                        f"Novo valor total: R$ {aditivo.novo_valor_total or 0}\n"
                        f"Nova data fim: {aditivo.nova_data_fim.strftime('%d/%m/%Y') if aditivo.nova_data_fim else 'Não informada'}\n\n"
                        f"Favor avaliar a minuta do aditivo.\n\n"
                        f"Atenciosamente,\n"
                        f"Sistema HIDROGestão"
                    ),
                    FROM_EMAIL,
                    destinatarios,
                    fail_silently=False,
                )

            messages.success(request, "Minuta do aditivo enviada para aprovação.")
            return redirect("detalhes_aditivo_contrato", pk=aditivo.pk)
    else:
        form = DocumentoAditivoContratoTerceiroForm(instance=aditivo)

    return render(
        request,
        "contratos/enviar_documento_aditivo.html",
        {"aditivo": aditivo, "contrato": aditivo.contrato, "form": form, "timeline_steps": build_addendum_timeline(aditivo)},
    )


@login_required
def avaliar_aditivo_contrato(request, pk):
    aditivo = get_object_or_404(AditivoContratoTerceiro, pk=pk)
    return redirect("detalhes_aditivo_contrato", pk=aditivo.pk)


@login_required
def avaliar_minuta_aditivo_contrato(request, pk):
    aditivo = get_object_or_404(AditivoContratoTerceiro, pk=pk)
    contrato = aditivo.contrato

    if request.user.grupo != "gerente_contrato":
        messages.error(request, "Somente a gerência de contrato pode avaliar a minuta do aditivo.")
        return redirect("detalhes_aditivo_contrato", pk=aditivo.pk)

    if not aditivo.tem_documento:
        messages.error(request, "A minuta do aditivo ainda não foi anexada pelo Suprimento.")
        return redirect("detalhes_aditivo_contrato", pk=aditivo.pk)

    if request.method == "POST":
        acao = request.POST.get("acao")
        justificativa = request.POST.get("justificativa", "").strip()

        if acao not in ["aprovar_minuta", "reprovar_minuta"]:
            messages.error(request, "Ação inválida para a minuta do aditivo.")
            return redirect("avaliar_minuta_aditivo_contrato", pk=aditivo.pk)

        if acao == "reprovar_minuta" and not justificativa:
            messages.error(request, "A justificativa é obrigatória para reprovar a minuta do aditivo.")
            return redirect("avaliar_minuta_aditivo_contrato", pk=aditivo.pk)

        if aditivo.status_lider != "pendente":
            messages.warning(request, "A minuta do aditivo já foi avaliada.")
            return redirect("detalhes_aditivo_contrato", pk=aditivo.pk)

        aditivo.status_lider = "aprovado" if acao == "aprovar_minuta" else "reprovado"
        aditivo.data_aprovacao_lider = timezone.now()
        aditivo.justificativa_reprovacao_lider = justificativa if acao == "reprovar_minuta" else None
        aditivo.save()

        destinatarios = list(
            User.objects.filter(grupo="suprimento", is_active=True)
            .exclude(email__isnull=True)
            .exclude(email__exact="")
            .values_list("email", flat=True)
            .distinct()
        )
        if destinatarios:
            if acao == "aprovar_minuta":
                assunto = f"Minuta de aditivo aprovada - Contrato {contrato.num_contrato}"
                mensagem = (
                    f"Olá, equipe de Suprimento!\n\n"
                    f"A minuta do aditivo do contrato {contrato.num_contrato} foi aprovada pela gerência de contrato.\n\n"
                    f"Favor anexar o documento do aditivo assinado.\n\n"
                    f"Atenciosamente,\n"
                    f"Sistema HIDROGestão"
                )
            else:
                assunto = f"Minuta de aditivo reprovada - Contrato {contrato.num_contrato}"
                mensagem = (
                    f"Olá, equipe de Suprimento!\n\n"
                    f"A minuta do aditivo do contrato {contrato.num_contrato} foi reprovada pela gerência de contrato.\n\n"
                    f"Justificativa: {justificativa}\n\n"
                    f"Favor anexar uma nova versão da minuta.\n\n"
                    f"Atenciosamente,\n"
                    f"Sistema HIDROGestão"
                )
            send_mail(assunto, mensagem, FROM_EMAIL, destinatarios, fail_silently=False)

        messages.success(request, "Ação registrada com sucesso.")
        return redirect("detalhes_aditivo_contrato", pk=aditivo.pk)

    return render(
        request,
        "contratos/avaliar_minuta_aditivo_contrato.html",
        {
            "aditivo": aditivo,
            "contrato": contrato,
            "timeline_steps": build_addendum_timeline(aditivo),
        },
    )


@login_required
def enviar_aditivo_assinado_contrato(request, pk):
    aditivo = get_object_or_404(AditivoContratoTerceiro, pk=pk)

    if request.user.grupo != "suprimento":
        messages.error(request, "Somente o Suprimento pode inserir o aditivo assinado.")
        return redirect("detalhes_aditivo_contrato", pk=aditivo.pk)

    if not aditivo.minuta_aprovada:
        messages.error(request, "A minuta do aditivo ainda não foi aprovada pela gerência de contrato.")
        return redirect("detalhes_aditivo_contrato", pk=aditivo.pk)

    if request.method == "POST":
        form = DocumentoAditivoAssinadoContratoTerceiroForm(request.POST, request.FILES, instance=aditivo)
        if form.is_valid():
            aditivo = form.save(commit=False)
            aditivo.documento_assinado_enviado_por = request.user
            aditivo.documento_assinado_enviado_em = timezone.now()
            aditivo.save()

            if aditivo.nova_data_fim:
                aditivo.contrato.data_fim = aditivo.nova_data_fim
            if aditivo.novo_valor_total is not None:
                aditivo.contrato.valor_total = aditivo.novo_valor_total
            aditivo.contrato.save(update_fields=["data_fim", "valor_total"])

            destinatarios = get_addendum_completion_notification_emails(aditivo)
            if destinatarios:
                assunto = f"Aditivo concluido - Contrato {aditivo.contrato.num_contrato}"
                mensagem = (
                    "Prezados,\n\n"
                    f"O processo de aditivo do contrato {aditivo.contrato.num_contrato} foi concluido com sucesso.\n\n"
                    f"Projeto: {aditivo.contrato.cod_projeto}\n"
                    f"Fornecedor: {aditivo.contrato.empresa_terceira}\n"
                    f"Novo valor total: R$ {aditivo.contrato.valor_total or 0}\n"
                    f"Nova data fim: {format_date_br(aditivo.contrato.data_fim)}\n\n"
                    "O documento assinado ja foi anexado no sistema e as informacoes do contrato foram atualizadas.\n\n"
                    "Atenciosamente,\n"
                    "Sistema HIDROGestão"
                )
                send_mail(assunto, mensagem, FROM_EMAIL, destinatarios, fail_silently=False)

            messages.success(request, "Documento do aditivo assinado inserido com sucesso.")
            return redirect("contrato_fornecedor_detalhe", pk=aditivo.contrato_id)
    else:
        form = DocumentoAditivoAssinadoContratoTerceiroForm(instance=aditivo)

    return render(
        request,
        "contratos/enviar_aditivo_assinado_contrato.html",
        {
            "aditivo": aditivo,
            "contrato": aditivo.contrato,
            "form": form,
            "timeline_steps": build_addendum_timeline(aditivo),
        },
    )


@login_required
def contrato_fornecedor_editar(request, pk):
    contrato = get_object_or_404(ContratoTerceiros, pk=pk)

    # PermissÃµes
    if request.user.grupo not in ["suprimento"]:
        messages.error(request, " Você não tem permissão para editar contratos.")
        return redirect("contrato_fornecedor_detalhe", pk=pk)

    if request.method == "POST":
        # IncluÃ­mos request.FILES para que o arquivo seja capturado
        form = ContratoFornecedorForm(request.POST, request.FILES, instance=contrato)
        print("POST:", request.POST)
        print("FILES:", request.FILES)
        if form.is_valid():
            # Salva o formulário e o arquivo enviado
            form.save()
            messages.success(request, "Contrato atualizado com sucesso!")
            return redirect("contrato_fornecedor_detalhe", pk=pk)
        else:
            messages.error(
                request, " Ocorreu um erro ao atualizar o contrato. Verifique os campos."
            )
    else:
        form = ContratoFornecedorForm(instance=contrato)

    return render(
        request,
        "contratos/contrato_fornecedor_detail_edit.html",
        {"form": form, "contrato": contrato},
    )


@login_required
def cliente_detalhe(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    contratos = cliente.contratos.all()  # pega todos os contratos do cliente

    if request.user.grupo in ["suprimento", "financeiro"]:
        if request.method == "POST":
            form = ClienteForm(request.POST, instance=cliente)
            if form.is_valid():
                form.save()
                messages.success(request, "Dados do Cliente atualizado com sucesso!")
                return redirect("lista_clientes")
            else:
                messages.error(request, " Ocorreu um erro ao atualizar o contrato. Verifique os campos e tente novamente.")
        else:
            form = ClienteForm(instance=cliente)
        return render(
            request,
            'clientes/cliente_detail_edit.html',
            {'form': form, 'cliente': cliente, 'contratos': contratos}
        )

    return render(
        request,
        'clientes/cliente_detail.html',
        {'cliente': cliente, 'contratos': contratos}
    )



@login_required
def fornecedor_detalhe(request, pk):
    fornecedor = get_object_or_404(EmpresaTerceira, pk=pk)
    indicadores_geral = Indicadores(empresa_terceira=fornecedor)
    if request.user.grupo in ["suprimento", "financeiro", "diretoria"]:
        contratos = fornecedor.contratos.all()
        os = OS.objects.filter(contrato__empresa_terceira=fornecedor)
    elif request.user.grupo == "coordenador":
        contratos = fornecedor.contratos.filter(coordenador=request.user)
        os = OS.objects.filter(contrato__empresa_terceira=fornecedor, coordenador=request.user)
    elif request.user.grupo == "lider_contrato":
        contratos = fornecedor.contratos.filter(lider_contrato=request.user)
        os = OS.objects.filter(contrato__empresa_terceira=fornecedor, lider_contrato=request.user)
    elif request.user.grupo == "gerente_contrato":
        contratos = fornecedor.contratos.filter(lider_contrato__grupo__in=['lider_contrato', 'gerente_contrato'])
        os = OS.objects.filter(contrato__empresa_terceira=fornecedor, lider_contrato__grupo__in=['lider_contrato', 'gerente_contrato'])
    elif request.user.grupo == "gerente":
        contratos = fornecedor.contratos.filter(coordenador__centros__in=request.user.centros.all())
        os = OS.objects.filter(contrato__empresa_terceira=fornecedor, coordenador__centros__in=request.user.centros.all())
    elif request.user.grupo == "gerente_lider":
        contratos = fornecedor.contratos.filter(
            coordenador__centros__in=request.user.centros.all()
        ).distinct()
        os = OS.objects.filter(
            contrato__empresa_terceira=fornecedor,
            coordenador__centros__in=request.user.centros.all()
        ).distinct()
    else:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    if request.user.grupo in ["suprimento", "financeiro"]:
        if request.method == "POST":
            form = FornecedorForm(request.POST, instance=fornecedor)
            if form.is_valid():
                form.save()
                messages.success(request, "Dados do Fornecedor atualizado com sucesso!")
                return redirect("lista_fornecedores")
            else:
                messages.error(request, " Ocorreu um erro ao atualizar os dados do Fornecedor. Verifique os campos e tente novamente.")
        else:
            form = FornecedorForm(instance=fornecedor)
        return render(
            request,
            'fornecedores/fornecedor_detail_edit.html',
            {
                'form': form,
                'fornecedor': fornecedor,
                'contratos': contratos,
                'indicadores_geral': indicadores_geral,
                'os': os.distinct(),
                }
        )

    return render(
        request,
        'fornecedores/fornecedor_detail.html',
        {
            'fornecedor': fornecedor,
            'contratos': contratos,
            'indicadores_geral': indicadores_geral,
            'os': os.distinct(),
            }
    )


@login_required
def nova_solicitacao_contrato(request):
    if request.user.grupo in ['lider_contrato', 'gerente_contrato', 'gerente_lider']:
        clientes = Cliente.objects.all().order_by('nome')
        if request.method == 'POST':
            form = SolicitacaoContratoForm(request.POST, user=request.user)
            if form.is_valid():
                solicitacao = form.save(commit=False)
                solicitacao.lider_contrato = request.user
                solicitacao.status = "Solicitação de contratação"
                nome = request.user.get_full_name()
                if not nome:
                    nome = request.user.username

                solicitacao.solicitante = nome
                solicitacao.save()

                suprimentos = User.objects.filter(grupo="suprimento").values_list("email", flat=True)

                if suprimentos:
                    assunto = "Nova solicitação de Contratação"
                    mensagem = (
                        f"O usuário {request.user.get_full_name() or request.user.username} "
                        f"deu início a uma solicitação de contratação.\n\n"
                        f"Detalhes da solicitação:\n"
                        f"- ID: {solicitacao.id}\n"
                        f"- Valor Disponível: {solicitacao.valor_disponivel}\n"
                        f"- Descrição: {solicitacao.descricao}\n\n"
                        "Acesse o sistema HIDROGestão para mais informações.\n"
                        "https://hidrogestao.pythonanywhere.com/"
                    )
                    try:
                        send_mail(
                            assunto, mensagem,
                            FROM_EMAIL,
                            list(suprimentos),
                            fail_silently=False,
                        )
                    except Exception as e:
                        messages.warning(request, f"Erro ao enviar e-mail para suprimentos: {e}")
                    try:
                        send_mail(
                            assunto, mensagem,
                            FROM_EMAIL,
                            [solicitacao.lider_contrato.email],
                            fail_silently=False,
                        )
                    except Exception as e:
                        messages.warning(request, f"Erro ao enviar e-mail para o líder de contrato: {e}")
                    try:
                        send_request_notification_to_management(assunto, mensagem)
                    except Exception as e:
                        messages.warning(request, f"Erro ao enviar e-mail para diretoria e gerente de contrato: {e}")

                messages.success(request, "Solicitação de contratação criada com sucesso! Por favor, cadastre os eventos que serão feitos pelo fornecedor.")
                return redirect('detalhes_solicitacao_contrato', pk=solicitacao.pk )
            else:
                messages.error(request, "Por favor, corrija os erros abaixo e tente novamente.")
        else:
            form = SolicitacaoContratoForm(user=request.user)
        return render(request, 'fornecedores/nova_solicitacao_contrato.html', {'form':form, 'clientes': clientes})
    else:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")


@login_required
def aprovar_solicitacao_contrato(request, pk):
    solicitacao = get_object_or_404(SolicitacaoContrato, pk=pk)

    # Somente gerente pode aprovar
    if request.user.grupo not in ["gerente_contrato", "diretoria"]:
        messages.error(request, "Você não tem permissão para aprovar.")
        return redirect('home')

    if request.method == "POST" and request.user.grupo == "gerente_contrato":
        acao = request.POST.get("acao")

        if usuario.grupo == "lider_contrato" or (
            usuario.grupo == "gerente_lider" and user_shares_center_with_coordinator(usuario, solicitacao.coordenador)
        ):
            if usuario.grupo == "lider_contrato" and usuario != solicitacao.lider_contrato:
                messages.error(request, "Você não tem permissão para isso.")
                return redirect("home")
            if acao == "aprovar":
                bm.status_coordenador = "aprovado"
                bm.data_aprovacao_coordenador = timezone.now()
                messages.success(request, "Minuta BM aprovada pelo líder.")
            elif acao == "reprovar":
                bm.status_coordenador = "reprovado"
                bm.data_aprovacao_coordenador = timezone.now()
                messages.warning(request, "Minuta BM reprovada pelo líder.")

        # Busca todos os usuários do grupo 'suprimento'
        emails_suprimentos = list(
            User.objects.filter(grupo="suprimento")
            .exclude(email__isnull=True)
            .exclude(email__exact="")
            .values_list("email", flat=True)
        )

        # Adiciona o e-mail do coordenador do projeto (se existir)
        if solicitacao.coordenador and solicitacao.coordenador.email:
            emails_suprimentos.append(solicitacao.coordenador.email)
        if solicitacao.lider_contrato and solicitacao.lider_contrato.email:
            emails_suprimentos.append(solicitacao.lider_contrato.email)

        assunto = ""
        mensagem = ""

        # --- Caso APROVADO ---
        if acao == "aprovar":
            solicitacao.aprovacao_fornecedor_gerente = "aprovado"
            solicitacao.aprocacao_fornecedor_gerente_em = timezone.now()
            solicitacao.save()

            messages.success(
                request,
                f"Fornecedor {solicitacao.fornecedor_escolhido.nome} foi aprovado pelo gerente de contrato."
            )

            assunto = f"Fornecedor {solicitacao.fornecedor_escolhido.nome} aprovado pela gerência de contrato"
            mensagem = (
                f"Prezados,\n\n"
                f"O gerente {request.user.username} "
                f"aprovou o fornecedor {solicitacao.fornecedor_escolhido.nome} "
                f"na solicitação #{solicitacao.id}.\n\n"
                f"Contrato: {solicitacao.contrato.cod_projeto}\n"
                f"Cliente: {solicitacao.contrato.cliente.nome}\n"
                f"Status atual: {solicitacao.status}\n\n"
                f"Acesse o sistema para mais detalhes."
            )

        # --- Caso REPROVADO ---
        elif acao == "reprovar":
            solicitacao.status = "Fornecedor reprovado pela gerência"
            solicitacao.aprovacao_fornecedor_gerente = "reprovado"
            solicitacao.aprocacao_fornecedor_gerente_em = timezone.now()
            solicitacao.save()

            messages.warning(request, "Fornecedor reprovado.")

            assunto = f"Fornecedor reprovado pela gerência de contrato"
            mensagem = (
                f"Prezados,\n\n"
                f"O gerente {request.user.get_full_name() or request.user.username} "
                f"reprovou o fornecedor selecionado na solicitação #{solicitacao.id}.\n\n"
                f"Contrato: {solicitacao.contrato.cod_projeto}\n"
                f"Cliente: {solicitacao.contrato.cliente.nome}\n\n"
                f"Acesse o sistema para mais detalhes."
            )

        else:
            messages.error(request, "Ação inválida.")
            return redirect('home')

        # Envia o e-mail para todos os destinatários (se houver)
        if emails_suprimentos:
            try:
                send_mail(
                    assunto,
                    mensagem,
                    FROM_EMAIL,
                    list(set(emails_suprimentos)),  # remove duplicados
                    fail_silently=False,
                )
            except Exception as e:
                messages.warning(request, f"Erro ao enviar e-mail para suprimentos: {e}")

        return redirect('lista_solicitacoes')

    elif request.method == "POST" and request.user.grupo == "diretoria":
        acao = request.POST.get("acao")

        # Busca todos os usuários do grupo 'suprimento'
        emails_suprimentos = list(
            User.objects.filter(grupo="suprimento")
            .exclude(email__isnull=True)
            .exclude(email__exact="")
            .values_list("email", flat=True)
        )

        # Adiciona o e-mail do coordenador do projeto (se existir)
        if solicitacao.coordenador and solicitacao.coordenador.email:
            emails_suprimentos.append(solicitacao.coordenador.email)
        if solicitacao.lider_contrato and solicitacao.lider_contrato.email:
            emails_suprimentos.append(solicitacao.lider_contrato.email)

        assunto = ""
        mensagem = ""

        # --- Caso APROVADO ---
        if acao == "aprovar":
            solicitacao.aprovacao_fornecedor_diretor = "aprovado"
            solicitacao.aprocacao_fornecedor_diretor_em = timezone.now()
            solicitacao.save()

            messages.success(
                request,
                f"Fornecedor {solicitacao.fornecedor_escolhido.nome} foi aprovado pela direção."
            )

            assunto = f"Fornecedor {solicitacao.fornecedor_escolhido.nome} aprovado pela direção"
            mensagem = (
                f"Prezados,\n\n"
                f"O diretor {request.user.username} "
                f"aprovou o fornecedor {solicitacao.fornecedor_escolhido.nome} "
                f"na solicitação #{solicitacao.id}.\n\n"
                f"Contrato: {solicitacao.contrato.cod_projeto}\n"
                f"Cliente: {solicitacao.contrato.cliente.nome}\n"
                f"Status atual: {solicitacao.status}\n\n"
                f"Acesse o sistema para mais detalhes."
            )

        # --- Caso REPROVADO ---
        elif acao == "reprovar":
            solicitacao.status = "Fornecedor reprovado pela diretoria"
            solicitacao.aprovacao_fornecedor_diretor = "reprovado"
            solicitacao.save()

            messages.warning(request, "Fornecedor reprovado.")

            assunto = f"Fornecedor reprovado pela direção"
            mensagem = (
                f"Prezados,\n\n"
                f"O diretor {request.user.get_full_name() or request.user.username} "
                f"reprovou o fornecedor selecionado na solicitação #{solicitacao.id}.\n\n"
                f"Contrato: {solicitacao.contrato.cod_projeto}\n"
                f"Cliente: {solicitacao.contrato.cliente.nome}\n\n"
                f"Acesse o sistema para mais detalhes."
            )

        else:
            messages.error(request, "Ação inválida.")
            return redirect('home')

        # Envia o e-mail para todos os destinatários (se houver)
        if emails_suprimentos:
            try:
                send_mail(
                    assunto,
                    mensagem,
                    FROM_EMAIL,
                    list(set(emails_suprimentos)),
                    fail_silently=False,
                )
            except Exception as e:
                messages.warning(request, f"Erro ao enviar e-mail para suprimentos: {e}")

        return redirect('lista_solicitacoes')

    return redirect('home')


@login_required
def nova_solicitacao_prospeccao(request):
    if request.user.grupo in ['lider_contrato', 'gerente_contrato', 'gerente_lider']:
        clientes = Cliente.objects.all().order_by('nome')
        if request.method == 'POST':
            form = SolicitacaoProspeccaoForm(request.POST, user=request.user)
            if form.is_valid():
                solicitacao = form.save(commit=False)
                solicitacao.lider_contrato = request.user
                solicitacao.status = "Solicitação de prospecção"
                nome = request.user.get_full_name()
                if not nome:
                    nome = request.user.username

                solicitacao.solicitante = nome
                solicitacao.save()

                suprimentos = User.objects.filter(grupo="suprimento").values_list("email", flat=True)

                if suprimentos:
                    assunto = "Nova Solicitação de Prospecção"
                    mensagem = (
                        f"O usuário {request.user.get_full_name() or request.user.username} "
                        f"solicitou uma prospecção.\n\n"
                        f"Detalhes da solicitação:\n"
                        f"- ID: {solicitacao.id}\n"
                        f"- Valor Disponível: {solicitacao.valor_disponivel}\n"
                        f"- Descrição: {solicitacao.descricao}\n\n"
                        "Acesse o sistema HIDROGestão para mais informações.\n"
                        "https://hidrogestao.pythonanywhere.com/"
                    )
                    try:
                        send_mail(
                            assunto, mensagem,
                            FROM_EMAIL,
                            list(suprimentos),
                            fail_silently=False,
                        )
                    except Exception as e:
                        messages.warning(request, f"Erro ao enviar e-mail para suprimentos: {e}")
                    try:
                        send_mail(
                            assunto, mensagem,
                            FROM_EMAIL,
                            [solicitacao.lider_contrato.email],
                            fail_silently=False,
                        )
                    except Exception as e:
                        messages.warning(request, f"Erro ao enviar e-mail para o líder de contrato: {e}")
                    try:
                        send_request_notification_to_management(assunto, mensagem)
                    except Exception as e:
                        messages.warning(request, f"Erro ao enviar e-mail para diretoria e gerente de contrato: {e}")
                messages.success(request, "Solicitação de prospecção criada com sucesso! Por favor, cadastre os eventos que serão feitos pelo fornecedor.")
                return redirect('lista_solicitacoes')
            else:
                messages.error(request, "Por favor, corrija os erros abaixo e tente novamente.")
        else:
            form = SolicitacaoProspeccaoForm(user=request.user)
        return render(request, 'fornecedores/nova_solicitacao.html', {'form':form, 'clientes':clientes})

    else:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")


@login_required
def nova_solicitacao_guarda_chuva(request):
    if request.user.grupo in ['lider_contrato', 'gerente_contrato', 'gerente_lider']:
        clientes = Cliente.objects.all().order_by('nome')
        if request.method == 'POST':
            form = SolicitacaoGuardaChuvaForm(request.POST, user=request.user)
            if form.is_valid():
                solicitacao = form.save(commit=False)
                solicitacao.lider_contrato = request.user
                solicitacao.status = "Solicitação de contratação"
                solicitacao.guarda_chuva = True
                nome = request.user.get_full_name()
                if not nome:
                    nome = request.user.username

                solicitacao.solicitante = nome
                solicitacao.save()

                suprimentos = User.objects.filter(grupo="suprimento").values_list("email", flat=True)
                assunto = "Nova Solicitação de contratação guarda-chuva"
                mensagem = (
                    f"O usuário {request.user.get_full_name() or request.user.username} "
                    f"solicitou uma contratação do tipo guarda-chuva.\n\n"
                    f"Detalhes da solicitação:\n"
                    f"- ID: {solicitacao.id}\n"
                    f"- Valor Disponível: {solicitacao.valor_disponivel}\n"
                    f"- Descrição: {solicitacao.descricao}\n\n"
                    "Acesse o sistema HIDROGestão para mais informações.\n"
                    "https://hidrogestao.pythonanywhere.com/"
                )

                if suprimentos:
                    try:
                        send_mail(
                            assunto, mensagem,
                            FROM_EMAIL,
                            list(suprimentos),
                            fail_silently=False,
                        )
                    except Exception as e:
                        messages.warning(request, f"Erro ao enviar e-mail para suprimentos: {e}")
                if solicitacao.lider_contrato and solicitacao.lider_contrato.email:
                    try:
                        send_mail(
                            assunto, mensagem,
                            FROM_EMAIL,
                            [solicitacao.lider_contrato.email],
                            fail_silently=False,
                        )
                    except Exception as e:
                        messages.warning(request, f"Erro ao enviar e-mail para o líder de contrato: {e}")
                try:
                    send_request_notification_to_management(assunto, mensagem)
                except Exception as e:
                    messages.warning(request, f"Erro ao enviar e-mail para diretoria e gerente de contrato: {e}")
                messages.success(request, "Solicitação de contratação criada com sucesso!")
                return redirect('lista_solicitacoes')
            else:
                messages.error(request, "Por favor, corrija os erros abaixo e tente novamente.")
        else:
            form = SolicitacaoGuardaChuvaForm(user=request.user)
        return render(request, 'fornecedores/nova_solicitacao_guarda_chuva.html', {'form':form, 'clientes':clientes})

    else:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")


@login_required
def add_contrato(request):
    if request.method == "POST":
        form = ContratoModalForm(request.POST)
        if form.is_valid():
            contrato = form.save(commit=False)
            contrato.status = 'ativo'
            contrato.lider_contrato = request.user
            contrato.save()

            suprimentos = User.objects.filter(grupo="suprimento", is_active=True).values_list("email", flat=True)

            if suprimentos:
                assunto = "Novo contrato cadastrado"
                mensagem = (
                    f"Um novo contrato foi cadastrado no sistema.\n\n"
                    f"Contrato: {contrato.cod_projeto}\n"
                    f"Cliente: {contrato.cliente}\n"
                    f"Objeto: {contrato.objeto}\n"
                    f"Cadastrado por: {request.user.get_full_name() or request.user.username}\n"
                    "Acesse o sistema HIDROGestão e complete o cadastro.\n"
                    "https://hidrogestao.pythonanywhere.com/"
                )
                try:
                    send_mail(
                        assunto, mensagem,
                        FROM_EMAIL,
                        list(suprimentos),
                        fail_silently=False,
                    )
                except Exception as e:
                    messages.warning(request, f"Erro ao enviar e-mail para suprimentos: {e}")

            return JsonResponse({
                "id": contrato.id,
                "nome": contrato.cod_projeto
            })
        else:
            return JsonResponse({"errors": form.errors}, status=400)


@login_required
def solicitar_os(request, contrato_id):
    if request.user.grupo not in ['gerente_lider', 'lider_contrato', 'gerente_contrato']:
        messages.error(request, "Você não tem permissão para isso.")
        return redirect("home")

    contrato = get_object_or_404(ContratoTerceiros, pk=contrato_id)

    if request.method == 'POST':
        form = SolicitacaoOrdemServicoForm(request.POST, user=request.user, contrato_fixo=contrato)
        if form.is_valid():
            os = form.save(commit=False)
            os.solicitante = request.user
            os.lider_contrato = request.user
            os.status = 'pendente_lider'
            os.contrato = contrato
            os.save()

            assunto = "Nova Solicitação de O.S."
            mensagem = (
                f"O usuário {request.user.get_full_name() or request.user.username} "
                f"solicitou uma Ordem de Serviço.\n\n"
                f"Detalhes da solicitação:\n"
                f"- ID: {os.id}\n"
                f"- Contrato: {os.cod_projeto}\n"
                f"- Valor Provisionado: {os.valor_previsto or 'Não Informado'}\n"
                f"- Descrição: {os.descricao}\n\n"
                "Acesse o sistema HIDROGestão para mais informações.\n"
                "https://hidrogestao.pythonanywhere.com/"
            )
            suprimentos = User.objects.filter(grupo="suprimento").values_list("email", flat=True)
            try:
                send_mail(
                    assunto, mensagem,
                    FROM_EMAIL,
                    list(suprimentos),
                    fail_silently=False,
                )
            except Exception as e:
                messages.warning(request, f"Erro ao enviar e-mail para suprimentos: {e}")

            try:
                send_mail(
                    assunto, mensagem,
                    FROM_EMAIL,
                    list(os.lider_contrato.email),
                    fail_silently=False,
                )
            except Exception as e:
                messages.warning(request, f"Erro ao enviar e-mail para o líder de contrato: {e}")

            messages.success(request, "Ordem de Serviço enviada para aprovação.")
            return redirect("home")
    else:
        form = SolicitacaoOrdemServicoForm(user=request.user)

    return render(request, 'fornecedores/solicitar_os.html', {'form': form, "contrato": contrato, "seleciona_contrato": False})


@login_required
def solicitar_os_com_contrato(request):
    if request.user.grupo not in ['gerente_lider', 'lider_contrato', 'gerente_contrato']:
        messages.error(request, "Você não tem permissão para isso.")
        return redirect("home")

    if request.method == 'POST':
        form = SolicitacaoOrdemServicoForm(request.POST, user=request.user, include_contrato=True)
        if form.is_valid():
            os = form.save(commit=False)
            os.solicitante = request.user
            os.lider_contrato = request.user
            os.status = 'pendente_lider'
            os.save()

            assunto = "Nova Solicitação de O.S."
            mensagem = (
                f"O usuário {request.user.get_full_name() or request.user.username} "
                f"solicitou uma Ordem de Serviço.\n\n"
                f"Detalhes da solicitação:\n"
                f"- ID: {os.id}\n"
                f"- Contrato: {os.contrato}\n"
                f"- Valor Provisionado: {os.valor_previsto or 'Não Informado'}\n"
                f"- Descrição: {os.descricao}\n\n"
                "Acesse o sistema HIDROGestão para mais informações.\n"
                "https://hidrogestao.pythonanywhere.com/"
            )

            if os.lider_contrato and os.lider_contrato.email:
                try:
                    send_mail(
                        assunto,
                        mensagem,
                        FROM_EMAIL,
                        [os.lider_contrato.email],
                        fail_silently=False,
                    )
                except Exception as e:
                    messages.warning(request, f"Erro ao enviar e-mail para o líder de contrato: {e}")

            messages.success(request, "Ordem de Serviço solicitada com sucesso!")
            return redirect("detalhe_ordem_servico", pk=os.pk)
    else:
        form = SolicitacaoOrdemServicoForm(user=request.user, include_contrato=True)

    return render(
        request,
        'fornecedores/solicitar_os.html',
        {"form": form, "contrato": None, "seleciona_contrato": True},
    )


@login_required
def editar_ordem_servico(request, pk):
    if request.user.grupo not in ['gerente', 'gerente_lider', 'gerente_contrato', 'suprimento']:
        messages.error(request, "Você não tem permissão para isso.")
        return redirect('home')

    os = get_object_or_404(SolicitacaoOrdemServico, pk=pk)

    if request.user.grupo == 'lider_contrato' and os.lider_contrato != request.user:
        messages.error(request, "Você não tem permissão para editar esta OS.")
        return redirect('detalhe_ordem_servico', pk=os.pk)

    if request.user.grupo == 'gerente_lider':
        if not user_shares_center_with_coordinator(request.user, os.coordenador):
            messages.error(request, "Você não tem permissão para editar esta OS.")
            return redirect('detalhe_ordem_servico', pk=os.pk)

    if os.status == 'finalizada':
        messages.warning(request, "Esta OS está finalizada e não pode ser editada.")
        return redirect('detalhe_ordem_servico', pk=os.pk)

    if os.status in ['aprovada', 'finalizada']:
        messages.warning(request, "Esta OS não pode mais ser editada.")
        return redirect('detalhe_ordem_servico', pk=os.pk)

    if request.method == 'POST':
        form = SolicitacaoOrdemServicoForm(request.POST, instance=os)
        if form.is_valid():
            form.save()
            messages.success(request, "Ordem de Serviço atualizada com sucesso!")
            return redirect('detalhe_ordem_servico', pk=os.pk)
    else:
        form = SolicitacaoOrdemServicoForm(instance=os)

    context = {
        'form': form,
        'os': os
    }

    return render(request, 'gestao_contratos/editar_ordem_servico.html', context)


@login_required
def aprovar_os_lider(request, pk, acao):
    os = get_object_or_404(SolicitacaoOrdemServico, pk=pk)

    if request.user.grupo not in ['lider_contrato', 'gerente_lider']:
        messages.error(request, "Você não tem permissão para esta ação.")
        return redirect('detalhe_ordem_servico', pk=os.pk)

    if request.user.grupo == 'lider_contrato' and os.lider_contrato != request.user:
        messages.error(request, "Você não é o líder responsável por esta OS.")
        return redirect('detalhe_ordem_servico', pk=os.pk)

    if request.user.grupo == 'gerente_lider' and not user_shares_center_with_coordinator(request.user, os.coordenador):
        messages.error(request, "Você não tem permissão para esta ação.")
        return redirect('detalhe_ordem_servico', pk=os.pk)

    if os.status != 'pendente_lider':
        messages.warning(request, "Esta OS não está pendente para o líder.")
        return redirect('detalhe_ordem_servico', pk=os.pk)

    if acao == 'aprovar':
        os.status = 'pendente_suprimento'
        os.aprovacao_lider = request.user.username
        os.aprovado_lider_em = timezone.now()
        os.save()
        messages.success(request, "Ordem de Serviço aprovada e enviada para a Gerência.")

        assunto = f"Ciência e Aprovação do Lider de Contrato - OS {os.id}"
        mensagem = (
            f"O usuário {request.user.get_full_name() or request.user.username} "
            f"tomou ciência e aprovou a Ordem de Serviço {os.id}.\n\n"
            "Acesse o sistema HIDROGestão para mais informações.\n"
            "https://hidrogestao.pythonanywhere.com/"
        )
        try:
            send_mail(
                assunto, mensagem,
                FROM_EMAIL,
                [os.solicitante.email],
                fail_silently=False,
            )
        except Exception as e:
            messages.warning(request, f"Erro ao enviar e-mail para líder técnico: {e}")

        emails_suprimentos = list(
            User.objects.filter(grupo="suprimento")
            .exclude(email__isnull=True)
            .exclude(email__exact="")
            .values_list("email", flat=True)
        )
        assunto = f"Necessidade de avaliação de Solicitação da OS {os.id}"
        mensagem = (
            f"O usuário {request.user.get_full_name() or request.user.username} "
            f"tomou ciência e aprovou a Ordem de Serviço {os.id}. "
            f"Solicitamos o desenvolvimento do contrato da OS.\n\n"
            f"Detalhes da O.S.:\n"
            f"- ID: {os.id}\n"
            f"- líder Técnico: {os.solicitante.username}\n"
            f"- Título: {os.titulo}\n"
            f"- Descrição: {os.descricao}\n"
            f"- Valor Provisionado: {os.valor_previsto or 'Não informado'}\n\n"
            "Acesse o sistema HIDROGestão para mais informações.\n"
            "https://hidrogestao.pythonanywhere.com/"
        )
        try:
            send_mail(
                assunto, mensagem,
                FROM_EMAIL,
                list(set(emails_suprimentos)),
                fail_silently=False,
            )
        except Exception as e:
            messages.warning(request, f"Erro ao enviar e-mail para a equipe de suprimentos: {e}")

    elif acao == 'reprovar':
        os.status = 'reprovada'
        os.aprovacao_lider = request.user.username
        os.aprovado_lider_em = timezone.now()
        os.save()
        messages.error(request, "Ordem de Serviço reprovada.")

        assunto = f"Ciência e Reprovação do Lider de Contrato - OS {os.id}"
        mensagem = (
            f"O usuário {request.user.get_full_name() or request.user.username} "
            f"tomou ciência, no entanto, a Ordem de Serviço {os.id} foi reprovada.\n\n"
            "Acesse o sistema HIDROGestão para mais informações.\n"
            "https://hidrogestao.pythonanywhere.com/"
        )
        try:
            send_mail(
                assunto, mensagem,
                FROM_EMAIL,
                list(os.solicitante.email),
                fail_silently=False,
            )
        except Exception as e:
            messages.warning(request, f"Erro ao enviar e-mail para líder técnico: {e}")

    else:
        messages.error(request, "Ação inválida.")

    return redirect('detalhe_ordem_servico', pk=os.pk)


@login_required
def upload_contrato_os(request, pk):
    os = get_object_or_404(SolicitacaoOrdemServico, pk=pk)

    if request.user.grupo != 'suprimento':
        messages.error(request, "Você não tem permissão para isso!")
        return redirect('detalhe_ordem_servico', pk=os.pk)

    if os.status != 'pendente_suprimento':
        messages.warning(
            request,
            "O contrato só pode ser anexado quando a OS estiver pendente do suprimento."
        )
        return redirect('detalhe_ordem_servico', pk=os.pk)

    if request.method == 'POST':
        form = UploadContratoOSForm(request.POST, request.FILES, instance=os)
        if form.is_valid():
            form.save()

            os.status = 'pendente_gerente'
            os.save()

            messages.success(request, "Contrato anexado com sucesso!")
            return redirect('detalhe_ordem_servico', pk=os.pk)
    else:
        form = UploadContratoOSForm(instance=os)

    return render(
        request,
        'gestao_contratos/upload_contrato_os.html',
        {'form': form, 'os': os}
    )


@login_required
def aprovar_os_gerente_contrato(request, pk, acao):
    os = get_object_or_404(SolicitacaoOrdemServico, pk=pk)

    if request.user.grupo != 'gerente_contrato':
        messages.error(request, "Você não tem permissão para esta ação.")
        return redirect('detalhe_ordem_servico', pk=os.pk)

    if os.status != 'pendente_gerente':
        messages.warning(request, "Esta OS não está pendente da Gerência de Contrato.")
        return redirect('detalhe_ordem_servico', pk=os.pk)

    if acao == 'aprovar':
        os.status = 'aprovada'
        os.aprovacao_lider = request.user.username
        os.aprovado_lider_em = timezone.now()
        os.save()
        messages.success(
            request,
            "Ordem de Serviço aprovada."
        )

        ordem_servico = OS.objects.create(
            contrato=os.contrato,
            solicitacao=os,
            cod_projeto=os.cod_projeto,
            coordenador=os.solicitante,
            lider_contrato=os.lider_contrato,
            titulo=os.titulo,
            descricao=os.descricao,
            valor=os.valor_previsto if os.valor_previsto else None,
            prazo_execucao=os.prazo_execucao if os.prazo_execucao else None,
            arquivo_os=os.arquivo_os if os.arquivo_os else None,
        )
        assunto = f"Aprovação da OS {os.id}"
        mensagem = (
            f"O gerente de contrato, {request.user.get_full_name() or request.user.username}, "
            f"aprovou a Ordem de Serviço {os.id}.\n\n"
            "Acesse o sistema HIDROGestão para mais informações.\n"
            "https://hidrogestao.pythonanywhere.com/"
        )
        try:
            send_mail(
                assunto, mensagem,
                FROM_EMAIL,
                [os.solicitante.email],
                fail_silently=False,
            )
        except Exception as e:
            messages.warning(request, f"Erro ao enviar e-mail para líder técnico: {e}")

        emails_suprimentos = list(
            User.objects.filter(grupo="suprimento")
            .exclude(email__isnull=True)
            .exclude(email__exact="")
            .values_list("email", flat=True)
        )
        try:
            send_mail(
                assunto, mensagem,
                FROM_EMAIL,
                list(set(emails_suprimentos)),
                fail_silently=False,
            )
        except Exception as e:
            messages.warning(request, f"Erro ao enviar e-mail para a equipe de suprimentos: {e}")



    elif acao == 'reprovar':
        os.status = 'reprovada'
        os.aprovacao_lider = request.user.username
        os.aprovado_lider_em = timezone.now()
        os.save()
        messages.error(request, "Ordem de Serviço reprovada pela Gerência de Contrato.")

        assunto = f"Reprovação da OS {os.id}"
        mensagem = (
            f"O gerente de contrato, {request.user.get_full_name() or request.user.username}, "
            f"reprovou a Ordem de Serviço {os.id}.\n\n"
            "Acesse o sistema HIDROGestão para mais informações.\n"
            "https://hidrogestao.pythonanywhere.com/"
        )
        try:
            send_mail(
                assunto, mensagem,
                FROM_EMAIL,
                [os.solicitante.email],
                fail_silently=False,
            )
        except Exception as e:
            messages.warning(request, f"Erro ao enviar e-mail para líder técnico: {e}")

        emails_suprimentos = list(
            User.objects.filter(grupo="suprimento")
            .exclude(email__isnull=True)
            .exclude(email__exact="")
            .values_list("email", flat=True)
        )
        try:
            send_mail(
                assunto, mensagem,
                FROM_EMAIL,
                list(set(emails_suprimentos)),
                fail_silently=False,
            )
        except Exception as e:
            messages.warning(request, f"Erro ao enviar e-mail para a equipe de suprimentos: {e}")

    else:
        messages.error(request, "Ação inválida.")

    return redirect('detalhe_ordem_servico', pk=os.pk)


@login_required
def lista_solicitacoes(request):
    #if request.user.grupo == 'coordenador' or request.user.grupo == 'financeiro':
    if request.user.grupo == 'coordenador':
        solicitacoes = SolicitacaoProspeccao.objects.filter(coordenador=request.user).exclude(status__in=["Onboarding", "Reprovada pelo suprimento"]).order_by('-data_solicitacao')
        solicitacoes_c = SolicitacaoContrato.objects.filter(coordenador=request.user).exclude(status__in=["Onboarding", "Reprovada pelo suprimento"]).order_by('-data_solicitacao')
        os = OS.objects.filter(coordenador=request.user).exclude(status__in=["finalizada", "aprovada", "reprovada"]).order_by('-criado_em')
    elif request.user.grupo == 'lider_contrato':
        solicitacoes = SolicitacaoProspeccao.objects.filter(lider_contrato=request.user).exclude(status__in=["Onboarding", "Reprovada pelo suprimento"]).order_by('-data_solicitacao')
        solicitacoes_c = SolicitacaoContrato.objects.filter(lider_contrato=request.user).exclude(status__in=["Onboarding", "Reprovada pelo suprimento"]).order_by('-data_solicitacao')
        os = OS.objects.filter(lider_contrato=request.user).exclude(status__in=["finalizada", "aprovada", "reprovada"]).order_by('-criado_em')
    elif request.user.grupo in ['suprimento', 'diretoria']:
        solicitacoes = SolicitacaoProspeccao.objects.all().exclude(status__in=["Onboarding", "Reprovada pelo suprimento"]).order_by('-data_solicitacao')
        solicitacoes_c = SolicitacaoContrato.objects.all().exclude(status__in=["Onboarding", "Reprovada pelo suprimento"]).order_by('-data_solicitacao')
        os = OS.objects.all().exclude(status__in=["finalizada", "aprovada", "reprovada"]).order_by('-criado_em')
    elif request.user.grupo in ['gerente', 'gerente_lider']:
        centros_do_gerente = request.user.centros.all()
        # filtra solicitações cujo solicitante tenha pelo menos um centro em comum
        solicitacoes = SolicitacaoProspeccao.objects.filter(
            coordenador__centros__in=centros_do_gerente
        ).exclude(status__in=["Onboarding", "Reprovada pelo suprimento"]).distinct().order_by('-data_solicitacao')
        solicitacoes_c = SolicitacaoContrato.objects.filter(coordenador__centros__in=centros_do_gerente).exclude(status__in=["Onboarding", "Reprovada pelo suprimento"]).order_by('-data_solicitacao')
        os = OS.objects.filter(coordenador__centros__in=centros_do_gerente).exclude(status__in=["finalizada", "aprovada", "reprovada"]).order_by('-criado_em')
    elif request.user.grupo == 'gerente_contrato':
        solicitacoes = SolicitacaoProspeccao.objects.filter(
            lider_contrato__grupo__in=['lider_contrato', 'gerente_contrato', 'gerente_lider']
        ).exclude(status__in=["Onboarding", "Reprovada pelo suprimento"]).distinct().order_by('-data_solicitacao')
        solicitacoes_c = SolicitacaoContrato.objects.filter(lider_contrato__grupo__in=['lider_contrato', 'gerente_contrato', 'gerente_lider']).exclude(status__in=["Onboarding", "Reprovada pelo suprimento"]).order_by('-data_solicitacao')
        os = OS.objects.filter(lider_contrato__grupo__in=['lider_contrato', 'gerente_contrato', 'gerente_lider']).exclude(status__in=["finalizada", "aprovada", "reprovada"]).order_by('-criado_em')
    else:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    lista_solicitacoes = []
    for s in solicitacoes:
        proposta_escolhida = PropostaFornecedor.objects.filter(
            solicitacao=s,
            fornecedor=s.fornecedor_escolhido
        ).first()

        contrato = DocumentoContratoTerceiro.objects.filter(solicitacao=s).first()
        lista_solicitacoes.append({
            "solicitacao": s,
            "fornecedor": s.fornecedor_escolhido,
            "proposta": proposta_escolhida,
            "contrato": contrato
        })

    for s in solicitacoes_c:
        proposta_escolhida = None

        contrato = None
        lista_solicitacoes.append({
            "solicitacao": s,
            "fornecedor": s.fornecedor_escolhido,
            "proposta": proposta_escolhida,
            "contrato": contrato
        })

    paginator = Paginator(solicitacoes, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    paginator_c = Paginator(solicitacoes_c, 10)
    page_number_c = request.GET.get('page_c')
    page_obj_c = paginator_c.get_page(page_number_c)

    paginator_os = Paginator(os, 10)
    page_number_os = request.GET.get('page_os')
    page_obj_os = paginator_os.get_page(page_number_os)

    aditivos_queryset = [
        item
        for item in get_visible_addendums_for_user(request.user)
        if not item.aprovado_totalmente and not item.solicitacao_reprovada_por_alguem
    ]
    paginator_aditivos = Paginator(aditivos_queryset, 10)
    page_number_aditivos = request.GET.get('page_ad')
    page_obj_aditivos = paginator_aditivos.get_page(page_number_aditivos)

    context = {
            'page_obj': page_obj,
            'page_obj_contrato': page_obj_c,
            "lista_solicitacoes": lista_solicitacoes,
            'ordens_servico_page': page_obj_os,
            'aditivos_page': page_obj_aditivos,
        }

    return render(request, 'gestao_contratos/lista_solicitacoes.html', context)


@login_required
def aprovar_solicitacao(request, pk, acao):
    if request.user.grupo not in ["suprimento"]:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    solicitacao = get_object_or_404(SolicitacaoProspeccao, pk=pk)
    if acao == "aprovar":
        if not solicitacao.guarda_chuva and not solicitacao.evento_set.exists():
            messages.error(request, "Cadastre ao menos um evento antes de aprovar esta solicitação.")
            return redirect("detalhes_solicitacao", pk=solicitacao.pk)
        solicitacao.status = "Aprovada pelo suprimento"
        solicitacao.aprovado = True
    elif acao == "reprovar":
        solicitacao.status = "Reprovada pelo suprimento"
        solicitacao.aprovado = False

        coordenador = solicitacao.coordenador
        assunto = "Solicitação reprovada"
        mensagem = (
            f"Olá, {coordenador.username}\n\n"
            f"A solicitação de prospecção foi reprovada. \n\n"
            "Por favor, entre no sistema HIDROGestão para mais informações.\n"
            "https://hidrogestao.pythonanywhere.com/"
        )
        try:
            send_mail(
                assunto, mensagem,
                FROM_EMAIL,
                [coordenador.email],
                fail_silently=False,
            )
        except Exception as e:
            messages.warning(request, f"Erro ao enviar e-mail para {coordenador.username}: {e}")

    solicitacao.data_aprovacao = timezone.now()
    solicitacao.aprovado_por = request.user
    solicitacao.save()

    return redirect('lista_solicitacoes')


@login_required
def triagem_fornecedores(request, pk):
    if request.user.grupo not in ["suprimento"]:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    condicoes_choices = PropostaFornecedor.CONDICOES_CHOICES
    solicitacao = get_object_or_404(SolicitacaoProspeccao, pk=pk, aprovado=True)
    fornecedores = EmpresaTerceira.objects.all().order_by('nome')

    # Mapeia fornecedores -> indicadores
    fornecedores_indicadores = {
        f.id: Indicadores.objects.filter(empresa_terceira=f) or None
        for f in fornecedores
    }

    # Mapeia propostas -> todos os fornecedores
    propostas_dict = {}
    for f in fornecedores:
        try:
            propostas_dict[f.id] = PropostaFornecedor.objects.get(solicitacao=solicitacao, fornecedor=f)
        except PropostaFornecedor.DoesNotExist:
            propostas_dict[f.id] = None

    if request.method == "POST":
        fornecedores_ids = request.POST.getlist("fornecedores")

        if "nenhum_fornecedor" in request.POST:
            solicitacao.fornecedores_selecionados.clear()
            solicitacao.triagem_realizada = False
            solicitacao.status = "Sem fornecedor adequado"
            solicitacao.save()
            messages.info(request, "Nenhum fornecedor foi considerado ideal.")
            return redirect("lista_solicitacoes")

        if fornecedores_ids:
            # Atualiza os fornecedores selecionados
            solicitacao.fornecedores_selecionados.set(fornecedores_ids)
            solicitacao.triagem_realizada = True
            solicitacao.status = "Triagem realizada"
            solicitacao.nenhum_fornecedor_ideal = False
            solicitacao.save()

            # Para cada fornecedor selecionado, salva/atualiza proposta
            for f_id in fornecedores_ids:
                fornecedor = get_object_or_404(EmpresaTerceira, pk=f_id)
                valor = request.POST.get(f"valor_{f_id}")
                prazo_validade = request.POST.get(f"prazo_{f_id}")
                condicao = request.POST.get(f"condicao_{f_id}")
                arquivo = request.FILES.get(f"arquivo_{f_id}")

                if valor or arquivo or condicao or prazo_validade:
                    proposta_obj, _ = PropostaFornecedor.objects.get_or_create(
                        solicitacao=solicitacao,
                        fornecedor=fornecedor,
                    )
                    if valor:
                        try:
                            valor_formatado = (
                                valor.replace('R$', '')
                                     .replace('.', '')
                                     .replace(',', '.')
                                     .strip()
                            )
                            proposta_obj.valor_proposta = Decimal(valor_formatado)
                        except:
                            messages.warning(request, f"Valor inválido para {fornecedor.nome}")
                    if prazo_validade:
                        proposta_obj.prazo_validade = prazo_validade
                    if arquivo:
                        proposta_obj.arquivo_proposta = arquivo
                    if condicao:
                        proposta_obj.condicao_pagamento = condicao
                    proposta_obj.save()

            # Notifica coordenador
            coordenador = solicitacao.coordenador
            #lider = solicitacao.lider_contrato
            assunto = "Triagem de fornecedores realizada"
            mensagem = (
                f"Olá, {coordenador.username}\n\n"
                f"A equipe de suprimentos realizou uma triagem de fornecedores para você. \n\n"
                "Por favor, entre no sistema HIDROGestão para selecionar sua escolha.\n"
                "https://hidrogestao.pythonanywhere.com/"
            )
            try:
                send_mail(
                    assunto, mensagem,
                    FROM_EMAIL,
                    [coordenador.email],
                    fail_silently=False,
                )
            except Exception as e:
                messages.warning(request, f"Erro ao enviar e-mail para {coordenador.username}: {e}")

            #try:
            #    send_mail(
            #        assunto, mensagem,
            #        FROM_EMAIL,
            #        [lider.email],
            #        fail_silently=False,
            #    )
            #except Exception as e:
            #    messages.warning(request, f"Erro ao enviar e-mail para {lider.username}: {e}")
            messages.success(request, "Triagem e propostas salvas com sucesso!")
            return redirect("lista_solicitacoes")

        else:
            solicitacao.fornecedores_selecionados.clear()
            solicitacao.triagem_realizada = False
            messages.warning(request, "Nenhum fornecedor foi selecionado.")
            return redirect("lista_solicitacoes")

    context = {
        'solicitacao': solicitacao,
        'fornecedores': fornecedores,
        'fornecedores_selecionados': solicitacao.fornecedores_selecionados.all(),
        'fornecedores_indicadores': fornecedores_indicadores,
        'propostas_dict': propostas_dict,
        'condicoes_choices': condicoes_choices,
    }
    return render(request, 'fornecedores/triagem_fornecedores.html', context)


@login_required
def nenhum_fornecedor_ideal(request, pk):
    solicitacao = get_object_or_404(SolicitacaoProspeccao, pk=pk)
    if request.user != solicitacao.lider_contrato:
        messages.error(request, "Você não tem permissão para essa ação.")
        return redirect("lista_solicitacoes")

    if request.method == "POST":
        solicitacao.nenhum_fornecedor_ideal = True
        solicitacao.fornecedores_selecionados.clear()
        solicitacao.triagem_realizada = False
        solicitacao.save()

        suprimentos = User.objects.filter(grupo="suprimento").values_list("email", flat=True)

        if suprimentos:
            assunto = "Triagem declarada ineficaz pelo líder de contrato"
            mensagem = (
                f"Olá,\n\n"
                f"O líder de contrato {solicitacao.lider_contrato.username} declarou que nenhum dos fornecedores é ideal."
                "Acesse o sistema HIDROGestão para mais informações.\n"
                "https://hidrogestao.pythonanywhere.com/"
            )
            try:
                send_mail(
                    assunto, mensagem,
                    FROM_EMAIL,
                    list(suprimentos),
                    fail_silently=False,
                )
            except Exception as e:
                messages.warning(request, f"Erro ao enviar e-mail para suprimentos: {e}")
        messages.success(request, "Solicitação atualizada: nenhum fornecedor é ideal.")
    return redirect("lista_solicitacoes")


@login_required
def detalhes_triagem_fornecedores(request, pk):
    solicitacao = get_object_or_404(
        SolicitacaoProspeccao, pk=pk, aprovado=True, triagem_realizada=True
    )
    fornecedores_selecionados = solicitacao.fornecedores_selecionados.all()

    # Mapeia propostas por fornecedor
    propostas_dict = {}
    propostas = PropostaFornecedor.objects.filter(solicitacao=solicitacao)
    for p in propostas:
        propostas_dict[p.fornecedor.id] = p

    if can_user_manage_supplier_choice(request.user, solicitacao) and request.method == "POST":
        escolhido_id = request.POST.get("fornecedor_escolhido")
        justificativa = request.POST.get("justificativa_fornecedor_escolhido", "").strip()

        if escolhido_id:
            if not justificativa:
                messages.warning(request, "Por favor, insira uma justificativa para a escolha do fornecedor.")
                return redirect('detalhes_triagem_fornecedores', pk=pk)

            fornecedor = get_object_or_404(EmpresaTerceira, pk=escolhido_id)
            proposta_escolhida = PropostaFornecedor.objects.filter(
                solicitacao=solicitacao,
                fornecedor=fornecedor,
            ).first()
            total_previsto = solicitacao.evento_set.aggregate(total=Sum("valor_previsto"))["total"] or Decimal("0.00")
            valor_proposta = (
                proposta_escolhida.valor_proposta
                if proposta_escolhida and proposta_escolhida.valor_proposta is not None
                else Decimal("0.00")
            )

            solicitacao.fornecedor_escolhido = fornecedor
            solicitacao.justificativa_fornecedor_escolhido = justificativa
            solicitacao.nenhum_fornecedor_ideal = False
            solicitacao.status = 'Fornecedor selecionado'
            solicitacao.aprovacao_gerente = "pendente"
            solicitacao.save()

            assunto = f"Aprovação necessária - Fornecedor escolhido para {solicitacao.contrato}"
            mensagem = (
                f"O usuário {request.user.get_full_name() or request.user.username} selecionou o fornecedor {fornecedor.nome}.\n"
                f"Justificativa: {justificativa}\n\n"
                f"É necessário que gerente de contrato e diretoria avaliem essa escolha.\n"
                f"Acesse o sistema HIDROGestão para mais informações:\n"
                f"https://hidrogestao.pythonanywhere.com/"
            )
            try:
                send_request_notification_to_management(assunto, mensagem)
            except Exception as e:
                messages.warning(request, f"Erro ao enviar e-mail para gerente de contrato e diretoria: {e}")

            messages.success(request, f"Fornecedor {fornecedor.nome} selecionado. Aguardando avaliação do gerente de contrato e da diretoria.")
            if total_previsto != valor_proposta:
                total_previsto_formatado = f"{total_previsto:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                valor_proposta_formatado = f"{valor_proposta:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                messages.warning(
                    request,
                    "Atenção: a soma dos valores previstos "
                    f"(R$ {total_previsto_formatado}) difere do valor total da proposta "
                    f"(R$ {valor_proposta_formatado})."
                )
        return redirect('lista_solicitacoes')

    context = {
        "solicitacao": solicitacao,
        "fornecedores_selecionados": fornecedores_selecionados,
        "propostas_dict": propostas_dict,
    }
    return render(request, "fornecedores/detalhes_triagem_fornecedores.html", context)


@login_required
def aprovar_fornecedor_gerente(request, pk):
    solicitacao = get_object_or_404(SolicitacaoProspeccao, pk=pk)

    # Somente gerente de contrato pode aprovar nesta etapa
    if request.user.grupo != "gerente_contrato":
        messages.error(request, "Você não tem permissão para aprovar.")
        return redirect('home')

    # Verifica se o coordenador já escolheu um fornecedor
    if not solicitacao.fornecedor_escolhido:
        messages.warning(request, "Coordenador ainda não escolheu um fornecedor.")
        return redirect('detalhes_triagem_fornecedores', pk=pk)

    if request.method == "POST":
        acao = request.POST.get("acao")
        justificativa = request.POST.get("justificativa", "")

        # Busca todos os usuários do grupo 'suprimento'
        emails_suprimentos = list(
            User.objects.filter(grupo="suprimento")
            .exclude(email__isnull=True)
            .exclude(email__exact="")
            .values_list("email", flat=True)
        )

        # Adiciona o e-mail do coordenador do projeto (se existir)
        if solicitacao.coordenador and solicitacao.coordenador.email:
            emails_suprimentos.append(solicitacao.coordenador.email)
        if solicitacao.lider_contrato and solicitacao.lider_contrato.email:
            emails_suprimentos.append(solicitacao.lider_contrato.email)

        assunto = ""
        mensagem = ""

        # --- Caso APROVADO ---
        if acao == "aprovar":

            solicitacao.aprovacao_fornecedor_gerente = "aprovado"
            solicitacao.aprocacao_fornecedor_gerente_em = timezone.now()
            solicitacao.save()

            messages.success(
                request,
                f"Fornecedor {solicitacao.fornecedor_escolhido.nome} aprovado pelo gerente de contrato."
            )

            if solicitacao.aprovacao_fornecedor_diretor == "aprovado":
                solicitacao.status = "Fornecedor aprovado"
                solicitacao.save()

                assunto = f"Fornecedor {solicitacao.fornecedor_escolhido.nome} aprovado pela gerência de contrato"
                mensagem = (
                    f"Prezados,\n\n"
                    f"O gerente {request.user.username} "
                    f"aprovou o fornecedor {solicitacao.fornecedor_escolhido.nome} "
                    f"na solicitação #{solicitacao.id}.\n\n"
                    f"Contrato: {solicitacao.contrato.cod_projeto}\n"
                    f"Cliente: {solicitacao.contrato.cliente.nome}\n"
                    f"Status atual: {solicitacao.status}\n\n"
                    f"Acesse o sistema para mais detalhes."
                )

                # Envia o e-mail para todos os destinatários (se houver)
                if emails_suprimentos:
                    try:
                        send_mail(
                            assunto,
                            mensagem,
                            FROM_EMAIL,
                            list(set(emails_suprimentos)),  # remove duplicados
                            fail_silently=False,
                        )
                    except Exception as e:
                        messages.warning(request, f"Erro ao enviar e-mail para suprimentos: {e}")

        # --- Caso REPROVADO ---
        elif acao == "reprovar":
            solicitacao.status = "Fornecedor reprovado pela gerência"
            solicitacao.aprovacao_fornecedor_gerente = "reprovado"
            solicitacao.aprocacao_fornecedor_gerente_em = timezone.now()
            solicitacao.fornecedor_escolhido = None
            solicitacao.triagem_realizada = False
            solicitacao.fornecedores_selecionados.clear()
            solicitacao.justificativa_gerencia = justificativa
            solicitacao.save()

            messages.warning(request, "Fornecedor reprovado. Nova triagem necessária pelo suprimento.")

            assunto = f"Fornecedor reprovado pela gerência de contrato"
            mensagem = (
                f"Prezados,\n\n"
                f"O gerente {request.user.get_full_name() or request.user.username} "
                f"reprovou o fornecedor selecionado na solicitação #{solicitacao.id}.\n\n"
                f"Contrato: {solicitacao.contrato.cod_projeto}\n"
                f"Cliente: {solicitacao.contrato.cliente.nome}\n\n"
                f"Uma nova triagem será necessária pelo setor de suprimentos.\n\n"
                f"Acesse o sistema para mais detalhes."
            )

            # Envia o e-mail para todos os destinatários (se houver)
            if emails_suprimentos:
                try:
                    send_mail(
                        assunto,
                        mensagem,
                        FROM_EMAIL,
                        list(set(emails_suprimentos)),  # remove duplicados
                        fail_silently=False,
                    )
                except Exception as e:
                    messages.warning(request, f"Erro ao enviar e-mail para suprimentos: {e}")

        else:
            messages.error(request, "Ação inválida.")
            return redirect('detalhes_triagem_fornecedores', pk=pk)

        return redirect('lista_solicitacoes')

    return redirect('detalhes_triagem_fornecedores', pk=pk)


@login_required
def aprovar_fornecedor_diretor(request, pk):
    solicitacao = get_object_or_404(SolicitacaoProspeccao, pk=pk)

    # Somente gerente pode aprovar
    if request.user.grupo != "diretoria":
        messages.error(request, "Você não tem permissão para aprovar.")
        return redirect('home')

    # Verifica se o coordenador já escolheu um fornecedor
    if not solicitacao.fornecedor_escolhido:
        messages.warning(request, "Coordenador ainda não escolheu um fornecedor.")
        return redirect('detalhes_triagem_fornecedores', pk=pk)

    if request.method == "POST":
        acao = request.POST.get("acao")
        justificativa = request.POST.get("justificativa", "")

        # Busca todos os usuários do grupo 'suprimento'
        emails_suprimentos = list(
            User.objects.filter(grupo="suprimento")
            .exclude(email__isnull=True)
            .exclude(email__exact="")
            .values_list("email", flat=True)
        )

        # Adiciona o e-mail do coordenador do projeto (se existir)
        if solicitacao.coordenador and solicitacao.coordenador.email:
            emails_suprimentos.append(solicitacao.coordenador.email)
        if solicitacao.lider_contrato and solicitacao.lider_contrato.email:
            emails_suprimentos.append(solicitacao.lider_contrato.email)

        assunto = ""
        mensagem = ""

        # --- Caso APROVADO ---
        if acao == "aprovar":
            solicitacao.aprovacao_fornecedor_diretor = "aprovado"
            solicitacao.aprocacao_fornecedor_diretor_em = timezone.now()
            solicitacao.save()

            messages.success(
                request,
                f"Fornecedor {solicitacao.fornecedor_escolhido.nome} aprovado pela diretoria de contrato."
            )

            if solicitacao.aprovacao_fornecedor_gerente == "aprovado":
                solicitacao.status = "Fornecedor aprovado"
                solicitacao.save()
                assunto = f"Fornecedor {solicitacao.fornecedor_escolhido.nome} aprovado pela diretoria de contrato"
                mensagem = (
                    f"Prezados,\n\n"
                    f"O diretor {request.user.username} "
                    f"aprovou o fornecedor {solicitacao.fornecedor_escolhido.nome} "
                    f"na solicitação #{solicitacao.id}.\n\n"
                    f"Contrato: {solicitacao.contrato.cod_projeto}\n"
                    f"Cliente: {solicitacao.contrato.cliente.nome}\n"
                    f"Status atual: {solicitacao.status}\n\n"
                    f"Acesse o sistema para mais detalhes."
                )

                # Envia o e-mail para todos os destinatários (se houver)
                if emails_suprimentos:
                    try:
                        send_mail(
                            assunto,
                            mensagem,
                            FROM_EMAIL,
                            list(set(emails_suprimentos)),  # remove duplicados
                            fail_silently=False,
                        )
                    except Exception as e:
                        messages.warning(request, f"Erro ao enviar e-mail para suprimentos: {e}")

        # --- Caso REPROVADO ---
        elif acao == "reprovar":
            solicitacao.status = "Fornecedor reprovado"
            solicitacao.aprovacao_fornecedor_diretor = "reprovado"
            solicitacao.aprocacao_fornecedor_diretor_em = timezone.now()
            solicitacao.fornecedor_escolhido = None
            solicitacao.triagem_realizada = False
            solicitacao.fornecedores_selecionados.clear()
            solicitacao.justificativa_diretoria = justificativa
            solicitacao.save()

            messages.warning(request, "Fornecedor reprovado. Nova triagem necessária pelo suprimento.")

            assunto = f"Fornecedor reprovado pela diretoria de contrato"
            mensagem = (
                f"Prezados,\n\n"
                f"O diretor {request.user.get_full_name() or request.user.username} "
                f"reprovou o fornecedor selecionado na solicitação #{solicitacao.id}.\n\n"
                f"Contrato: {solicitacao.contrato.cod_projeto}\n"
                f"Cliente: {solicitacao.contrato.cliente.nome}\n\n"
                f"Uma nova triagem será necessária pelo setor de suprimentos.\n\n"
                f"Acesse o sistema para mais detalhes."
            )

            # Envia o e-mail para todos os destinatários (se houver)
            if emails_suprimentos:
                try:
                    send_mail(
                        assunto,
                        mensagem,
                        FROM_EMAIL,
                        list(set(emails_suprimentos)),  # remove duplicados
                        fail_silently=False,
                    )
                except Exception as e:
                    messages.warning(request, f"Erro ao enviar e-mail para suprimentos: {e}")

        else:
            messages.error(request, "Ação inválida.")
            return redirect('detalhes_triagem_fornecedores', pk=pk)

        return redirect('lista_solicitacoes')

    return redirect('detalhes_triagem_fornecedores', pk=pk)


@login_required
def detalhes_solicitacao_contrato(request, pk):
    # Busca a solicitação
    solicitacao = get_object_or_404(SolicitacaoContrato, pk=pk)

    if request.user.grupo == "gerente_lider" and not user_shares_center_with_coordinator(request.user, solicitacao.coordenador):
        messages.error(request, "Você não tem permissão para isso.")
        return redirect("home")

    status_order = [
        "Solicitação de contratação",
        "Fornecedor aprovado",
        "Planejamento do Contrato",
        "Aprovação Final",
        SIGNED_FILES_PENDING_STATUS,
        "Onboarding",
    ]

    if request.method == "POST":
        acao = request.POST.get("acao")
        justificativa = request.POST.get("justificativa", "")

        if acao == "aprovar" and not solicitacao.guarda_chuva and not solicitacao.evento_set.exists():
            messages.error(request, "Cadastre ao menos um evento antes de avançar esta solicitação.")
            return redirect("detalhes_solicitacao_contrato", pk=solicitacao.pk)

        # GERENTE DE CONTRATO
        if request.user.grupo == "gerente_contrato":
            if acao == "aprovar":
                solicitacao.aprovacao_fornecedor_gerente = "aprovado"
                solicitacao.aprocacao_fornecedor_gerente_em = timezone.now()
            elif acao == "reprovar":
                if not justificativa:
                    messages.error(request, "A justificativa é obrigatÃ³ria para reprovar.")
                    return redirect("detalhes_solicitacao_contrato", pk=solicitacao.pk)
                solicitacao.aprovacao_fornecedor_gerente = "reprovado"
                solicitacao.aprocacao_fornecedor_gerente_em = timezone.now()
                solicitacao.justificativa_gerencia = justificativa

        # DIRETORIA
        elif request.user.grupo == "diretoria":
            if acao == "aprovar":
                solicitacao.aprovacao_fornecedor_diretor = "aprovado"
                solicitacao.aprocacao_fornecedor_diretor_em = timezone.now()
            elif acao == "reprovar":
                if not justificativa:
                    messages.error(request, "A justificativa é obrigatÃ³ria para reprovar.")
                    return redirect("detalhes_solicitacao_contrato", pk=solicitacao.pk)
                solicitacao.aprovacao_fornecedor_diretor = "reprovado"
                solicitacao.aprocacao_fornecedor_diretor_em = timezone.now()
                solicitacao.justificativa_diretoria = justificativa

        if solicitacao.aprovacao_fornecedor_gerente == "aprovado"  and solicitacao.aprovacao_fornecedor_diretor == "aprovado":
            solicitacao.status = "Fornecedor aprovado"
            solicitacao.save()
            suprimentos = User.objects.filter(grupo="suprimento").values_list("email", flat=True)

            if suprimentos:
                assunto = f"Aprovação de Solicitação de Contratação #{solicitacao.id}"
                mensagem = (
                    f"O HIDROGestão informa que a Solicitação de Contrato #{solicitacao.id} e o Fornecedor {solicitacao.fornecedor_escolhido}\n"
                    f"foram aprovados pelo Gerente de Contrato e pela Diretoria.\n\n"

                    f"Dessa forma, o processo encontra-se liberado para a etapa de Suprimentos, \n"
                    f"cabendo a este setor a elaboração e inserção no sistema da \n"
                    f"Minuta do BM (Boletim de Medição) e da Minuta do Contrato. \n\n"

                    f"Acesse o sistema HIDROGestão para mais informações.\n"
                    f"https://hidrogestao.pythonanywhere.com/"
                )
                try:
                    send_mail(
                        assunto, mensagem,
                        FROM_EMAIL,
                        list(suprimentos),
                        fail_silently=False,
                    )
                except Exception as e:
                    messages.warning(request, f"Erro ao enviar e-mail para suprimentos: {e}")
        elif solicitacao.aprovacao_fornecedor_gerente == "reprovado" or solicitacao.aprovacao_fornecedor_diretor == "reprovado":
            solicitacao.status = "Solicitação de contratação"
        solicitacao.save()
        return redirect("detalhes_solicitacao_contrato", pk=solicitacao.pk)

    fornecedor_escolhido = solicitacao.fornecedor_escolhido

    proposta_escolhida = None
    indicadores = None

    if fornecedor_escolhido:
        # Busca a proposta do fornecedor escolhido nesta solicitação
        proposta_escolhida = None

        # Busca indicadores do fornecedor escolhido
        indicadores = Indicadores.objects.filter(
            empresa_terceira=fornecedor_escolhido
        )

    eventos = solicitacao.evento_set.all()

    current_index = status_order.index(solicitacao.status)+1 if solicitacao.status in status_order else 0
    progress_percent = "{:.2f}".format((current_index / (len(status_order))) * 100)

    context = {
        "solicitacao": solicitacao,
        "fornecedor_escolhido": fornecedor_escolhido,
        "proposta_escolhida": proposta_escolhida,
        "indicadores": indicadores,
        "status_order": status_order,
        "current_index": current_index,
        "progress_percent": progress_percent,
        "eventos": eventos,
    }

    return render(request, "gestao_contratos/detalhes_solicitacao_contratacao.html", context)


@login_required
def detalhes_solicitacao(request, pk):
    # Busca a solicitação
    solicitacao = get_object_or_404(SolicitacaoProspeccao, pk=pk)

    if request.user.grupo == "gerente_lider" and not user_shares_center_with_coordinator(request.user, solicitacao.coordenador):
        messages.error(request, "Você não tem permissão para isso.")
        return redirect("home")

    status_order = [
        "Solicitação de prospecção",
        "Aprovada pelo suprimento",
        "Triagem realizada",
        "Fornecedor selecionado",
        "Fornecedor aprovado",
        "Planejamento do Contrato",
        "Aprovação Final",
        SIGNED_FILES_PENDING_STATUS,
        "Onboarding",
    ]

    fornecedor_escolhido = solicitacao.fornecedor_escolhido
    fornecedores_selecionados = solicitacao.fornecedores_selecionados.all()

    proposta_escolhida = None
    indicadores = None

    if fornecedor_escolhido:
        # Busca a proposta do fornecedor escolhido nesta solicitação
        proposta_escolhida = PropostaFornecedor.objects.filter(
            solicitacao=solicitacao,
            fornecedor=fornecedor_escolhido
        ).first()

        # Busca indicadores do fornecedor escolhido
        indicadores = Indicadores.objects.filter(
            empresa_terceira=fornecedor_escolhido
        )

    eventos = solicitacao.evento_set.all()

    current_index = status_order.index(solicitacao.status)+1 if solicitacao.status in status_order else 0
    progress_percent = "{:.2f}".format((current_index / (len(status_order))) * 100)

    # Propostas dos fornecedores selecionados
    propostas_dict = {}
    for f in fornecedores_selecionados:
        proposta = PropostaFornecedor.objects.filter(
            solicitacao=solicitacao,
            fornecedor=f
        ).first()
        propostas_dict[f.id] = proposta


    # Indicadores dos fornecedores selecionados
    fornecedores_indicadores = {
        f.id: Indicadores.objects.filter(empresa_terceira=f)
        for f in fornecedores_selecionados
    }

    context = {
        "solicitacao": solicitacao,
        "fornecedor_escolhido": fornecedor_escolhido,
        "proposta_escolhida": proposta_escolhida,
        "fornecedores_selecionados": fornecedores_selecionados,
        "propostas_dict": propostas_dict,
        "fornecedores_indicadores": fornecedores_indicadores,
        "indicadores": indicadores,
        "status_order": status_order,
        "current_index": current_index,
        "progress_percent": progress_percent,
        "eventos": eventos,
    }

    return render(request, "gestao_contratos/detalhes_solicitacao.html", context)


@login_required
def detalhe_os(request, pk):
    if request.user.grupo not in ['suprimento', 'lider_contrato', 'coordenador', 'gerente', 'gerente_lider', 'gerente_contrato']:
        messages.error(request, "Você não tem permissão para isso.")
        return redirect('home')

    os = get_object_or_404(SolicitacaoOrdemServico, pk=pk)

    if request.user.grupo == 'lider_contrato' and os.lider_contrato != request.user:
        messages.error(request, "Você não tem permissão para visualizar esta OS.")
        return redirect('lista_solicitacoes')

    if request.user.grupo == 'gerente_lider':
        if not user_shares_center_with_coordinator(request.user, os.coordenador):
            messages.error(request, "Você não tem permissão para visualizar esta OS.")
            return redirect('lista_solicitacoes')

    if request.user.grupo == 'coordenador' and os.solicitante != request.user:
        messages.error(request, "Você não tem permissão para visualizar esta OS.")
        return redirect('lista_solicitacoes')

    status_order = [
        'Solicitação de OS',
        'Pendente líder',
        'Pendente Gerente',
        'Pendente Suprimento',
        'Aprovada',
    ]

    status_map = {
        'solicitacao_os': 0,
        'pendente_lider': 1,
        'pendente_gerente': 2,
        'pendente_suprimento': 3,
        'aprovada': 4,
    }

    current_index = status_map.get(os.status, 0)

    total_steps = len(status_order)

    if total_steps > 1:
        progress_percent = int((current_index / (total_steps - 1)) * 100)
    else:
        progress_percent = 0

    context = {
        'os': os,
        'status_order': status_order,
        'current_index': current_index,
        'total_steps': len(status_order),
        'progress_percent': progress_percent,
    }

    return render(request, 'gestao_contratos/detalhe_os.html', context)


@login_required
def lista_ordens_servico(request):
    if request.user.grupo in ['suprimento', 'financeiro', 'diretoria']:
        os = OS.objects.all()
    elif request.user.grupo == 'coordenador':
        os = OS.objects.filter(coordenador=request.user)
    elif request.user.grupo == 'lider_contrato':
        os = OS.objects.filter(lider_contrato=request.user)
    elif request.user.grupo == 'gerente':
        os = OS.objects.filter(coordenador__centros__in=request.user.centros.all())
    elif request.user.grupo == 'gerente_lider':
        os = OS.objects.filter(
            coordenador__centros__in=request.user.centros.all()
        ).distinct()
    elif request.user.grupo == 'gerente_contrato':
        os = OS.objects.filter(lider_contrato__grupo='lider_contrato')
    else:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    data_limite = timezone.now().date() - timedelta(days=60)
    os = os.exclude(status='finalizada', data_pagamento__lte=data_limite)
    os = os.exclude(status='cancelada', prazo_execucao__lte=data_limite)

    search_query = request.GET.get('search', '').strip()

    if search_query:
        os = os.filter(
            Q(cod_projeto__cod_projeto__icontains=search_query) |
            Q(coordenador__username__icontains=search_query) |
            Q(lider_contrato__username__icontains=search_query) |
            Q(cod_projeto__cliente__nome__icontains=search_query) |
            Q(titulo__icontains=search_query) |
            Q(contrato__num_contrato__icontains=search_query)
        ).order_by('-criado_em')

    os = os.order_by('-criado_em')
    paginator = Paginator(os, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
    }

    return render(request, 'gestao_contratos/lista_ordens_servico.html', context)

@login_required
def propostas_fornecedores(request, pk):
    solicitacao = get_object_or_404(SolicitacaoProspeccao, pk=pk)

    fornecedores = solicitacao.fornecedores_selecionados.all()
    PropostaFormSet = modelformset_factory( PropostaFornecedor, form=PropostaFornecedorForm, extra=0, can_delete=False)

    for f in fornecedores:
        PropostaFornecedor.objects.get_or_create(solicitacao=solicitacao, fornecedor=f)

    queryset = PropostaFornecedor.objects.filter(solicitacao=solicitacao)
    if request.method == "POST":
        formset = PropostaFormSet(request.POST, request.FILES, queryset=queryset)
        if formset.is_valid():
            formset.save()
            messages.success(request, "Propostas salvas com sucesso!")
            return redirect("lista_solicitacoes")
    else:
        formset = PropostaFormSet(queryset=queryset)

    context = {"solicitacao": solicitacao, "formset": formset}

    return render(request, "fornecedores/propostas_fornecedores.html", context)


@login_required
def cadastrar_propostas(request, pk):
    solicitacao = get_object_or_404(SolicitacaoProspeccao, pk=pk, aprovado=True)

    # Apenas comercial pode cadastrar propostas
    if request.user.grupo != "suprimento":
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    fornecedores = solicitacao.fornecedores_selecionados.all()

    if request.method == "POST":
        for fornecedor in fornecedores:
            valor = request.POST.get(f"valor_{fornecedor.id}")
            prazo = request.POST.get(f"prazo_{fornecedor.id}")
            arquivo = request.FILES.get(f"arquivo_{fornecedor.id}")

            proposta, created = PropostaFornecedor.objects.get_or_create(
                solicitacao=solicitacao,
                fornecedor=fornecedor,
            )

            if valor:
                proposta.valor_proposta = valor
            if prazo:
                proposta.prazo_validade = prazo
            if arquivo:
                proposta.arquivo_proposta = arquivo

            proposta.save()

        messages.success(request, "Propostas salvas com sucesso!")
        return redirect("lista_solicitacoes")

    context = {
        "solicitacao": solicitacao,
        "fornecedores": fornecedores,
    }
    return render(request, "fornecedores/cadastrar_propostas.html", context)


@login_required
#@user_passes_test(is_financeiro)
def elaboracao_contrato(request):
    solicitacoes = SolicitacaoProspeccao.objects.filter(
        fornecedor_escolhido__isnull=False, aprovacao_fornecedor_gerente="aprovado"
    ).select_related("fornecedor_escolhido").exclude(status="Onboarding")

    lista_solicitacoes = []
    for s in solicitacoes:
        proposta_escolhida = PropostaFornecedor.objects.filter(
            solicitacao=s,
            fornecedor=s.fornecedor_escolhido
        ).first()

        contrato = DocumentoContratoTerceiro.objects.filter(solicitacao=s).first()
        lista_solicitacoes.append({
            "solicitacao": s,
            "fornecedor": s.fornecedor_escolhido,
            "proposta": proposta_escolhida,
            "contrato": contrato
        })

    context = {"lista_solicitacoes": lista_solicitacoes}
    return render(request, "contratos/elaboracao_contrato.html", context)


@login_required
#@user_passes_test(is_financeiro)
def cadastrar_contrato(request, solicitacao_id):
    if request.user.grupo not in ["suprimento"]:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")
    solicitacao = get_object_or_404(
        SolicitacaoProspeccao,
        id=solicitacao_id,
        fornecedor_escolhido__isnull=False
    )

    contrato_existente = DocumentoContratoTerceiro.objects.filter(solicitacao=solicitacao).first()
    documento_bm = DocumentoBM.objects.filter(solicitacao=solicitacao).first()
    proposta_escolhida = PropostaFornecedor.objects.filter(
        solicitacao=solicitacao,
        fornecedor=solicitacao.fornecedor_escolhido,
    ).first()
    allow_signed_upload = bool(
        contrato_existente
        and documento_bm
        and documento_bm.status_gerente == "aprovado"
        and solicitacao.aprovacao_gerencia is True
    )
    if request.method == "POST":
        form = DocumentoContratoTerceiroForm(request.POST, request.FILES, instance=contrato_existente)

        if form.is_valid():
            should_notify_minuta = bool(request.FILES.get("arquivo_contrato"))
            is_signed_file_upload_only = bool(request.FILES.get("arquivo_contrato_assinado")) and not should_notify_minuta
            contrato = form.save(commit=False)
            contrato.solicitacao = solicitacao
            if not is_signed_file_upload_only:
                if DocumentoBM.objects.filter(solicitacao=solicitacao).exists():
                    solicitacao.status = "Aprovação Final"
                else:
                    solicitacao.status = "Planejamento do Contrato"
                solicitacao.save()
            contrato.prazo_inicio = solicitacao.data_inicio
            contrato.prazo_fim = solicitacao.data_fim

            # mantém arquivo antigo se não foi enviado novo
            if not request.FILES.get("arquivo_contrato") and contrato_existente:
                contrato.arquivo_contrato = contrato_existente.arquivo_contrato
            if not request.FILES.get("arquivo_contrato_assinado") and contrato_existente:
                contrato.arquivo_contrato_assinado = contrato_existente.arquivo_contrato_assinado

            contrato.save()
            criar_contrato_se_aprovado(solicitacao)

            gerente = User.objects.filter(grupo="gerente_contrato").values_list("email", flat=True).distinct()

            if should_notify_minuta and gerente:
                assunto = "Foi anexado uma nova minuta de contrato"
                mensagem = (
                    f"Olá,\n\n"
                    f"A equipe de Suprimento anexou uma nova minuta de contrato para análise.\n\n"
                    "por favor, acesse o sistema HIDROGestão para avaliar a referente minuta.\n"
                    "https://hidrogestao.pythonanywhere.com/"
                )
                try:
                    send_mail(
                        assunto, mensagem,
                        FROM_EMAIL,
                        list(gerente),
                        fail_silently=False,
                    )
                except Exception as e:
                    messages.warning(request, f"Erro ao enviar e-mail para gerente: {e}")
            messages.success(request, "Contrato salvo com sucesso!")
            return redirect("detalhes_solicitacao", pk=solicitacao.pk)
        else:
            messages.error(request, f"Erro ao salvar contrato: {form.errors}")
    else:
        form = DocumentoContratoTerceiroForm(instance=contrato_existente)

    context = {
        "form": form,
        "solicitacao": solicitacao,
        "fornecedor": solicitacao.fornecedor_escolhido,
        "contrato": contrato_existente,
        "documento_bm": documento_bm,
        "proposta_escolhida": proposta_escolhida,
        "allow_signed_upload": allow_signed_upload,
        "generation_url": reverse_lazy("gerar_minuta_contrato_docm", kwargs={"solicitacao_id": solicitacao.id}),
    }
    return render(request, "fornecedores/cadastrar_contrato.html", context)


def cadastrar_minuta_contrato(request, solicitacao_id):
    if request.user.grupo not in ["suprimento"]:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")
    solicitacao = get_object_or_404(
        SolicitacaoContrato,
        id=solicitacao_id,
        fornecedor_escolhido__isnull=False
    )

    contrato_existente = DocumentoContratoTerceiro.objects.filter(solicitacao_contrato=solicitacao).first()
    documento_bm = DocumentoBM.objects.filter(solicitacao_contrato=solicitacao).first()
    proposta_escolhida = PropostaFornecedor.objects.filter(
        solicitacao_contrato=solicitacao,
        fornecedor=solicitacao.fornecedor_escolhido,
    ).first()
    allow_signed_upload = bool(
        contrato_existente
        and (
            solicitacao.guarda_chuva
            or (documento_bm and documento_bm.status_gerente == "aprovado")
        )
        and solicitacao.aprovacao_gerencia is True
    )
    if request.method == "POST":
        form = DocumentoContratoTerceiroForm(request.POST, request.FILES, instance=contrato_existente)

        if form.is_valid():
            should_notify_minuta = bool(request.FILES.get("arquivo_contrato"))
            is_signed_file_upload_only = bool(request.FILES.get("arquivo_contrato_assinado")) and not should_notify_minuta
            contrato = form.save(commit=False)
            contrato.solicitacao_contrato = solicitacao
            contrato.prazo_inicio = solicitacao.data_inicio
            contrato.prazo_fim = solicitacao.data_fim
            if solicitacao.valor_provisionado:
                contrato.valor_total = solicitacao.valor_provisionado
            if not is_signed_file_upload_only:
                solicitacao.aprovacao_gerencia = False
                if not solicitacao.guarda_chuva and hasattr(solicitacao, "minuta_boletins_medicao_contrato"):
                    solicitacao.status = "Aprovação Final"
                else:
                    solicitacao.status = "Planejamento do Contrato"
                solicitacao.save()
            # mantém arquivo antigo se não foi enviado novo
            if not request.FILES.get("arquivo_contrato") and contrato_existente:
                contrato.arquivo_contrato = contrato_existente.arquivo_contrato
            if not request.FILES.get("arquivo_contrato_assinado") and contrato_existente:
                contrato.arquivo_contrato_assinado = contrato_existente.arquivo_contrato_assinado

            contrato.save()
            criar_contrato_se_aprovado_minuta(solicitacao)

            gerente = User.objects.filter(grupo="gerente_contrato").values_list("email", flat=True).distinct()

            if should_notify_minuta and gerente:
                assunto = "Foi anexado uma nova minuta de contrato"
                mensagem = (
                    f"Olá,\n\n"
                    f"A equipe de Suprimento anexou uma nova minuta de contrato para análise.\n\n"
                    "por favor, acesse o sistema HIDROGestão para avaliar a referente minuta.\n"
                    "https://hidrogestao.pythonanywhere.com/"
                )
                try:
                    send_mail(
                        assunto, mensagem,
                        FROM_EMAIL,
                        list(gerente),
                        fail_silently=False,
                    )
                except Exception as e:
                    messages.warning(request, f"Erro ao enviar e-mail para gerente: {e}")
            messages.success(request, "Contrato salvo com sucesso!")
            return redirect("lista_solicitacoes")
        else:
            messages.error(request, f"Erro ao salvar contrato: {form.errors}")
    else:
        form = DocumentoContratoTerceiroForm(instance=contrato_existente)

    context = {
        "form": form,
        "solicitacao": solicitacao,
        "fornecedor": solicitacao.fornecedor_escolhido,
        "contrato": contrato_existente,
        "documento_bm": documento_bm,
        "proposta_escolhida": proposta_escolhida,
        "allow_signed_upload": allow_signed_upload,
        "generation_url": reverse_lazy("gerar_minuta_contrato_contratacao_docm", kwargs={"solicitacao_id": solicitacao.id}),
    }
    return render(request, "fornecedores/cadastrar_contrato.html", context)


def build_contract_docm_replacements(documento_contrato, fornecedor, contrato_projeto, proposta=None, guarda_chuva=False):
    contrato_codigo = contrato_projeto.cod_projeto if contrato_projeto else "-"
    data_inicio = documento_contrato.prazo_inicio
    data_fim = documento_contrato.prazo_fim
    valor_total = documento_contrato.valor_total or (proposta.valor_proposta if proposta else Decimal("0.00"))
    proposal_document = "-"
    if proposta and proposta.arquivo_proposta:
        proposal_document = os.path.basename(proposta.arquivo_proposta.name)
    contract_type_line = "☐ ESPECÍFICO    ☒ GUARDA-CHUVA" if guarda_chuva else "☒ ESPECÍFICO    ☐ GUARDA-CHUVA"

    return {
        "__nome_empresa_terceira__": fornecedor.nome or "-",
        "__nome_empresa__": fornecedor.nome or "-",
        "__cpf_cnpj__": fornecedor.cpf_cnpj or "-",
        "__endereco__": fornecedor.endereco or "-",
        "__numero__": fornecedor.numero or "-",
        "__bairro__": fornecedor.bairro or "-",
        "__municipio__": fornecedor.municipio or "-",
        "__estado__": fornecedor.estado or "-",
        "__cep__": fornecedor.cep or "-",
        "__telefone__": fornecedor.telefone or "-",
        "__email__": fornecedor.email or "-",
        "__ponto_focal__": fornecedor.ponto_focal or "-",
        "__telefone_focal__": fornecedor.telefone_focal or "-",
        "__email_focal__": fornecedor.email_focal or "-",
        "__informacoes_bancarias__": fornecedor.informacoes_bancarias or "-",
        "__descricao__": documento_contrato.objeto or "-",
        "__ valor_proposta__": format_currency_br(valor_total, with_symbol=True),
        "__valor_proposta_extenso__": decimal_to_money_words_pt_br(valor_total),
        "__contrato__": contrato_codigo or "-",
        "__dias_totais__": calculate_inclusive_days(data_inicio, data_fim),
        "__data_inicio__": format_date_br(data_inicio),
        "__data_fim__": format_date_br(data_fim),
        "__documento_proposta__": proposal_document,
        "__numero_revisao__": "00",
        "__data_hoje__": format_date_br(timezone.localdate()),
        "__cod_contrato__": documento_contrato.numero_contrato or "-",
        "☒ ESPECÍFICO    ☐ GUARDA-CHUVA": contract_type_line,
    }


def build_addendum_docm_replacements(aditivo):
    contrato = aditivo.contrato
    fornecedor = contrato.empresa_terceira
    contrato_codigo = contrato.cod_projeto.cod_projeto if contrato.cod_projeto else "-"
    ordem_aditivo = contrato.aditivos.filter(pk__lte=aditivo.pk).count() or 1
    ordem_label = f"{ordem_aditivo}º"
    descricao = contrato.objeto or aditivo.motivo or "-"
    contract_type_line = "☐ ESPECÍFICO    ☒ GUARDA-CHUVA" if contrato.guarda_chuva else "☒ ESPECÍFICO    ☐ GUARDA-CHUVA"

    return {
        "__ordem_aditivo__": ordem_label,
        "__ordem_adtivo__": ordem_label,
        "__cpf_cnpj__": fornecedor.cpf_cnpj or "-",
        "__numero_contrato__": contrato.num_contrato or "-",
        "__data_fim__": format_date_br(aditivo.data_fim_anterior or contrato.data_fim),
        "__data_fim_aditivo__": format_date_br(aditivo.nova_data_fim),
        "__descricao__": descricao,
        "__novo_valor _total__": format_currency_br(aditivo.novo_valor_total or contrato.valor_total or Decimal("0.00"), with_symbol=True),
        "__novo_valor _total_extenso__": decimal_to_money_words_pt_br(aditivo.novo_valor_total or contrato.valor_total or Decimal("0.00")),
        "__contrato__": contrato_codigo or "-",
        "__informacoes_bancarias__": fornecedor.informacoes_bancarias or "-",
        "__endereco__": fornecedor.endereco or "-",
        "__numero__": fornecedor.numero or "-",
        "__complemento__": "-",
        "__bairro__": fornecedor.bairro or "-",
        "__municipio__": fornecedor.municipio or "-",
        "__estado__": fornecedor.estado or "-",
        "__cep__": fornecedor.cep or "-",
        "__telefone__": fornecedor.telefone or "-",
        "__email__": fornecedor.email or "-",
        "__ponto_focal__": fornecedor.ponto_focal or "-",
        "__telefone_focal__": fornecedor.telefone_focal or "-",
        "__email_focal__": fornecedor.email_focal or "-",
        "__dias_totais_novo__": calculate_inclusive_days(contrato.data_inicio, aditivo.nova_data_fim or contrato.data_fim),
        "__data_inicio__": format_date_br(contrato.data_inicio),
        "__nova_data_fim__": format_date_br(aditivo.nova_data_fim),
        "__data_hoje_completo__": format_date_long_br(timezone.localdate()),
        "__nome_empresa__": fornecedor.nome or "-",
        "__numero_revisao__": "00",
        "__data_hoje__": format_date_br(timezone.localdate()),
        "☒ ESPECÍFICO    ☐ GUARDA-CHUVA": contract_type_line,
    }


@login_required
def gerar_minuta_contrato_docm(request, solicitacao_id):
    if request.user.grupo != "suprimento":
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    solicitacao = get_object_or_404(
        SolicitacaoProspeccao,
        id=solicitacao_id,
        fornecedor_escolhido__isnull=False,
    )
    contrato_existente = DocumentoContratoTerceiro.objects.filter(solicitacao=solicitacao).first()
    proposta_escolhida = PropostaFornecedor.objects.filter(
        solicitacao=solicitacao,
        fornecedor=solicitacao.fornecedor_escolhido,
    ).first()

    if request.method != "POST":
        return redirect("cadastrar_contrato", solicitacao_id=solicitacao_id)
    if not CONTRACT_TEMPLATE_DOCM_PATH.exists():
        messages.error(request, f"Modelo de contrato não encontrado em {CONTRACT_TEMPLATE_DOCM_PATH}.")
        return redirect("cadastrar_contrato", solicitacao_id=solicitacao_id)

    form = DocumentoContratoTerceiroForm(request.POST, instance=contrato_existente)
    documento_contrato = build_contract_document_preview(request.POST, contrato_existente)
    documento_contrato.solicitacao = solicitacao
    documento_contrato.prazo_inicio = solicitacao.data_inicio
    documento_contrato.prazo_fim = solicitacao.data_fim
    proposta = get_selected_supplier_proposal(
        solicitacao=solicitacao,
        fornecedor=solicitacao.fornecedor_escolhido,
    )
    generated_file = replace_placeholders_in_docm(
        CONTRACT_TEMPLATE_DOCM_PATH,
        build_contract_docm_replacements(
            documento_contrato,
            solicitacao.fornecedor_escolhido,
            solicitacao.contrato,
            proposta=proposta,
            guarda_chuva=solicitacao.guarda_chuva,
        ),
    )
    response = HttpResponse(generated_file, content_type="application/vnd.ms-word.document.macroEnabled.12")
    response["Content-Disposition"] = f'attachment; filename="Contrato_{documento_contrato.numero_contrato or solicitacao.id}.docm"'
    return response


@login_required
def gerar_minuta_contrato_contratacao_docm(request, solicitacao_id):
    if request.user.grupo != "suprimento":
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    solicitacao = get_object_or_404(
        SolicitacaoContrato,
        id=solicitacao_id,
        fornecedor_escolhido__isnull=False,
    )
    contrato_existente = DocumentoContratoTerceiro.objects.filter(solicitacao_contrato=solicitacao).first()
    proposta_escolhida = PropostaFornecedor.objects.filter(
        solicitacao_contrato=solicitacao,
        fornecedor=solicitacao.fornecedor_escolhido,
    ).first()

    if request.method != "POST":
        return redirect("cadastrar_minuta_contrato", solicitacao_id=solicitacao_id)
    if not CONTRACT_TEMPLATE_DOCM_PATH.exists():
        messages.error(request, f"Modelo de contrato não encontrado em {CONTRACT_TEMPLATE_DOCM_PATH}.")
        return redirect("cadastrar_minuta_contrato", solicitacao_id=solicitacao_id)

    form = DocumentoContratoTerceiroForm(request.POST, instance=contrato_existente)
    documento_contrato = build_contract_document_preview(request.POST, contrato_existente)
    documento_contrato.solicitacao_contrato = solicitacao
    documento_contrato.prazo_inicio = solicitacao.data_inicio
    documento_contrato.prazo_fim = solicitacao.data_fim
    if solicitacao.valor_provisionado:
        documento_contrato.valor_total = solicitacao.valor_provisionado
    proposta = get_selected_supplier_proposal(
        solicitacao_contrato=solicitacao,
        fornecedor=solicitacao.fornecedor_escolhido,
    )
    generated_file = replace_placeholders_in_docm(
        CONTRACT_TEMPLATE_DOCM_PATH,
        build_contract_docm_replacements(
            documento_contrato,
            solicitacao.fornecedor_escolhido,
            solicitacao.contrato,
            proposta=proposta,
            guarda_chuva=solicitacao.guarda_chuva,
        ),
    )
    response = HttpResponse(generated_file, content_type="application/vnd.ms-word.document.macroEnabled.12")
    response["Content-Disposition"] = f'attachment; filename="Contrato_{documento_contrato.numero_contrato or solicitacao.id}.docm"'
    return response


@login_required
def gerar_aditivo_contrato_docm(request, pk):
    aditivo = get_object_or_404(AditivoContratoTerceiro, pk=pk)

    if request.user.grupo != "suprimento":
        messages.error(request, "Somente o Suprimento pode gerar o documento do aditivo.")
        return redirect("contrato_fornecedor_detalhe", pk=aditivo.contrato_id)
    if not ADDENDUM_TEMPLATE_DOCM_PATH.exists():
        messages.error(request, f"Modelo de aditivo não encontrado em {ADDENDUM_TEMPLATE_DOCM_PATH}.")
        return redirect("enviar_documento_aditivo_contrato", pk=aditivo.pk)

    generated_file = replace_placeholders_in_docm(
        ADDENDUM_TEMPLATE_DOCM_PATH,
        build_addendum_docm_replacements(aditivo),
    )
    response = HttpResponse(generated_file, content_type="application/vnd.ms-word.document.macroEnabled.12")
    response["Content-Disposition"] = f'attachment; filename="Aditivo_{aditivo.contrato.num_contrato or aditivo.pk}.docm"'
    return response


def criar_contrato_se_aprovado(solicitacao):
    try:
        bm = solicitacao.minuta_boletins_medicao
    except DocumentoBM.DoesNotExist:
        return None

    bm_aprovado = bm.status_gerente == 'aprovado'
    contrato_aprovado = solicitacao.aprovacao_gerencia is True

    contrato_existente = ContratoTerceiros.objects.filter(prospeccao=solicitacao).first()
    if bm_aprovado and contrato_aprovado and not contrato_existente:
        documento = DocumentoContratoTerceiro.objects.filter(solicitacao=solicitacao).first()
        if not documento:
            return None
        if not documento.arquivo_contrato_assinado:
            update_signed_files_pending_status(solicitacao, bm=bm, documento=documento)
            return None
        proposta = PropostaFornecedor.objects.filter(
            solicitacao=solicitacao,
            fornecedor=solicitacao.fornecedor_escolhido
        ).first()

        contrato = ContratoTerceiros.objects.create(
            cod_projeto=solicitacao.contrato,
            prospeccao=solicitacao,
            lider_contrato=solicitacao.lider_contrato,
            num_contrato=documento.numero_contrato if documento else None,
            empresa_terceira=solicitacao.fornecedor_escolhido,
            coordenador=solicitacao.coordenador,
            data_inicio=documento.prazo_inicio if documento else None,
            data_fim=documento.prazo_fim if documento else None,
            valor_total=documento.valor_total if documento else 0,
            objeto=documento.objeto if documento else "",
            condicao_pagamento=proposta.condicao_pagamento if proposta else None,
            status="ativo",
            num_contrato_arquivo = documento.arquivo_contrato_assinado,
            observacao=documento.observacao if documento else None,
        )
        Evento.objects.filter(
            prospeccao=solicitacao,
            contrato_terceiro__isnull=True
        ).update(contrato_terceiro=contrato)

        solicitacao.status = "Onboarding"
        solicitacao.save()

        notify_contract_process_completed(solicitacao, contrato)

        return contrato

    return None


def criar_contrato_se_aprovado_minuta(solicitacao):
    bm = None
    if not solicitacao.guarda_chuva:
        try:
            bm = solicitacao.minuta_boletins_medicao_contrato
        except DocumentoBM.DoesNotExist:
            return None

    bm_aprovado = True if solicitacao.guarda_chuva else bm.aprovado_gerente
    contrato_aprovado = solicitacao.aprovacao_gerencia is True

    contrato_existente = ContratoTerceiros.objects.filter(solicitacao=solicitacao).first()
    if bm_aprovado and contrato_aprovado and not contrato_existente:
        documento = DocumentoContratoTerceiro.objects.filter(solicitacao_contrato=solicitacao).first()
        if not documento:
            return None
        if not documento.arquivo_contrato_assinado:
            update_signed_files_pending_status(solicitacao, bm=bm, documento=documento)
            return None
        proposta = PropostaFornecedor.objects.filter(
            solicitacao_contrato=solicitacao,
            fornecedor=solicitacao.fornecedor_escolhido
        ).first()


        contrato = ContratoTerceiros.objects.create(
            cod_projeto=solicitacao.contrato if solicitacao.contrato else None,
            solicitacao=solicitacao,
            lider_contrato=solicitacao.lider_contrato,
            guarda_chuva=solicitacao.guarda_chuva,
            num_contrato=documento.numero_contrato if documento else None,
            empresa_terceira=solicitacao.fornecedor_escolhido,
            coordenador=solicitacao.coordenador,
            data_inicio=documento.prazo_inicio if documento else None,
            data_fim=documento.prazo_fim if documento else None,
            valor_total=documento.valor_total if documento else 0,
            objeto=documento.objeto if documento else "",
            condicao_pagamento=proposta.condicao_pagamento if proposta else None,
            status="ativo",
            num_contrato_arquivo = documento.arquivo_contrato_assinado,
            observacao=documento.observacao if documento else None,
        )
        Evento.objects.filter(
            solicitacao_contrato=solicitacao,
            contrato_terceiro__isnull=True
        ).update(contrato_terceiro=contrato)

        solicitacao.status = "Onboarding"
        solicitacao.save()

        notify_contract_process_completed(solicitacao, contrato)

        return contrato

    return None




@login_required
def detalhes_contrato(request, pk):
    solicitacao = get_object_or_404(SolicitacaoProspeccao, pk=pk)

    contrato_doc = getattr(solicitacao, "contrato_relacionado", None)
    fornecedor_escolhido = solicitacao.fornecedor_escolhido
    proposta_escolhida = None
    if fornecedor_escolhido:
        proposta_escolhida = PropostaFornecedor.objects.filter(
            solicitacao=solicitacao, fornecedor=fornecedor_escolhido
        ).first()
    documento_bm = getattr(solicitacao, "minuta_boletins_medicao", None)
    fornecedores_selecionados = solicitacao.fornecedores_selecionados.all()
    revisoes = solicitacao.revisoes.all()
    origem = solicitacao.solicitacao_origem

    if request.method == "POST" and request.user.grupo == "gerente_contrato" and contrato_doc:
        acao = request.POST.get("acao")
        justificativa = request.POST.get("justificativa", "")

        if acao == "aprovar":
            solicitacao.aprovacao_gerencia = True
            solicitacao.reprovacao_gerencia = False
            solicitacao.justificativa_gerencia = ""
            messages.success(request, "Documento do contrato aprovado pela gerência.")
        elif acao == "reprovar":
            solicitacao.aprovacao_gerencia = False
            solicitacao.reprovacao_gerencia = True
            solicitacao.justificativa_gerencia = justificativa
            messages.warning(request, "Documento do contrato reprovado pela gerência.")
        else:
            messages.error(request, "Ação inválida.")

        solicitacao.save()

        # Tenta criar o contrato caso BM e documento do contrato estejam aprovados
        criar_contrato_se_aprovado(solicitacao)

        return redirect("lista_solicitacoes")

    return render(request, "gestao_contratos/detalhes_contrato.html", {
        "solicitacao": solicitacao,
        "contrato_doc": contrato_doc,
        "documento_bm": documento_bm,
        "fornecedor_escolhido": fornecedor_escolhido,
        "proposta_escolhida": proposta_escolhida,
        "fornecedores_selecionados": fornecedores_selecionados,
        "revisoes": revisoes,
        "origem": origem,
        "signed_files_pending_status": SIGNED_FILES_PENDING_STATUS,
    })


@login_required
def detalhes_minuta_contrato(request, pk):
    solicitacao = get_object_or_404(SolicitacaoContrato, pk=pk)

    contrato_doc = getattr(solicitacao, "minuta_contrato", None)
    fornecedor_escolhido = solicitacao.fornecedor_escolhido
    proposta_escolhida = None
    if fornecedor_escolhido:
        proposta_escolhida = PropostaFornecedor.objects.filter(
            solicitacao_contrato=solicitacao,
            fornecedor=fornecedor_escolhido,
        ).first()
    documento_bm = getattr(solicitacao, "minuta_boletins_medicao_contrato", None)

    if request.method == "POST" and request.user.grupo == "gerente_contrato" and contrato_doc:
        acao = request.POST.get("acao")
        justificativa = request.POST.get("justificativa", "")

        if acao == "aprovar":
            solicitacao.aprovacao_gerencia = True
            solicitacao.reprovacao_gerencia = False
            solicitacao.justificativa_gerencia = ""
            messages.success(request, "Documento do contrato aprovado pela gerência.")
        elif acao == "reprovar":
            if not justificativa:
                messages.warning(request, "Por favor, insira uma justificativa para a reprovação.")
                return redirect('detalhes_minuta_contrato', pk=pk)
            solicitacao.aprovacao_gerencia = False
            solicitacao.reprovacao_gerencia = True
            solicitacao.justificativa_gerencia = justificativa
            messages.warning(request, "Documento do contrato reprovado pela gerência.")

            suprimentos = User.objects.filter(grupo="suprimento").exclude(email__isnull=True).exclude(email__exact="")
            lista_emails = [u.email for u in suprimentos]
            if lista_emails:
                assunto = f"Solicitação #{solicitacao.id} - Minuta de Contrato Reprovada"
                mensagem = (
                    f"Olá, equipe de Suprimentos!\n\n"
                    f"A Gerência de Contratos reprovou a minuta do contrato:\n\n"
                    f"Solicitação: {solicitacao.id}\n"
                    f"Fornecedor: {solicitacao.fornecedor_escolhido}\n"
                    f"Justificativa da reprovação: {solicitacao.justificativa_gerencia}\n"
                    "Acesse o sistema HIDROGestão para mais informações.\n"
                    "https://hidrogestao.pythonanywhere.com/"
                )
                try:
                    send_mail(
                        assunto, mensagem,
                        FROM_EMAIL,
                        lista_emails,
                        fail_silently=False
                    )
                except Exception as e:
                    messages.warning(request, f"Erro ao enviar e-mail para suprimentos: {e}")

            diretoria = User.objects.filter(grupo="diretoria").exclude(email__isnull=True).exclude(email__exact="")
            lista_emails = [u.email for u in diretoria]

            if solicitacao.lider_contrato and solicitacao.lider_contrato.email:
                lista_emails.append(solicitacao.lider_contrato.email)
            lista_emails = list(set(lista_emails))

            if lista_emails:
                assunto = f"Solicitação #{solicitacao.id} - Minuta de Contrato Reprovada"
                mensagem = (
                    f"Olá, \n\n"
                    f"A Gerência de Contratos reprovou a minuta do contrato:\n\n"
                    f"Solicitação: {solicitacao.id}\n"
                    f"Fornecedor: {solicitacao.fornecedor_escolhido}\n"
                    f"Justificativa da reprovação: {solicitacao.justificativa_gerencia}\n"
                    "A minuta do contrato será revisada e passará por uma nova avaliação."
                    "Acesse o sistema HIDROGestão para mais informações.\n"
                    "https://hidrogestao.pythonanywhere.com/"
                )
                try:
                    send_mail(
                        assunto, mensagem,
                        FROM_EMAIL,
                        lista_emails,
                        fail_silently=False
                    )
                except Exception as e:
                    messages.warning(request, f"Erro ao enviar e-mail para suprimentos: {e}")

        else:
            messages.error(request, "Ação inválida.")

        solicitacao.save()

        # Tenta criar o contrato caso BM e documento do contrato estejam aprovados
        criar_contrato_se_aprovado_minuta(solicitacao)

        return redirect("lista_solicitacoes")

    return render(request, "gestao_contratos/detalhes_minuta_contrato.html", {
        "solicitacao": solicitacao,
        "contrato_doc": contrato_doc,
        "documento_bm": documento_bm,
        "fornecedor_escolhido": fornecedor_escolhido,
        "proposta_escolhida": proposta_escolhida,
        "signed_files_pending_status": SIGNED_FILES_PENDING_STATUS,
    })


@login_required
def renegociar_valor(request, pk):
    solicitacao = get_object_or_404(SolicitacaoProspeccao, pk=pk)
    proposta = PropostaFornecedor.objects.filter(solicitacao=solicitacao).first()

    if not proposta:
        return render(request, "erro.html", {"mensagem": "Nenhuma proposta encontrada para esta solicitação."})

    if request.method == "POST":
        novo_valor = request.POST.get("valor_proposta")
        if novo_valor:
            proposta.valor_proposta = novo_valor
            proposta.save()
            return redirect("detalhes_contrato", pk=solicitacao.pk)

    return render(request, "gestao_contratos/renegociar_valor.html", {"solicitacao": solicitacao, "proposta": proposta})


@login_required
def renegociar_prazo(request, pk):
    solicitacao = get_object_or_404(SolicitacaoProspeccao, pk=pk)
    contrato = DocumentoContratoTerceiro.objects.filter(solicitacao=solicitacao).first()

    if not contrato:
        return render(request, "erro.html", {"mensagem": "Nenhum contrato encontrado para esta solicitação."})

    if request.method == "POST":
        prazo_inicio = request.POST.get("prazo_inicio")
        prazo_fim = request.POST.get("prazo_fim")

        if prazo_inicio and prazo_fim:
            contrato.prazo_inicio = prazo_inicio
            contrato.prazo_fim = prazo_fim
            contrato.save()
            return redirect("detalhes_contrato", pk=solicitacao.pk)

    return render(request, "gestao_contratos/renegociar_prazo.html", {"solicitacao": solicitacao, "contrato": contrato})


@login_required
def nova_prospeccao(request, pk):
    solicitacao = get_object_or_404(SolicitacaoProspeccao, pk=pk)

    if request.method == "POST":
        nova_solicitacao = SolicitacaoProspeccao.objects.create(
            contrato=f"{solicitacao.contrato} - Revisão {timezone.now().strftime('%d/%m/%Y %H:%M')}",
            descricao=solicitacao.descricao,
            criado_por=request.user,
            status="Em Prospecção",
            solicitacao_origem=solicitacao
        )
        return redirect("detalhes_contrato", pk=nova_solicitacao.pk)

    return render(request, "nova_prospeccao.html", {"solicitacao": solicitacao})


@login_required
def inserir_minuta_bm(request, pk):
    solicitacao = get_object_or_404(SolicitacaoProspeccao, pk=pk)

    if request.user.grupo != 'suprimento':
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    # Cria ou recupera a minuta ligada Ã  solicitação
    documento_bm, created = DocumentoBM.objects.get_or_create(solicitacao=solicitacao)
    contrato_doc = getattr(solicitacao, "contrato_relacionado", None)
    allow_signed_upload = bool(
        contrato_doc
        and documento_bm.status_gerente == "aprovado"
        and solicitacao.aprovacao_gerencia is True
    )

    if request.method == "POST":
        form = DocumentoBMForm(request.POST, request.FILES, instance=documento_bm)
        if form.is_valid():
            documento_salvo = form.save(commit=False)
            if not request.FILES.get("minuta_boletim") and documento_bm.minuta_boletim:
                documento_salvo.minuta_boletim = documento_bm.minuta_boletim
            if not request.FILES.get("minuta_boletim_assinado") and documento_bm.minuta_boletim_assinado:
                documento_salvo.minuta_boletim_assinado = documento_bm.minuta_boletim_assinado
            documento_salvo.save()
            if hasattr(solicitacao, "contrato_relacionado"):
                solicitacao.status = "Aprovação Final"
            else:
                solicitacao.status = "Planejamento do Contrato"
            solicitacao.save()
            criar_contrato_se_aprovado(solicitacao)
            messages.success(request, "Minuta do Boletim de Medição enviada com sucesso!")
            return redirect('lista_solicitacoes')
    else:
        form = DocumentoBMForm(instance=documento_bm)

    return render(request, 'fornecedores/inserir_minuta_bm.html', {
        'solicitacao': solicitacao,
        'form': form,
        'documento_bm': documento_bm,
        'allow_signed_upload': allow_signed_upload,
    })


@login_required
def inserir_minuta_bm_contrato(request, pk):
    solicitacao = get_object_or_404(SolicitacaoContrato, pk=pk)

    if request.user.grupo != 'suprimento':
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")
    if solicitacao.guarda_chuva:
        messages.warning(request, "Solicitações guarda-chuva não exigem minuta de BM.")
        return redirect("detalhes_solicitacao_contrato", pk=solicitacao.pk)

    # Cria ou recupera a minuta ligada Ã  solicitação
    documento_bm, created = DocumentoBM.objects.get_or_create(solicitacao_contrato=solicitacao)
    contrato_doc = getattr(solicitacao, "minuta_contrato", None)
    allow_signed_upload = bool(
        contrato_doc
        and documento_bm.status_gerente == "aprovado"
        and solicitacao.aprovacao_gerencia is True
    )

    if request.method == "POST":
        form = DocumentoBMForm(request.POST, request.FILES, instance=documento_bm)
        if form.is_valid():
            documento_salvo = form.save(commit=False)
            if not request.FILES.get("minuta_boletim") and documento_bm.minuta_boletim:
                documento_salvo.minuta_boletim = documento_bm.minuta_boletim
            if not request.FILES.get("minuta_boletim_assinado") and documento_bm.minuta_boletim_assinado:
                documento_salvo.minuta_boletim_assinado = documento_bm.minuta_boletim_assinado
            documento_salvo.save()
            if hasattr(solicitacao, "minuta_contrato"):
                solicitacao.status = "Planejamento do Contrato"
            solicitacao.save()
            criar_contrato_se_aprovado_minuta(solicitacao)
            messages.success(request, "Minuta do Boletim de Medição enviada com sucesso!")
            return redirect('lista_solicitacoes')
    else:
        form = DocumentoBMForm(instance=documento_bm)

    return render(request, 'fornecedores/inserir_minuta_bm.html', {
        'solicitacao': solicitacao,
        'form': form,
        'documento_bm': documento_bm,
        'allow_signed_upload': allow_signed_upload,
    })


@login_required
def detalhe_bm(request, pk):
    if request.user.grupo not in ["gerente_contrato", "lider_contrato", "gerente_lider"]:
        messages.error(request, "Você não tem permissão para isso.")
        return redirect("home")
    bm = get_object_or_404(DocumentoBM, pk=pk)
    solicitacao = bm.solicitacao
    usuario = request.user

    if usuario.grupo == "gerente_lider" and not user_shares_center_with_coordinator(usuario, solicitacao.coordenador):
        messages.error(request, "Você não tem permissão para isso.")
        return redirect("home")

    if request.method == "POST":
        acao = request.POST.get("acao")

        # Avaliação do coordenador
        if usuario.grupo == "lider_contrato" or (
            usuario.grupo == "gerente_lider" and user_shares_center_with_coordinator(usuario, solicitacao.coordenador)
        ):
            if usuario.grupo == "lider_contrato" and usuario != solicitacao.lider_contrato:
                messages.error(request, "Você não tem permissão para isso.")
                return redirect("home")
            if acao == "aprovar":
                bm.status_coordenador = "aprovado"
                bm.data_aprovacao_coordenador = timezone.now()
                messages.success(request, "Minuta BM aprovada pelo coordenador.")
            elif acao == "reprovar":
                bm.status_coordenador = "reprovado"
                bm.data_aprovacao_coordenador = timezone.now()
                messages.warning(request, "Minuta BM reprovada pelo coordenador.")

        # Avaliação do gerente
        elif usuario.grupo == "gerente_contrato":
            if acao == "aprovar":
                bm.status_gerente = "aprovado"
                bm.data_aprovacao_gerente = timezone.now()
                messages.success(request, "Minuta BM aprovada pelo gerente.")
            elif acao == "reprovar":
                bm.status_gerente = "reprovado"
                bm.data_aprovacao_gerente = timezone.now()
                solicitacao.status = "Planejamento do Contrato"
                solicitacao.save()
                messages.warning(request, "Minuta BM reprovada pelo gerente.")

        bm.save()
        # Tenta criar o contrato caso BM e documento do contrato estejam aprovados
        criar_contrato_se_aprovado(solicitacao)

        return redirect("lista_solicitacoes")

    return render(request, "gestao_contratos/detalhe_bm.html", {
        "bm": bm,
        "solicitacao": solicitacao
    })


@login_required
def detalhe_bm_contrato(request, pk):
    if request.user.grupo not in ["gerente_contrato", "lider_contrato", "gerente_lider"]:
        messages.error(request, "Você não tem permissão para isso.")
        return redirect("home")
    bm = get_object_or_404(DocumentoBM, pk=pk)
    solicitacao = bm.solicitacao_contrato
    usuario = request.user

    if usuario.grupo == "gerente_lider" and not user_shares_center_with_coordinator(usuario, solicitacao.coordenador):
        messages.error(request, "Você não tem permissão para isso.")
        return redirect("home")

    if request.method == "POST":
        acao = request.POST.get("acao")

        if usuario.grupo == "lider_contrato" or (
            usuario.grupo == "gerente_lider" and user_shares_center_with_coordinator(usuario, solicitacao.coordenador)
        ):
            if usuario.grupo == "lider_contrato" and usuario != solicitacao.lider_contrato:
                messages.error(request, "Você não tem permissão para isso.")
                return redirect("home")
            if acao == "aprovar":
                bm.status_coordenador = "aprovado"
                bm.data_aprovacao_coordenador = timezone.now()
                messages.success(request, "Minuta BM aprovada pelo líder.")
            elif acao == "reprovar":
                bm.status_coordenador = "reprovado"
                bm.data_aprovacao_coordenador = timezone.now()
                messages.warning(request, "Minuta BM reprovada pelo líder.")

        elif usuario.grupo == "gerente_contrato":
            if acao == "aprovar":
                bm.status_gerente = "aprovado"
                bm.data_aprovacao_gerente = timezone.now()
                messages.success(request, "Minuta BM aprovada pelo gerente.")
            elif acao == "reprovar":
                bm.status_gerente = "reprovado"
                bm.data_aprovacao_gerente = timezone.now()
                messages.warning(request, "Minuta BM reprovada pelo gerente.")

        bm.save()
        criar_contrato_se_aprovado_minuta(solicitacao)

        return redirect("lista_solicitacoes")

    return render(request, "gestao_contratos/detalhe_bm.html", {
        "bm": bm,
        "solicitacao": solicitacao
    })


@login_required
def aprovar_bm(request, pk, papel):
    bm = get_object_or_404(DocumentoBM, pk=pk)

    if papel == "coordenador" and request.user.groups.filter(name="Coordenador de Contrato").exists():
        bm.status_coordenador = "aprovado"
        bm.data_aprovacao_coordenador = timezone.now()
    elif papel == "gerente" and request.user.groups.filter(name="Gerente de Contrato").exists():
        bm.status_gerente = "aprovado"
        bm.data_aprovacao_gerente = timezone.now()
    else:
        messages.error(request, "Você não tem permissão para aprovar este documento.")
        return redirect("lista_solicitacoes")

    bm.save()
    return redirect("detalhe_bm", pk=bm.pk)


@login_required
def reprovar_bm(request, pk, papel):
    bm = get_object_or_404(DocumentoBM, pk=pk)

    if papel == "coordenador" and request.user.groups.filter(name="Coordenador de Contrato").exists():
        bm.status_coordenador = "reprovado"
    elif papel == "gerente" and request.user.groups.filter(name="Gerente de Contrato").exists():
        bm.status_gerente = "reprovado"
    else:
        messages.error(request, "Você não tem permissão para reprovar este documento.")
        return redirect("lista_solicitacoes")

    bm.save()

    # Se alguém reprovou, exigir novo upload de minuta.
    if bm.reprovado_por_alguem:
        messages.warning(request, "A minuta foi reprovada. Suprimentos deve reenviar um novo BM.")

        suprimentos = User.objects.filter(grupo="suprimento").values_list("email", flat=True)
        contrato_referencia = None
        if bm.solicitacao and bm.solicitacao.contrato:
            contrato_referencia = bm.solicitacao.contrato
        elif bm.solicitacao_contrato and bm.solicitacao_contrato.contrato:
            contrato_referencia = bm.solicitacao_contrato.contrato

        if suprimentos:
            assunto = "Reprovação do Boletim de Medição"
            mensagem = (
                f"O Boletim de Medição {bm.id}, referente ao contrato {contrato_referencia or 'não identificado'} "
                f"foi reprovado.\n\n"
                "Acesse o sistema HIDROGestão para mais informações.\n"
                "https://hidrogestao.pythonanywhere.com/"
            )
            try:
                send_mail(
                    assunto, mensagem,
                    FROM_EMAIL,
                    list(suprimentos),
                    fail_silently=False,
                )
            except Exception as e:
                messages.warning(request, f"Erro ao enviar e-mail para suprimentos: {e}")

    return redirect("detalhe_bm", pk=bm.pk)


@login_required
def cadastrar_evento(request, pk):
    if request.user.grupo not in ["suprimento", "coordenador", "gerente", "gerente_lider", "gerente_contrato", "lider_contrato"]:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")
    solicitacao = get_object_or_404(SolicitacaoProspeccao, pk=pk)

    if request.method == "POST":
        form = EventoPrevisaoForm(request.POST)
        if form.is_valid():
            evento = form.save(commit=False)
            evento.prospeccao = solicitacao
            evento.empresa_terceira = solicitacao.fornecedor_escolhido  # ajuste se o campo for diferente
            evento.save()
            return redirect("detalhes_solicitacao", pk=pk)
    else:
        form = EventoPrevisaoForm()

    return render(request, "gestao_contratos/cadastrar_evento.html", {
        "form": form,
        "solicitacao": solicitacao,
    })


@login_required
def buscar_proxima_data_pagamento(request):
    data_prevista = request.GET.get("data_prevista")

    if not data_prevista:
        return JsonResponse({"data": None})

    proxima_data = CalendarioPagamento.objects.filter(
        data_pagamento__gte=data_prevista
    ).order_by("data_pagamento").first()

    if proxima_data:
        return JsonResponse({
            "data": proxima_data.data_pagamento.strftime("%Y-%m-%d")
        })

    return JsonResponse({"data": None})


@login_required
def duplicar_evento(request, pk):
    evento = get_object_or_404(Evento, pk=pk)

    evento.pk = None
    evento.realizado = False
    evento.valor_pago = None
    evento.data_pagamento = None

    if evento.data_prevista:
        evento.data_prevista = evento.data_prevista + timedelta(days=30)

    if evento.data_prevista_pagamento:
        evento.data_prevista_pagamento = evento.data_prevista_pagamento + timedelta(days=30)

    evento.save()

    messages.success(request, "Evento duplicado com sucesso!")

    return redirect('detalhes_solicitacao', pk=evento.prospeccao.id)



@login_required
def cadastrar_evento_solicitacao(request, pk):
    if request.user.grupo not in ["suprimento", "coordenador", "gerente", "gerente_lider", "gerente_contrato", "lider_contrato"]:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")
    solicitacao = get_object_or_404(SolicitacaoContrato, pk=pk)

    if request.method == "POST":
        form = EventoPrevisaoForm(request.POST)
        if form.is_valid():
            evento = form.save(commit=False)
            evento.solicitacao_contrato = solicitacao
            evento.empresa_terceira = solicitacao.fornecedor_escolhido
            evento.save()
            return redirect("detalhes_solicitacao_contrato", pk=pk)
    else:
        form = EventoPrevisaoForm()

    return render(request, "gestao_contratos/cadastrar_evento_contrato.html", {
        "form": form,
        "solicitacao": solicitacao,
    })

@login_required
def duplicar_evento_solicitacao(request, pk):
    evento = get_object_or_404(Evento, pk=pk)

    evento.pk = None
    evento.realizado = False
    evento.valor_pago = None
    evento.data_pagamento = None

    if evento.data_prevista:
        evento.data_prevista = evento.data_prevista + timedelta(days=30)

    if evento.data_prevista_pagamento:
        evento.data_prevista_pagamento = evento.data_prevista_pagamento + timedelta(days=30)

    evento.save()

    messages.success(request, "Evento duplicado com sucesso!")

    return redirect('detalhes_solicitacao_contrato', pk=evento.solicitacao_contrato.id)


@login_required
def cadastrar_evento_contrato(request, pk):
    if request.user.grupo not in ["suprimento", "coordenador", "gerente", "gerente_contrato", "lider_contrato"]:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    contrato = get_object_or_404(ContratoTerceiros, pk=pk)
    #solicitacao = contrato.prospeccao

    if request.method == "POST":
        form = EventoPrevisaoForm(request.POST)
        if form.is_valid():
            evento = form.save(commit=False)
            #evento.prospeccao = solicitacao
            evento.contrato_terceiro = contrato
            evento.empresa_terceira = contrato.empresa_terceira
            evento.save()
            return redirect("contrato_fornecedor_detalhe", pk=contrato.pk)
    else:
        form = EventoPrevisaoForm()

    return render(request, "gestao_contratos/cadastrar_evento_contrato.html", {
        "form": form,
        "solicitacao": contrato, #fiz isso para utilizar o mesmo html
        #"solicitacao": solicitacao,
    })


@login_required
def duplicar_evento_contrato(request, pk):
    evento = get_object_or_404(Evento, pk=pk)

    evento.pk = None
    evento.realizado = False
    evento.valor_pago = None
    evento.data_pagamento = None

    if evento.data_prevista:
        evento.data_prevista = evento.data_prevista + timedelta(days=30)

    if evento.data_prevista_pagamento:
        evento.data_prevista_pagamento = evento.data_prevista_pagamento + timedelta(days=30)

    evento.save()

    messages.success(request, "Evento duplicado com sucesso!")

    return redirect('contrato_fornecedor_detalhe', pk=evento.contrato_terceiro.id)


@login_required
def editar_evento(request, pk):
    if request.user.grupo not in ["suprimento", "coordenador", "gerente", "gerente_lider", "gerente_contrato", "lider_contrato"]:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    evento = get_object_or_404(Evento, pk=pk)
    if request.method == "POST":
        form = EventoPrevisaoForm(request.POST, request.FILES, instance=evento)
        if form.is_valid():
            form.save()
            return redirect("detalhes_solicitacao", pk=evento.prospeccao.id)
    else:
        form = EventoPrevisaoForm(instance=evento)
    return render(request, "gestao_contratos/editar_evento.html", {"form": form, "evento": evento})


@login_required
def editar_evento_contrato(request, pk):
    if request.user.grupo not in ["suprimento", "coordenador", "gerente", "gerente_lider", "gerente_contrato", "lider_contrato"]:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    evento = get_object_or_404(Evento, pk=pk)
    if request.method == "POST":
        form = EventoPrevisaoForm(request.POST, request.FILES, instance=evento)
        if form.is_valid():
            form.save()
            return redirect("contrato_fornecedor_detalhe", pk=evento.contrato_terceiro.pk)
    else:
        form = EventoPrevisaoForm(instance=evento)
    return render(request, "gestao_contratos/editar_evento_contrato.html", {"form": form, "evento": evento})


@login_required
def excluir_evento(request, pk):
    if request.user.grupo not in ["suprimento", "gerente_lider", "lider_contrato", "gerente_contrato"]:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    evento = get_object_or_404(Evento, pk=pk)

    if evento.prospeccao:
        pk_origem = evento.prospeccao.id
    elif evento.solicitacao_contrato:
        pk_origem = evento.solicitacao_contrato.id
    else:
        pk_origem = None

    if request.method == "POST":
        evento.delete()

        if evento.prospeccao:
            return redirect("detalhes_solicitacao", pk=pk_origem)
        elif evento.solicitacao_contrato:
            return redirect("detalhes_solicitacao_contrato", pk=pk_origem)

        else:
            return redirect("home")

    return render(request, "gestao_contratos/excluir_evento.html", {
        "evento": evento,
        "pk_origem": pk_origem
    })


@login_required
def excluir_evento_contrato(request, pk):
    if request.user.grupo not in ["suprimento", "lider_contrato", "gerente_contrato", "gerente_lider"]:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    evento = get_object_or_404(Evento, pk=pk)
    if request.method == "POST":
        contrato_id = evento.contrato_terceiro.pk
        evento.delete()
        return redirect("contrato_fornecedor_detalhe", pk=contrato_id)
    return render(request, "gestao_contratos/excluir_evento_contrato.html", {"evento": evento})


@login_required
def registrar_entrega(request, pk):
    evento = get_object_or_404(Evento, pk=pk)
    contrato = evento.contrato_terceiro
    can_manage_delivery = can_user_manage_event_delivery(request.user, contrato)
    boletins = evento.boletins_medicao.all()
    notas_fiscais = evento.nota_fiscal.all()

    if request.user.grupo != "suprimento" and not can_manage_delivery:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    boletins_detalhados = []
    has_reprovacao_coordenador = False
    has_reprovacao_gerente = False
    tem_reprovacao_diretor = False

    for bm in boletins:
        operational_pending = bm_is_operationally_pending(bm)
        can_approve_pagamento = (
            request.user.grupo == "diretoria"
            and bm_has_operational_approval(bm)
        )

        if bm_has_operational_approval(bm):
            row_class = "table-success"
        elif bm.status_coordenador == "reprovado" or bm.status_gerente == "reprovado":
            row_class = "table-danger"
        else:
            row_class = "table-warning"

        if bm.status_coordenador == "reprovado":
            has_reprovacao_coordenador = True
        if bm.status_gerente == "reprovado":
            has_reprovacao_gerente = True
        if bm.aprovacao_pagamento == "reprovado":
            tem_reprovacao_diretor = True

        boletins_detalhados.append({
            "bm": bm,
            "row_class": row_class,
            "operational_pending": operational_pending,
            "can_approve_pagamento": can_approve_pagamento,
        })

    if request.method == "POST":
        if request.user.grupo != "suprimento" and not can_manage_delivery:
            messages.error(request, "Você não tem permissão para registrar a entrega deste evento.")
            return redirect("contrato_fornecedor_detalhe", pk=contrato.pk)

        form = EventoEntregaForm(request.POST, request.FILES, instance=evento)

        if form.is_valid():
            form.save()
            messages.success(request, "Entrega Registrada com sucesso")
            return redirect('contrato_fornecedor_detalhe', pk=contrato.pk)
        else:
            messages.error(request, "Erro ao registrar a entrega!")
    else:
        form = EventoEntregaForm(instance=evento)

    return render(request, "eventos/registrar_entrega.html", {
        "form": form,
        "evento": evento,
        "contrato": contrato,
        "can_manage_delivery": can_manage_delivery,
        "can_approve_bm_as_lider": request.user.grupo == "lider_contrato" and request.user == contrato.lider_contrato or request.user.grupo == "gerente_lider" and user_shares_center_with_coordinator(request.user, contrato.coordenador),
        "can_approve_bm_as_gerente": request.user.grupo == "gerente_contrato",
        "boletins_detalhados": boletins_detalhados,
        "has_reprovacao_coordenador": has_reprovacao_coordenador,
        "has_reprovacao_gerente": has_reprovacao_gerente,
        "tem_reprovacao_diretor": tem_reprovacao_diretor,
        "notas_fiscais": notas_fiscais,
    })


@login_required
def avaliar_bm(request, bm_id):
    bm = get_object_or_404(BM, id=bm_id)
    usuario = request.user
    contrato = bm.contrato
    operational_pending = bm_is_operationally_pending(bm)

    if usuario.grupo not in ["lider_contrato", "gerente_lider", "gerente_contrato", "diretoria"]:
        messages.error(request, "⚠ Você não tem permissão para isso.")
        return redirect("home")

    acao = request.POST.get("acao")
    justificativa = request.POST.get("justificativa", "").strip()

    if acao not in ["aprovar", "reprovar", "aprovar_pagamento", "reprovar_pagamento"]:
        return JsonResponse({
            "success": False,
            "error": "Ação inválida.",
        }, status=400)

    if usuario.grupo in ["lider_contrato", "gerente_lider"]:
        if not can_user_manage_event_delivery(usuario, contrato):
            return JsonResponse({
                "success": False,
                "error": "Você não tem permissão para avaliar este BM."
            }, status=403)

        if acao not in ["aprovar", "reprovar"]:
            return JsonResponse({
                "success": False,
                "error": "Ação inválida para o perfil de liderança."
            }, status=400)

        if not operational_pending:
            return JsonResponse({
                "success": False,
                "error": "Este BM já recebeu uma avaliação operacional."
            }, status=400)

        bm.status_coordenador = "aprovado" if acao == "aprovar" else "reprovado"
        bm.data_aprovacao_coordenador = timezone.now()

        if acao == "reprovar":
            bm.justificativa_reprovacao_coordenador = justificativa or "Sem justificativa informada."
        else:
            bm.justificativa_reprovacao_coordenador = None

            usuarios_gerencia = User.objects.filter(grupo="gerente_contrato")
            lista_emails_gerencia = [u.email for u in usuarios_gerencia if u.email]
            if lista_emails_gerencia:
                assunto = f"BM aprovado pela liderança - Contrato {bm.contrato.num_contrato}"
                mensagem = (
                    f"Olá,\n\n"
                    f"O BM {bm.numero_bm} do contrato {bm.contrato.num_contrato} "
                    f"foi aprovado por {usuario.get_full_name() or usuario.username}.\n\n"
                    f"Projeto: {bm.contrato.cod_projeto}\n"
                    f"Evento: {bm.evento.descricao}\n"
                    f"Valor BM: R$ {bm.valor_pago}\n\n"
                    f"Seu registro adicional não é obrigatório, mas esta aprovação foi comunicada para acompanhamento.\n\n"
                    f"Atenciosamente,\n"
                    f"Sistema HIDROGestão"
                )
                send_mail(
                    assunto,
                    mensagem,
                    FROM_EMAIL,
                    lista_emails_gerencia,
                    fail_silently=False,
                )

    elif usuario.grupo == "gerente_contrato":
        if acao not in ["aprovar", "reprovar"]:
            return JsonResponse({
                "success": False,
                "error": "Ação inválida para a gerência de contrato."
            }, status=400)

        if not operational_pending:
            return JsonResponse({
                "success": False,
                "error": "Este BM já recebeu uma avaliação operacional."
            }, status=400)

        bm.status_gerente = "aprovado" if acao == "aprovar" else "reprovado"
        bm.data_aprovacao_gerente = timezone.now()

        if acao == "reprovar":
            bm.justificativa_reprovacao_gerente = justificativa or "Sem justificativa informada."
        else:
            bm.justificativa_reprovacao_gerente = None

    elif usuario.grupo == "diretoria":
        if acao not in ["aprovar_pagamento", "reprovar_pagamento"]:
            return JsonResponse({
                "success": False,
                "error": "Ação inválida para a diretoria."
            }, status=400)

        if not bm_has_operational_approval(bm):
            return JsonResponse({
                "success": False,
                "error": "O BM ainda não recebeu aprovação operacional de líder ou gerente de contrato."
            }, status=400)

        if acao == "aprovar_pagamento":
            bm.aprovacao_pagamento = "aprovado"
            bm.data_aprovacao_diretor = timezone.now()
            bm.justificativa_reprovacao_diretor = None

            usuarios_destino = User.objects.filter(grupo__in=["suprimento", "financeiro"])
            lista_emails = list(dict.fromkeys(u.email for u in usuarios_destino if u.email))

            if lista_emails:
                assunto = f"Pagamento Aprovado pela Diretoria – BM {bm.id}"
                mensagem = (
                    f"Olá, equipe!\n\n"
                    f"A diretoria APROVOU o pagamento do BM abaixo:\n\n"
                    f"Projeto: {bm.contrato.cod_projeto}\n"
                    f"Contrato: {bm.contrato.num_contrato} - {bm.contrato.empresa_terceira}\n"
                    f"Evento: {bm.evento.descricao}\n"
                    f"Valor BM: R$ {bm.valor_pago}\n\n"
                    f"Atenciosamente,\n"
                    f"Sistema HIDROGestão"
                )

                send_mail(
                    assunto,
                    mensagem,
                    FROM_EMAIL,
                    lista_emails,
                    fail_silently=False,
                )

        elif acao == "reprovar_pagamento":
            bm.aprovacao_pagamento = "reprovado"
            bm.data_aprovacao_diretor = timezone.now()
            bm.justificativa_reprovacao_diretor = justificativa or "Sem justificativa informada."

    bm.save()

    if usuario.grupo == "gerente_contrato":
        try:
            status_coord = bm.status_coordenador
            status_ger = bm.status_gerente

            if status_ger != "pendente":
                usuarios_suprimento = User.objects.filter(grupo="suprimento")
                lista_emails = [u.email for u in usuarios_suprimento if u.email]

                if lista_emails:
                    if status_ger == "aprovado":
                        assunto = f"BM aprovado - Contrato {bm.contrato.num_contrato}"
                        mensagem = (
                            f"Olá, equipe de Suprimentos!\n\n"
                            f"O Boletim de Medição foi APROVADO pelo gerente de contrato.\n\n"
                            f"Projeto: {bm.contrato.cod_projeto}\n"
                            f"Contrato: {bm.contrato.num_contrato} - {bm.contrato.empresa_terceira}\n"
                            f"Evento: {bm.evento.descricao}\n"
                            f"Valor BM: R$ {bm.valor_pago}\n\n"
                            f"Atenciosamente,\n"
                            f"Sistema HIDROGestão"
                        )
                    else:
                        assunto = f"BM reprovado - Contrato {bm.contrato.num_contrato}"
                        mensagem = (
                            f"Olá, equipe de Suprimentos!\n\n"
                            f"O Boletim de Medição foi REPROVADO.\n\n"
                            f"Projeto: {bm.contrato.cod_projeto}\n"
                            f"Contrato: {bm.contrato.num_contrato} - {bm.contrato.empresa_terceira}\n"
                            f"Evento: {bm.evento.descricao}\n\n"
                            f"Justificativas:\n"
                            f"- Líder: {bm.justificativa_reprovacao_coordenador or 'Aprovou'}\n"
                            f"- Gerente: {bm.justificativa_reprovacao_gerente or 'Aprovou'}\n\n"
                            f"Atenciosamente,\n"
                            f"Sistema HIDROGestão"
                        )

                    send_mail(
                        assunto,
                        mensagem,
                        FROM_EMAIL,
                        lista_emails,
                        fail_silently=False,
                    )

        except Exception as e:
            print("Erro ao enviar e-mail de avaliação:", e)
            messages.error(request, "⚠ Não foi possível enviar o e-mail para Suprimentos.")

    return JsonResponse({
        "success": True,
        "status_coordenador": bm.status_coordenador,
        "status_gerente": bm.status_gerente,
        "justificativa_reprovacao_coordenador": bm.justificativa_reprovacao_coordenador,
        "justificativa_reprovacao_gerente": bm.justificativa_reprovacao_gerente,
        "aprovacao_pagamento": bm.aprovacao_pagamento,
        "justificativa_reprovacao_diretor": bm.justificativa_reprovacao_diretor,
    })



@login_required
def previsao_pagamentos(request):
    if request.user.grupo not in ["suprimento", "gerente", "gerente_lider", "diretoria", "financeiro", "gerente_contrato"]:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")
    form = FiltroPrevisaoForm(request.GET or None, user=request.user)
    pagamentos = []
    total_previsto = 0
    total_pago = 0
    grafico_html = None
    grafico_barra = None
    grafico_barras = None
    grafico_barras_lider_contrato = None
    grafico_barras_projeto = None
    bms = []
    total_bm_pago = None
    total_bm_previsto = None

    if form.is_valid():
        data_limite = form.cleaned_data["data_limite"]
        data_inicial = form.cleaned_data.get("data_inicial")
        coordenador = form.cleaned_data.get("coordenador")

        hoje = timezone.now().date()
        usuario = request.user
        data_inicio_filtro = data_inicial or hoje

        filtros_base = Q(
            Q(data_prevista_pagamento__range=[data_inicio_filtro, data_limite]) |
            Q(data_pagamento__range=[data_inicio_filtro, data_limite])
        )

        filtros_os = Q(
            Q(data_pagamento__range=[data_inicio_filtro, data_limite])
        )

        eventos = filter_payment_events_for_user(usuario, filtros_base)
        os_queryset = filter_payment_os_for_user(usuario, filtros_os)

        if usuario.grupo not in ["suprimento", "diretoria", "financeiro", "gerente", "gerente_lider", "gerente_contrato"]:
            return redirect('home')

        # Aplica o filtro do coordenador se foi selecionado
        if coordenador:
            eventos = eventos.filter(contrato_terceiro__coordenador=coordenador)
            os_queryset = os_queryset.filter(coordenador=coordenador)

        eventos = eventos.order_by('data_prevista_pagamento', 'data_pagamento')
        os_queryset = os_queryset.order_by('data_pagamento')

        pagamentos_eventos = [
            {
                "tipo": "Evento",
                "projeto": evento.contrato_terceiro.cod_projeto.cod_projeto if evento.contrato_terceiro and evento.contrato_terceiro.cod_projeto else "",
                "coordenador": evento.contrato_terceiro.coordenador.username if evento.contrato_terceiro and evento.contrato_terceiro.coordenador else "",
                "fornecedor": evento.empresa_terceira.nome if evento.empresa_terceira else "",
                "data_prevista": evento.data_prevista_pagamento,
                "valor_previsto": Decimal(evento.valor_previsto or 0),
                "data_pagamento": evento.data_pagamento,
                "valor_pago": Decimal(evento.valor_pago or 0),
            }
            for evento in eventos
        ]

        pagamentos_os = [
            {
                "tipo": "OS",
                "projeto": ordem_servico.cod_projeto.cod_projeto if ordem_servico.cod_projeto else "",
                "coordenador": ordem_servico.coordenador.username if ordem_servico.coordenador else "",
                "fornecedor": ordem_servico.contrato.empresa_terceira.nome if ordem_servico.contrato and ordem_servico.contrato.empresa_terceira else "",
                "data_prevista": ordem_servico.data_pagamento,
                "valor_previsto": Decimal(ordem_servico.valor or 0),
                "data_pagamento": ordem_servico.data_pagamento,
                "valor_pago": Decimal(ordem_servico.valor_pago or 0),
            }
            for ordem_servico in os_queryset
        ]

        pagamentos = pagamentos_eventos + pagamentos_os
        pagamentos.sort(
            key=lambda item: (
                item["data_prevista"] or date.min,
                item["data_pagamento"] or date.min,
                item["tipo"],
                item["projeto"],
            )
        )

        total_previsto_eventos = sum(item['valor_previsto'] for item in pagamentos_eventos)
        total_previsto_os = os_queryset.aggregate(
            total=Coalesce(Sum("valor"), Decimal("0.00"))
        )["total"]

        total_previsto = total_previsto_eventos + total_previsto_os

        total_pago_eventos = sum(item['valor_pago'] for item in pagamentos_eventos)
        total_pago_os = os_queryset.aggregate(
            total=Coalesce(Sum("valor_pago"), Decimal("0.00"))
        )["total"]

        total_pago = total_pago_eventos + total_pago_os


        # ==== GRÁFICO 1: LINHA ACUMULADA (EVENTOS + OS) ====

        from collections import defaultdict

        acumulado_previsto_por_data = defaultdict(Decimal)
        acumulado_pago_por_data = defaultdict(Decimal)

        # EVENTOS
        for e in eventos:
            if e.data_prevista_pagamento:
                acumulado_previsto_por_data[e.data_prevista_pagamento] += e.valor_previsto or 0
            if e.data_pagamento:
                acumulado_pago_por_data[e.data_pagamento] += e.valor_pago or 0

        # OS
        for os in os_queryset:
            if os.data_pagamento:
                acumulado_previsto_por_data[os.data_pagamento] += os.valor or 0
                acumulado_pago_por_data[os.data_pagamento] += os.valor_pago or 0

        datas = sorted(set(acumulado_previsto_por_data.keys()) | set(acumulado_pago_por_data.keys()))

        datas_prevista = []
        acumulado_previsto = []
        datas_pago = []
        acumulado_pago = []

        total_prev = Decimal("0.00")
        total_pg = Decimal("0.00")

        for d in datas:
            total_prev += acumulado_previsto_por_data.get(d, 0)
            total_pg += acumulado_pago_por_data.get(d, 0)

            datas_prevista.append(d)
            acumulado_previsto.append(total_prev)

            datas_pago.append(d)
            acumulado_pago.append(total_pg)

        if datas:
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=datas_prevista,
                y=acumulado_previsto,
                mode='lines+markers',
                name='Previsto',
                line=dict(color='blue', width=3),
                marker=dict(size=6),
                hovertemplate="Data: %{x|%d/%m/%Y}<br>Valor acumulado: R$ %{y:,.2f}<extra></extra>",
            ))
            fig.add_trace(go.Scatter(
                x=datas_pago,
                y=acumulado_pago,
                mode='lines+markers',
                name='Pago',
                line=dict(color='green', width=3),
                marker=dict(size=6),
                hovertemplate="Data: %{x|%d/%m/%Y}<br>Valor acumulado: R$ %{y:,.2f}<extra></extra>",
            ))
            fig.update_layout(
                title='Previsão x Pagamentos Acumulados (Eventos + OS)',
                xaxis=dict(
                    title='Data',
                    tickformat='%d/%m/%Y',
                    hoverformat='%d/%m/%Y',
                ),
                yaxis_title='Valor Acumulado (R$)',
                hovermode='x unified'
            )

            grafico_html = plot(fig, auto_open=False, output_type='div')


        # ==== GRÁFICO 2: POR COORDENADOR ====
        calendario = list(CalendarioPagamento.objects.order_by('data_pagamento')
                          .values_list('data_pagamento', flat=True))

        coordenadores = list(set(
            eventos.values_list('contrato_terceiro__coordenador__username', flat=True)
        ))

        fig_barra = go.Figure()
        for coord in coordenadores:
            y_previstos = []
            data_inicio = None

            for data_fim in calendario:
                filtro_periodo = Q(data_prevista_pagamento__lte=data_fim)
                if data_inicio:
                    filtro_periodo &= Q(data_prevista_pagamento__gt=data_inicio)

                if request.user.grupo == 'gerente_contrato':
                    eventos_previsto = Evento.objects.filter(
                        filtro_periodo,
                        contrato_terceiro__coordenador__username=coord,
                        contrato_terceiro__lider_contrato__grupo='lider_contrato'
                    )
                else:
                    eventos_previsto = Evento.objects.filter(
                        filtro_periodo,
                        contrato_terceiro__coordenador__username=coord
                    )

                total_previsto_periodo = eventos_previsto.aggregate(
                    total=Coalesce(Sum('valor_previsto'), Decimal('0.00'))
                )['total']

                y_previstos.append(total_previsto_periodo)
                data_inicio = data_fim

            fig_barra.add_trace(go.Bar(
                name=f"{coord or 'Sem Coordenador'}",
                x=calendario,
                y=y_previstos
            ))

        fig_barra.update_layout(
            barmode='stack',
            title="Pagamentos Previsto (por Coordenador, conforme calendário de pagamento)",
            xaxis_title="Data do Calendário",
            yaxis_title="Valor Previsto (R$)",
            template="plotly_white",
            height=500,
            legend_title="Coordenador"
        )

        grafico_barras = plot(fig_barra, output_type='div')

        # ==== GRÁFICO 2.1: POR LÍDER DE CONTRATO ====
        lideres_contrato = list(set(
            eventos.values_list('contrato_terceiro__lider_contrato__username', flat=True)
        ))

        fig_barra_lider = go.Figure()
        for lider in lideres_contrato:
            y_previstos = []
            data_inicio = None

            for data_fim in calendario:
                filtro_periodo = Q(data_prevista_pagamento__lte=data_fim)
                if data_inicio:
                    filtro_periodo &= Q(data_prevista_pagamento__gt=data_inicio)

                eventos_previsto = eventos.filter(
                    filtro_periodo,
                    contrato_terceiro__lider_contrato__username=lider,
                )

                total_previsto_periodo = eventos_previsto.aggregate(
                    total=Coalesce(Sum('valor_previsto'), Decimal('0.00'))
                )['total']

                y_previstos.append(total_previsto_periodo)
                data_inicio = data_fim

            fig_barra_lider.add_trace(go.Bar(
                name=f"{lider or 'Sem Líder de Contrato'}",
                x=calendario,
                y=y_previstos
            ))

        fig_barra_lider.update_layout(
            barmode='stack',
            title="Pagamentos Previsto (por Líder de Contrato, conforme calendário de pagamento)",
            xaxis_title="Data do Calendário",
            yaxis_title="Valor Previsto (R$)",
            template="plotly_white",
            height=500,
            legend_title="Líder de Contrato"
        )

        grafico_barras_lider_contrato = plot(fig_barra_lider, output_type='div')

        # ==== GRÁFICO 2.2: POR PROJETO ====
        calendario = list(CalendarioPagamento.objects.order_by('data_pagamento')
                          .values_list('data_pagamento', flat=True))

        # filtra os projetos conforme coordenador (ou todos se não houver filtro)
        if coordenador:
            if request.user.grupo == 'gerente_contrato':
                projetos = list(Evento.objects.filter(
                    contrato_terceiro__coordenador=coordenador, contrato_terceiro__isnull=False,
                    contrato_terceiro__lider_contrato__grupo='lider_contrato'
                ).values_list('contrato_terceiro__cod_projeto__cod_projeto', flat=True))
            else:
                projetos = list(Evento.objects.filter(
                    contrato_terceiro__coordenador=coordenador, contrato_terceiro__isnull=False,
                ).values_list('contrato_terceiro__cod_projeto__cod_projeto', flat=True))
        else:
            if request.user.grupo == 'gerente_contrato':
                projetos = list(Evento.objects.filter(
                    contrato_terceiro__lider_contrato__grupo='lider_contrato', contrato_terceiro__isnull=False,
                ).values_list('contrato_terceiro__cod_projeto__cod_projeto', flat=True))
            else:
                projetos = list(Evento.objects.values_list(
                    'contrato_terceiro__cod_projeto__cod_projeto', flat=True))

        projetos = list(set(projetos))

        fig_barra_proj = go.Figure()
        for proj in projetos:
            y_previstos = []
            data_inicio = None

            for data_fim in calendario:
                filtro_periodo = Q(data_prevista_pagamento__lte=data_fim)
                if data_inicio:
                    filtro_periodo &= Q(data_prevista_pagamento__gt=data_inicio)

                filtro_base = Q(contrato_terceiro__cod_projeto__cod_projeto=proj)
                if coordenador:
                    filtro_base &= Q(contrato_terceiro__coordenador=coordenador)

                eventos_previsto = Evento.objects.filter(filtro_periodo & filtro_base, contrato_terceiro__isnull=False)

                total_previsto_periodo = eventos_previsto.aggregate(
                    total=Coalesce(Sum('valor_previsto'), Decimal('0.00'))
                )['total']

                y_previstos.append(total_previsto_periodo)
                data_inicio = data_fim

            fig_barra_proj.add_trace(go.Bar(
                name=f"{proj or 'Sem Projeto'}",
                x=calendario,
                y=y_previstos
            ))

        fig_barra_proj.update_layout(
            barmode='stack',
            title="Pagamentos Previsto (por Projeto, conforme calendário de pagamento)",
            xaxis_title="Data do Calendário",
            yaxis_title="Valor Previsto (R$)",
            template="plotly_white",
            height=500,
            legend_title="Projeto"
        )

        grafico_barras_projeto = plot(fig_barra_proj, output_type='div')

        # ==== GRÁFICO 3: PREVISTO x PAGO (EVENTOS + OS) ====

        calendario = list(
            CalendarioPagamento.objects.order_by("data_pagamento")
            .values_list("data_pagamento", flat=True)
        )

        y_previstos = []
        y_pagos = []
        data_inicio = None

        for data_fim in calendario:
            filtro_prev_evento = Q(data_prevista_pagamento__lte=data_fim)
            filtro_pago_evento = Q(data_pagamento__lte=data_fim)
            filtro_os = Q(data_pagamento__lte=data_fim)

            if data_inicio:
                filtro_prev_evento &= Q(data_prevista_pagamento__gt=data_inicio)
                filtro_pago_evento &= Q(data_pagamento__gt=data_inicio)
                filtro_os &= Q(data_pagamento__gt=data_inicio)

            if coordenador:
                filtro_prev_evento &= Q(contrato_terceiro__coordenador=coordenador)
                filtro_pago_evento &= Q(contrato_terceiro__coordenador=coordenador)
                filtro_os &= Q(coordenador=coordenador)

            if request.user.grupo == 'gerente_contrato':
                filtro_prev_evento &= Q(contrato_terceiro__lider_contrato__grupo='lider_contrato')
                filtro_pago_evento &= Q(contrato_terceiro__lider_contrato__grupo='lider_contrato')
                filtro_os &= Q(lider_contrato__grupo='lider_contrato')

            total_prev_eventos = Evento.objects.filter(filtro_prev_evento, contrato_terceiro__isnull=False,).aggregate(
                total=Coalesce(Sum("valor_previsto"), Decimal("0.00"))
            )["total"]

            total_pago_eventos = Evento.objects.filter(filtro_pago_evento, contrato_terceiro__isnull=False,).aggregate(
                total=Coalesce(Sum("valor_pago"), Decimal("0.00"))
            )["total"]

            total_prev_os = os_queryset.filter(filtro_os).aggregate(
                total=Coalesce(Sum("valor"), Decimal("0.00"))
            )["total"]

            total_pago_os = os_queryset.filter(filtro_os).aggregate(
                total=Coalesce(Sum("valor_pago"), Decimal("0.00"))
            )["total"]

            y_previstos.append(total_prev_eventos + total_prev_os)
            y_pagos.append(total_pago_eventos + total_pago_os)

            data_inicio = data_fim

        fig_barra_final = go.Figure(data=[
            go.Bar(name="Previsto", x=calendario, y=y_previstos, marker_color="orange"),
            go.Bar(name="Pago", x=calendario, y=y_pagos, marker_color="green"),
        ])

        fig_barra_final.update_layout(
            barmode="group",
            title="Pagamentos Previsto x Pago (Eventos + OS)",
            xaxis_title="Data",
            yaxis_title="Valor (R$)",
            template="plotly_white",
            height=500,
        )

        grafico_barra = plot(fig_barra_final, output_type="div")


        # ==== TABELA DE BMs ====
        # === FILTRO DE BM (Pagamento OU PerÃ­odo de Medição) ===
        filtro_bm = request.GET.get("filtro_bm", "pagamento")

        if filtro_bm == "medicao":
            # filtra pelo perÃ­odo de medição definido no form
            bms = BM.objects.filter(
                data_inicial_medicao__date__gte=data_inicio_filtro,
                data_final_medicao__date__lte=data_limite
            )
        else:
            # filtro padrão: data da aprovação operacional
            bms = BM.objects.filter(
                Q(data_aprovacao_coordenador__date__range=[data_inicio_filtro, data_limite])
                | Q(data_aprovacao_gerente__date__range=[data_inicio_filtro, data_limite])
            ).filter(
                bm_operational_approval_query()
            )

        bms = filter_payment_bms_for_user(usuario, bms)

        bms = bms.select_related(
            'contrato__cod_projeto',
            'contrato__coordenador',
            'evento'
        ).order_by('-data_pagamento')

        # Se o coordenador foi selecionado, filtra também os BMs
        if coordenador:
            bms = bms.filter(contrato__coordenador=coordenador)

        # Determina status de aprovação (para exibir na tabela)
        for bm in bms:
            # Define o status geral
            if bm_has_operational_approval(bm):
                bm.status_geral = 'Aprovado'
                bm.falta_aprovar = '-'
            elif bm.status_coordenador == 'reprovado' or bm.status_gerente == 'reprovado':
                bm.status_geral = 'Reprovado'
                bm.falta_aprovar = '-'
            else:
                bm.status_geral = 'Aguardando aprovação'
                faltando = []
                if bm.status_coordenador != 'aprovado':
                    faltando.append('Líder / Gerente-Líder')
                if bm.status_gerente != 'aprovado':
                    faltando.append('Gerente de Contrato')
                bm.falta_aprovar = ', '.join(faltando)

        bm_approval_audit_map = build_bm_approval_audit_map(bms)

        for bm in bms:
            approval_info = bm_approval_audit_map.get(bm.id, {})
            bm.aprovacao_coordenador_audit = approval_info.get("coordenador")
            bm.aprovacao_gerente_audit = approval_info.get("gerente")

        total_bm_pago = sum(bm.valor_pago or 0 for bm in bms)
        total_bm_previsto = sum(
            (bm.evento.valor_previsto if bm.evento and bm.evento.valor_previsto else 0)
            for bm in bms
        )

    return render(request, "gestao_contratos/previsao_pagamentos.html", {
        "form": form,
        "pagamentos": pagamentos,
        "total_previsto": total_previsto,
        "total_pago": total_pago,
        "grafico_html": grafico_html,
        "grafico_barras": grafico_barras,
        "grafico_barras_lider_contrato": grafico_barras_lider_contrato,
        "grafico_barras_projeto": grafico_barras_projeto,
        "grafico_barra": grafico_barra,
        "bms": bms,
        "total_bm_pago": total_bm_pago,
        "total_bm_previsto": total_bm_previsto,
    })


@login_required
def download_bms_aprovados(request):
    if request.user.grupo not in ["suprimento", "diretoria", "gerente", "gerente_lider", "financeiro", "gerente_contrato"]:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    data_inicial_str = request.GET.get("data_inicial")
    data_limite_str = request.GET.get("data_limite")
    filtro_bm = request.GET.get("filtro_bm", "pagamento")  # pagto ou medicao
    coordenador = request.GET.get("coordenador")

    if not data_limite_str:
        messages.error(request, "Data limite é obrigatÃ³ria para o download.")
        return redirect("previsao_pagamentos")


    try:
        data_inicial = datetime.strptime(data_inicial_str, "%Y-%m-%d").date() if data_inicial_str else timezone.now().date()
        data_limite = datetime.strptime(data_limite_str, "%Y-%m-%d").date()
    except ValueError:
        messages.error(request, "Datas inválidas no formato. Use YYYY-MM-DD.")
        return redirect("previsao_pagamentos")

    bms_aprovados = BM.objects.filter(
        bm_operational_approval_query()
    )

    # === FILTRAR POR PAGAMENTO OU MEDIÃ‡ÃƒO ===
    if filtro_bm == "medicao":
        bms_aprovados = bms_aprovados.filter(
            data_inicial_medicao__date__gte=data_inicial,
            data_final_medicao__date__lte=data_limite
        )
    else:
        bms_aprovados = bms_aprovados.filter(
            Q(data_aprovacao_coordenador__date__range=[data_inicial, data_limite])
            | Q(data_aprovacao_gerente__date__range=[data_inicial, data_limite])
        )

    bms_aprovados = filter_payment_bms_for_user(request.user, bms_aprovados)

    # === FILTRAR POR COORDENADOR, SE APLICADO ===
    if coordenador:
        bms_aprovados = bms_aprovados.filter(contrato__coordenador=coordenador)

    # === APENAS COM ARQUIVO ===
    bms_aprovados = bms_aprovados.exclude(arquivo_bm='')

    if not bms_aprovados.exists():
        messages.error(request, "Nenhum BM aprovado encontrado para download nesse perÃ­odo.")
        return redirect("previsao_pagamentos")

    # === CRIAÃ‡ÃƒO DO ZIP ===
    buffer_zip = BytesIO()
    with zipfile.ZipFile(buffer_zip, "w", zipfile.ZIP_DEFLATED) as zipf:
        for bm in bms_aprovados:
            if not bm.arquivo_bm:
                continue

            try:
                with bm.arquivo_bm.open('rb') as f:
                    conteudo = f.read()

                nome_projeto = (
                    bm.contrato.cod_projeto.cod_projeto
                    if bm.contrato and bm.contrato.cod_projeto
                    else "SemProjeto"
                )
                nome_original = bm.arquivo_bm.name.split('/')[-1]
                nome_arquivo_zip = f"{nome_projeto}_BM{bm.numero_bm}_{nome_original}"

                zipf.writestr(nome_arquivo_zip, conteudo)

            except Exception as e:
                print(f"⚠ Erro ao adicionar BM {bm.id} ao ZIP: {e}")

    buffer_zip.seek(0)

    nome_arquivo_zip = f"BMs_Aprovados_{data_inicial.strftime('%Y-%m-%d')}_a_{data_limite.strftime('%Y-%m-%d')}.zip"

    response = HttpResponse(buffer_zip.getvalue(), content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename=\"{nome_arquivo_zip}\"'
    return response



@login_required
def exportar_previsao_pagamentos_excel(request):
    if request.user.grupo not in ["suprimento", "diretoria", "gerente", "gerente_lider", "financeiro", "gerente_contrato"]:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    form = FiltroPrevisaoForm(request.GET or None, user=request.user)

    if not form.is_valid():
        messages.error(request, "Preencha os filtros corretamente antes de exportar.")
        return redirect('previsao_pagamentos')

    data_limite = form.cleaned_data["data_limite"]
    data_inicial = form.cleaned_data.get("data_inicial")
    coordenador = form.cleaned_data.get("coordenador")

    hoje = timezone.now().date()
    usuario = request.user
    data_inicio_filtro = data_inicial or hoje

    filtros_base = Q(
        Q(data_prevista_pagamento__range=[data_inicio_filtro, data_limite]) |
        Q(data_pagamento__range=[data_inicio_filtro, data_limite])
    )

    eventos = filter_payment_events_for_user(usuario, filtros_base)
    os_queryset = filter_payment_os_for_user(
        usuario,
        Q(data_pagamento__range=[data_inicio_filtro, data_limite])
    )

    if usuario.grupo not in ["suprimento", "diretoria", "financeiro", "gerente", "gerente_lider", "gerente_contrato"]:
        return redirect('home')

    if coordenador:
        eventos = eventos.filter(contrato_terceiro__coordenador=coordenador, contrato_terceiro__isnull=False)
        os_queryset = os_queryset.filter(coordenador=coordenador)

    eventos = eventos.order_by('data_prevista_pagamento', 'data_pagamento')
    os_queryset = os_queryset.order_by('data_pagamento')

    pagamentos_exportacao = []
    for evento in eventos:
        pagamentos_exportacao.append({
            "tipo": "Evento",
            "data_prevista": evento.data_prevista_pagamento,
            "projeto": evento.contrato_terceiro.cod_projeto.cod_projeto if evento.contrato_terceiro and evento.contrato_terceiro.cod_projeto else "",
            "fornecedor": evento.empresa_terceira.nome if evento.empresa_terceira else "",
            "coordenador": evento.contrato_terceiro.coordenador.username if evento.contrato_terceiro and evento.contrato_terceiro.coordenador else "",
            "valor_previsto": Decimal(evento.valor_previsto or 0),
            "data_pagamento": evento.data_pagamento,
            "valor_pago": Decimal(evento.valor_pago or 0),
        })

    for ordem_servico in os_queryset:
        pagamentos_exportacao.append({
            "tipo": "OS",
            "data_prevista": ordem_servico.data_pagamento,
            "projeto": ordem_servico.cod_projeto.cod_projeto if ordem_servico.cod_projeto else "",
            "fornecedor": ordem_servico.contrato.empresa_terceira.nome if ordem_servico.contrato and ordem_servico.contrato.empresa_terceira else "",
            "coordenador": ordem_servico.coordenador.username if ordem_servico.coordenador else "",
            "valor_previsto": Decimal(ordem_servico.valor or 0),
            "data_pagamento": ordem_servico.data_pagamento,
            "valor_pago": Decimal(ordem_servico.valor_pago or 0),
        })

    pagamentos_exportacao.sort(
        key=lambda item: (
            item["data_prevista"] or date.min,
            item["data_pagamento"] or date.min,
            item["tipo"],
            item["projeto"],
        )
    )

    # === Criar workbook ===
    wb = Workbook()
    ws = wb.active
    ws.title = "Previsão de Pagamentos"

    # Planilha 1 - Dados completos
    headers = [
        "Tipo", "Data Prevista", "Projeto", "Fornecedor", "Coordenador",
        "Valor Previsto", "Data Pagamento", "Valor Pago"
    ]
    ws.append(headers)

    for item in pagamentos_exportacao:
        ws.append([
            item["tipo"],
            item["data_prevista"].strftime("%d/%m/%Y") if item["data_prevista"] else '',
            item["projeto"],
            item["fornecedor"],
            item["coordenador"],
            float(item["valor_previsto"]),
            item["data_pagamento"].strftime("%d/%m/%Y") if item["data_pagamento"] else '',
            float(item["valor_pago"])
        ])

    # === Planilha 2 - Por Coordenador (com datas) ===
    ws2 = wb.create_sheet("Por Coordenador")
    coordenadores = sorted({item["coordenador"] for item in pagamentos_exportacao if item["coordenador"]})
    ws2.append(["Data Prevista"] + coordenadores)

    datas = sorted({item["data_prevista"] for item in pagamentos_exportacao if item["data_prevista"]})

    for data in datas:
        linha = [data.strftime("%d/%m/%Y") if data else ""]
        for coord in coordenadores:
            total = sum(
                item["valor_previsto"]
                for item in pagamentos_exportacao
                if item["data_prevista"] == data and item["coordenador"] == coord
            )
            linha.append(float(total or Decimal("0")))
        ws2.append(linha)

    # Gráfico
    chart1 = BarChart()
    chart1.title = "Valores Previsto por Coordenador (por Data)"
    data = Reference(ws2, min_col=2, max_col=len(coordenadores) + 1, min_row=1, max_row=len(datas) + 1)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=len(datas) + 1)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.x_axis.title = "Data Prevista"
    chart1.y_axis.title = "Valor (R$)"
    chart1.height = 10
    chart1.width = 20
    ws2.add_chart(chart1, "H2")

    # === Planilha 3 - Por Projeto (com datas) ===
    ws3 = wb.create_sheet("Por Projeto")
    projetos = sorted({item["projeto"] for item in pagamentos_exportacao if item["projeto"]})
    ws3.append(["Data Prevista"] + projetos)

    for data in datas:
        linha = [data.strftime("%d/%m/%Y") if data else ""]
        for proj in projetos:
            total = sum(
                item["valor_previsto"]
                for item in pagamentos_exportacao
                if item["data_prevista"] == data and item["projeto"] == proj
            )
            linha.append(float(total or Decimal("0")))
        ws3.append(linha)

    chart2 = BarChart()
    chart2.title = "Valores Previsto por Projeto (por Data)"
    data2 = Reference(ws3, min_col=2, max_col=len(projetos) + 1, min_row=1, max_row=len(datas) + 1)
    cats2 = Reference(ws3, min_col=1, min_row=2, max_row=len(datas) + 1)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.x_axis.title = "Data Prevista"
    chart2.y_axis.title = "Valor (R$)"
    chart2.height = 10
    chart2.width = 20
    ws3.add_chart(chart2, "H2")

    # === Planilha 4 - Acumulado (com datas) ===
    ws4 = wb.create_sheet("Acumulado")
    ws4.append(["Data Prevista", "Acumulado (R$)"])

    acumulado = 0
    for data in datas:
        total = sum(
            item["valor_previsto"]
            for item in pagamentos_exportacao
            if item["data_prevista"] == data
        ) or Decimal("0")
        acumulado += total
        ws4.append([data.strftime("%d/%m/%Y"), float(acumulado)])

    chart3 = LineChart()
    chart3.title = "Acumulado por Data Prevista"
    data3 = Reference(ws4, min_col=2, min_row=1, max_row=len(datas) + 1)
    cats3 = Reference(ws4, min_col=1, min_row=2, max_row=len(datas) + 1)
    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(cats3)
    chart3.x_axis.title = "Data Prevista"
    chart3.y_axis.title = "Valor Acumulado (R$)"
    chart3.height = 10
    chart3.width = 20
    for serie in chart3.series:
        serie.marker.symbol = "circle"
        serie.marker.size = 6
        serie.graphicalProperties.line.width = 20000

    ws4.add_chart(chart3, "E2")

    # === Exportar Excel ===
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="previsao_pagamentos_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    )
    wb.save(response)
    return response

@login_required
def ranking_fornecedores(request):
    if request.user.grupo not in ["suprimento", "diretoria", "gerente", "gerente_lider", "financeiro", "gerente_contrato"]:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    user = request.user
    contratos_ativos = ContratoTerceiros.objects.filter(status='ativo')

    if user.grupo in ['gerente', 'gerente_lider']:
        # Pega os centros de trabalho do gerente
        centros_gerente = user.centros.all()

        # Filtra contratos cujos coordenadores tenham centros em comum
        coordenadores_mesmos_centros = User.objects.filter(
            grupo='coordenador',
            centros__in=centros_gerente
        ).distinct()

        contratos_ativos = contratos_ativos.filter(
            Q(coordenador__in=coordenadores_mesmos_centros) |
            Q(coordenadores__in=coordenadores_mesmos_centros)
        ).distinct()

    elif user.grupo in ['suprimento', 'diretoria', 'financeiro']:
        # vê todos os contratos ativos
        pass

    elif user.grupo == 'gerente_contrato':
        contratos_ativos = contratos_ativos.filter(lider_contrato=user)

    else:
        # Outros grupos não visualizam nada
        contratos_ativos = ContratoTerceiros.objects.none()

    dados = []
    fornecedores = EmpresaTerceira.objects.filter(contratos__in=contratos_ativos).distinct()

    for fornecedor in fornecedores:
        contratos = contratos_ativos.filter(empresa_terceira=fornecedor)
        eventos = Evento.objects.filter(
            empresa_terceira=fornecedor,
            contrato_terceiro__in=contratos,
        )
        projetos_ativos = sorted(
            {
                contrato.cod_projeto.cod_projeto
                for contrato in contratos.select_related('cod_projeto')
                if contrato.cod_projeto
            }
        )

        valor_total_contratos = contratos.aggregate(total=Sum('valor_total'))['total'] or 0
        valor_previsto = eventos.aggregate(total=Sum('valor_previsto'))['total'] or 0
        valor_pago = eventos.aggregate(total=Sum('valor_pago'))['total'] or 0

        percentual_execucao = (valor_pago / valor_previsto * 100) if valor_previsto > 0 else 0

        dados.append({
            'id': fornecedor.id,  # ðŸ‘ˆ Adiciona o ID
            'fornecedor': fornecedor.nome,
            'projetos_ativos': projetos_ativos,
            'valor_total_contratos': valor_total_contratos,
            'valor_previsto': valor_previsto,
            'valor_pago': valor_pago,
            'percentual_execucao': percentual_execucao,
        })

    # Ordenação hierárquica
    dados = sorted(
        dados,
        key=lambda x: (
            x['valor_total_contratos'],
            x['valor_previsto'],
            x['valor_pago']
        ),
        reverse=True
    )

    context = {'dados': dados}
    return render(request, 'gestao_contratos/ranking_fornecedores.html', context)



@login_required
def cadastrar_bm(request, contrato_id, evento_id):
    if request.user.grupo not in ["suprimento"]:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    contrato = get_object_or_404(ContratoTerceiros, id=contrato_id)
    evento = get_object_or_404(Evento, id=evento_id)

    if request.method == "POST":
        form = BMForm(request.POST, request.FILES)
        if form.is_valid():
            bm = form.save(commit=False)
            bm.contrato = contrato
            bm.evento = evento

            if not bm.valor_pago:
                bm.valor_pago = evento.valor_previsto

            bm.save()

            try:
                emails = set()

                # COORDENADOR
                coordenador = contrato.coordenador
                if coordenador and coordenador.email:
                    emails.add(coordenador.email)

                # GERENTES
                if coordenador:
                    coordenador_centros = coordenador.centros.all()

                    gerentes = User.objects.filter(
                        grupo="gerente",
                        centros__in=coordenador_centros
                    ).distinct()

                    for gerente in gerentes:
                        if gerente.email:
                            emails.add(gerente.email)

                # Se existir alguém para enviar
                if emails:
                    assunto = f"BM cadastrado para avaliação - Projeto {contrato.cod_projeto}"
                    mensagem = (
                        f"Olá,\n\n"
                        f"Um novo Boletim de Medição foi cadastrado e está aguardando avaliação.\n\n"
                        f"Projeto: {contrato.cod_projeto}\n"
                        f"Contrato: {contrato.num_contrato} - {contrato.empresa_terceira}\n"
                        f"Evento: {evento.descricao}\n"
                        f"Valor previsto: R$ {evento.valor_previsto}\n"
                        f"Valor informado no BM: R$ {bm.valor_pago}\n\n"
                        f"Acesse o sistema para aprovar ou reprovar o BM.\n\n"
                        f"Atenciosamente,\n"
                        f"Sistema HIDROGestão"
                    )

                    send_mail(
                        assunto,
                        mensagem,
                        FROM_EMAIL,
                        list(emails),
                        fail_silently=False,
                    )

            except Exception as e:
                print("Erro ao enviar e-mail:", e)
                messages.error(request, "⚠ Não foi possível enviar o e-mail aos gestores.")
            return JsonResponse({"success": True})

        else:
            return JsonResponse({
                "success": False,
                "errors": form.errors
            })

    else:
        form = BMForm(initial={
            "valor_pago": evento.valor_previsto,
            "data_pagamento": evento.data_pagamento.strftime('%Y-%m-%d') if evento.data_pagamento else None,
        })

    form = BMForm(initial={
        "valor_pago": evento.valor_previsto,
        "data_pagamento": (
            evento.data_pagamento.strftime('%Y-%m-%d')
            if evento.data_pagamento else None
        ),
    })

    return render(request, "bm/cadastrar_bm_popup.html", {
        "form": form,
        "contrato": contrato,
        "evento": evento,
    })


@login_required
def cadastrar_nf(request, evento_id):
    if request.user.grupo not in ["suprimento", "financeiro"]:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    evento = get_object_or_404(Evento, id=evento_id)
    contrato = get_object_or_404(ContratoTerceiros, id=evento.contrato_terceiro.id)

    def atualizar_evento_com_base_na_nf(nf_obj):
        if not nf_obj.financeiro_autorizou:
            return
        evento.valor_pago = nf_obj.valor_pago
        evento.data_pagamento = nf_obj.data_pagamento
        evento.justificativa = nf_obj.observacao
        evento.save(update_fields=["valor_pago", "data_pagamento", "justificativa"])

    if request.method == "POST":
        form = NFForm(request.POST, request.FILES, evento=evento)
        if form.is_valid():
            nf = form.save(commit=False)
            nf.contrato = contrato
            nf.evento = evento
            nf.save()
            atualizar_evento_com_base_na_nf(nf)

            # ==============================
            # ENVIO DE E-MAIL AUTOMÁTICO
            # ==============================
            if request.user.grupo == "financeiro":

                # E-mails do grupo suprimento
                emails_suprimento = list(User.objects.filter(grupo='suprimento', is_active=True).values_list("email", flat=True))

                # Coordenador do contrato
                coordenador = contrato.coordenador
                email_coordenador = [coordenador.email] if coordenador and coordenador.email else []

                assunto = f"NF cadastrada para o evento #{evento.id}"
                mensagem = (
                    f"A nota fiscal do evento '{evento.descricao or 'Sem descrição'}'\n"
                    f"do contrato: {contrato}\n\n"
                    f"Foi cadastrada por: {request.user.get_full_name() or request.user.username} (Financeiro).\n\n"
                    f"Data de pagamento: {nf.data_pagamento}\n"
                    f"Valor: R$ {nf.valor_pago}"
                )

                # Enviar e-mail para suprimento
                try:
                    send_mail(
                        assunto, mensagem,
                        FROM_EMAIL,
                        emails_suprimento,
                        fail_silently=False,
                    )
                except Exception as e:
                    messages.warning(request, f"Erro ao enviar e-mail para suprimentos: {e}")

                try:
                    send_mail(
                        assunto, mensagem,
                        FROM_EMAIL,
                        email_coordenador,
                        fail_silently=False,
                    )
                except Exception as e:
                    messages.warning(request, f"Erro ao enviar e-mail para o coordenador: {e}")

            return JsonResponse({"success": True})
        else:
            return JsonResponse({"success": False, "errors": form.errors})

    else:
        form = NFForm(
            evento=evento,
            initial={
                "valor_pago": evento.valor_previsto,
                "data_pagamento": evento.data_pagamento or None
            }
        )

    return render(request, "nf/cadastrar_nf_popup.html", {
        "form": form,
        "contrato": contrato,
        "evento": evento,
    })


@login_required
def editar_bm(request, bm_id):
    bm = get_object_or_404(BM, id=bm_id)

    if request.user.grupo != "suprimento":
        return JsonResponse({"success": False, "error": "Sem permissão."}, status=403)

    if request.method == "POST":
        form = BMForm(request.POST, request.FILES, instance=bm)
        if form.is_valid():
            form.save()
            return JsonResponse({"success": True})
        else:
            return JsonResponse({
                "success": False,
                "errors": form.errors
            })

    else:
        form = BMForm(instance=bm)

    return render(request, "bm/editar_bm.html", {"form": form, "bm": bm})


@login_required
def editar_nf(request, nf_id):
    if request.user.grupo not in ["suprimento", "financeiro"]:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    nf = get_object_or_404(NF, id=nf_id)
    evento = nf.evento

    if request.method == "POST":
        form = NFForm(request.POST, request.FILES, evento=evento, instance=nf)
        if form.is_valid():
            nf = form.save()
            if nf.financeiro_autorizou:
                evento.valor_pago = nf.valor_pago
                evento.data_pagamento = nf.data_pagamento
                evento.justificativa = nf.observacao
                evento.save(update_fields=["valor_pago", "data_pagamento", "justificativa"])
            return JsonResponse({"success": True})
        return JsonResponse({"success": False, "errors": form.errors})

    else:
        form = NFForm(
            evento=evento,
            instance=nf,
            initial={
                "valor_pago": nf.valor_pago,
                "data_pagamento": nf.data_pagamento,
            }
        )

    return render(request, "nf/editar_nf_popup.html", {
        "form": form,
        "nf": nf,
        "evento": evento,
        "contrato": nf.contrato,
    })




@login_required
def deletar_bm(request, pk):
    bm = get_object_or_404(BM, pk=pk)

    if request.user.grupo != "suprimento":
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    evento_id = bm.evento.id
    bm.delete()

    messages.success(request, "BM apagado com sucesso!")
    return redirect("registrar_entrega", pk=evento_id)


@login_required
def deletar_nf(request, nf_id):
    if request.user.grupo not in ["suprimento", "financeiro"]:
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    nf = get_object_or_404(NF, id=nf_id)
    evento_id = nf.evento.id


    nf.delete()
    messages.success(request, "NF apagada com sucesso!")
    if request.user.grupo == 'financeiro':
        return redirect("detalhes_entrega", evento_id=evento_id)
    return redirect("registrar_entrega", pk=evento_id)


@login_required
def detalhes_entrega(request, evento_id):
    evento = get_object_or_404(Evento, id=evento_id)
    contrato = evento.contrato_terceiro

    if request.user.grupo not in ["suprimento", "financeiro", "diretoria"] and not can_user_manage_event_delivery(request.user, contrato):
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    bms = BM.objects.filter(evento=evento).order_by('-data_pagamento')
    notas_fiscais = evento.nota_fiscal.all()

    for bm in bms:
        bm.operational_pending = bm_is_operationally_pending(bm)
        bm.can_approve_pagamento = (
            request.user.grupo == "diretoria"
            and bm_has_operational_approval(bm)
        )

    fornecedor = None
    if hasattr(evento, 'contrato_terceiro') and evento.contrato_terceiro.empresa_terceira:
        fornecedor = evento.contrato_terceiro.empresa_terceira

    tem_reprovacao_coordenador = bms.filter(status_coordenador='reprovado').exists()
    tem_reprovacao_gerente = bms.filter(status_gerente='reprovado').exists()
    tem_reprovacao_diretor = bms.filter(aprovacao_pagamento='reprovado').exists()

    return render(request, 'contratos/detalhes_entrega.html', {
        'evento': evento,
        'fornecedor': fornecedor,
        'can_approve_bm_as_lider': request.user.grupo == 'lider_contrato' and request.user == contrato.lider_contrato or request.user.grupo == 'gerente_lider' and user_shares_center_with_coordinator(request.user, contrato.coordenador),
        'can_approve_bm_as_gerente': request.user.grupo == 'gerente_contrato',
        'bms': bms,
        'tem_reprovacao_coordenador': tem_reprovacao_coordenador,
        'tem_reprovacao_gerente': tem_reprovacao_gerente,
        'tem_reprovacao_diretor': tem_reprovacao_diretor,
        'notas_fiscais': notas_fiscais,
    })

@login_required
def avaliar_evento_fornecedor(request, evento_id):
    evento = get_object_or_404(Evento, id=evento_id)
    contrato = evento.contrato_terceiro
    fornecedor = contrato.empresa_terceira

    if not can_user_manage_event_delivery(request.user, contrato):
        messages.error(request, "Você não tem permissão para isso!")
        return redirect("home")

    avaliacao_existente = AvaliacaoFornecedor.objects.filter(evento=evento).first()

    if request.method == "POST":
        if avaliacao_existente:
            return JsonResponse({"success": False, "error": "Evento já avaliado."})

        nota_gestao = request.POST.get("nota_gestao")
        nota_tecnica = request.POST.get("nota_tecnica")
        nota_entrega = request.POST.get("nota_entrega")
        comentario = request.POST.get("comentario")

        AvaliacaoFornecedor.objects.create(
            empresa_terceira=fornecedor,
            contrato_terceiro=contrato,
            evento=evento,
            area_avaliadora=request.user.grupo,
            avaliador=request.user,
            nota_gestao=nota_gestao,
            nota_tecnica=nota_tecnica,
            nota_entrega=nota_entrega,
            comentario=comentario,
        )

        return redirect('contrato_fornecedor_detalhe', pk=contrato.pk)

    return render(request, "eventos/avaliar_evento.html", {
        "evento": evento,
        "contrato": contrato,
        "fornecedor": fornecedor,
        "avaliacao": avaliacao_existente,
    })


@login_required
def detalhes_os(request, pk):
    if request.user.grupo not in ['suprimento', 'lider_contrato', 'coordenador', 'gerente', 'gerente_lider', 'gerente_contrato']:
        messages.error(request, "Você não tem permissão para isso.")
        return redirect('home')

    os = get_object_or_404(OS, pk=pk)

    if request.user.grupo == 'lider_contrato' and os.lider_contrato != request.user:
        messages.error(request, "Você não tem permissão para visualizar esta OS.")
        return redirect('lista_solicitacoes')

    if request.user.grupo == 'gerente_lider':
        if not user_shares_center_with_coordinator(request.user, os.coordenador):
            messages.error(request, "Você não tem permissão para visualizar esta OS.")
            return redirect('lista_solicitacoes')

    if request.user.grupo == 'coordenador' and request.user != os.coordenador:
        messages.error(request, "Você não tem permissão para visualizar esta OS.")
        return redirect('lista_solicitacoes')

    fornecedor = os.contrato.empresa_terceira
    indicadores, _ = Indicadores.objects.get_or_create(
        empresa_terceira=fornecedor
    )

    context = {
        'os': os,
        'indicadores': indicadores,
    }

    return render(request, "gestao_contratos/os.html", context)


@login_required
def registrar_entrega_os(request, pk):
    os = get_object_or_404(OS, pk=pk)

    if not can_user_manage_os_delivery(request.user, os):
        messages.error(request, "Você não tem permissão para isso.")
        return redirect("detalhes_os", pk=os.id)

    if os.status == "cancelada":
        messages.error(request, "Não é possível registrar entrega de uma OS cancelada.")
        return redirect("detalhes_os", pk=os.id)

    if request.method == "POST":
        form = RegistroEntregaOSForm(request.POST, instance=os)
        if form.is_valid():
            entrega = form.save(commit=False)

            if entrega.data_entrega and os.prazo_execucao:
                entrega.com_atraso = entrega.data_entrega > os.prazo_execucao

            entrega.realizado = True
            entrega.status = "finalizada"
            entrega.save()

            messages.success(request, "Entrega da OS registrada com sucesso.")
            return redirect("detalhes_os", pk=os.id)
    else:
        form = RegistroEntregaOSForm(instance=os)

    return render(
        request,
        "gestao_contratos/registrar_entrega_os.html",
        {
            "os": os,
            "form": form,
        }
    )
